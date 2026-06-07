from pathlib import Path

from openpyxl import Workbook, load_workbook

import daily_ops_app as app


ROOT = Path(__file__).resolve().parent
SHEIN_DIR = ROOT / "shein数据源表"
OUTPUT = ROOT / "outputs" / f"{app.today_code()}-Shein申报价异常-V1.xlsx"
SHEIN_FILES = None

STORE_ORDER = ["琪琪", "童话", "加加", "宝宝", "牛牛"]
DEFAULT_OWNERS = {"琪琪": "胡娟", "童话": "胡娟", "加加": "洁琳", "宝宝": "洁琳", "牛牛": "胡娟"}
SIZE_ORDER = {"XXS": 0, "XS": 1, "S": 2, "M": 3, "MD": 4, "L": 5, "XL": 6, "XXL": 7}


def store_from_file(path):
    for store in STORE_ORDER:
        if store in path.stem:
            return store
    return ""


def latest_store_file(store):
    candidates = sorted(
        [path for path in SHEIN_DIR.glob("*.xlsx") if store in path.stem],
        key=lambda path: (path.stat().st_mtime, path.name),
    )
    return candidates[-1] if candidates else None


def source_files():
    if SHEIN_FILES:
        return [Path(path) for path in SHEIN_FILES]
    files = [latest_store_file(store) for store in STORE_ORDER]
    return [path for path in files if path and path.exists()]


def sku_size_rank(code):
    text = app.norm(code).upper().split("@", 1)[0]
    size = text.rsplit("-", 1)[-1] if "-" in text else ""
    return SIZE_ORDER.get(size, 99)


def sort_rows(rows):
    store_rank = {store: idx for idx, store in enumerate(STORE_ORDER)}
    return sorted(
        rows,
        key=lambda row: (
            store_rank.get(row["店铺"], 99),
            app.norm(row["SKC"]),
            app.norm(row["商家编码"]).upper().rsplit("-", 1)[0],
            sku_size_rank(row["商家编码"]),
            app.norm(row["商家编码"]).upper(),
        ),
    )


def main():
    OUTPUT.parent.mkdir(exist_ok=True)
    files = source_files()
    if not files:
        raise FileNotFoundError("未找到 Shein 数据源")

    erp = app.load_erp_price_map()
    owners = {**DEFAULT_OWNERS, **app.load_owners()}
    summary = {
        store: {
            "skc": set(),
            "sales7": 0,
            "sales30": 0,
            "active30_skc": set(),
            "hot_skc": set(),
            "cost_skc": set(),
            "wholesale_skc": set(),
        }
        for store in STORE_ORDER
    }
    below_cost = []
    below_wholesale = []
    source_rows = 0
    skipped_zero = 0
    skipped_erp = 0

    for path in files:
        store = store_from_file(path)
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        ws.reset_dimensions()
        iterator = ws.iter_rows(values_only=True)
        headers = app.header_map(next(iterator, []))
        price_header = "供货价" if "供货价" in headers else "采购价"
        for row_number, row in enumerate(iterator, start=2):
            source_rows += 1
            sku_raw = app.norm(app.cell(row, headers, "商家SKU"))
            sku = app.sku_key(sku_raw)
            if not sku:
                continue
            price = app.to_number(app.cell(row, headers, price_header))
            if price == 0:
                skipped_zero += 1
                continue
            info = erp.get(sku)
            if not info:
                skipped_erp += 1
                continue

            skc = app.norm(app.cell(row, headers, "SKC"))
            sales7 = app.to_number(app.cell(row, headers, "近7天销量"))
            sales30 = app.to_number(app.cell(row, headers, "近30天销量"))
            tag = app.norm(app.cell(row, headers, "商品标签"))
            item = summary[store]
            if skc:
                item["skc"].add(skc)
                if sales30 > 0:
                    item["active30_skc"].add(skc)
                if any(word in tag for word in ["高销款", "爆", "旺"]):
                    item["hot_skc"].add(skc)
            item["sales7"] += sales7
            item["sales30"] += sales30

            base = {
                "店铺": store,
                "SKC": skc,
                "商家编码": sku,
                "货品名称": info["货品名称"],
                "货品规格": info["规格名称"],
                "申报价": price,
                "成本价": info["成本价"],
                "批发价": info["批发价"],
                "批发价80%": round(info["批发价"] * 0.8, 2),
                "负责人": owners.get(store, DEFAULT_OWNERS.get(store, "")),
                "7天销量": sales7,
                "30天销量": sales30,
                "源SKU": sku_raw,
                "源文件": path.name,
                "源行": row_number,
            }
            if info["成本价"] and price < info["成本价"]:
                below_cost.append(base)
                item["cost_skc"].add(skc)
            if info["批发价"] and price < info["批发价"] * 0.8:
                below_wholesale.append(base)
                item["wholesale_skc"].add(skc)
        wb.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "总览表"
    ws.append(["店铺", "负责人", "在售SKC数量", "7天销量", "30天销量", "30天内有销量的SKC数", "爆旺款数量", "亏损SKC数量", "破价SKC数量"])
    for store in STORE_ORDER:
        item = summary[store]
        ws.append([
            store,
            owners.get(store, DEFAULT_OWNERS.get(store, "")),
            len(item["skc"]),
            item["sales7"],
            item["sales30"],
            len(item["active30_skc"]),
            len(item["hot_skc"]),
            len(item["cost_skc"]),
            len(item["wholesale_skc"]),
        ])

    app.write_rows(wb, "低于成本价", sort_rows(below_cost), ["店铺", "SKC", "商家编码", "货品名称", "货品规格", "申报价", "成本价", "负责人", "7天销量", "30天销量", "源SKU", "源文件", "源行"])
    app.write_rows(wb, "低于批发价80%", sort_rows(below_wholesale), ["店铺", "SKC", "商家编码", "货品名称", "货品规格", "申报价", "成本价", "批发价", "批发价80%", "负责人", "7天销量", "30天销量", "源SKU", "源文件", "源行"])

    check = wb.create_sheet("数据校验")
    for row in [
        ["检查项", "结果"],
        ["Shein源文件", "；".join(path.name for path in files)],
        ["源明细行数", source_rows],
        ["申报价为0跳过", skipped_zero],
        ["未匹配ERP跳过", skipped_erp],
        ["低于成本价明细", len(below_cost)],
        ["低于批发价80%明细", len(below_wholesale)],
    ]:
        check.append(row)

    for sheet in wb.worksheets:
        app.style_basic_sheet(sheet)
    wb.save(OUTPUT)
    print(OUTPUT)
    print(f"LOW_COST={len(below_cost)}")
    print(f"LOW_WHOLESALE_80={len(below_wholesale)}")
    print(f"SKIPPED_ERP={skipped_erp}")


if __name__ == "__main__":
    main()
