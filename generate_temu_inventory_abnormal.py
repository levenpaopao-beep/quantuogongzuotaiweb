import os
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import update_shein_summary_30d_skc as xlsx


ROOT = Path(os.environ.get("DAILYWORK_ROOT", Path(__file__).resolve().parent))
TEMU_DIR = ROOT / "temu数据源表"
ERP_DIR = ROOT / "erp数据源"
OWNER_FILE = ROOT / "店铺负责人对应表.xlsx"
OUTPUT = ROOT / "TEMU仓备库存异常清单_20260603.xlsx"
ERP_FILES = None
COMBO_FILE = None

STORE_ORDER = [
    ("1", "一弟"),
    ("2", "二弟"),
    ("3", "三弟"),
    ("4", "四弟"),
    ("5", "五弟"),
    ("6", "六弟"),
    ("7", "七弟"),
    ("8", "八弟"),
    ("9", "九弟"),
    ("10", "十弟"),
    ("11", "十一"),
    ("12", "十二弟"),
    ("13", "十三弟"),
    ("15", "十五弟"),
]
STORE_ALIASES = {"弟弟": "一弟", "十一弟": "十一", "十二": "十二弟", "十三": "十三弟", "十五": "十五弟"}
STORE_CODE = {name: code for code, name in STORE_ORDER}
STORE_RANK = {name: i for i, (_, name) in enumerate(STORE_ORDER)}
SIZE_ORDER = {"XXS": 0, "XS": 1, "S": 2, "M": 3, "MD": 4, "L": 5, "XL": 6, "XXL": 7}


def norm(value):
    return xlsx.norm_text(value)


def num(value):
    return xlsx.to_number(value)


def sku_key(value):
    text = norm(value).upper()
    if "@" in text:
        text = text.split("@", 1)[0]
    return text.strip()


def canonical_store(value):
    text = norm(value)
    return STORE_ALIASES.get(text, text)


def size_rank(code):
    text = norm(code).upper()
    parts = re.split(r"[-_\s]+", text)
    for part in reversed(parts):
        if part in SIZE_ORDER:
            return SIZE_ORDER[part]
    return 99


def header_map(row):
    return {norm(v): i for i, v in enumerate(row) if norm(v)}


def latest_files(pattern):
    files = sorted(TEMU_DIR.glob(pattern), key=lambda p: p.name)
    date_re = re.compile(r"(20\d{6})")
    by_date = defaultdict(list)
    for path in files:
        m = date_re.search(path.name)
        by_date[m.group(1) if m else "00000000"].append(path)
    if not by_date:
        return []
    latest = max(by_date)
    return sorted(by_date[latest])


def load_owner_map():
    owners = {}
    if not OWNER_FILE.exists():
        return owners
    from openpyxl import load_workbook

    wb = load_workbook(OWNER_FILE, data_only=True, read_only=True)
    ws = wb.active
    headers = [norm(cell.value) for cell in ws[1]]
    if "店铺" not in headers or "业务" not in headers:
        return owners
    store_i = headers.index("店铺") + 1
    owner_i = headers.index("业务") + 1
    for row in range(2, ws.max_row + 1):
        store = canonical_store(ws.cell(row, store_i).value)
        owner = norm(ws.cell(row, owner_i).value)
        if store and owner:
            owners[store] = owner
    return owners


def load_erp_records():
    records = {}
    erp_rows = 0
    for path in (ERP_FILES or sorted(ERP_DIR.glob("erp产品基础信息表*.xlsx"))):
        rows = xlsx.read_xlsx_rows(path)
        if not rows:
            continue
        hm = header_map(rows[0])
        if not {"商家编码（新）", "货品名称", "规格名称"}.issubset(hm):
            continue
        for row in rows[1:]:
            erp_rows += 1
            code = norm(row[hm["商家编码（新）"]] if hm["商家编码（新）"] < len(row) else "")
            key = sku_key(code)
            if key and key not in records:
                records[key] = {
                    "商家编码": code,
                    "货品名称": norm(row[hm["货品名称"]] if hm["货品名称"] < len(row) else ""),
                    "货品规格": norm(row[hm["规格名称"]] if hm["规格名称"] < len(row) else ""),
                }

    combo_rows = 0
    combo = COMBO_FILE or (ERP_DIR / "erp产品组合装基础信息表.xlsx")
    if combo.exists():
        rows = xlsx.read_xlsx_rows(combo)
        hm = header_map(rows[0]) if rows else {}
        code_i = hm.get("商家编码（新）")
        name_i = hm.get("组合装名称")
        spec_i = hm.get("组合装简称")
        if code_i is not None:
            for row in rows[1:]:
                combo_rows += 1
                code = norm(row[code_i] if code_i < len(row) else "")
                key = sku_key(code)
                if key and key not in records:
                    records[key] = {
                        "商家编码": code,
                        "货品名称": norm(row[name_i] if name_i is not None and name_i < len(row) else ""),
                        "货品规格": norm(row[spec_i] if spec_i is not None and spec_i < len(row) else ""),
                    }
    return records, erp_rows, combo_rows


def load_hot_skc():
    hot = defaultdict(set)
    files = latest_files("*Temu爆旺款*.xlsx")
    for path in files:
        rows = xlsx.read_xlsx_rows(path)
        if not rows:
            continue
        hm = header_map(rows[0])
        skc_i = hm.get("SKC")
        store_i = hm.get("店铺")
        if skc_i is None or store_i is None:
            continue
        for row in rows[1:]:
            store = canonical_store(row[store_i] if store_i < len(row) else "")
            skc = norm(row[skc_i] if skc_i < len(row) else "")
            if store in STORE_CODE and skc:
                hot[store].add(skc)
    return hot, [p.name for p in files]


def read_temu_rows(erp, owners):
    groups = defaultdict(list)
    summary = {name: {"skc": set()} for _, name in STORE_ORDER}
    source_rows = 0
    skipped_unmatched = 0
    files = latest_files("*Temu仓库销售情况导出*.xlsx")
    for path in files:
        rows = xlsx.read_xlsx_rows(path)
        if not rows:
            continue
        hm = header_map(rows[0])
        required = ["SKC", "SKU货号", "7天销量", "30天销量", "可用", "店铺"]
        if not all(k in hm for k in required):
            raise RuntimeError(f"{path.name} 缺少必要字段")
        for ridx, row in enumerate(rows[1:], start=2):
            source_rows += 1
            store = canonical_store(row[hm["店铺"]] if hm["店铺"] < len(row) else "")
            if store not in STORE_CODE:
                continue
            skc = norm(row[hm["SKC"]] if hm["SKC"] < len(row) else "")
            sku = norm(row[hm["SKU货号"]] if hm["SKU货号"] < len(row) else "")
            stock = num(row[hm["可用"]] if hm["可用"] < len(row) else "")
            sales30 = num(row[hm["30天销量"]] if hm["30天销量"] < len(row) else "")
            sales7 = num(row[hm["7天销量"]] if hm["7天销量"] < len(row) else "")
            if skc:
                summary[store]["skc"].add(skc)
            erp_row = erp.get(sku_key(sku))
            if not erp_row:
                skipped_unmatched += 1
                continue
            groups[(store, skc)].append(
                {
                    "店铺编码": STORE_CODE[store],
                    "店铺": store,
                    "SKC": skc,
                    "商家编码": erp_row["商家编码"],
                    "货品名称": erp_row["货品名称"],
                    "货品规格": erp_row["货品规格"],
                    "仓备可用": stock,
                    "30天销量": sales30,
                    "7天销量": sales7,
                    "负责人": owners.get(store, ""),
                    "源SKU": sku,
                    "源文件": path.name,
                    "源行": ridx,
                }
            )
    return groups, summary, source_rows, skipped_unmatched, [p.name for p in files]


def expand_group_alerts(groups, multiplier):
    rows = []
    alert_skc = defaultdict(set)
    alert_groups = set()
    for group_key, items in groups.items():
        link_stock_total = sum(item["仓备可用"] for item in items)
        if link_stock_total <= 30:
            continue
        if not any(item["仓备可用"] > item["30天销量"] * multiplier for item in items):
            continue
        alert_groups.add(group_key)
        store = group_key[0]
        for item in items:
            out = dict(item)
            out["链接备货总数量"] = link_stock_total
            out["触发规则"] = f"组内任一尺码仓备可用 > 30天销量 * {multiplier}，且SKC链接备货总数量 > 30"
            rows.append(out)
            if out["SKC"]:
                alert_skc[store].add(out["SKC"])
    return rows, alert_skc, alert_groups


def sort_rows(rows):
    return sorted(
        rows,
        key=lambda r: (
            STORE_RANK.get(r["店铺"], 99),
            norm(r["SKC"]),
            re.sub(r"[-_\s]+(XXS|XS|S|M|MD|L|XL|XXL)$", "", norm(r["商家编码"]).upper()),
            size_rank(r["商家编码"]),
            norm(r["商家编码"]).upper(),
        ),
    )


def setup_sheet(ws, title, headers, color):
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="44546A")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"


def style_sheet(ws):
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")
    for col in range(1, ws.max_column + 1):
        max_len = 8
        for row in range(1, min(ws.max_row, 300) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 36)
    ws.sheet_view.showGridLines = False


def write_detail(wb, sheet_name, title, rows, color):
    headers = [
        "店铺编码", "店铺", "SKC", "商家编码", "货品名称", "货品规格", "仓备可用",
        "链接备货总数量", "30天销量", "7天销量", "负责人", "源SKU", "源文件", "源行", "触发规则",
    ]
    ws = wb.create_sheet(sheet_name)
    setup_sheet(ws, title, headers, color)
    for row in sort_rows(rows):
        ws.append([row[h] for h in headers])
    for row in ws.iter_rows(min_row=3, min_col=7, max_col=10):
        for cell in row:
            cell.number_format = "0"
    style_sheet(ws)


def build_workbook(summary, hot, owners, rows_2x, skc_2x, rows_1x, skc_1x, source_rows, erp_rows, combo_rows, skipped, groups_2x, groups_1x, source_files, hot_files):
    wb = Workbook()
    ws = wb.active
    ws.title = "总览表"
    headers = ["店铺编码", "店铺", "负责人", "在售SKC数量", "爆旺款数量", "仓备大于30天销量2倍SKC数", "仓备大于30天销量SKC数"]
    setup_sheet(ws, "TEMU仓备库存异常清单", headers, "5B9BD5")
    for code, store in STORE_ORDER:
        ws.append([
            code,
            store,
            owners.get(store, ""),
            len(summary[store]["skc"]),
            len(hot.get(store, set())),
            len(skc_2x.get(store, set())),
            len(skc_1x.get(store, set())),
        ])
    style_sheet(ws)

    write_detail(wb, "大于30天销量2倍", "组内任一尺码Temu可用库存 > 30天销量 * 2，列出整个SKC", rows_2x, "C00000")
    write_detail(wb, "大于30天销量", "组内任一尺码Temu可用库存 > 30天销量，列出整个SKC", rows_1x, "ED7D31")

    check = wb.create_sheet("数据校验")
    for row in [
        ["项目", "数值"],
        ["Temu销售源文件", "；".join(source_files)],
        ["Temu爆旺款源文件", "；".join(hot_files)],
        ["Temu源明细行数", source_rows],
        ["ERP基础表读取行数", erp_rows],
        ["ERP组合装读取行数", combo_rows],
        ["未匹配ERP/组合装跳过行数", skipped],
        ["大于30天销量2倍明细行数（SKC展开后）", len(rows_2x)],
        ["大于30天销量明细行数（SKC展开后）", len(rows_1x)],
        ["大于30天销量2倍触发SKC组数", len(groups_2x)],
        ["大于30天销量触发SKC组数", len(groups_1x)],
        ["库存口径", "Temu源表字段：可用"],
        ["触发门槛", "同一店铺同一SKC的链接备货总数量 > 30"],
        ["展开口径", "组内任一尺码命中预警，则列出该SKC内所有可匹配ERP的码数"],
        ["排序规则", "店铺编码、SKC、商家编码尺码（XXS-XS-S-M-MD-L-XL-XXL）"],
    ]:
        check.append(row)
    style_sheet(check)
    wb.save(OUTPUT)


if __name__ == "__main__":
    owners = load_owner_map()
    erp, erp_rows, combo_rows = load_erp_records()
    hot, hot_files = load_hot_skc()
    groups, summary, source_rows, skipped, source_files = read_temu_rows(erp, owners)
    rows_2x, skc_2x, groups_2x = expand_group_alerts(groups, 2)
    rows_1x, skc_1x, groups_1x = expand_group_alerts(groups, 1)
    build_workbook(
        summary, hot, owners, rows_2x, skc_2x, rows_1x, skc_1x,
        source_rows, erp_rows, combo_rows, skipped, groups_2x, groups_1x, source_files, hot_files
    )
    print(OUTPUT)
    print("source", source_files)
    print("hot", hot_files)
    print("2x rows", len(rows_2x), "groups", len(groups_2x))
    print("1x rows", len(rows_1x), "groups", len(groups_1x))
    for code, store in STORE_ORDER:
        print(code, store, len(summary[store]["skc"]), len(hot.get(store, set())), len(skc_2x.get(store, set())), len(skc_1x.get(store, set())))
