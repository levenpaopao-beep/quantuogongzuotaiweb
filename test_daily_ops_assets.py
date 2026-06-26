import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import daily_ops_assets
import daily_ops_app


class AssetSnapshotStoreTest(unittest.TestCase):
    def test_upsert_summary_export_and_import_archive(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            db_path = root / "important_assets.sqlite"
            store = daily_ops_assets.AssetSnapshotStore(db_path)

            store.upsert_metric("2026-05-25", "temu_hot_skc", 8, source_file="may.xlsx")
            store.upsert_metric("2025-06-25", "temu_hot_skc", 5, source_file="last-year.xlsx")
            store.upsert_metric("2026-06-25", "temu_hot_skc", 12, source_file="today.xlsx")
            store.upsert_metric("2026-06-25", "shein_hot_skc", 6, source_file="shein.xlsx")
            store.upsert_metric("2026-05-25", "temu_hot_skc", 3, source_file="may-owner.xlsx", owner="胡娟")
            store.upsert_metric("2026-06-25", "temu_hot_skc", 4, source_file="today-owner.xlsx", owner="胡娟")

            summary = store.overview(["temu_hot_skc", "shein_hot_skc"], anchor_date="2026-06-25")
            self.assertEqual(summary["temu_hot_skc"]["value"], 12)
            self.assertEqual(summary["temu_hot_skc"]["previous_month"]["value"], 8)
            self.assertEqual(summary["temu_hot_skc"]["previous_month"]["delta"], 4)
            self.assertEqual(summary["temu_hot_skc"]["previous_year"]["value"], 5)
            self.assertEqual(summary["temu_hot_skc"]["previous_year"]["delta"], 7)
            self.assertEqual(summary["shein_hot_skc"]["value"], 6)
            self.assertIsNone(summary["shein_hot_skc"]["previous_month"]["value"])

            owner_summary = store.overview(["temu_hot_skc", "shein_hot_skc"], anchor_date="2026-06-25", owner="胡娟")
            self.assertEqual(owner_summary["temu_hot_skc"]["value"], 4)
            self.assertEqual(owner_summary["temu_hot_skc"]["previous_month"]["value"], 3)
            self.assertIsNone(owner_summary["shein_hot_skc"]["value"])

            archive = root / "资产快照导出.xlsx"
            result = store.export_archive(archive)
            self.assertEqual(result["rows"], 6)
            wb = load_workbook(archive, read_only=True, data_only=True)
            try:
                headers = [cell.value for cell in next(wb.active.iter_rows(max_row=1))]
                self.assertEqual(headers[:7], ["日期", "指标", "数量", "来源文件", "更新时间", "指标名称", "范围类型"])
            finally:
                wb.close()

            imported = daily_ops_assets.AssetSnapshotStore(root / "imported.sqlite")
            imported_result = imported.import_archive(archive)
            self.assertEqual(imported_result["rows"], 6)
            imported_summary = imported.overview(["temu_hot_skc"], anchor_date="2026-06-25")
            self.assertEqual(imported_summary["temu_hot_skc"]["previous_year"]["value"], 5)
            imported_owner_summary = imported.overview(["temu_hot_skc"], anchor_date="2026-06-25", owner="胡娟")
            self.assertEqual(imported_owner_summary["temu_hot_skc"]["value"], 4)

    def test_initialize_from_simple_workbook(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "初始化资产.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["日期", "指标", "数量", "来源文件"])
            ws.append(["2026-06-25", "temu_hot_skc", 11, "temu.xlsx"])
            ws.append(["2026-06-25", "shein_hot_skc", 7, "shein.xlsx"])
            wb.save(source)

            store = daily_ops_assets.AssetSnapshotStore(root / "assets.sqlite")
            result = store.import_archive(source)
            self.assertEqual(result["rows"], 2)
            summary = store.overview(["temu_hot_skc", "shein_hot_skc"], anchor_date="2026-06-25")
            self.assertEqual(summary["temu_hot_skc"]["value"], 11)
            self.assertEqual(summary["shein_hot_skc"]["value"], 7)

    def test_app_asset_overview_export_and_import_payloads(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            db_path = root / "important_assets.sqlite"
            archive = root / "archive.xlsx"
            old_db = daily_ops_app.IMPORTANT_ASSETS_DB
            old_output = daily_ops_app.OUTPUT_DIR
            try:
                daily_ops_app.IMPORTANT_ASSETS_DB = db_path
                daily_ops_app.OUTPUT_DIR = root
                daily_ops_app.record_asset_metric("2026-06-25", "temu_hot_skc", 9, "temu.xlsx")
                daily_ops_app.record_asset_metric("2026-06-25", "temu_hot_skc", 4, "temu-owner.xlsx", owner="胡娟")
                with patch.object(daily_ops_app, "count_temu_hot_skc", return_value={"value": None, "source_file": "", "source": "missing"}), \
                     patch.object(daily_ops_app, "count_shein_hot_skc", return_value={"value": None, "source_file": "", "source": "missing"}):
                    overview = daily_ops_app.asset_overview("2026-06-25")
                    owner_overview = daily_ops_app.asset_overview("2026-06-25", role="owner", user="胡娟")
                self.assertEqual(overview["metrics"]["temu_hot_skc"]["value"], 9)
                self.assertEqual(owner_overview["metrics"]["temu_hot_skc"]["value"], 4)
                self.assertEqual(overview["metrics"]["shein_hot_skc"]["value"], None)

                exported = daily_ops_app.export_asset_archive(archive)
                self.assertEqual(exported["rows"], 2)

                daily_ops_app.IMPORTANT_ASSETS_DB = root / "imported.sqlite"
                imported = daily_ops_app.import_asset_archive(archive)
                self.assertEqual(imported["rows"], 2)
                with patch.object(daily_ops_app, "count_temu_hot_skc", return_value={"value": None, "source_file": "", "source": "missing"}), \
                     patch.object(daily_ops_app, "count_shein_hot_skc", return_value={"value": None, "source_file": "", "source": "missing"}):
                    self.assertEqual(daily_ops_app.asset_overview("2026-06-25")["metrics"]["temu_hot_skc"]["value"], 9)
                    self.assertEqual(daily_ops_app.asset_overview("2026-06-25", role="owner", user="胡娟")["metrics"]["temu_hot_skc"]["value"], 4)
            finally:
                daily_ops_app.IMPORTANT_ASSETS_DB = old_db
                daily_ops_app.OUTPUT_DIR = old_output

    def test_asset_overview_persists_latest_file_fallback_as_daily_snapshot(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            db_path = root / "important_assets.sqlite"
            old_db = daily_ops_app.IMPORTANT_ASSETS_DB
            try:
                daily_ops_app.IMPORTANT_ASSETS_DB = db_path
                with patch.object(daily_ops_app, "count_temu_hot_skc", return_value={"value": 91, "source_file": "temu.xlsx", "source": "latest_file"}), \
                     patch.object(daily_ops_app, "count_shein_hot_skc", return_value={"value": 6, "source_file": "shein.xlsx", "source": "latest_file"}):
                    overview = daily_ops_app.asset_overview("2026-06-26")

                self.assertEqual(overview["metrics"]["temu_hot_skc"]["value"], 91)
                store = daily_ops_assets.AssetSnapshotStore(db_path)
                self.assertEqual(store.value_for("temu_hot_skc", "2026-06-26")["value"], 91)
                self.assertEqual(store.value_for("shein_hot_skc", "2026-06-26")["value"], 6)
            finally:
                daily_ops_app.IMPORTANT_ASSETS_DB = old_db


if __name__ == "__main__":
    unittest.main()
