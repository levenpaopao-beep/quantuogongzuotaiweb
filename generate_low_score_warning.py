import html
import json
import re
import shutil
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

import update_shein_summary_30d_skc as raw_xlsx


ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "outputs"
ARCHIVE_DIR = ROOT / "低分预警历史归档"
HISTORY_STATE_FILE = ROOT / "基础数据库" / "low_score_history_state.json"
OWNER_FILE = ROOT / "店铺负责人对应表.xlsx"

CURRENT_FILES = None
HISTORY_FILES = None
TEMU_SALES_FILES = None
TEMU_HOT_FILES = None
ERP_FILES = None
OUTPUT = None

LOW_SCORE_THRESHOLD = 60

STORE_CODE_MAP = {
    "弟弟": "1",
    "一弟": "1",
    "二弟": "2",
    "三弟": "3",
    "四弟": "4",
    "五弟": "5",
    "六弟": "6",
    "七弟": "7",
    "八弟": "8",
    "九弟": "9",
    "十弟": "10",
    "十一弟": "11",
    "十一": "11",
    "十二弟": "12",
    "十二": "12",
    "十三弟": "13",
    "十三": "13",
    "十五弟": "15",
    "十五": "15",
}
STORE_ALIASES = {
    "一弟": "弟弟",
    "十一": "十一弟",
    "十二": "十二弟",
    "十三": "十三弟",
    "十五": "十五弟",
    "九弟（喵喵）": "九弟",
    "十三（节日）": "十三弟",
    "十五弟（毛衣）": "十五弟",
}

OUTPUT_HEADERS = [
    "SPU",
    "SKC",
    "货品名称",
    "店铺品质分情况",
    "所属店铺",
    "产品负责人",
    "是否下架",
    "平台仓库库存",
    "是否爆旺款",
    "是否本周新增低分",
    "30天销量",
    "填表人",
    "填表时间",
]


def norm(value):
    return html.unescape(str(value or "")).strip()


def to_number(value):
    text = norm(value).replace(",", "").replace("，", "")
    if text.endswith("%"):
        text = text[:-1]
    text = re.sub(r"[^\d.\-]", "", text)
    if text in {"", "-", ".", "-."}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def header_map(row):
    return {norm(value): idx for idx, value in enumerate(row) if norm(value)}


def cell(row, headers, *names):
    for name in names:
        idx = headers.get(name)
        if idx is not None and idx < len(row):
            return row[idx]
    return ""


def canonical_store(value):
    text = norm(value)
    return STORE_ALIASES.get(text, text)


def store_display(value):
    text = canonical_store(value)
    return STORE_CODE_MAP.get(text, text)


def style_basic_sheet(ws):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if ws.max_row >= 1:
        for item in ws[1]:
            item.fill = header_fill
            item.font = header_font
            item.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for item in row:
            item.border = border
            item.alignment = Alignment(vertical="center", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        width = 10
        for row in range(1, min(ws.max_row, 300) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                width = max(width, len(str(value)) + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(width, 40)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False


def read_rows(path):
    return raw_xlsx.read_xlsx_rows(path)


def load_owner_map(owner_file=OWNER_FILE):
    owners = {}
    if not owner_file.exists():
        return owners
    wb = load_workbook(owner_file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return owners
    headers = header_map(rows[0])
    for row in rows[1:]:
        store = canonical_store(cell(row, headers, "店铺", "店铺名称", "店铺名"))
        owner = norm(cell(row, headers, "业务", "负责人"))
        if store and owner:
            owners[store] = owner
    return owners


def normalize_current_row(row, headers):
    spu = norm(cell(row, headers, "SPU", "spu"))
    if not spu:
        return None
    score = to_number(cell(row, headers, "店铺品质分情况", "品质分情况"))
    if score >= LOW_SCORE_THRESHOLD:
        return None
    return {
        "SPU": spu,
        "店铺品质分情况": norm(cell(row, headers, "店铺品质分情况", "品质分情况")),
        "所属店铺_原始": norm(cell(row, headers, "所属店铺", "店铺")),
        "填表人": norm(cell(row, headers, "填表人")),
        "填表时间": norm(cell(row, headers, "填表时间")),
    }


def load_low_score_rows(paths):
    rows = []
    source_files = []
    for path in paths or []:
        path = Path(path)
        if not path.exists():
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        table = list(ws.iter_rows(values_only=True))
        wb.close()
        if not table:
            continue
        headers = header_map(table[0])
        for row in table[1:]:
            item = normalize_current_row(row, headers)
            if item:
                rows.append(item)
        source_files.append(path.name)
    return rows, source_files


def merge_current_rows(rows):
    merged = OrderedDict()
    for row in rows:
        spu = norm(row.get("SPU"))
        if not spu:
            continue
        if spu not in merged:
            merged[spu] = dict(row)
            continue
        for key, value in row.items():
            if not merged[spu].get(key) and value:
                merged[spu][key] = value
    return list(merged.values())


def classify_spu_status(spu, history_spus, sales_index):
    sales = sales_index.get(spu, {})
    in_sales = bool(sales)
    return {
        "是否本周新增低分": "历史持续低分产品" if spu in history_spus else "本周新增低分产品",
        "是否下架": "在售" if in_sales else "已下架",
    }


def load_history_spus(paths):
    rows, _ = load_low_score_rows(paths)
    return {norm(row.get("SPU")) for row in merge_current_rows(rows) if norm(row.get("SPU"))}


def load_erp_name_map(erp_files):
    names = {}
    for path in erp_files or []:
        table = read_rows(path)
        if not table:
            continue
        headers = header_map(table[0])
        for row in table[1:]:
            spu = norm(cell(row, headers, "货品编码", "SPU"))
            name = norm(cell(row, headers, "货品名称", "组合装名称"))
            if spu and name and spu not in names:
                names[spu] = name
    return names


def load_sales_index(sales_files, owners):
    index = {}
    for path in sales_files or []:
        table = read_rows(path)
        if not table:
            continue
        headers = header_map(table[0])
        required = {"SPU", "SKC", "30天销量", "可用"}
        if not required.issubset(headers):
            continue
        for row in table[1:]:
            spu = norm(cell(row, headers, "SPU"))
            if not spu:
                continue
            store_raw = canonical_store(cell(row, headers, "店铺"))
            store = store_display(store_raw)
            owner = owners.get(store_raw, "")
            skc = norm(cell(row, headers, "SKC"))
            sku_code = norm(cell(row, headers, "SKU货号"))
            product_code = sku_code.split("@", 1)[0].split("-", 1)[0].strip() if sku_code else ""
            sales30 = to_number(cell(row, headers, "30天销量"))
            stock = to_number(cell(row, headers, "可用", "平台仓库库存"))
            item = index.setdefault(
                spu,
                {
                    "skc_set": OrderedDict(),
                    "store_set": OrderedDict(),
                    "owner_set": OrderedDict(),
                    "30天销量": 0.0,
                    "平台仓库库存": 0.0,
                    "在售": True,
                    "ERP货品编码": "",
                },
            )
            if skc:
                item["skc_set"][skc] = True
            if store:
                item["store_set"][store] = True
            if owner:
                item["owner_set"][owner] = True
            if product_code and not item["ERP货品编码"]:
                item["ERP货品编码"] = product_code
            item["30天销量"] += sales30
            item["平台仓库库存"] += stock
    for spu, item in index.items():
        item["SKC"] = "、".join(item["skc_set"].keys())
        item["所属店铺"] = "、".join(item["store_set"].keys())
        item["产品负责人"] = "、".join(item["owner_set"].keys())
    return index


def load_hot_spus(hot_files):
    spus = set()
    for path in hot_files or []:
        table = read_rows(path)
        if not table:
            continue
        headers = header_map(table[0])
        if "SPU" not in headers:
            continue
        for row in table[1:]:
            spu = norm(cell(row, headers, "SPU"))
            if spu:
                spus.add(spu)
    return spus


def build_output_rows(current_rows, history_spus, sales_index, erp_names, hot_spus):
    output_rows = []
    for row in current_rows:
        spu = norm(row.get("SPU"))
        sales = sales_index.get(spu, {})
        status = classify_spu_status(spu, history_spus, sales_index)
        in_sales = bool(sales)
        product_code = norm(sales.get("ERP货品编码", "")) if in_sales else ""
        product_name = erp_names.get(spu) or (erp_names.get(product_code) if product_code else "") or "无ERP匹配信息"
        output_rows.append(
            {
                "SPU": spu,
                "SKC": sales.get("SKC", "") if in_sales else "",
                "货品名称": product_name,
                "店铺品质分情况": row.get("店铺品质分情况", ""),
                "所属店铺": sales.get("所属店铺", row.get("所属店铺_原始", "")) if in_sales else "",
                "产品负责人": sales.get("产品负责人", "") if in_sales else "",
                "是否下架": status["是否下架"],
                "平台仓库库存": int(sales.get("平台仓库库存", 0)) if in_sales else "",
                "是否爆旺款": "是" if in_sales and spu in hot_spus else ("否" if in_sales else ""),
                "是否本周新增低分": status["是否本周新增低分"],
                "30天销量": int(sales.get("30天销量", 0)) if in_sales else "",
                "填表人": row.get("填表人", ""),
                "填表时间": row.get("填表时间", ""),
            }
        )
    output_rows.sort(
        key=lambda item: (
            0 if item["是否本周新增低分"] == "本周新增低分产品" else 1,
            0 if item["是否下架"] == "在售" else 1,
            str(item["所属店铺"]),
            str(item["SPU"]),
        )
    )
    return output_rows


def workbook_stats(rows):
    total = len(rows)
    new_count = sum(1 for row in rows if row["是否本周新增低分"] == "本周新增低分产品")
    history_count = total - new_count
    off_shelf = sum(1 for row in rows if row["是否下架"] == "已下架")
    on_sale = total - off_shelf
    hot_count = sum(1 for row in rows if row["是否爆旺款"] == "是")
    return {
        "total": total,
        "new_count": new_count,
        "history_count": history_count,
        "off_shelf": off_shelf,
        "on_sale": on_sale,
        "hot_count": hot_count,
    }


def write_workbook(rows, output_path, meta):
    wb = Workbook()
    ws = wb.active
    ws.title = "低分预警明细"
    ws.append(OUTPUT_HEADERS)
    for row in rows:
        ws.append([row.get(col, "") for col in OUTPUT_HEADERS])

    stats = workbook_stats(rows)
    stat_ws = wb.create_sheet("统计")
    stat_ws.append(["统计项", "数量"])
    stat_ws.append(["低分产品总数", stats["total"]])
    stat_ws.append(["本周新增低分产品", stats["new_count"]])
    stat_ws.append(["历史持续低分产品", stats["history_count"]])
    stat_ws.append(["在售待处理产品", stats["on_sale"]])
    stat_ws.append(["已下架产品", stats["off_shelf"]])
    stat_ws.append(["低分爆旺款产品", stats["hot_count"]])

    note_ws = wb.create_sheet("说明")
    note_ws.append(["项目", "说明"])
    notes = [
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("低分阈值", f"品质分小于 {LOW_SCORE_THRESHOLD}"),
        ("新增判定", "本周有、上周无=本周新增低分产品；本周有、上周也有=历史持续低分产品"),
        ("下架判定", "Temu销售表查不到SPU=已下架；查得到=在售"),
        ("本周输入文件", "、".join(meta.get("current_files", []))),
        ("历史对比文件", "、".join(meta.get("history_files", [])) or "无历史文件，默认全部视为本周新增"),
        ("Temu销售表", "、".join(meta.get("sales_files", []))),
        ("Temu爆旺款表", "、".join(meta.get("hot_files", []))),
        ("ERP基础信息", "、".join(meta.get("erp_files", []))),
    ]
    for item in notes:
        note_ws.append(list(item))

    for sheet in wb.worksheets:
        style_basic_sheet(sheet)
    wb.save(output_path)


def read_history_state():
    if not HISTORY_STATE_FILE.exists():
        return {"archives": []}
    try:
        return json.loads(HISTORY_STATE_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {"archives": []}


def save_history_state(payload):
    HISTORY_STATE_FILE.parent.mkdir(exist_ok=True)
    HISTORY_STATE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def archive_current_files(current_files):
    ARCHIVE_DIR.mkdir(exist_ok=True)
    state = read_history_state()
    archives = [item for item in state.get("archives", []) if isinstance(item, dict)]
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    archived_paths = []
    for idx, path in enumerate(current_files, start=1):
        src = Path(path)
        if not src.exists():
            continue
        name = f"{stamp}-{idx:02d}-{src.name}"
        target = ARCHIVE_DIR / name
        shutil.copy2(src, target)
        archived_paths.append(str(target))
    if archived_paths:
        archives.append(
            {
                "archived_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "source_paths": [str(Path(path)) for path in current_files],
                "archive_paths": archived_paths,
            }
        )
        state["archives"] = archives[-12:]
        save_history_state(state)
    return archived_paths


def latest_distinct_history_files(current_files):
    state = read_history_state()
    archives = list(state.get("archives", []))
    current_names = sorted(Path(path).name for path in current_files)
    for item in reversed(archives):
        archive_paths = [Path(path) for path in item.get("archive_paths", []) if Path(path).exists()]
        if not archive_paths:
            continue
        source_names = sorted(Path(path).name for path in item.get("source_paths", []))
        if source_names == current_names:
            continue
        return archive_paths
    return []


def generate_report(current_files, history_files, sales_files, hot_files, erp_files, output_path):
    current_rows_raw, current_source_files = load_low_score_rows(current_files)
    current_rows = merge_current_rows(current_rows_raw)
    history_spus = load_history_spus(history_files)
    owners = load_owner_map()
    sales_index = load_sales_index(sales_files, owners)
    hot_spus = load_hot_spus(hot_files)
    erp_names = load_erp_name_map(erp_files)
    rows = build_output_rows(current_rows, history_spus, sales_index, erp_names, hot_spus)
    write_workbook(
        rows,
        output_path,
        {
            "current_files": current_source_files,
            "history_files": [Path(path).name for path in history_files],
            "sales_files": [Path(path).name for path in sales_files],
            "hot_files": [Path(path).name for path in hot_files],
            "erp_files": [Path(path).name for path in erp_files],
        },
    )
    return {
        "rows": len(rows),
        "current_files": current_source_files,
        "history_files": [Path(path).name for path in history_files],
        "new_count": sum(1 for row in rows if row["是否本周新增低分"] == "本周新增低分产品"),
        "off_shelf": sum(1 for row in rows if row["是否下架"] == "已下架"),
    }


def main():
    current_files = [Path(path) for path in (CURRENT_FILES or [])]
    if not current_files:
        raise FileNotFoundError("未提供低分预警输入表")
    history_files = [Path(path) for path in (HISTORY_FILES or latest_distinct_history_files(current_files))]
    sales_files = [Path(path) for path in (TEMU_SALES_FILES or [])]
    hot_files = [Path(path) for path in (TEMU_HOT_FILES or [])]
    erp_files = [Path(path) for path in (ERP_FILES or [])]
    if not sales_files:
        raise FileNotFoundError("未提供Temu销售表")
    if not hot_files:
        raise FileNotFoundError("未提供Temu爆旺款表")
    if not erp_files:
        raise FileNotFoundError("未提供ERP基础信息")
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = OUTPUT or (OUTPUT_DIR / f"{datetime.now():%y%m%d}-店铺低分产品预警表-V1.xlsx")
    result = generate_report(current_files, history_files, sales_files, hot_files, erp_files, output_path)
    archive_current_files(current_files)
    return result


if __name__ == "__main__":
    print(main())
