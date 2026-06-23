import unittest
import os
import json
import io
import zipfile
import threading
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from urllib.parse import urlparse
from unittest.mock import patch

from openpyxl import load_workbook

import daily_ops_tasks
import daily_ops_app
import daily_ops_cli
import daily_ops_desktop_adapter


class OperationTaskStoreTest(unittest.TestCase):
    def test_upsert_filter_submit_review_and_export_tasks(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")

            rows = [
                {
                    "platform": "Temu",
                    "task_type": "爆旺冲突",
                    "store": "7",
                    "owner": "小琴",
                    "merchant_code": "3303197351-L",
                    "skc": "4173751716",
                    "spu": "9685510192",
                    "product_name": "篮球队服-红",
                    "system_action": "停止补货、售完下架",
                    "source_report": "Temu爆旺款重复预警",
                    "source_file": "hot.xlsx",
                    "source_row": 2,
                },
                {
                    "platform": "Temu",
                    "task_type": "低分预警",
                    "store": "5",
                    "owner": "洁琳",
                    "merchant_code": "",
                    "skc": "69443952334",
                    "spu": "6232074702",
                    "product_name": "天蓝带帽雨衣",
                    "system_action": "低分仍在售，需处理",
                    "source_report": "店铺低分产品预警",
                    "source_file": "low.xlsx",
                    "source_row": 27,
                },
            ]

            result = store.upsert_generated_tasks(rows)
            self.assertEqual(result["created"], 2)
            self.assertEqual(result["updated"], 0)
            generated = store.list_tasks(role="owner", user="小琴")[0]
            self.assertEqual(generated["history"][0]["event"], "系统生成")
            self.assertEqual(generated["history"][0]["actor"], "系统")
            self.assertEqual(generated["history"][0]["action"], "生成待处理任务")
            self.assertIn("Temu爆旺款重复预警", generated["history"][0]["remark"])

            repeat = store.upsert_generated_tasks(rows)
            self.assertEqual(repeat["created"], 0)
            self.assertEqual(repeat["updated"], 2)

            xiaoqin_tasks = store.list_tasks(role="owner", user="小琴")
            self.assertEqual(len(xiaoqin_tasks), 1)
            self.assertEqual(xiaoqin_tasks[0]["store"], "7")
            self.assertEqual(xiaoqin_tasks[0]["status"], daily_ops_tasks.STATUS_PENDING_OWNER)
            self.assertEqual(store.list_tasks(role="owner", user=""), [])

            with self.assertRaises(ValueError):
                store.submit_owner_action(
                    xiaoqin_tasks[0]["id"],
                    actor="洁琳",
                    action="已下架",
                    remark="不能处理其他负责人的任务",
                )

            submitted = store.submit_owner_action(
                xiaoqin_tasks[0]["id"],
                actor="小琴",
                action="已下架",
                remark="已在后台处理",
                proof="后台截图：https://example.test/proof-a",
            )
            self.assertEqual(submitted["status"], daily_ops_tasks.STATUS_PENDING_REVIEW)
            self.assertEqual(submitted["owner_action"], "已下架")
            self.assertEqual(submitted["owner_proof"], "后台截图：https://example.test/proof-a")

            reviewed = store.review_task(
                submitted["id"],
                admin="管理员",
                decision="通过",
                remark="同意处理",
            )
            self.assertEqual(reviewed["status"], daily_ops_tasks.STATUS_APPROVED)
            self.assertEqual(reviewed["admin_decision"], "通过")
            self.assertEqual(len(reviewed["history"]), 3)

            export_path = store.export_tasks(root / "导出.xlsx", filters={"role": "owner", "user": "小琴", "status": "已通过"}, now=datetime(2026, 6, 22, 12, 0, 0))
            workbook = load_workbook(export_path, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["任务台账", "操作记录", "负责人汇总", "状态汇总", "管理员待办队列", "导出口径"])
                ws = workbook["任务台账"]
                self.assertEqual(ws.max_row, 3)
                headers = [cell.value for cell in ws[1]]
                self.assertIn("任务状态", headers)
                self.assertIn("店长处理动作", headers)
                self.assertIn("管理员审核结果", headers)
                self.assertIn("店长处理凭证", headers)
                self.assertIn("任务生成批次", headers)
                self.assertIn("是否超时", headers)
                self.assertIn("超时天数", headers)
                proof_col = headers.index("店长处理凭证") + 1
                self.assertEqual(ws.cell(row=2, column=proof_col).value, "后台截图：https://example.test/proof-a")
                log_ws = workbook["操作记录"]
                log_headers = [cell.value for cell in log_ws[1]]
                self.assertIn("事件", log_headers)
                self.assertIn("操作人", log_headers)
                self.assertIn("处理凭证", log_headers)
                self.assertIn("当前状态", log_headers)
                self.assertIn("下一步处理人", log_headers)
                self.assertIn("下一步动作", log_headers)
                events = [log_ws.cell(row=row, column=log_headers.index("事件") + 1).value for row in range(2, log_ws.max_row + 1)]
                self.assertIn("系统生成", events)
                self.assertIn("店长提交", events)
                self.assertIn("管理员审核", events)
                log_rows_by_event = {
                    log_ws.cell(row=row, column=log_headers.index("事件") + 1).value: row
                    for row in range(2, log_ws.max_row + 1)
                }
                review_row = log_rows_by_event["管理员审核"]
                submit_row = log_rows_by_event["店长提交"]
                self.assertEqual(log_ws.cell(row=submit_row, column=log_headers.index("处理凭证") + 1).value, "后台截图：https://example.test/proof-a")
                self.assertEqual(log_ws.cell(row=review_row, column=log_headers.index("当前状态") + 1).value, daily_ops_tasks.STATUS_APPROVED)
                self.assertEqual(log_ws.cell(row=review_row, column=log_headers.index("下一步处理人") + 1).value, "管理员")
                self.assertEqual(log_ws.cell(row=review_row, column=log_headers.index("下一步动作") + 1).value, "标记完成或归档")
                owner_ws = workbook["负责人汇总"]
                owner_headers = [cell.value for cell in owner_ws[1]]
                self.assertEqual(owner_headers, ["负责人", "任务总数", "待店长处理", "待管理员审核", "超时未处理", "返工任务", "已通过", "已驳回", "已完成"])
                owner_rows = {owner_ws.cell(row=row, column=1).value: row for row in range(2, owner_ws.max_row + 1)}
                self.assertEqual(owner_ws.cell(row=owner_rows["小琴"], column=2).value, 1)
                self.assertEqual(owner_ws.cell(row=owner_rows["小琴"], column=7).value, 1)
                summary_ws = workbook["状态汇总"]
                summary = {
                    summary_ws.cell(row=row, column=1).value: summary_ws.cell(row=row, column=2).value
                    for row in range(2, summary_ws.max_row + 1)
                }
                self.assertEqual(summary["任务总数"], 2)
                self.assertIn("超时未处理", summary)
                self.assertEqual(summary["下一步处理人：管理员"], 1)
                self.assertEqual(summary["下一步处理人：店长"], 1)
                self.assertEqual(summary["下一步动作：标记完成或归档"], 1)
                self.assertEqual(summary["下一步动作：填写处理结果"], 1)
                self.assertEqual(summary["来源批次：Temu爆旺款重复预警 / hot.xlsx"], 1)
                self.assertEqual(summary["来源批次：店铺低分产品预警 / low.xlsx"], 1)
                criteria_ws = workbook["导出口径"]
                criteria = {
                    criteria_ws.cell(row=row, column=1).value: criteria_ws.cell(row=row, column=2).value
                    for row in range(2, criteria_ws.max_row + 1)
                }
                self.assertEqual(criteria["role"], "owner")
                self.assertEqual(criteria["user"], "小琴")
                self.assertEqual(criteria["status"], "已通过")
                self.assertEqual(criteria["rows"], 2)
            finally:
                workbook.close()

    def test_owner_submission_requires_remark_or_proof_for_traceability(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {
                    "platform": "Temu",
                    "task_type": "爆旺冲突",
                    "store": "7",
                    "owner": "小琴",
                    "merchant_code": "A-001",
                    "source_report": "r",
                    "source_row": 1,
                }
            ])
            task = store.list_tasks(role="owner", user="小琴")[0]

            with self.assertRaises(ValueError) as ctx:
                store.submit_owner_action(task["id"], actor="小琴", action="已下架", remark="", proof="")

            self.assertIn("处理依据", str(ctx.exception))
            submitted = store.submit_owner_action(task["id"], actor="小琴", action="已下架", remark="", proof="后台截图：https://example.test/proof")
            self.assertEqual(submitted["status"], daily_ops_tasks.STATUS_PENDING_REVIEW)

    def test_generated_tasks_are_deduped_within_same_batch_but_not_across_weekly_files(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            first_week = {
                "platform": "Temu",
                "task_type": "爆旺冲突",
                "store": "7",
                "owner": "小琴",
                "merchant_code": "A-001",
                "skc": "SKC1",
                "product_name": "红色球衣",
                "system_action": "立即下架",
                "source_report": "Temu爆旺款重复预警",
                "source_file": "2026-W25-爆旺.xlsx",
                "source_sheet": "具体店铺操作表",
                "source_row": 2,
            }
            second_week = {**first_week, "source_file": "2026-W26-爆旺.xlsx"}

            self.assertEqual(store.upsert_generated_tasks([first_week])["created"], 1)
            self.assertEqual(store.upsert_generated_tasks([first_week])["updated"], 1)
            self.assertEqual([item["event"] for item in store.list_tasks()[0]["history"]], ["系统生成"])
            second_result = store.upsert_generated_tasks([second_week])
            self.assertEqual(second_result["created"], 1)
            self.assertEqual(second_result["total"], 2)

    def test_report_sync_records_task_generation_batch_for_traceability(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            report = root / "250623-Temu爆旺款重复预警-V1-120000.xlsx"
            workbook = daily_ops_tasks.Workbook()
            ws = workbook.active
            ws.title = "具体店铺操作表"
            ws.append(["商家编码", "货品名称", "skc", "所属店铺", "负责人", "冲突类型", "处理意见"])
            ws.append(["A-001", "红色球衣", "SKC1", "7", "小琴", "平销款冲突爆款", "立即下架"])
            workbook.save(report)

            task_db = root / "operation_tasks.json"
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db):
                sync = daily_ops_app.sync_report_tasks("temu_hot", report)

            self.assertEqual(sync["created"], 1)
            task = daily_ops_tasks.OperationTaskStore(task_db).list_tasks()[0]
            self.assertEqual(task["source_batch_id"], "250623-Temu爆旺款重复预警-V1-120000")
            self.assertIn("任务生成批次", task["history"][0]["remark"])

    def test_task_summary_counts_unassigned_tasks_for_admin_followup(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "7", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_file": "a.xlsx", "source_row": 1},
                {"platform": "Shein", "task_type": "爆旺冲突", "store": "琪琪", "owner": "", "merchant_code": "B", "source_report": "r", "source_file": "b.xlsx", "source_row": 2},
            ])

            summary = store.summary()
            self.assertEqual(summary["total"], 2)
            self.assertEqual(summary["unassigned"], 1)
            self.assertEqual(summary["by_owner"], {"小琴": 1})

    def test_open_only_filter_hides_completed_tasks_but_keeps_active_followups(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "7", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_file": "a.xlsx", "source_row": 1},
                {"platform": "Temu", "task_type": "低分预警", "store": "8", "owner": "小琴", "merchant_code": "B", "source_report": "r", "source_file": "b.xlsx", "source_row": 2},
            ])

            rows = store.list_tasks(role="owner", user="小琴")
            done_task = store.submit_owner_action(rows[0]["id"], actor="小琴", action="已处理", remark="后台已处理")
            done_task = store.review_task(done_task["id"], admin="管理员", decision="通过", remark="同意")
            store.mark_done(done_task["id"], actor="管理员", remark="后台已确认")
            active_task = store.submit_owner_action(rows[1]["id"], actor="小琴", action="继续观察", remark="后台继续观察")
            store.review_task(active_task["id"], admin="管理员", decision="通过", remark="同意")

            open_rows = store.list_tasks(role="owner", user="小琴", open_only="1")
            self.assertEqual([row["status"] for row in open_rows], [daily_ops_tasks.STATUS_APPROVED])
            self.assertEqual(open_rows[0]["merchant_code"], "B")
            self.assertEqual(store.summary(open_rows)["total"], 1)

    def test_concurrent_owner_submissions_do_not_overwrite_each_other(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "7", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
                {"platform": "Temu", "task_type": "低分预警", "store": "8", "owner": "洁琳", "merchant_code": "B", "source_report": "r", "source_row": 2},
            ])
            tasks = {row["owner"]: row["id"] for row in store.list_tasks()}
            original_load = store.load
            barrier = threading.Barrier(2)

            def racing_load():
                payload = original_load()
                if threading.current_thread() is not threading.main_thread():
                    try:
                        barrier.wait(timeout=0.2)
                    except threading.BrokenBarrierError:
                        pass
                return payload

            store.load = racing_load
            errors = []

            def submit(owner, task_id):
                try:
                    store.submit_owner_action(task_id, actor=owner, action="已处理", remark=owner)
                except Exception as exc:
                    errors.append(exc)

            threads = [
                threading.Thread(target=submit, args=("小琴", tasks["小琴"])),
                threading.Thread(target=submit, args=("洁琳", tasks["洁琳"])),
            ]
            for thread in threads:
                thread.start()
            for thread in threads:
                thread.join()

            self.assertEqual(errors, [])
            rows = {row["owner"]: row for row in store.list_tasks()}
            self.assertEqual(rows["小琴"]["status"], daily_ops_tasks.STATUS_PENDING_REVIEW)
            self.assertEqual(rows["洁琳"]["status"], daily_ops_tasks.STATUS_PENDING_REVIEW)

    def test_tasks_expose_next_handler_and_action_for_workflow_routing(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            task_db.write_text(json.dumps({"tasks": [
                {
                    "id": "unassigned",
                    "platform": "Temu",
                    "task_type": "爆旺冲突",
                    "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                    "store": "7",
                    "owner": "",
                    "product_name": "未分配商品",
                    "created_at": "2026-06-22 09:00:00",
                    "updated_at": "2026-06-22 09:00:00",
                    "history": [],
                },
                {
                    "id": "owner-overdue",
                    "platform": "Temu",
                    "task_type": "低分预警",
                    "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                    "store": "8",
                    "owner": "小琴",
                    "product_name": "超时待处理商品",
                    "created_at": "2026-06-18 09:00:00",
                    "updated_at": "2026-06-18 09:00:00",
                    "history": [],
                },
                {
                    "id": "owner-normal",
                    "platform": "Temu",
                    "task_type": "滞销处理",
                    "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                    "store": "9",
                    "owner": "洁琳",
                    "product_name": "待店长商品",
                    "created_at": "2026-06-22 09:00:00",
                    "updated_at": "2026-06-22 09:00:00",
                    "history": [],
                },
                {
                    "id": "review",
                    "platform": "Temu",
                    "task_type": "议价审核",
                    "status": daily_ops_tasks.STATUS_PENDING_REVIEW,
                    "store": "10",
                    "owner": "小琴",
                    "product_name": "待审核商品",
                    "owner_submitted_at": "2026-06-22 09:00:00",
                    "created_at": "2026-06-22 08:00:00",
                    "updated_at": "2026-06-22 09:00:00",
                    "history": [],
                },
                {
                    "id": "approved",
                    "platform": "Temu",
                    "task_type": "爆旺冲突",
                    "status": daily_ops_tasks.STATUS_APPROVED,
                    "store": "11",
                    "owner": "洁琳",
                    "product_name": "已通过商品",
                    "created_at": "2026-06-22 08:00:00",
                    "updated_at": "2026-06-22 10:00:00",
                    "history": [],
                },
            ]}, ensure_ascii=False), encoding="utf-8")

            rows = {
                row["id"]: row
                for row in daily_ops_tasks.OperationTaskStore(task_db).list_tasks(now=datetime(2026, 6, 22, 12, 0, 0))
            }
            ordered_rows = daily_ops_tasks.OperationTaskStore(task_db).list_tasks(now=datetime(2026, 6, 22, 12, 0, 0))
            self.assertEqual([row["id"] for row in ordered_rows], ["unassigned", "owner-overdue", "approved", "review", "owner-normal"])
            self.assertEqual(rows["unassigned"]["next_handler"], "管理员")
            self.assertEqual(rows["unassigned"]["next_action"], "指派负责人")
            self.assertEqual(rows["unassigned"]["priority"], "高")
            self.assertEqual(rows["unassigned"]["priority_reason"], "未分配负责人")
            self.assertEqual(rows["owner-overdue"]["next_handler"], "管理员")
            self.assertEqual(rows["owner-overdue"]["next_action"], "跟进超时店长处理")
            self.assertEqual(rows["owner-overdue"]["priority"], "高")
            self.assertEqual(rows["owner-overdue"]["priority_reason"], "超时未处理")
            self.assertEqual(rows["owner-normal"]["next_handler"], "店长")
            self.assertEqual(rows["owner-normal"]["next_action"], "填写处理结果")
            self.assertEqual(rows["owner-normal"]["priority"], "普通")
            self.assertEqual(rows["review"]["next_handler"], "管理员")
            self.assertEqual(rows["review"]["next_action"], "审核通过或驳回")
            self.assertEqual(rows["review"]["priority"], "中")
            self.assertEqual(rows["review"]["priority_reason"], "待管理员审核")
            self.assertEqual(rows["approved"]["next_handler"], "管理员")
            self.assertEqual(rows["approved"]["next_action"], "标记完成或归档")
            self.assertEqual(rows["approved"]["priority"], "中")
            self.assertEqual(rows["approved"]["priority_reason"], "待完成确认")

            admin_queue = daily_ops_tasks.OperationTaskStore(task_db).list_tasks(
                next_handler="管理员",
                now=datetime(2026, 6, 22, 12, 0, 0),
            )
            self.assertEqual({row["id"] for row in admin_queue}, {"unassigned", "owner-overdue", "review", "approved"})
            owner_queue = daily_ops_tasks.OperationTaskStore(task_db).list_tasks(
                next_handler="店长",
                now=datetime(2026, 6, 22, 12, 0, 0),
            )
            self.assertEqual([row["id"] for row in owner_queue], ["owner-normal"])
            high_priority = daily_ops_tasks.OperationTaskStore(task_db).list_tasks(
                priority="高",
                now=datetime(2026, 6, 22, 12, 0, 0),
            )
            self.assertEqual({row["id"] for row in high_priority}, {"unassigned", "owner-overdue"})

            export_path = daily_ops_tasks.OperationTaskStore(task_db).export_tasks(root / "导出.xlsx", now=datetime(2026, 6, 22, 12, 0, 0))
            workbook = load_workbook(export_path, read_only=True, data_only=True)
            try:
                ws = workbook["任务台账"]
                headers = [cell.value for cell in ws[1]]
                self.assertIn("下一步处理人", headers)
                self.assertIn("下一步动作", headers)
                self.assertIn("处理优先级", headers)
                self.assertIn("优先级原因", headers)
                next_handler_col = headers.index("下一步处理人") + 1
                next_action_col = headers.index("下一步动作") + 1
                priority_col = headers.index("处理优先级") + 1
                product_rows = {ws.cell(row=row, column=headers.index("货品名称") + 1).value: row for row in range(2, ws.max_row + 1)}
                self.assertEqual(ws.cell(row=product_rows["未分配商品"], column=next_handler_col).value, "管理员")
                self.assertEqual(ws.cell(row=product_rows["未分配商品"], column=next_action_col).value, "指派负责人")
                self.assertEqual(ws.cell(row=product_rows["未分配商品"], column=priority_col).value, "高")
                self.assertEqual(ws.cell(row=product_rows["待审核商品"], column=next_action_col).value, "审核通过或驳回")
            finally:
                workbook.close()

            export_path = daily_ops_tasks.OperationTaskStore(task_db).export_tasks(root / "高优先级.xlsx", tasks=high_priority, filters={"priority": "高"}, now=datetime(2026, 6, 22, 12, 0, 0))
            workbook = load_workbook(export_path, read_only=True, data_only=True)
            try:
                ws = workbook["任务台账"]
                self.assertEqual(ws.max_row, 3)
                criteria_ws = workbook["导出口径"]
                criteria = {
                    criteria_ws.cell(row=row, column=1).value: criteria_ws.cell(row=row, column=2).value
                    for row in range(2, criteria_ws.max_row + 1)
                }
                self.assertEqual(criteria["priority"], "高")
            finally:
                workbook.close()

    def test_task_summary_breaks_status_counts_down_by_owner(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "7", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
                {"platform": "Temu", "task_type": "低分预警", "store": "8", "owner": "小琴", "merchant_code": "B", "source_report": "r", "source_row": 2},
                {"platform": "Shein", "task_type": "爆旺冲突", "store": "琪琪", "owner": "洁琳", "merchant_code": "C", "source_report": "r", "source_row": 3},
                {"platform": "Temu", "task_type": "滞销处理", "store": "9", "owner": "", "merchant_code": "D", "source_report": "r", "source_row": 4},
            ])
            xiaoqin_rows = store.list_tasks(role="owner", user="小琴")
            store.submit_owner_action(xiaoqin_rows[0]["id"], actor="小琴", action="已处理", remark="后台已处理")
            store.submit_owner_action(xiaoqin_rows[1]["id"], actor="小琴", action="已处理", remark="后台已处理")
            store.review_task(xiaoqin_rows[0]["id"], admin="管理员", decision="驳回", remark="缺截图")
            store.submit_owner_action(xiaoqin_rows[0]["id"], actor="小琴", action="已补图", remark="已补充截图")
            store.review_task(xiaoqin_rows[0]["id"], admin="管理员", decision="通过", remark="复核通过")

            summary = store.summary()
            self.assertEqual(summary["owner_status"]["小琴"]["total"], 2)
            self.assertEqual(summary["owner_status"]["小琴"]["by_status"]["待管理员审核"], 1)
            self.assertEqual(summary["owner_status"]["小琴"]["by_status"]["已通过"], 1)
            self.assertEqual(summary["owner_status"]["小琴"]["reworked"], 1)
            self.assertEqual(summary["owner_status"]["洁琳"]["by_status"]["待店长处理"], 1)
            self.assertEqual(summary["owner_status"]["洁琳"]["reworked"], 0)
            self.assertEqual(summary["owner_status"]["未分配"]["total"], 1)
            self.assertEqual(summary["owner_status"]["未分配"]["by_status"]["待店长处理"], 1)
            self.assertEqual(summary["by_next_handler"], {"管理员": 3, "店长": 1})
            self.assertEqual(summary["by_next_action"]["审核通过或驳回"], 1)
            self.assertEqual(summary["by_next_action"]["标记完成或归档"], 1)
            self.assertEqual(summary["by_next_action"]["指派负责人"], 1)
            self.assertEqual(summary["by_next_action"]["填写处理结果"], 1)

    def test_task_summary_counts_overdue_owner_and_review_work(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            task_db.write_text(json.dumps({
                "tasks": [
                    {
                        "id": "owner-old",
                        "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                        "owner": "小琴",
                        "created_at": "2026-06-18 09:00:00",
                        "updated_at": "2026-06-18 09:00:00",
                    },
                    {
                        "id": "review-old",
                        "status": daily_ops_tasks.STATUS_PENDING_REVIEW,
                        "owner": "小琴",
                        "owner_submitted_at": "2026-06-20 08:00:00",
                        "updated_at": "2026-06-20 08:00:00",
                        "history": [
                            {"event": "管理员审核", "action": "驳回", "remark": "缺截图"},
                        ],
                    },
                    {
                        "id": "owner-fresh",
                        "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                        "owner": "洁琳",
                        "created_at": "2026-06-21 09:00:00",
                        "updated_at": "2026-06-21 09:00:00",
                    },
                ]
            }, ensure_ascii=False), encoding="utf-8")

            summary = daily_ops_tasks.OperationTaskStore(task_db).summary(now=datetime(2026, 6, 22, 12, 0, 0))
            self.assertEqual(summary["overdue"]["total"], 2)
            self.assertEqual(summary["overdue"]["by_status"][daily_ops_tasks.STATUS_PENDING_OWNER], 1)
            self.assertEqual(summary["overdue"]["by_status"][daily_ops_tasks.STATUS_PENDING_REVIEW], 1)
            self.assertEqual(summary["owner_status"]["小琴"]["overdue"], 2)
            self.assertEqual(summary["owner_status"]["洁琳"]["overdue"], 0)

    def test_task_summary_builds_admin_action_queue(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            task_db.write_text(json.dumps({
                "tasks": [
                    {
                        "id": "unassigned",
                        "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                        "owner": "",
                        "created_at": "2026-06-22 09:00:00",
                        "updated_at": "2026-06-22 09:00:00",
                    },
                    {
                        "id": "review-old",
                        "status": daily_ops_tasks.STATUS_PENDING_REVIEW,
                        "owner": "小琴",
                        "owner_submitted_at": "2026-06-20 08:00:00",
                        "updated_at": "2026-06-20 08:00:00",
                    },
                    {
                        "id": "owner-old",
                        "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                        "owner": "洁琳",
                        "created_at": "2026-06-18 09:00:00",
                        "updated_at": "2026-06-18 09:00:00",
                    },
                    {
                        "id": "rejected-old",
                        "status": daily_ops_tasks.STATUS_REJECTED,
                        "owner": "小琴",
                        "admin_reviewed_at": "2026-06-18 09:00:00",
                        "updated_at": "2026-06-18 09:00:00",
                    },
                    {
                        "id": "review-fresh",
                        "status": daily_ops_tasks.STATUS_PENDING_REVIEW,
                        "owner": "洁琳",
                        "owner_submitted_at": "2026-06-22 09:00:00",
                        "updated_at": "2026-06-22 09:00:00",
                    },
                    {
                        "id": "approved",
                        "status": daily_ops_tasks.STATUS_APPROVED,
                        "owner": "小琴",
                        "updated_at": "2026-06-22 08:00:00",
                    },
                    {
                        "id": "done",
                        "status": daily_ops_tasks.STATUS_DONE,
                        "owner": "小琴",
                        "updated_at": "2026-06-22 08:00:00",
                    },
                ]
            }, ensure_ascii=False), encoding="utf-8")

            summary = daily_ops_tasks.OperationTaskStore(task_db).summary(now=datetime(2026, 6, 22, 12, 0, 0))

            queue = summary["admin_queue"]
            self.assertEqual(
                [item["action"] for item in queue],
                ["指派负责人", "处理超时审核", "跟进超时店长处理", "跟进驳回返工超时", "审核通过或驳回", "标记完成或归档"],
            )
            self.assertEqual([item["count"] for item in queue], [1, 1, 1, 1, 1, 1])
            self.assertEqual(queue[0]["priority"], "高")
            self.assertEqual(queue[0]["filters"], {"unassigned": "1", "open_only": "1"})
            self.assertEqual(queue[1]["filters"], {"status": daily_ops_tasks.STATUS_PENDING_REVIEW, "overdue": "1", "open_only": "1"})
            self.assertEqual(queue[-1]["priority"], "中")

    def test_task_export_includes_admin_action_queue_sheet(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            task_db.write_text(json.dumps({
                "tasks": [
                    {
                        "id": "unassigned",
                        "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                        "owner": "",
                        "created_at": "2026-06-22 09:00:00",
                        "updated_at": "2026-06-22 09:00:00",
                    },
                    {
                        "id": "review-old",
                        "status": daily_ops_tasks.STATUS_PENDING_REVIEW,
                        "owner": "小琴",
                        "owner_submitted_at": "2026-06-20 08:00:00",
                        "updated_at": "2026-06-20 08:00:00",
                    },
                    {
                        "id": "approved",
                        "status": daily_ops_tasks.STATUS_APPROVED,
                        "owner": "洁琳",
                        "updated_at": "2026-06-22 08:00:00",
                    },
                ]
            }, ensure_ascii=False), encoding="utf-8")

            export_path = daily_ops_tasks.OperationTaskStore(task_db).export_tasks(root / "导出.xlsx", now=datetime(2026, 6, 22, 12, 0, 0))
            workbook = load_workbook(export_path, read_only=True, data_only=True)
            try:
                self.assertIn("管理员待办队列", workbook.sheetnames)
                ws = workbook["管理员待办队列"]
                headers = [cell.value for cell in ws[1]]
                self.assertEqual(headers, ["处理动作", "优先级", "任务数量", "筛选条件"])
                rows = {
                    ws.cell(row=row, column=1).value: row
                    for row in range(2, ws.max_row + 1)
                }
                self.assertEqual(ws.cell(row=rows["指派负责人"], column=3).value, 1)
                self.assertEqual(ws.cell(row=rows["指派负责人"], column=4).value, "unassigned=1; open_only=1")
                self.assertEqual(ws.cell(row=rows["处理超时审核"], column=2).value, "高")
                self.assertEqual(ws.cell(row=rows["处理超时审核"], column=4).value, "status=待管理员审核; overdue=1; open_only=1")
                self.assertEqual(ws.cell(row=rows["标记完成或归档"], column=3).value, 1)
            finally:
                workbook.close()

    def test_rejected_tasks_become_overdue_when_owner_does_not_rework(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            task_db.write_text(json.dumps({
                "tasks": [
                    {
                        "id": "rejected-old",
                        "platform": "Temu",
                        "task_type": "议价审核",
                        "status": daily_ops_tasks.STATUS_REJECTED,
                        "store": "7",
                        "owner": "小琴",
                        "product_name": "驳回未返工商品",
                        "created_at": "2026-06-18 09:00:00",
                        "updated_at": "2026-06-19 09:00:00",
                        "admin_reviewed_at": "2026-06-19 09:00:00",
                        "history": [
                            {"event": "管理员审核", "action": "驳回", "remark": "缺截图"},
                        ],
                    },
                ]
            }, ensure_ascii=False), encoding="utf-8")
            store = daily_ops_tasks.OperationTaskStore(task_db)
            now = datetime(2026, 6, 22, 12, 0, 0)

            row = store.list_tasks(now=now)[0]
            self.assertTrue(daily_ops_tasks.task_overdue(row, now))
            self.assertEqual(row["next_handler"], "管理员")
            self.assertEqual(row["next_action"], "跟进驳回返工超时")
            self.assertEqual(store.summary([row], now=now)["overdue"]["total"], 1)

    def test_task_export_marks_overdue_tasks_for_followup(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            task_db.write_text(json.dumps({
                "tasks": [
                    {
                        "id": "owner-old",
                        "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                        "owner": "小琴",
                        "product_name": "红色球衣",
                        "created_at": "2026-06-18 09:00:00",
                        "updated_at": "2026-06-18 09:00:00",
                    },
                    {
                        "id": "owner-fresh",
                        "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                        "owner": "洁琳",
                        "product_name": "蓝色球衣",
                        "created_at": "2026-06-21 09:00:00",
                        "updated_at": "2026-06-21 09:00:00",
                    },
                ]
            }, ensure_ascii=False), encoding="utf-8")

            export_path = daily_ops_tasks.OperationTaskStore(task_db).export_tasks(root / "导出.xlsx", now=datetime(2026, 6, 22, 12, 0, 0))
            workbook = load_workbook(export_path, read_only=True, data_only=True)
            try:
                ws = workbook["任务台账"]
                headers = [cell.value for cell in ws[1]]
                overdue_col = headers.index("是否超时") + 1
                overdue_days_col = headers.index("超时天数") + 1
                product_col = headers.index("货品名称") + 1
                rows = {
                    ws.cell(row=row, column=product_col).value: row
                    for row in range(2, ws.max_row + 1)
                }
                self.assertEqual(ws.cell(row=rows["红色球衣"], column=overdue_col).value, "是")
                self.assertEqual(ws.cell(row=rows["红色球衣"], column=overdue_days_col).value, 4)
                self.assertEqual(ws.cell(row=rows["蓝色球衣"], column=overdue_col).value, "否")
                self.assertEqual(ws.cell(row=rows["蓝色球衣"], column=overdue_days_col).value, 1)
            finally:
                workbook.close()

    def test_task_export_counts_rejected_overdue_days_from_admin_rejection_time(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            task_db.write_text(json.dumps({
                "tasks": [
                    {
                        "id": "rejected-old",
                        "status": daily_ops_tasks.STATUS_REJECTED,
                        "owner": "小琴",
                        "product_name": "驳回未返工商品",
                        "created_at": "2026-06-18 09:00:00",
                        "updated_at": "2026-06-19 09:00:00",
                        "admin_reviewed_at": "2026-06-19 09:00:00",
                        "history": [
                            {"event": "管理员审核", "action": "驳回", "remark": "缺截图"},
                        ],
                    },
                ]
            }, ensure_ascii=False), encoding="utf-8")

            export_path = daily_ops_tasks.OperationTaskStore(task_db).export_tasks(root / "导出.xlsx", now=datetime(2026, 6, 22, 12, 0, 0))
            workbook = load_workbook(export_path, read_only=True, data_only=True)
            try:
                ws = workbook["任务台账"]
                headers = [cell.value for cell in ws[1]]
                overdue_col = headers.index("是否超时") + 1
                overdue_days_col = headers.index("超时天数") + 1
                next_action_col = headers.index("下一步动作") + 1
                self.assertEqual(ws.cell(row=2, column=overdue_col).value, "是")
                self.assertEqual(ws.cell(row=2, column=overdue_days_col).value, 3)
                self.assertEqual(ws.cell(row=2, column=next_action_col).value, "跟进驳回返工超时")
            finally:
                workbook.close()

    def test_task_export_owner_summary_includes_overdue_counts(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            task_db.write_text(json.dumps({
                "tasks": [
                    {
                        "id": "owner-old",
                        "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                        "owner": "小琴",
                        "created_at": "2026-06-18 09:00:00",
                        "updated_at": "2026-06-18 09:00:00",
                    },
                    {
                        "id": "review-old",
                        "status": daily_ops_tasks.STATUS_PENDING_REVIEW,
                        "owner": "小琴",
                        "owner_submitted_at": "2026-06-20 08:00:00",
                        "updated_at": "2026-06-20 08:00:00",
                        "history": [
                            {"event": "管理员审核", "action": "驳回", "remark": "缺截图"},
                        ],
                    },
                    {
                        "id": "fresh",
                        "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                        "owner": "洁琳",
                        "created_at": "2026-06-21 09:00:00",
                        "updated_at": "2026-06-21 09:00:00",
                    },
                ]
            }, ensure_ascii=False), encoding="utf-8")

            export_path = daily_ops_tasks.OperationTaskStore(task_db).export_tasks(root / "导出.xlsx", now=datetime(2026, 6, 22, 12, 0, 0))
            workbook = load_workbook(export_path, read_only=True, data_only=True)
            try:
                owner_ws = workbook["负责人汇总"]
                headers = [cell.value for cell in owner_ws[1]]
                overdue_col = headers.index("超时未处理") + 1
                reworked_col = headers.index("返工任务") + 1
                owner_rows = {owner_ws.cell(row=row, column=1).value: row for row in range(2, owner_ws.max_row + 1)}
                self.assertEqual(owner_ws.cell(row=owner_rows["小琴"], column=overdue_col).value, 2)
                self.assertEqual(owner_ws.cell(row=owner_rows["小琴"], column=reworked_col).value, 1)
                self.assertEqual(owner_ws.cell(row=owner_rows["洁琳"], column=overdue_col).value, 0)
                self.assertEqual(owner_ws.cell(row=owner_rows["洁琳"], column=reworked_col).value, 0)
            finally:
                workbook.close()

    def test_task_listing_can_filter_only_overdue_tasks(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            task_db.write_text(json.dumps({
                "tasks": [
                    {
                        "id": "owner-old",
                        "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                        "owner": "小琴",
                        "product_name": "红色球衣",
                        "created_at": "2026-06-18 09:00:00",
                        "updated_at": "2026-06-18 09:00:00",
                    },
                    {
                        "id": "fresh",
                        "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                        "owner": "小琴",
                        "product_name": "蓝色球衣",
                        "created_at": "2026-06-21 09:00:00",
                        "updated_at": "2026-06-21 09:00:00",
                    },
                ]
            }, ensure_ascii=False), encoding="utf-8")

            rows = daily_ops_tasks.OperationTaskStore(task_db).list_tasks(
                role="owner",
                user="小琴",
                overdue="1",
                now=datetime(2026, 6, 22, 12, 0, 0),
            )
            self.assertEqual([row["product_name"] for row in rows], ["红色球衣"])

    def test_task_listing_can_filter_unassigned_tasks(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "低分预警", "store": "7", "owner": "", "merchant_code": "A", "source_report": "r", "source_row": 1},
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "8", "owner": "小琴", "merchant_code": "B", "source_report": "r", "source_row": 2},
            ])

            rows = store.list_tasks(unassigned="1")
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["store"], "7")
            self.assertEqual(rows[0]["owner"], "")

    def test_owner_listing_requires_named_owner(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "低分预警", "store": "7", "owner": "", "merchant_code": "A", "source_report": "r", "source_row": 1},
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "8", "owner": "小琴", "merchant_code": "B", "source_report": "r", "source_row": 2},
            ])

            self.assertEqual(store.list_tasks(role="owner", user=""), [])

    def test_admin_can_assign_unassigned_task_to_owner(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {
                    "platform": "Temu",
                    "task_type": "低分预警",
                    "store": "7",
                    "owner": "",
                    "merchant_code": "A-001",
                    "product_name": "红色球衣",
                    "system_action": "补充负责人后处理",
                    "source_report": "店铺低分产品预警",
                    "source_file": "low.xlsx",
                    "source_row": 2,
                }
            ])
            task = store.list_tasks()[0]

            with self.assertRaises(ValueError):
                store.assign_task(task["id"], actor="管理员", owner="", remark="")

            assigned = store.assign_task(task["id"], actor="管理员", owner="小琴", remark="按店铺负责人补齐")
            self.assertEqual(assigned["owner"], "小琴")
            self.assertEqual(assigned["history"][-1]["event"], "任务指派")
            self.assertEqual(assigned["history"][-1]["action"], "指派给 小琴")

            owner_rows = store.list_tasks(role="owner", user="小琴")
            self.assertEqual(len(owner_rows), 1)
            self.assertEqual(store.summary()["unassigned"], 0)

    def test_reimport_without_owner_keeps_manual_assignment(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            source = {
                "platform": "Shein",
                "task_type": "爆旺冲突",
                "store": "琪琪",
                "owner": "",
                "merchant_code": "B-001",
                "product_name": "宠物背带",
                "system_action": "下架重复铺货",
                "source_report": "Shein爆旺款重复预警",
                "source_file": "shein-hot.xlsx",
                "source_sheet": "具体店铺操作表",
                "source_row": 2,
            }
            store.upsert_generated_tasks([source])
            task = store.list_tasks()[0]
            store.assign_task(task["id"], actor="管理员", owner="洁琳", remark="手动补负责人")

            result = store.upsert_generated_tasks([{**source, "system_action": "继续下架重复铺货"}])

            self.assertEqual(result["updated"], 1)
            updated = store.list_tasks()[0]
            self.assertEqual(updated["owner"], "洁琳")
            self.assertEqual(updated["system_action"], "继续下架重复铺货")
            self.assertEqual(updated["history"][-2]["event"], "任务指派")
            self.assertEqual(updated["history"][-1]["event"], "系统更新")
            self.assertIn("系统建议动作", updated["history"][-1]["remark"])

    def test_reimport_does_not_move_submitted_task_to_another_owner(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            source = {
                "platform": "Temu",
                "task_type": "价格异常",
                "store": "7",
                "owner": "小琴",
                "merchant_code": "A-001",
                "product_name": "红色球衣",
                "system_action": "低于成本价",
                "source_report": "Temu申报价异常",
                "source_file": "price.xlsx",
                "source_sheet": "低于成本价",
                "source_row": 2,
            }
            store.upsert_generated_tasks([source])
            task = store.list_tasks()[0]
            store.submit_owner_action(task["id"], actor="小琴", action="已调价", remark="后台已处理")

            result = store.upsert_generated_tasks([{**source, "owner": "洁琳", "system_action": "低于成本价，继续处理"}])

            self.assertEqual(result["updated"], 1)
            updated = store.list_tasks()[0]
            self.assertEqual(updated["owner"], "小琴")
            self.assertEqual(updated["status"], daily_ops_tasks.STATUS_PENDING_REVIEW)
            self.assertEqual(updated["owner_action"], "已调价")
            self.assertEqual(updated["system_action"], "低于成本价，继续处理")
            self.assertEqual(updated["history"][-1]["event"], "系统更新")
            self.assertEqual(updated["history"][-1]["actor"], "系统")
            self.assertEqual(updated["history"][-1]["action"], "更新任务明细")
            self.assertIn("系统建议动作", updated["history"][-1]["remark"])

    def test_reimport_does_not_overwrite_manual_assignment_with_source_owner(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            source = {
                "platform": "Shein",
                "task_type": "库存异常",
                "store": "琪琪",
                "owner": "",
                "merchant_code": "B-001",
                "product_name": "宠物背带",
                "system_action": "仓备大于30天销量2倍",
                "source_report": "Shein仓备库存异常",
                "source_file": "inventory.xlsx",
                "source_sheet": "仓备大于30天销量2倍",
                "source_row": 2,
            }
            store.upsert_generated_tasks([source])
            task = store.list_tasks()[0]
            store.assign_task(task["id"], actor="管理员", owner="洁琳", remark="按负责人配置补齐")

            result = store.upsert_generated_tasks([{**source, "owner": "小琴", "system_action": "仓备仍异常"}])

            self.assertEqual(result["updated"], 1)
            updated = store.list_tasks()[0]
            self.assertEqual(updated["owner"], "洁琳")
            self.assertEqual(updated["system_action"], "仓备仍异常")
            self.assertEqual(updated["history"][-2]["event"], "任务指派")
            self.assertEqual(updated["history"][-1]["event"], "系统更新")
            self.assertIn("系统建议动作", updated["history"][-1]["remark"])

    def test_reimport_does_not_modify_completed_task_archive(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            source = {
                "platform": "Temu",
                "task_type": "低分预警",
                "store": "7",
                "owner": "小琴",
                "merchant_code": "A-001",
                "product_name": "红色球衣",
                "system_action": "低分仍在售，需处理",
                "source_report": "店铺低分产品预警",
                "source_file": "low.xlsx",
                "source_sheet": "低分明细",
                "source_row": 2,
            }
            store.upsert_generated_tasks([source])
            task = store.list_tasks()[0]
            submitted = store.submit_owner_action(task["id"], actor="小琴", action="已下架", remark="后台已下架")
            approved = store.review_task(submitted["id"], admin="管理员", decision="通过", remark="同意")
            done = store.mark_done(approved["id"], actor="管理员", remark="后台已确认")
            history_count = len(done["history"])

            result = store.upsert_generated_tasks([{**source, "system_action": "导入后仍低分"}])

            self.assertEqual(result["updated"], 0)
            archived = store.list_tasks()[0]
            self.assertEqual(archived["status"], daily_ops_tasks.STATUS_DONE)
            self.assertEqual(archived["system_action"], "低分仍在售，需处理")
            self.assertEqual(len(archived["history"]), history_count)

    def test_unassigned_task_must_be_assigned_before_owner_submit(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {
                    "platform": "Shein",
                    "task_type": "爆旺冲突",
                    "store": "琪琪",
                    "owner": "",
                    "merchant_code": "B-001",
                    "product_name": "宠物背带",
                    "system_action": "下架重复铺货",
                    "source_report": "Shein爆旺款重复预警",
                    "source_row": 2,
                }
            ])
            task = store.list_tasks()[0]

            with self.assertRaises(ValueError):
                store.submit_owner_action(task["id"], actor="洁琳", action="已下架", remark="后台已处理")

            assigned = store.assign_task(task["id"], actor="管理员", owner="洁琳", remark="补负责人")
            submitted = store.submit_owner_action(assigned["id"], actor="洁琳", action="已下架", remark="后台已处理")
            self.assertEqual(submitted["status"], daily_ops_tasks.STATUS_PENDING_REVIEW)
            self.assertEqual(submitted["owner"], "洁琳")

    def test_task_status_flow_rejects_invalid_operations(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {
                    "platform": "Temu",
                    "task_type": "爆旺冲突",
                    "store": "7",
                    "owner": "小琴",
                    "merchant_code": "A-001",
                    "product_name": "红色球衣",
                    "system_action": "立即下架",
                    "source_report": "Temu爆旺款重复预警",
                    "source_row": 2,
                }
            ])
            task = store.list_tasks()[0]

            with self.assertRaises(ValueError):
                store.submit_owner_action(task["id"], actor="小琴", action="", remark="只写备注")

            with self.assertRaises(ValueError):
                store.review_task(task["id"], admin="管理员", decision="通过", remark="未提交就审核")

            submitted = store.submit_owner_action(task["id"], actor="小琴", action="已下架", remark="后台已处理")
            with self.assertRaises(ValueError):
                store.review_task(submitted["id"], admin="管理员", decision="驳回", remark="")
            with self.assertRaises(ValueError):
                store.review_task(submitted["id"], admin="管理员", decision="通过", remark="")
            reviewed = store.review_task(submitted["id"], admin="管理员", decision="通过", remark="同意")
            self.assertEqual(reviewed["status"], daily_ops_tasks.STATUS_APPROVED)

            with self.assertRaises(ValueError):
                store.submit_owner_action(task["id"], actor="小琴", action="改为继续观察", remark="")

            with self.assertRaises(ValueError):
                store.review_task(task["id"], admin="管理员", decision="同意", remark="非法审核结果")

    def test_rejected_tasks_keep_rework_context_after_owner_resubmits(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {
                    "platform": "Temu",
                    "task_type": "议价审核",
                    "store": "7",
                    "owner": "小琴",
                    "merchant_code": "A-001",
                    "product_name": "红色球衣",
                    "system_action": "确认是否接受议价",
                    "source_report": "Temu议价审核表",
                    "source_row": 2,
                }
            ])
            task = store.list_tasks()[0]
            submitted = store.submit_owner_action(task["id"], actor="小琴", action="同意议价", remark="已处理")
            rejected = store.review_task(submitted["id"], admin="管理员", decision="驳回", remark="缺少后台截图")
            self.assertEqual(rejected["rejection_count"], 1)
            self.assertEqual(rejected["last_rejection_reason"], "缺少后台截图")
            self.assertEqual(rejected["next_action"], "按驳回原因重新处理")

            resubmitted = store.submit_owner_action(rejected["id"], actor="小琴", action="补截图后同意议价", remark="已补充")
            self.assertEqual(resubmitted["status"], daily_ops_tasks.STATUS_PENDING_REVIEW)
            self.assertEqual(resubmitted["rejection_count"], 1)
            self.assertEqual(resubmitted["last_rejection_reason"], "缺少后台截图")
            store.upsert_generated_tasks([
                {
                    "platform": "Temu",
                    "task_type": "议价审核",
                    "store": "8",
                    "owner": "小琴",
                    "merchant_code": "A-002",
                    "product_name": "蓝色球衣",
                    "system_action": "确认是否接受议价",
                    "source_report": "Temu议价审核表",
                    "source_row": 3,
                }
            ])

            reworked_rows = store.list_tasks(reworked="1")
            self.assertEqual([row["product_name"] for row in reworked_rows], ["红色球衣"])

            export_path = store.export_tasks(root / "导出.xlsx", tasks=reworked_rows, filters={"reworked": "1"})
            workbook = load_workbook(export_path, read_only=True, data_only=True)
            try:
                ws = workbook["任务台账"]
                headers = [cell.value for cell in ws[1]]
                rejection_count_col = headers.index("驳回次数") + 1
                last_reason_col = headers.index("最近驳回原因") + 1
                self.assertEqual(ws.max_row, 2)
                self.assertEqual(ws.cell(row=2, column=rejection_count_col).value, 1)
                self.assertEqual(ws.cell(row=2, column=last_reason_col).value, "缺少后台截图")
                criteria_ws = workbook["导出口径"]
                criteria = {
                    criteria_ws.cell(row=row, column=1).value: criteria_ws.cell(row=row, column=2).value
                    for row in range(2, criteria_ws.max_row + 1)
                }
                self.assertEqual(criteria["reworked"], "1")
            finally:
                workbook.close()

    def test_admin_can_batch_review_pending_tasks_atomically(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "7", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
                {"platform": "Temu", "task_type": "低分预警", "store": "8", "owner": "小琴", "merchant_code": "B", "source_report": "r", "source_row": 2},
            ])
            rows = store.list_tasks(role="owner", user="小琴")
            submitted = [
                store.submit_owner_action(row["id"], actor="小琴", action="已处理", remark="后台已处理")
                for row in rows
            ]

            result = store.review_tasks(
                [row["id"] for row in submitted],
                admin="管理员",
                decision="通过",
                remark="批量确认",
            )

            self.assertEqual(result["count"], 2)
            self.assertEqual([row["status"] for row in result["tasks"]], [daily_ops_tasks.STATUS_APPROVED, daily_ops_tasks.STATUS_APPROVED])
            for row in store.list_tasks():
                self.assertEqual(row["admin_decision"], "通过")
                self.assertEqual(row["history"][-1]["event"], "管理员批量审核")
                self.assertEqual(row["history"][-1]["remark"], "批量确认")

    def test_batch_review_rejects_empty_or_non_pending_selection(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "7", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
            ])
            task = store.list_tasks()[0]

            with self.assertRaises(ValueError):
                store.review_tasks([], admin="管理员", decision="通过", remark="")

            with self.assertRaises(ValueError):
                store.review_tasks([task["id"]], admin="管理员", decision="通过", remark="未提交")

    def test_admin_can_mark_approved_task_done_once(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {
                    "platform": "Temu",
                    "task_type": "议价审核",
                    "store": "7",
                    "owner": "小琴",
                    "merchant_code": "A-001",
                    "product_name": "红色球衣",
                    "system_action": "同意议价",
                    "source_report": "Temu议价回复",
                    "source_row": 2,
                }
            ])
            task = store.list_tasks()[0]

            with self.assertRaises(ValueError):
                store.mark_done(task["id"], actor="管理员", remark="未审核不能完成")

            submitted = store.submit_owner_action(task["id"], actor="小琴", action="已处理", remark="后台已处理")
            reviewed = store.review_task(submitted["id"], admin="管理员", decision="通过", remark="同意")
            with self.assertRaises(ValueError):
                store.mark_done(reviewed["id"], actor="管理员", remark="")
            done = store.mark_done(reviewed["id"], actor="管理员", remark="后台已确认")

            self.assertEqual(done["status"], daily_ops_tasks.STATUS_DONE)
            self.assertEqual(done["completed_by"], "管理员")
            self.assertEqual(done["completed_remark"], "后台已确认")
            self.assertTrue(done["completed_at"])
            self.assertEqual(done["history"][-1]["event"], "标记完成")
            self.assertEqual(done["history"][-1]["action"], daily_ops_tasks.STATUS_DONE)
            self.assertEqual(done["history"][-1]["remark"], "后台已确认")

            export_path = store.export_tasks(root / "导出.xlsx")
            workbook = load_workbook(export_path, read_only=True, data_only=True)
            try:
                ws = workbook["任务台账"]
                headers = [cell.value for cell in ws[1]]
                completed_by_col = headers.index("完成确认人") + 1
                completed_at_col = headers.index("完成时间") + 1
                completed_remark_col = headers.index("完成说明") + 1
                self.assertEqual(ws.cell(row=2, column=completed_by_col).value, "管理员")
                self.assertTrue(ws.cell(row=2, column=completed_at_col).value)
                self.assertEqual(ws.cell(row=2, column=completed_remark_col).value, "后台已确认")

                history_ws = workbook["操作记录"]
                history_headers = [cell.value for cell in history_ws[1]]
                status_after_col = history_headers.index("动作后状态") + 1
                event_col = history_headers.index("事件") + 1
                status_by_event = {
                    history_ws.cell(row=row, column=event_col).value: history_ws.cell(row=row, column=status_after_col).value
                    for row in range(2, history_ws.max_row + 1)
                }
                self.assertEqual(status_by_event["系统生成"], daily_ops_tasks.STATUS_PENDING_OWNER)
                self.assertEqual(status_by_event["店长提交"], daily_ops_tasks.STATUS_PENDING_REVIEW)
                self.assertEqual(status_by_event["管理员审核"], daily_ops_tasks.STATUS_APPROVED)
                self.assertEqual(status_by_event["标记完成"], daily_ops_tasks.STATUS_DONE)
            finally:
                workbook.close()

            with self.assertRaises(ValueError):
                store.mark_done(task["id"], actor="管理员", remark="重复完成")

    def test_owner_directory_summarizes_task_owners_and_stores(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "7", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
                {"platform": "Temu", "task_type": "低分预警", "store": "9", "owner": "小琴", "merchant_code": "B", "source_report": "r", "source_row": 2},
                {"platform": "Shein", "task_type": "爆旺冲突", "store": "琪琪", "owner": "洁琳", "merchant_code": "C", "source_report": "r", "source_row": 3},
                {"platform": "Temu", "task_type": "滞销处理", "store": "", "owner": "", "merchant_code": "D", "source_report": "r", "source_row": 4},
            ])

            owners = store.owner_directory()
            self.assertEqual([row["owner"] for row in owners], ["小琴", "洁琳"])
            self.assertEqual(owners[0]["stores"], ["7", "9"])
            self.assertEqual(owners[0]["task_count"], 2)
            self.assertEqual(owners[1]["stores"], ["琪琪"])
            self.assertEqual(owners[1]["platforms"], ["Shein"])

    def test_daily_ops_app_syncs_report_output_into_tasks(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            report = root / "hot.xlsx"
            workbook = daily_ops_tasks.Workbook()
            ws = workbook.active
            ws.title = "具体店铺操作表"
            ws.append(["商家编码", "货品名称", "skc", "所属店铺", "负责人", "冲突类型", "处理意见"])
            ws.append(["3303197351-L", "篮球队服-红", "4173751716", "7", "小琴", "平销款冲突爆款", "立即下架！"])
            workbook.save(report)

            output_dir = root / "outputs"
            task_db = root / "operation_tasks.json"
            with patch.object(daily_ops_app, "OUTPUT_DIR", output_dir), \
                 patch.object(daily_ops_app, "TASK_DB_PATH", task_db):
                result = daily_ops_app.sync_report_tasks("temu_hot", report)
                self.assertEqual(result["created"], 1)

                tasks = daily_ops_app.list_operation_tasks(role="owner", user="小琴")
                self.assertEqual(len(tasks), 1)
                self.assertEqual(tasks[0]["task_type"], "爆旺冲突")

                submitted = daily_ops_app.submit_operation_task(tasks[0]["id"], "小琴", "已下架", "已处理")
                self.assertEqual(submitted["status"], daily_ops_tasks.STATUS_PENDING_REVIEW)

                reviewed = daily_ops_app.review_operation_task(tasks[0]["id"], "管理员", "通过", "同意")
                self.assertEqual(reviewed["status"], daily_ops_tasks.STATUS_APPROVED)

                exported = daily_ops_app.export_operation_tasks()
                self.assertTrue((output_dir / exported["file"]).exists())
                self.assertEqual(exported["rows"], 1)

    def test_report_task_mapping_keeps_review_details_for_manual_decision(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)

            low_score = root / "low.xlsx"
            workbook = daily_ops_tasks.Workbook()
            ws = workbook.active
            ws.title = "低分明细"
            ws.append(["店铺", "负责人", "SKC", "货品名称", "品质分", "是否已下架", "是否本周新增低分"])
            ws.append(["7", "小琴", "SKC-L1", "低分彩虹雨衣", 2.8, "否", "是"])
            workbook.save(low_score)

            low_rows = daily_ops_tasks.rows_from_report_workbook("low_score_warning", "店铺低分产品预警", low_score)
            self.assertEqual(len(low_rows), 1)
            self.assertIn("品质分：2.8", low_rows[0]["task_detail"])
            self.assertIn("是否已下架：否", low_rows[0]["task_detail"])

            bargain = root / "bargain.xlsx"
            workbook = daily_ops_tasks.Workbook()
            ws = workbook.active
            ws.title = "议价回复"
            ws.append(["店铺", "负责人", "SKC", "货品名称", "建议价格", "是否通过"])
            ws.append(["8", "洁琳", "SKC-B1", "议价猫窝", 19.9, "不同意"])
            workbook.save(bargain)

            bargain_rows = daily_ops_tasks.rows_from_report_workbook("temu_bargain", "Temu议价回复", bargain)
            self.assertEqual(len(bargain_rows), 1)
            self.assertIn("建议价格：19.9", bargain_rows[0]["task_detail"])
            self.assertIn("是否通过：不同意", bargain_rows[0]["task_detail"])

            slow = root / "slow.xlsx"
            workbook = daily_ops_tasks.Workbook()
            ws = workbook.active
            ws.title = "滞销处理"
            ws.append(["店铺", "负责人", "SPU", "货品名称", "预警类型", "建议动作"])
            ws.append(["9", "小琴", "SPU-S1", "滞销狗窝", "新品滞销", "暂停补货"])
            workbook.save(slow)

            slow_rows = daily_ops_tasks.rows_from_report_workbook("temu_slow", "Temu滞销动销预警", slow)
            self.assertEqual(len(slow_rows), 1)
            self.assertIn("预警类型：新品滞销", slow_rows[0]["task_detail"])

    def test_price_and_inventory_reports_create_operation_tasks(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)

            price = root / "price.xlsx"
            workbook = daily_ops_tasks.Workbook()
            ws = workbook.active
            ws.title = "低于成本价"
            ws.append(["店铺", "SKC", "商家编码", "货品名称", "申报价", "成本价", "负责人", "7天销量", "30天销量"])
            ws.append(["7", "SKC-P1", "SKU-P1", "破价猫窝", 18.5, 20, "小琴", 3, 9])
            workbook.save(price)

            price_rows = daily_ops_tasks.rows_from_report_workbook("temu_price", "Temu申报价异常", price)
            self.assertEqual(len(price_rows), 1)
            self.assertEqual(price_rows[0]["task_type"], "价格异常")
            self.assertEqual(price_rows[0]["system_action"], "低于成本价")
            self.assertIn("申报价：18.5", price_rows[0]["task_detail"])
            self.assertIn("成本价：20", price_rows[0]["task_detail"])

            inventory = root / "inventory.xlsx"
            workbook = daily_ops_tasks.Workbook()
            ws = workbook.active
            ws.title = "仓备大于30天销量2倍"
            ws.append(["店铺", "SKC", "商家编码", "货品名称", "仓备可用", "30天销量", "7天销量", "负责人"])
            ws.append(["琪琪", "SKC-I1", "SKU-I1", "库存猫抓板", 120, 20, 4, "洁琳"])
            workbook.save(inventory)

            inventory_rows = daily_ops_tasks.rows_from_report_workbook("shein_inventory", "Shein仓备库存异常", inventory)
            self.assertEqual(len(inventory_rows), 1)
            self.assertEqual(inventory_rows[0]["platform"], "Shein")
            self.assertEqual(inventory_rows[0]["task_type"], "库存异常")
            self.assertEqual(inventory_rows[0]["system_action"], "仓备大于30天销量2倍")
            self.assertIn("仓备可用：120", inventory_rows[0]["task_detail"])
            self.assertIn("30天销量：20", inventory_rows[0]["task_detail"])

    def test_store_owner_mapping_fills_report_tasks_without_owner_column(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            report = root / "shein_hot.xlsx"
            workbook = daily_ops_tasks.Workbook()
            ws = workbook.active
            ws.title = "具体店铺操作表"
            ws.append(["商家编码", "货品名称", "SKC", "店铺", "处理意见"])
            ws.append(["S-001", "宠物背带", "SKC-S1", "琪琪", "下架重复铺货"])
            workbook.save(report)

            task_db = root / "operation_tasks.json"
            owner_map = root / "store_owner_map.json"
            admin = daily_ops_app.login_operator("admin", "管理员", "")
            owner = daily_ops_app.login_operator("owner", "洁琳", "")
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db), \
                 patch.object(daily_ops_app, "STORE_OWNER_MAP_FILE", owner_map):
                status, _content_type, body = daily_ops_app.handle_store_owners_api(
                    "POST_SAVE",
                    {"X-Operator-Token": admin["token"]},
                    {"assignments": [{"platform": "Shein", "store": "琪琪", "owner": "洁琳"}]},
                )
                self.assertEqual(status, 200)
                self.assertEqual(json.loads(body)["assignments"][0]["owner"], "洁琳")

                sync = daily_ops_app.sync_report_tasks("shein_hot", report)
                self.assertEqual(sync["created"], 1)

                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "GET",
                    {"X-Operator-Token": owner["token"]},
                    {},
                )
                payload = json.loads(body)
                self.assertEqual(status, 200)
                self.assertEqual(len(payload["tasks"]), 1)
                self.assertEqual(payload["tasks"][0]["owner"], "洁琳")
                self.assertIn("店铺负责人配置", payload["tasks"][0]["history"][0]["remark"])
                self.assertEqual(payload["summary"]["unassigned"], 0)

    def test_store_owner_mapping_api_requires_admin_to_save(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        owner = daily_ops_app.login_operator("owner", "小琴", "")
        admin = daily_ops_app.login_operator("admin", "管理员", "")
        with TemporaryDirectory() as tmp:
            owner_map = Path(tmp) / "store_owner_map.json"
            with patch.object(daily_ops_app, "STORE_OWNER_MAP_FILE", owner_map):
                status, _content_type, body = daily_ops_app.handle_store_owners_api("GET", {})
                self.assertEqual(status, 401)
                self.assertIn("请先登录", json.loads(body)["error"])

                status, _content_type, body = daily_ops_app.handle_store_owners_api(
                    "GET",
                    {"X-Operator-Token": owner["token"]},
                )
                self.assertEqual(status, 403)

                status, _content_type, body = daily_ops_app.handle_store_owners_api(
                    "GET",
                    {"X-Operator-Token": admin["token"]},
                )
                self.assertEqual(status, 200)

                status, _content_type, body = daily_ops_app.handle_store_owners_api(
                    "POST_SAVE",
                    {"X-Operator-Token": owner["token"]},
                    {"assignments": [{"platform": "Temu", "store": "7", "owner": "小琴"}]},
                )
                self.assertEqual(status, 403)

                status, _content_type, body = daily_ops_app.handle_store_owners_api(
                    "POST_SAVE",
                    {"X-Operator-Token": admin["token"]},
                    {"assignments": [
                        {"platform": "Temu", "store": "7", "owner": "小琴"},
                        {"platform": "Shein", "store": "琪琪", "owner": "洁琳"},
                    ]},
                )
                payload = json.loads(body)
                self.assertEqual(status, 200)
                self.assertEqual([row["store"] for row in payload["assignments"]], ["7", "琪琪"])

    def test_saving_store_owner_mapping_assigns_existing_unassigned_tasks(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        admin = daily_ops_app.login_operator("admin", "管理员", "")
        owner = daily_ops_app.login_operator("owner", "小琴", "")
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            owner_map = root / "store_owner_map.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "低分预警", "store": "7", "owner": "", "merchant_code": "A", "source_report": "r", "source_row": 1},
                {"platform": "Shein", "task_type": "爆旺冲突", "store": "琪琪", "owner": "", "merchant_code": "B", "source_report": "r", "source_row": 2},
            ])
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db), \
                 patch.object(daily_ops_app, "STORE_OWNER_MAP_FILE", owner_map):
                status, _content_type, body = daily_ops_app.handle_store_owners_api(
                    "POST_SAVE",
                    {"X-Operator-Token": admin["token"]},
                    {"assignments": [{"platform": "Temu", "store": "7", "owner": "小琴"}]},
                )
                payload = json.loads(body)
                self.assertEqual(status, 200)
                self.assertEqual(payload["assigned_existing"], 1)

                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "GET",
                    {"X-Operator-Token": owner["token"]},
                    {},
                )
                owner_payload = json.loads(body)
                self.assertEqual(status, 200)
                self.assertEqual(len(owner_payload["tasks"]), 1)
                self.assertEqual(owner_payload["tasks"][0]["store"], "7")
                self.assertEqual(owner_payload["summary"]["unassigned"], 0)

    def test_desktop_adapter_saving_store_owners_assigns_existing_unassigned_tasks(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            owner_map = root / "store_owner_map.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "价格异常", "store": "7", "owner": "", "merchant_code": "A", "source_report": "r", "source_row": 1},
            ])

            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db), \
                 patch.object(daily_ops_app, "STORE_OWNER_MAP_FILE", owner_map):
                result = daily_ops_desktop_adapter.save_store_owners([
                    {"platform": "Temu", "store": "7", "owner": "小琴"},
                ])

                self.assertEqual(result["assigned_existing"], 1)
                owner_rows = daily_ops_desktop_adapter.operation_tasks(role="owner", user="小琴")["tasks"]
                self.assertEqual(len(owner_rows), 1)
                self.assertEqual(owner_rows[0]["owner"], "小琴")

    def test_desktop_store_owner_save_records_actual_admin_for_auto_assignment(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            owner_map = root / "store_owner_map.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "价格异常", "store": "7", "owner": "", "merchant_code": "A", "source_report": "r", "source_row": 1},
            ])
            payload = {
                "role": "admin",
                "user": "运营主管",
                "assignments": [{"platform": "Temu", "store": "7", "owner": "小琴"}],
            }

            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db), \
                 patch.object(daily_ops_app, "STORE_OWNER_MAP_FILE", owner_map), \
                 patch("sys.stdin", io.StringIO(json.dumps(payload, ensure_ascii=False))):
                result = daily_ops_cli.command(["save-store-owners"])

            self.assertTrue(result["ok"])
            task = daily_ops_tasks.OperationTaskStore(task_db).list_tasks()[0]
            self.assertEqual(task["history"][-1]["event"], "自动指派")
            self.assertEqual(task["history"][-1]["actor"], "运营主管")
            self.assertEqual(task["history"][-1]["status_after"], daily_ops_tasks.STATUS_PENDING_OWNER)

    def test_local_web_page_exposes_operation_task_workflow(self):
        html = daily_ops_app.HTML_PAGE
        self.assertIn("/api/tasks", html)
        self.assertIn("/api/tasks/submit", html)
        self.assertIn("/api/tasks/review", html)
        self.assertIn("/api/tasks/batch-review", html)
        self.assertIn("/api/tasks/done", html)
        self.assertIn("/api/tasks/assign", html)
        self.assertIn("/api/tasks/export", html)
        self.assertIn("/api/store-owners", html)
        self.assertIn("任务台账", html)
        self.assertIn("管理员审核", html)
        self.assertIn("批量通过", html)
        self.assertIn("标记完成", html)
        self.assertIn("指派负责人", html)
        self.assertIn("店铺负责人配置", html)
        self.assertIn("assigned_existing", html)
        self.assertIn("补齐", html)
        self.assertIn("taskActionButtons", html)
        self.assertIn("canSubmitOwnerTask", html)
        self.assertIn("canReviewTask", html)
        self.assertIn("canMarkDoneTask", html)
        self.assertIn("canAssignTask", html)
        self.assertIn("待店长处理", html)
        self.assertIn("已驳回", html)
        self.assertIn("operatorSession.role === 'owner'", html)
        self.assertIn("店长只能填写自己负责的任务", html)
        self.assertIn("data-admin-only", html)
        self.assertIn("data-admin-only=\"task-review\"", html)
        self.assertIn("applyRoleVisibility", html)
        self.assertIn("switchTab('tasks')", html)
        self.assertIn("item.stores || []", html)
        self.assertIn("负责人待办", html)
        self.assertIn("renderOwnerTaskSummary", html)
        self.assertIn("applyOwnerSummaryFilter", html)
        self.assertIn("data-owner-index", html)
        self.assertIn("管理员待办队列", html)
        self.assertIn("renderAdminTaskQueue", html)
        self.assertIn("applyAdminQueueFilter", html)
        self.assertIn("data-queue-index", html)
        self.assertIn("admin_queue", html)
        self.assertIn("owner_status", html)
        self.assertIn("超时未处理", html)
        self.assertIn("overdue", html)
        self.assertIn("taskOverdue", html)
        self.assertIn("只看超时", html)
        self.assertIn("taskOpenOnly", html)
        self.assertIn("只看未完成", html)
        self.assertIn("defaultOpenTasksForOwner", html)
        self.assertIn("open_only", html)
        self.assertIn("taskUnassigned", html)
        self.assertIn("只看未分配", html)
        self.assertIn("taskPlatform", html)
        self.assertIn("Temu", html)
        self.assertIn("Shein", html)
        self.assertIn("platform", html)
        self.assertIn("先指派负责人", html)
        self.assertIn("驳回原因", html)
        self.assertIn("管理员审核必须填写说明", html)
        self.assertIn("批量审核必须填写说明", html)
        self.assertIn("备注或处理凭证至少填一个", html)
        self.assertIn("完成确认说明", html)
        self.assertIn("标记完成必须填写确认说明", html)
        for text in ["来源", "source_report", "source_file", "source_row", "task_detail"]:
            self.assertIn(text, html)

    def test_electron_bridge_exposes_operation_task_workflow(self):
        root = Path(__file__).resolve().parent
        preload = (root / "electron" / "preload.js").read_text(encoding="utf-8")
        main = (root / "electron" / "main.js").read_text(encoding="utf-8")
        for text in ["tasks", "submitTask", "reviewTask", "batchReviewTasks", "doneTask", "assignTask", "exportTasks", "storeOwners", "saveStoreOwners"]:
            self.assertIn(text, preload)
        for text in ["api:tasks", "api:submit-task", "api:review-task", "api:batch-review-tasks", "api:done-task", "api:assign-task", "api:export-tasks", "api:store-owners", "api:save-store-owners"]:
            self.assertIn(text, main)
        self.assertIn("payload.open_only", main)

    def test_electron_renderer_exposes_admin_task_queue(self):
        root = Path(__file__).resolve().parent
        html = (root / "electron" / "renderer.html").read_text(encoding="utf-8")
        js = (root / "electron" / "renderer.js").read_text(encoding="utf-8")

        self.assertIn("adminTaskQueue", html)
        self.assertIn("renderAdminTaskQueue", js)
        self.assertIn("applyAdminQueueFilter", js)
        self.assertIn("data-queue-index", js)
        self.assertIn("admin_queue", js)
        self.assertIn("管理员待办队列", js)

    def test_desktop_adapter_scopes_task_summary_like_task_rows(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "1", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
                {"platform": "Temu", "task_type": "低分预警", "store": "2", "owner": "洁琳", "merchant_code": "B", "source_report": "r", "source_row": 2},
            ])
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db):
                payload = daily_ops_desktop_adapter.operation_tasks(role="owner", user="小琴")
                self.assertEqual(len(payload["tasks"]), 1)
                self.assertEqual(payload["summary"]["total"], 1)
                self.assertEqual(payload["summary"]["by_owner"], {"小琴": 1})

    def test_electron_renderer_exposes_operation_task_center(self):
        root = Path(__file__).resolve().parent
        html = (root / "electron" / "renderer.html").read_text(encoding="utf-8")
        js = (root / "electron" / "renderer.js").read_text(encoding="utf-8")
        css = (root / "electron" / "renderer.css").read_text(encoding="utf-8")
        for text in ["任务中心", "任务台账", "店长填写", "管理员审核", "批量通过", "批量驳回", "标记完成", "指派负责人", "店铺负责人配置", "导出任务"]:
            self.assertIn(text, html + js)
        for text in ["价格异常", "库存异常"]:
            self.assertIn(text, html)
            self.assertIn(text, daily_ops_app.HTML_PAGE)
        for text in ["renderTaskCenter", "loadTasks", "submitTask", "reviewTask", "batchReviewTasks", "doneTask", "assignTask", "loadStoreOwners", "saveStoreOwners", "exportTasks", "taskActionButtons", "renderOwnerTaskSummary"]:
            self.assertIn(text, js)
        self.assertIn("canSubmitOwnerTask", js)
        self.assertIn("canReviewTask", js)
        self.assertIn("canMarkDoneTask", js)
        self.assertIn("canAssignTask", js)
        self.assertIn('task.status === "待店长处理"', js)
        self.assertIn('task.status === "已驳回"', js)
        self.assertIn('task.status === "待管理员审核"', js)
        self.assertIn('task.status === "已通过"', js)
        self.assertIn("owner_status", js)
        self.assertIn("负责人待办", html + js)
        self.assertIn("applyOwnerSummaryFilter", js)
        self.assertIn("data-owner-index", js)
        self.assertIn("超时未处理", js)
        self.assertIn("overdue", js)
        self.assertIn("taskOverdue", html)
        self.assertIn("只看超时", html)
        self.assertIn("taskOpenOnly", html)
        self.assertIn("只看未完成", html)
        self.assertIn("defaultOpenTasksForOwner", js)
        self.assertIn("open_only", js)
        self.assertIn('data-admin-only="task-review"', html)
        self.assertIn("applyRoleVisibility", js)
        self.assertIn("taskUnassigned", html)
        self.assertIn("只看未分配", html)
        self.assertIn("taskPlatform", html)
        self.assertIn("Temu", html)
        self.assertIn("Shein", html)
        self.assertIn('platform: $("#taskPlatform")?.value || ""', js)
        self.assertIn("先指派负责人", js)
        self.assertIn("驳回原因", js)
        self.assertIn("管理员审核必须填写说明", js)
        self.assertIn("批量审核必须填写说明", js)
        self.assertIn("备注或处理凭证至少填一个", js)
        self.assertIn("完成确认说明", js)
        self.assertIn("标记完成必须填写确认说明", js)
        for text in ["operator.role === \"owner\"", "店长只能填写自己负责的任务"]:
            self.assertIn(text, js)
        for text in ["来源", "source_report", "source_file", "source_row", "task_detail"]:
            self.assertIn(text, js)
        self.assertIn("未分配", html + js + daily_ops_app.HTML_PAGE)
        self.assertIn("assigned_existing", js)
        self.assertIn("已补齐", js)
        self.assertIn("function showTaskError", js)
        self.assertGreaterEqual(js.count("showTaskError(error)"), 8)
        for text in ["task-summary", "task-table", "task-actions"]:
            self.assertIn(text, css)

    def test_task_center_exposes_online_history_view(self):
        root = Path(__file__).resolve().parent
        html = daily_ops_app.HTML_PAGE
        renderer = (root / "electron" / "renderer.js").read_text(encoding="utf-8")
        for text in ["showTaskHistory", "查看记录", "操作记录", "处理凭证", "动作后状态", "动作后下一步"]:
            self.assertIn(text, html)
            self.assertIn(text, renderer)

    def test_electron_renderer_persists_operator_identity_for_tasks(self):
        root = Path(__file__).resolve().parent
        html = (root / "electron" / "renderer.html").read_text(encoding="utf-8")
        js = (root / "electron" / "renderer.js").read_text(encoding="utf-8")
        self.assertIn("operatorRole", html)
        self.assertIn("operatorUser", html)
        self.assertIn("saveOperator", js)
        self.assertIn("localStorage", js)
        self.assertIn("currentOperator", js)

    def test_lan_host_configuration_keeps_local_default_and_allows_lan_binding(self):
        with patch.dict(os.environ, {}, clear=True):
            self.assertEqual(daily_ops_app.configured_host(), "127.0.0.1")
        with patch.dict(os.environ, {"DAILY_OPS_HOST": "0.0.0.0"}):
            self.assertEqual(daily_ops_app.configured_host(), "0.0.0.0")
        self.assertIn("局域网", daily_ops_app.access_hint("0.0.0.0", 8876))

    def test_operator_session_enforces_owner_scope_and_admin_review(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        owner_session = daily_ops_app.login_operator("owner", "小琴", "")
        admin_session = daily_ops_app.login_operator("admin", "管理员", "")

        owner = daily_ops_app.operator_from_token(owner_session["token"])
        admin = daily_ops_app.operator_from_token(admin_session["token"])

        self.assertEqual(owner["role"], "owner")
        self.assertEqual(owner["user"], "小琴")
        self.assertEqual(daily_ops_app.scoped_task_filters(owner, {"role": "admin", "user": "洁琳"}), {"role": "owner", "user": "小琴"})
        self.assertEqual(daily_ops_app.scoped_task_filters(admin, {"role": "owner", "user": "洁琳"}), {"role": "owner", "user": "洁琳"})
        self.assertTrue(daily_ops_app.can_review_tasks(admin))
        self.assertFalse(daily_ops_app.can_review_tasks(owner))

    def test_operator_can_logout_and_invalidate_session_token(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        session = daily_ops_app.login_operator("owner", "小琴", "")
        headers = {"X-Operator-Token": session["token"]}

        status, _content_type, body = daily_ops_app.handle_session_logout(headers)
        self.assertEqual(status, 200)
        self.assertTrue(json.loads(body)["ok"])

        status, _content_type, body = daily_ops_app.handle_tasks_api("GET", headers, {})
        self.assertEqual(status, 401)
        self.assertIn("请先登录", json.loads(body)["error"])

    def test_operator_session_expires_after_configured_seconds(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        with patch.dict(os.environ, {"DAILY_OPS_SESSION_SECONDS": "1"}):
            session = daily_ops_app.login_operator("owner", "小琴", "")
            self.assertEqual(daily_ops_app.operator_from_token(session["token"])["user"], "小琴")
            daily_ops_app.OPERATOR_SESSIONS[session["token"]]["login_at"] = "2026-06-22 11:59:58"
            with patch.object(daily_ops_app, "datetime") as fake_datetime:
                fake_datetime.now.return_value = datetime(2026, 6, 22, 12, 0, 0)
                fake_datetime.strptime = datetime.strptime
                fake_datetime.side_effect = lambda *args, **kwargs: datetime(*args, **kwargs)
                with self.assertRaises(PermissionError) as ctx:
                    daily_ops_app.operator_from_token(session["token"])
            self.assertIn("登录已过期", str(ctx.exception))
            self.assertNotIn(session["token"], daily_ops_app.OPERATOR_SESSIONS)

    def test_lan_mode_requires_admin_password_configuration(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        with patch.dict(os.environ, {"DAILY_OPS_HOST": "0.0.0.0"}, clear=True):
            with self.assertRaises(PermissionError):
                daily_ops_app.login_operator("admin", "管理员", "")

        with patch.dict(os.environ, {"DAILY_OPS_HOST": "0.0.0.0", "DAILY_OPS_ADMIN_PASSWORD": "safe-pass"}, clear=True):
            with self.assertRaises(PermissionError):
                daily_ops_app.login_operator("admin", "管理员", "wrong")
            session = daily_ops_app.login_operator("admin", "管理员", "safe-pass")
            self.assertEqual(session["role"], "admin")

        with patch.dict(os.environ, {}, clear=True):
            session = daily_ops_app.login_operator("admin", "管理员", "")
            self.assertEqual(session["role"], "admin")

    def test_owner_password_can_be_required_for_lan_store_manager_login(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        with patch.dict(os.environ, {"DAILY_OPS_HOST": "0.0.0.0", "DAILY_OPS_ADMIN_PASSWORD": "admin-pass", "DAILY_OPS_OWNER_PASSWORD": "owner-pass"}, clear=True):
            with self.assertRaises(PermissionError):
                daily_ops_app.login_operator("owner", "小琴", "")
            with self.assertRaises(PermissionError):
                daily_ops_app.login_operator("owner", "小琴", "wrong")
            session = daily_ops_app.login_operator("owner", "小琴", "owner-pass")
            self.assertEqual(session["role"], "owner")
            self.assertEqual(session["user"], "小琴")

        with patch.dict(os.environ, {"DAILY_OPS_HOST": "0.0.0.0", "DAILY_OPS_ADMIN_PASSWORD": "admin-pass"}, clear=True):
            session = daily_ops_app.login_operator("owner", "小琴", "")
            self.assertEqual(session["role"], "owner")

    def test_owner_login_must_match_existing_owner_directory_when_available(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            owner_map = root / "store_owner_map.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "低分预警", "store": "7", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
            ])
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db), \
                 patch.object(daily_ops_app, "STORE_OWNER_MAP_FILE", owner_map):
                with self.assertRaises(PermissionError):
                    daily_ops_app.login_operator("owner", "洁琳", "")
                session = daily_ops_app.login_operator("owner", "小琴", "")
                self.assertEqual(session["user"], "小琴")

            empty_task_db = root / "empty_tasks.json"
            with patch.object(daily_ops_app, "TASK_DB_PATH", empty_task_db), \
                 patch.object(daily_ops_app, "STORE_OWNER_MAP_FILE", owner_map):
                session = daily_ops_app.login_operator("owner", "临时店长", "")
                self.assertEqual(session["user"], "临时店长")

    def test_http_task_handlers_require_session_and_scope_owner(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        owner = daily_ops_app.login_operator("owner", "小琴", "")
        admin = daily_ops_app.login_operator("admin", "管理员", "")
        owner_headers = {"X-Operator-Token": owner["token"]}
        admin_headers = {"X-Operator-Token": admin["token"]}

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "1", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
                {"platform": "Temu", "task_type": "低分预警", "store": "2", "owner": "洁琳", "merchant_code": "B", "source_report": "r", "source_row": 2},
            ])
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db):
                status, _content_type, body = daily_ops_app.handle_tasks_api("GET", {}, {})
                self.assertEqual(status, 401)

                status, _content_type, body = daily_ops_app.handle_tasks_api("GET", owner_headers, {"role": "admin", "user": "洁琳"})
                payload = json.loads(body)
                self.assertEqual(status, 200)
                self.assertEqual(len(payload["tasks"]), 1)
                self.assertEqual(payload["tasks"][0]["owner"], "小琴")
                self.assertEqual(payload["summary"]["total"], 1)
                self.assertEqual(payload["summary"]["by_owner"], {"小琴": 1})

                owner_task = payload["tasks"][0]
                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_SUBMIT",
                    owner_headers,
                    {"id": owner_task["id"], "action": "已下架", "remark": "已处理"},
                )
                self.assertEqual(status, 200)

                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_REVIEW",
                    owner_headers,
                    {"id": owner_task["id"], "decision": "通过", "remark": "复核通过"},
                )
                self.assertEqual(status, 403)

                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_REVIEW",
                    admin_headers,
                    {"id": owner_task["id"], "decision": "通过", "remark": "复核通过"},
                )
                self.assertEqual(status, 200)

                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_DONE",
                    owner_headers,
                    {"id": owner_task["id"], "remark": "店长不能完成"},
                )
                self.assertEqual(status, 403)

                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_DONE",
                    admin_headers,
                    {"id": owner_task["id"], "remark": "管理员确认完成"},
                )
                self.assertEqual(status, 200)
                done_payload = json.loads(body)
                self.assertEqual(done_payload["task"]["status"], daily_ops_tasks.STATUS_DONE)

    def test_admin_cannot_submit_owner_action(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        admin = daily_ops_app.login_operator("admin", "管理员", "")
        admin_headers = {"X-Operator-Token": admin["token"]}
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "低分预警", "store": "1", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
            ])
            task_id = store.list_tasks()[0]["id"]
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db):
                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_SUBMIT",
                    admin_headers,
                    {"id": task_id, "action": "管理员代填", "remark": "不应允许"},
                )
            self.assertEqual(status, 403)
            self.assertIn("只有店长可以填写处理结果", json.loads(body)["error"])

    def test_desktop_admin_cannot_submit_owner_action(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "低分预警", "store": "1", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
            ])
            task_id = store.list_tasks()[0]["id"]
            admin_payload = {"role": "admin", "user": "管理员", "id": task_id, "action": "管理员代填", "remark": "不应允许"}
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db), \
                 patch("sys.stdin", io.StringIO(json.dumps(admin_payload, ensure_ascii=False))):
                with self.assertRaises(PermissionError) as ctx:
                    daily_ops_cli.command(["submit-task"])
            self.assertIn("只有店长可以填写处理结果", str(ctx.exception))

    def test_http_task_assign_requires_admin_and_updates_owner_scope(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        owner = daily_ops_app.login_operator("owner", "小琴", "")
        admin = daily_ops_app.login_operator("admin", "管理员", "")
        owner_headers = {"X-Operator-Token": owner["token"]}
        admin_headers = {"X-Operator-Token": admin["token"]}

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "低分预警", "store": "1", "owner": "", "merchant_code": "A", "source_report": "r", "source_file": "a.xlsx", "source_row": 1},
            ])
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db):
                task_id = store.list_tasks()[0]["id"]
                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_ASSIGN",
                    owner_headers,
                    {"id": task_id, "owner": "小琴", "remark": "店长不能指派"},
                )
                self.assertEqual(status, 403)

                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_ASSIGN",
                    admin_headers,
                    {"id": task_id, "owner": "小琴", "remark": "补齐负责人"},
                )
                self.assertEqual(status, 200)
                payload = json.loads(body)
                self.assertEqual(payload["task"]["owner"], "小琴")

                status, _content_type, body = daily_ops_app.handle_tasks_api("GET", owner_headers, {})
                owner_payload = json.loads(body)
                self.assertEqual(len(owner_payload["tasks"]), 1)
                self.assertEqual(owner_payload["summary"]["unassigned"], 0)

    def test_desktop_task_commands_enforce_operator_role(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "低分预警", "store": "1", "owner": "", "merchant_code": "A", "source_report": "r", "source_file": "a.xlsx", "source_row": 1},
            ])

            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db):
                task_id = store.list_tasks()[0]["id"]
                owner_payload = {"role": "owner", "user": "小琴", "id": task_id, "owner": "小琴", "remark": "店长不能指派"}
                with patch("sys.stdin", io.StringIO(json.dumps(owner_payload, ensure_ascii=False))):
                    with self.assertRaises(PermissionError):
                        daily_ops_cli.command(["assign-task"])

                admin_payload = {"role": "admin", "user": "管理员", "id": task_id, "owner": "小琴", "remark": "补齐负责人"}
                with patch("sys.stdin", io.StringIO(json.dumps(admin_payload, ensure_ascii=False))):
                    result = daily_ops_cli.command(["assign-task"])
                self.assertTrue(result["ok"])
                self.assertEqual(result["data"]["owner"], "小琴")

    def test_desktop_task_query_and_export_scope_owner_payload(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            output_dir = root / "outputs"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "低分预警", "store": "1", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "2", "owner": "洁琳", "merchant_code": "B", "source_report": "r", "source_row": 2},
            ])
            owner_payload = {
                "role": "owner",
                "user": "小琴",
                "filters": {"role": "admin", "user": "", "open_only": "1"},
            }
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db), \
                 patch.object(daily_ops_app, "OUTPUT_DIR", output_dir):
                with patch("sys.stdin", io.StringIO(json.dumps(owner_payload, ensure_ascii=False))):
                    listed = daily_ops_cli.command(["tasks"])
                self.assertTrue(listed["ok"])
                self.assertEqual([row["owner"] for row in listed["data"]["tasks"]], ["小琴"])

                with patch("sys.stdin", io.StringIO(json.dumps(owner_payload, ensure_ascii=False))):
                    exported = daily_ops_cli.command(["export-tasks"])
                self.assertTrue(exported["ok"])
                self.assertEqual(exported["data"]["rows"], 1)

    def test_desktop_admin_commands_enforce_operator_role(self):
        owner_payload = {"role": "owner", "user": "小琴", "assignments": [], "path": "/tmp/backup.zip"}
        admin_payload = {"role": "admin", "user": "管理员", "assignments": [], "path": "/tmp/backup.zip"}
        guarded_commands = [
            (["generate-weekly"], daily_ops_desktop_adapter.generate_weekly_reports, {"ran": True}),
            (["generate-report", "temu_hot", "V1.1"], daily_ops_desktop_adapter.generate_report, {"file": "hot.xlsx"}),
            (["save-rules"], daily_ops_desktop_adapter.save_rules, {"saved": True}),
            (["search", "SKU-A", "80"], daily_ops_desktop_adapter.search, [{"content": "SKU-A"}]),
            (["export-search", "SKU-A", "80"], daily_ops_desktop_adapter.export_search, {"file": "search.xlsx"}),
            (["save-store-owners"], daily_ops_desktop_adapter.save_store_owners, {"assignments": []}),
            (["create-backup"], daily_ops_desktop_adapter.create_backup, {"file": "backup.zip"}),
            (["restore-backup"], daily_ops_desktop_adapter.restore_backup, {"count": 1}),
        ]

        for argv, target, return_value in guarded_commands:
            with self.subTest(command=argv[0]):
                with patch("sys.stdin", io.StringIO(json.dumps(owner_payload, ensure_ascii=False))), \
                     patch.object(daily_ops_desktop_adapter, target.__name__, return_value=return_value):
                    with self.assertRaises(PermissionError):
                        daily_ops_cli.command(argv)

                with patch("sys.stdin", io.StringIO(json.dumps(admin_payload, ensure_ascii=False))), \
                     patch.object(daily_ops_desktop_adapter, target.__name__, return_value=return_value):
                    result = daily_ops_cli.command(argv)
                self.assertTrue(result["ok"])

    def test_desktop_store_owner_directory_requires_admin_role(self):
        owner_payload = {"role": "owner", "user": "小琴"}
        admin_payload = {"role": "admin", "user": "管理员"}
        with patch("sys.stdin", io.StringIO(json.dumps(owner_payload, ensure_ascii=False))), \
             patch.object(daily_ops_desktop_adapter, "store_owners", return_value={"assignments": []}):
            with self.assertRaises(PermissionError):
                daily_ops_cli.command(["store-owners"])

        with patch("sys.stdin", io.StringIO(json.dumps(admin_payload, ensure_ascii=False))), \
             patch.object(daily_ops_desktop_adapter, "store_owners", return_value={"assignments": []}):
            result = daily_ops_cli.command(["store-owners"])
        self.assertTrue(result["ok"])

    def test_desktop_source_import_commands_enforce_operator_role(self):
        owner_payload = {"role": "owner", "user": "小琴"}
        admin_payload = {"role": "admin", "user": "管理员"}
        source_commands = [
            (["import-source", "temu_platform", "/tmp/source.xlsx"], daily_ops_desktop_adapter.import_source_files, {"count": 1}),
            (["finish-upload", "temu_platform"], daily_ops_desktop_adapter.finish_upload, {"rows": 1}),
            (["clear-upload", "temu_platform"], daily_ops_desktop_adapter.clear_upload, {"cleared": True}),
        ]

        for argv, target, return_value in source_commands:
            with self.subTest(command=argv[0]):
                with patch("sys.stdin", io.StringIO(json.dumps(owner_payload, ensure_ascii=False))), \
                     patch.object(daily_ops_desktop_adapter, target.__name__, return_value=return_value):
                    with self.assertRaises(PermissionError):
                        daily_ops_cli.command(argv)

                with patch("sys.stdin", io.StringIO(json.dumps(admin_payload, ensure_ascii=False))), \
                     patch.object(daily_ops_desktop_adapter, target.__name__, return_value=return_value):
                    result = daily_ops_cli.command(argv)
                self.assertTrue(result["ok"])

    def test_completed_task_cannot_be_reassigned(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        admin = daily_ops_app.login_operator("admin", "管理员", "")
        admin_headers = {"X-Operator-Token": admin["token"]}
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "低分预警", "store": "1", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_file": "a.xlsx", "source_row": 1},
            ])
            task = store.list_tasks()[0]
            submitted = store.submit_owner_action(task["id"], actor="小琴", action="已处理", remark="后台已处理")
            approved = store.review_task(submitted["id"], admin="管理员", decision="通过", remark="同意")
            store.mark_done(approved["id"], actor="管理员", remark="后台已确认")

            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db):
                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_ASSIGN",
                    admin_headers,
                    {"id": task["id"], "owner": "洁琳", "remark": "完成后不应再改负责人"},
                )
                self.assertEqual(status, 400)
                self.assertIn("已完成任务不能重新指派", json.loads(body)["error"])

    def test_http_batch_review_requires_admin_and_updates_all_selected_tasks(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        owner = daily_ops_app.login_operator("owner", "小琴", "")
        admin = daily_ops_app.login_operator("admin", "管理员", "")
        owner_headers = {"X-Operator-Token": owner["token"]}
        admin_headers = {"X-Operator-Token": admin["token"]}

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "1", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
                {"platform": "Temu", "task_type": "低分预警", "store": "2", "owner": "洁琳", "merchant_code": "B", "source_report": "r", "source_row": 2},
            ])
            pending = [store.submit_owner_action(row["id"], actor=row["owner"], action="已处理", remark="后台已处理") for row in store.list_tasks()]
            ids = [row["id"] for row in pending]
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db):
                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_BATCH_REVIEW",
                    owner_headers,
                    {"ids": ids, "decision": "通过", "remark": "店长不能批量审核"},
                )
                self.assertEqual(status, 403)

                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_BATCH_REVIEW",
                    admin_headers,
                    {"ids": ids, "decision": "通过", "remark": "批量确认"},
                )
                self.assertEqual(status, 200)
                payload = json.loads(body)
                self.assertEqual(payload["count"], 2)
                self.assertEqual([row["status"] for row in payload["tasks"]], [daily_ops_tasks.STATUS_APPROVED, daily_ops_tasks.STATUS_APPROVED])

    def test_batch_reject_requires_admin_remark(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = daily_ops_tasks.OperationTaskStore(root / "tasks.json")
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "1", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
            ])
            task = store.list_tasks()[0]
            submitted = store.submit_owner_action(task["id"], actor="小琴", action="已处理", remark="后台已处理")

            with self.assertRaises(ValueError):
                store.review_tasks([submitted["id"]], admin="管理员", decision="驳回", remark="")

            rejected = store.review_tasks([submitted["id"]], admin="管理员", decision="驳回", remark="处理截图不完整")
            self.assertEqual(rejected["tasks"][0]["status"], daily_ops_tasks.STATUS_REJECTED)
            self.assertEqual(rejected["tasks"][0]["admin_remark"], "处理截图不完整")

    def test_http_task_export_can_filter_overdue_tasks(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        admin = daily_ops_app.login_operator("admin", "管理员", "")
        admin_headers = {"X-Operator-Token": admin["token"]}

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            output_dir = root / "outputs"
            task_db.write_text(json.dumps({
                "tasks": [
                    {
                        "id": "owner-old",
                        "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                        "owner": "小琴",
                        "product_name": "红色球衣",
                        "created_at": "2026-06-18 09:00:00",
                        "updated_at": "2026-06-18 09:00:00",
                    },
                    {
                        "id": "fresh",
                        "status": daily_ops_tasks.STATUS_PENDING_OWNER,
                        "owner": "洁琳",
                        "product_name": "蓝色球衣",
                        "created_at": "2026-06-21 09:00:00",
                        "updated_at": "2026-06-21 09:00:00",
                    },
                ]
            }, ensure_ascii=False), encoding="utf-8")

            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db), \
                 patch.object(daily_ops_app, "OUTPUT_DIR", output_dir), \
                 patch.object(daily_ops_tasks, "datetime") as fake_datetime:
                fake_datetime.now.return_value = datetime(2026, 6, 22, 12, 0, 0)
                fake_datetime.strptime.side_effect = datetime.strptime
                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_EXPORT",
                    admin_headers,
                    {"overdue": "1"},
                )
            self.assertEqual(status, 200)
            payload = json.loads(body)
            self.assertEqual(payload["rows"], 1)

            workbook = load_workbook(output_dir / payload["file"], read_only=True, data_only=True)
            try:
                ws = workbook["任务台账"]
                headers = [cell.value for cell in ws[1]]
                product_col = headers.index("货品名称") + 1
                self.assertEqual(ws.cell(row=2, column=product_col).value, "红色球衣")
                criteria_ws = workbook["导出口径"]
                criteria = {
                    criteria_ws.cell(row=row, column=1).value: criteria_ws.cell(row=row, column=2).value
                    for row in range(2, criteria_ws.max_row + 1)
                }
                self.assertEqual(criteria["overdue"], "1")
            finally:
                workbook.close()

    def test_http_task_export_can_filter_unassigned_tasks(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        admin = daily_ops_app.login_operator("admin", "管理员", "")
        admin_headers = {"X-Operator-Token": admin["token"]}

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            output_dir = root / "outputs"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "低分预警", "store": "7", "owner": "", "merchant_code": "A", "product_name": "未分配雨衣", "source_report": "r", "source_row": 1},
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "8", "owner": "小琴", "merchant_code": "B", "product_name": "已分配球衣", "source_report": "r", "source_row": 2},
            ])

            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db), \
                 patch.object(daily_ops_app, "OUTPUT_DIR", output_dir):
                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_EXPORT",
                    admin_headers,
                    {"unassigned": "1"},
                )
            self.assertEqual(status, 200)
            payload = json.loads(body)
            self.assertEqual(payload["rows"], 1)

            workbook = load_workbook(output_dir / payload["file"], read_only=True, data_only=True)
            try:
                ws = workbook["任务台账"]
                headers = [cell.value for cell in ws[1]]
                product_col = headers.index("货品名称") + 1
                self.assertEqual(ws.cell(row=2, column=product_col).value, "未分配雨衣")
                criteria_ws = workbook["导出口径"]
                criteria = {
                    criteria_ws.cell(row=row, column=1).value: criteria_ws.cell(row=row, column=2).value
                    for row in range(2, criteria_ws.max_row + 1)
                }
                self.assertEqual(criteria["unassigned"], "1")
            finally:
                workbook.close()

    def test_task_query_payload_keeps_web_filter_controls(self):
        payload = daily_ops_app.task_query_payload({
            "role": ["admin"],
            "user": [""],
            "status": [""],
            "task_type": ["低分预警"],
            "store": ["7"],
            "platform": ["Temu"],
            "next_handler": ["管理员"],
            "open_only": ["1"],
            "overdue": ["1"],
            "unassigned": ["1"],
            "reworked": ["1"],
        })

        self.assertEqual(payload["platform"], "Temu")
        self.assertEqual(payload["next_handler"], "管理员")
        self.assertEqual(payload["open_only"], "1")
        self.assertEqual(payload["overdue"], "1")
        self.assertEqual(payload["unassigned"], "1")
        self.assertEqual(payload["reworked"], "1")

    def test_admin_workflow_apis_require_admin_session(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        owner = daily_ops_app.login_operator("owner", "小琴", "")
        admin = daily_ops_app.login_operator("admin", "管理员", "")

        status, _content_type, body = daily_ops_app.handle_admin_api("上传数据源", {})
        self.assertEqual(status, 401)
        self.assertIn("请先登录", json.loads(body)["error"])

        status, _content_type, body = daily_ops_app.handle_admin_api("生成报表", {"X-Operator-Token": owner["token"]})
        self.assertEqual(status, 403)
        self.assertIn("只有管理员", json.loads(body)["error"])

        status, _content_type, body = daily_ops_app.handle_admin_api("生成报表", {"X-Operator-Token": admin["token"]})
        self.assertEqual(status, 200)
        self.assertEqual(json.loads(body)["operator"]["role"], "admin")

    def test_rules_api_requires_admin_session(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        owner = daily_ops_app.login_operator("owner", "小琴", "")
        admin = daily_ops_app.login_operator("admin", "管理员", "")

        status, _content_type, body = daily_ops_app.handle_rules_api("GET", {}, None)
        self.assertEqual(status, 401)
        self.assertIn("请先登录", json.loads(body)["error"])

        status, _content_type, body = daily_ops_app.handle_rules_api(
            "GET",
            {"X-Operator-Token": owner["token"]},
            None,
        )
        self.assertEqual(status, 403)
        self.assertIn("只有管理员", json.loads(body)["error"])

        with patch.object(daily_ops_app, "load_rules", return_value={"hot_item": {"keywords": ["爆"]}}):
            status, _content_type, body = daily_ops_app.handle_rules_api(
                "GET",
                {"X-Operator-Token": admin["token"]},
                None,
            )

        self.assertEqual(status, 200)
        self.assertEqual(json.loads(body)["rules"], {"hot_item": {"keywords": ["爆"]}})

    def test_search_api_requires_admin_session(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        owner = daily_ops_app.login_operator("owner", "小琴", "")
        admin = daily_ops_app.login_operator("admin", "管理员", "")

        status, _content_type, body = daily_ops_app.handle_search_api("GET", {}, {"q": "A", "limit": "10"})
        self.assertEqual(status, 401)
        self.assertIn("请先登录", json.loads(body)["error"])

        status, _content_type, body = daily_ops_app.handle_search_api(
            "GET",
            {"X-Operator-Token": owner["token"]},
            {"q": "A", "limit": "10"},
        )
        self.assertEqual(status, 403)
        self.assertIn("只有管理员", json.loads(body)["error"])

        with patch.object(daily_ops_app, "search_database", return_value=[{"content": "A"}]):
            status, _content_type, body = daily_ops_app.handle_search_api(
                "GET",
                {"X-Operator-Token": admin["token"]},
                {"q": "A", "limit": "10"},
            )

        payload = json.loads(body)
        self.assertEqual(status, 200)
        self.assertEqual(payload["rows"], [{"content": "A"}])

    def test_end_to_end_task_workflow_exports_traceable_ledger(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            report = root / "hot.xlsx"
            workbook = daily_ops_tasks.Workbook()
            ws = workbook.active
            ws.title = "具体店铺操作表"
            ws.append(["商家编码", "货品名称", "skc", "所属店铺", "负责人", "冲突类型", "处理意见"])
            ws.append(["A-001", "红色球衣", "SKC1", "7", "小琴", "平销款冲突爆款", "立即下架"])
            ws.append(["B-001", "蓝色球衣", "SKC2", "8", "洁琳", "平销款冲突爆款", "继续观察"])
            workbook.save(report)

            task_db = root / "operation_tasks.json"
            output_dir = root / "outputs"
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db), \
                 patch.object(daily_ops_app, "OUTPUT_DIR", output_dir):
                sync = daily_ops_app.sync_report_tasks("temu_hot", report)
                self.assertEqual(sync["created"], 2)

                owner = daily_ops_app.login_operator("owner", "小琴", "")
                admin = daily_ops_app.login_operator("admin", "管理员", "")
                owner_headers = {"X-Operator-Token": owner["token"]}
                admin_headers = {"X-Operator-Token": admin["token"]}

                status, _content_type, body = daily_ops_app.handle_tasks_api("GET", owner_headers, {"role": "admin", "user": "洁琳"})
                self.assertEqual(status, 200)
                owner_payload = json.loads(body)
                self.assertEqual(len(owner_payload["tasks"]), 1)
                self.assertEqual(owner_payload["tasks"][0]["owner"], "小琴")

                task_id = owner_payload["tasks"][0]["id"]
                status, _content_type, _body = daily_ops_app.handle_tasks_api(
                    "POST_SUBMIT",
                    owner_headers,
                    {"id": task_id, "action": "已下架", "remark": "后台已处理"},
                )
                self.assertEqual(status, 200)

                status, _content_type, _body = daily_ops_app.handle_tasks_api(
                    "POST_REVIEW",
                    admin_headers,
                    {"id": task_id, "decision": "通过", "remark": "同意"},
                )
                self.assertEqual(status, 200)

                status, _content_type, body = daily_ops_app.handle_tasks_api(
                    "POST_EXPORT",
                    admin_headers,
                    {"role": "owner", "user": "小琴", "next_handler": "管理员"},
                )
                self.assertEqual(status, 200)
                exported = json.loads(body)
                self.assertEqual(exported["rows"], 1)
                self.assertEqual(exported["history_rows"], 3)

                exported_path = output_dir / exported["file"]
                exported_book = load_workbook(exported_path, read_only=True, data_only=True)
                try:
                    self.assertEqual(exported_book.sheetnames, ["任务台账", "操作记录", "负责人汇总", "状态汇总", "管理员待办队列", "导出口径"])
                    self.assertEqual(exported_book["任务台账"].max_row, 2)
                    self.assertEqual(exported_book["操作记录"].max_row, 4)
                    self.assertEqual(exported_book["负责人汇总"].max_row, 2)
                    self.assertEqual(exported_book["状态汇总"].cell(row=2, column=1).value, "任务总数")
                    self.assertEqual(exported_book["状态汇总"].cell(row=2, column=2).value, 1)
                    criteria_ws = exported_book["导出口径"]
                    criteria = {
                        criteria_ws.cell(row=row, column=1).value: criteria_ws.cell(row=row, column=2).value
                        for row in range(2, criteria_ws.max_row + 1)
                    }
                    self.assertEqual(criteria["role"], "owner")
                    self.assertEqual(criteria["user"], "小琴")
                    self.assertEqual(criteria["next_handler"], "管理员")
                    self.assertEqual(criteria["rows"], 1)
                    self.assertEqual(criteria["history_rows"], 3)
                finally:
                    exported_book.close()

    def test_local_web_page_exposes_operator_login(self):
        html = daily_ops_app.HTML_PAGE
        self.assertIn("/api/session/login", html)
        self.assertIn("/api/session/logout", html)
        self.assertIn("/api/owners", html)
        self.assertIn("operatorToken", html)
        self.assertIn("登录身份", html)
        self.assertIn("退出身份", html)
        self.assertIn("logoutOperator", html)
        self.assertIn("ownerOptions", html)
        self.assertIn("店长入口", html)
        self.assertIn("ownerEntryLink", html)
        self.assertIn("applyEntryParams", html)
        self.assertIn("URLSearchParams(window.location.search)", html)
        self.assertIn("role=owner", html)
        self.assertIn("url.search = ''", html)
        self.assertIn("请以店长身份登录", html)
        self.assertIn("taskRole.disabled = operatorSession.role === 'owner'", html)
        self.assertIn("taskRole.disabled = false", html)
        self.assertIn("function clearOperatorSession", html)
        self.assertIn("error.status = r.status", html)
        self.assertIn("if(error.status === 401) clearOperatorSession()", html)
        self.assertIn("正在登录", html)
        self.assertIn("登录失败", html)
        self.assertIn("catch(e)", html)
        self.assertIn("function showTaskError", html)
        self.assertGreaterEqual(html.count("showTaskError(e)"), 6)

    def test_web_download_links_include_operator_token(self):
        html = daily_ops_app.HTML_PAGE
        self.assertIn("function authDownload", html)
        self.assertIn("searchParams.set('token', operatorToken)", html)
        self.assertIn('href="${authDownload(f.download)}"', html)
        self.assertIn('href="${authDownload(res.download)}"', html)

    def test_download_requires_logged_in_operator(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            output_dir = root / "outputs"
            output_dir.mkdir()
            (output_dir / "台账.xlsx").write_bytes(b"demo")
            handler = daily_ops_app.DailyOpsHandler.__new__(daily_ops_app.DailyOpsHandler)
            handler.headers = {}

            with patch.object(daily_ops_app, "OUTPUT_DIR", output_dir):
                with self.assertRaises(PermissionError):
                    handler.handle_download(urlparse("/download?path=%E5%8F%B0%E8%B4%A6.xlsx"))

    def test_owner_download_requires_file_grant(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        owner = daily_ops_app.login_operator("owner", "小琴", "")
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            output_dir = root / "outputs"
            output_dir.mkdir()
            (output_dir / "全部店铺.xlsx").write_bytes(b"demo")
            handler = daily_ops_app.DailyOpsHandler.__new__(daily_ops_app.DailyOpsHandler)
            handler.headers = {}

            with patch.object(daily_ops_app, "OUTPUT_DIR", output_dir):
                with self.assertRaises(PermissionError):
                    handler.handle_download(urlparse(f"/download?path=%E5%85%A8%E9%83%A8%E5%BA%97%E9%93%BA.xlsx&token={owner['token']}"))

    def test_owner_task_export_grants_download_to_that_session(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        daily_ops_app.DOWNLOAD_GRANTS.clear()
        owner = daily_ops_app.login_operator("owner", "小琴", "")
        owner_headers = {"X-Operator-Token": owner["token"]}
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            output_dir = root / "outputs"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "7", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
            ])

            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db), \
                 patch.object(daily_ops_app, "OUTPUT_DIR", output_dir):
                status, _content_type, body = daily_ops_app.handle_tasks_api("POST_EXPORT", owner_headers, {})

            self.assertEqual(status, 200)
            exported = json.loads(body)
            self.assertIn(exported["file"], daily_ops_app.DOWNLOAD_GRANTS[owner["token"]])

    def test_task_export_filename_includes_key_filter_context(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            output_dir = root / "outputs"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "7", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
            ])
            task = store.list_tasks()[0]
            submitted = store.submit_owner_action(task["id"], actor="小琴", action="已下架", remark="后台已下架")
            store.review_task(submitted["id"], admin="管理员", decision="通过", remark="同意")

            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db), \
                 patch.object(daily_ops_app, "OUTPUT_DIR", output_dir):
                exported = daily_ops_app.export_operation_tasks(
                    role="owner",
                    user="小琴",
                    status=daily_ops_tasks.STATUS_APPROVED,
                    platform="Temu",
                    task_type="爆旺冲突",
                )

            self.assertIn("小琴", exported["file"])
            self.assertIn("已通过", exported["file"])
            self.assertIn("Temu", exported["file"])
            self.assertIn("爆旺冲突", exported["file"])

    def test_owner_directory_api_is_available_before_login(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        owner = daily_ops_app.login_operator("owner", "小琴", "")
        admin = daily_ops_app.login_operator("admin", "管理员", "")
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "tasks.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "7", "owner": "小琴", "merchant_code": "A", "source_report": "r", "source_row": 1},
                {"platform": "Shein", "task_type": "爆旺冲突", "store": "琪琪", "owner": "洁琳", "merchant_code": "B", "source_report": "r", "source_row": 2},
            ])
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db):
                status, _content_type, body = daily_ops_app.handle_owners_api()
                payload = json.loads(body)
                self.assertEqual(status, 200)
                self.assertEqual(payload["owners"][0]["owner"], "小琴")
                self.assertNotIn("stores", payload["owners"][0])
                self.assertEqual(payload["owners"][1]["owner"], "洁琳")

                status, _content_type, body = daily_ops_app.handle_owners_api({"X-Operator-Token": owner["token"]})
                payload = json.loads(body)
                self.assertEqual(status, 200)
                self.assertNotIn("stores", payload["owners"][0])

                status, _content_type, body = daily_ops_app.handle_owners_api({"X-Operator-Token": admin["token"]})
                payload = json.loads(body)
                self.assertEqual(status, 200)
                self.assertEqual(payload["owners"][0]["stores"], ["7"])

    def test_backup_exports_only_operational_state_and_can_restore(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            base = root / "基础数据库"
            outputs = root / "outputs"
            temu = root / "temu数据源表"
            base.mkdir()
            outputs.mkdir()
            temu.mkdir()
            (base / "operation_tasks.json").write_text('{"tasks":[{"id":"t1"}]}', encoding="utf-8")
            (base / "report_rules.json").write_text('{"slow_moving":{}}', encoding="utf-8")
            (base / "data_source_manifest.json").write_text('{"categories":{}}', encoding="utf-8")
            (base / "store_owner_map.json").write_text('{"assignments":[{"store":"7","owner":"小琴"}]}', encoding="utf-8")
            (outputs / "should-not-backup.xlsx").write_bytes(b"xlsx")
            (temu / "source.xlsx").write_bytes(b"xlsx")

            with patch.object(daily_ops_app, "ROOT", root), \
                 patch.object(daily_ops_app, "TASK_DB_PATH", base / "operation_tasks.json"), \
                 patch.object(daily_ops_app, "RULES_FILE", base / "report_rules.json"), \
                 patch.object(daily_ops_app, "DATA_SOURCE_MANIFEST", base / "data_source_manifest.json"), \
                 patch.object(daily_ops_app, "STORE_OWNER_MAP_FILE", base / "store_owner_map.json"), \
                 patch.object(daily_ops_app, "OUTPUT_DIR", outputs), \
                 patch.object(daily_ops_app, "TEMU_DIR", temu):
                result = daily_ops_app.create_operational_backup()
                backup_path = Path(result["path"])
                self.assertTrue(backup_path.exists())
                with zipfile.ZipFile(backup_path) as zf:
                    names = set(zf.namelist())
                    self.assertIn("基础数据库/operation_tasks.json", names)
                    self.assertIn("基础数据库/report_rules.json", names)
                    self.assertIn("基础数据库/data_source_manifest.json", names)
                    self.assertIn("基础数据库/store_owner_map.json", names)
                    self.assertIn("temu数据源表/source.xlsx", names)
                    self.assertNotIn("outputs/should-not-backup.xlsx", names)

                (base / "operation_tasks.json").write_text('{"tasks":[]}', encoding="utf-8")
                restored = daily_ops_app.restore_operational_backup(backup_path)
                self.assertIn("基础数据库/operation_tasks.json", restored["restored"])
                self.assertEqual(json.loads((base / "operation_tasks.json").read_text(encoding="utf-8"))["tasks"][0]["id"], "t1")

    def test_source_group_status_tracks_upload_batch_metadata(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source_dir = root / "temu数据源表"
            source_dir.mkdir()
            manifest = root / "data_source_manifest.json"
            source = source_dir / "20260622-Temu销售.xlsx"
            workbook = daily_ops_tasks.Workbook()
            ws = workbook.active
            ws.append(["店铺", "SKC", "销量"])
            ws.append(["7", "SKC1", 10])
            workbook.save(source)

            weekly_groups = {
                "temu_platform": {
                    "name": "Temu 销售表",
                    "description": "Temu 仓库销售情况导出表",
                    "patterns": ["*.xlsx"],
                    "folder": source_dir,
                }
            }
            upload_targets = {"temu_platform": ("Temu平台表", source_dir)}

            with patch.object(daily_ops_app, "DATA_SOURCE_MANIFEST", manifest), \
                 patch.object(daily_ops_app, "WEEKLY_SOURCE_GROUPS", weekly_groups), \
                 patch.object(daily_ops_app, "UPLOAD_TARGETS", upload_targets):
                pending = daily_ops_app.record_uploaded_source("temu_platform", source)
                self.assertTrue(pending["pending"])
                self.assertIn("batch_id", pending)

                pending_status = daily_ops_app.source_group_status()[0]
                self.assertEqual(pending_status["status"], "待结束上传")
                self.assertEqual(pending_status["pending_count"], 1)
                self.assertEqual(pending_status["pending_files"], ["20260622-Temu销售.xlsx"])
                self.assertEqual(pending_status["pending_batch_id"], pending["batch_id"])

                finished = daily_ops_app.finish_upload_batch("temu_platform")
                self.assertEqual(finished["batch_id"], pending["batch_id"])
                self.assertEqual(finished["count"], 1)
                self.assertEqual(finished["rows"], 1)

                active_status = daily_ops_app.source_group_status()[0]
                self.assertEqual(active_status["batch_id"], pending["batch_id"])
                self.assertEqual(active_status["uploaded_at"], finished["uploaded_at"])
                self.assertEqual(active_status["batch_files"], ["20260622-Temu销售.xlsx"])
                self.assertEqual(active_status["total_rows"], 1)

    def test_backup_entrypoints_are_exposed(self):
        html = daily_ops_app.HTML_PAGE
        cli = (Path(__file__).resolve().parent / "daily_ops_cli.py").read_text(encoding="utf-8")
        adapter = (Path(__file__).resolve().parent / "daily_ops_desktop_adapter.py").read_text(encoding="utf-8")
        self.assertIn("/api/backup/create", html)
        self.assertIn("/api/backup/restore", html)
        self.assertIn("生成备份", html)
        self.assertIn("restore-backup", cli)
        self.assertIn("open_only", cli)
        self.assertIn("create_backup", adapter)

    def test_report_generation_surfaces_task_sync_summary(self):
        root = Path(__file__).resolve().parent
        html = daily_ops_app.HTML_PAGE
        renderer = (root / "electron" / "renderer.js").read_text(encoding="utf-8")
        for text in ["taskSyncSummary", "新增任务", "更新任务", "导入明细"]:
            self.assertIn(text, html)
            self.assertIn(text, renderer)
        self.assertIn("result.task_sync", html)
        self.assertIn("result.task_sync", renderer)

    def test_weekly_report_generation_returns_total_task_sync_summary(self):
        seen_reports = []

        def fake_run_report(report_id, _version):
            seen_reports.append(report_id)
            if report_id == "temu_inventory":
                raise ValueError("缺少库存表")
            return {
                "file": f"{report_id}.xlsx",
                "download": f"/download?path={report_id}.xlsx",
                "task_sync": {
                    "created": 2 if report_id == "temu_hot" else 0,
                    "updated": 1 if report_id == "temu_slow" else 0,
                    "imported_rows": 3 if report_id in {"temu_hot", "temu_slow"} else 0,
                },
            }

        with patch.object(daily_ops_app, "run_report", fake_run_report):
            result = daily_ops_app.run_weekly_reports()

        self.assertIn("temu_bargain", seen_reports)
        self.assertEqual(result["summary"]["total"], 9)
        self.assertEqual(result["summary"]["ok"], 8)
        self.assertEqual(result["summary"]["failed"], 1)
        self.assertEqual(result["task_sync"], {"created": 2, "updated": 1, "imported_rows": 6})

    def test_status_summarizes_existing_tasks_by_report(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            task_db = root / "operation_tasks.json"
            store = daily_ops_tasks.OperationTaskStore(task_db)
            store.upsert_generated_tasks([
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "7", "owner": "小琴", "merchant_code": "A", "source_report": "Temu爆旺款重复预警", "source_row": 1},
                {"platform": "Temu", "task_type": "爆旺冲突", "store": "8", "owner": "洁琳", "merchant_code": "B", "source_report": "Temu爆旺款重复预警", "source_row": 2},
                {"platform": "Shein", "task_type": "爆旺冲突", "store": "琪琪", "owner": "洁琳", "merchant_code": "C", "source_report": "Shein爆旺款重复预警", "source_row": 3},
            ])
            with patch.object(daily_ops_app, "TASK_DB_PATH", task_db):
                status = daily_ops_app.data_status()
                self.assertEqual(status["report_tasks"]["temu_hot"]["total"], 2)
                self.assertEqual(status["report_tasks"]["temu_hot"]["by_status"][daily_ops_tasks.STATUS_PENDING_OWNER], 2)
                self.assertEqual(status["report_tasks"]["shein_hot"]["total"], 1)

    def test_status_api_hides_operational_details_without_admin_session(self):
        daily_ops_app.OPERATOR_SESSIONS.clear()
        owner = daily_ops_app.login_operator("owner", "小琴", "")
        admin = daily_ops_app.login_operator("admin", "管理员", "")

        full_status = {
            "version": "v2.0",
            "temu_files": 1,
            "shein_files": 1,
            "erp_files": 1,
            "source_groups": [{"key": "temu_platform", "latest": {"name": "temu.xlsx"}}],
            "outputs": [{"name": "运营任务台账.xlsx"}],
            "database": {"exists": True, "tables": 1, "rows": 2, "path": "/tmp/db.sqlite"},
            "tasks": {"total": 3, "by_owner": {"小琴": 1}, "owner_status": {"小琴": {"total": 1}}},
            "report_tasks": {"temu_hot": {"total": 1}},
            "reports": {"temu_hot": {"name": "Temu爆旺款重复预警"}},
            "rules": {"hot_item": {}},
            "upload_targets": {"temu_platform": "Temu平台表"},
        }
        with patch.object(daily_ops_app, "data_status", return_value=full_status):
            status, _content_type, body = daily_ops_app.handle_status_api({})
            public_payload = json.loads(body)
            self.assertEqual(status, 200)
            self.assertEqual(public_payload["outputs"], [])
            self.assertEqual(public_payload["source_groups"], [])
            self.assertEqual(public_payload["database"]["path"], "")
            self.assertEqual(public_payload["reports"], {})
            self.assertEqual(public_payload["rules"], {})
            self.assertEqual(public_payload["tasks"], {})

            status, _content_type, body = daily_ops_app.handle_status_api({"X-Operator-Token": owner["token"]})
            owner_payload = json.loads(body)
            self.assertEqual(status, 200)
            self.assertEqual(owner_payload["outputs"], [])
            self.assertEqual(owner_payload["source_groups"], [])
            self.assertEqual(owner_payload["tasks"], {})

            status, _content_type, body = daily_ops_app.handle_status_api({"X-Operator-Token": admin["token"]})
            admin_payload = json.loads(body)
            self.assertEqual(status, 200)
            self.assertEqual(admin_payload["outputs"], [{"name": "运营任务台账.xlsx"}])
            self.assertEqual(admin_payload["source_groups"][0]["latest"]["name"], "temu.xlsx")
            self.assertEqual(admin_payload["database"]["path"], "/tmp/db.sqlite")
            self.assertEqual(admin_payload["tasks"]["by_owner"], {"小琴": 1})

    def test_report_queue_surfaces_persisted_task_counts(self):
        root = Path(__file__).resolve().parent
        html = daily_ops_app.HTML_PAGE
        renderer = (root / "electron" / "renderer.js").read_text(encoding="utf-8")
        for text in ["reportTaskSummary", "已生成任务", "report_tasks"]:
            self.assertIn(text, html)
            self.assertIn(text, renderer)

    def test_upload_batch_metadata_is_visible_in_web_and_desktop(self):
        root = Path(__file__).resolve().parent
        html = daily_ops_app.HTML_PAGE
        renderer = (root / "electron" / "renderer.js").read_text(encoding="utf-8")
        for text in ["batch_id", "uploaded_at", "pending_files", "批次", "上传时间"]:
            self.assertIn(text, html)
            self.assertIn(text, renderer)


if __name__ == "__main__":
    unittest.main()
