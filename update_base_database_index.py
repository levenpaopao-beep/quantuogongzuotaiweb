import hashlib
import re
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path.cwd()
DB_PATH = ROOT / "基础数据库" / "project_base_data.sqlite"
INDEX_XLSX = ROOT / "基础数据库" / "数据目录.xlsx"
README = ROOT / "基础数据库" / "README_先查这里.txt"

FOLDERS = [
    {
        "path": ROOT / "shein数据源表",
        "domain": "shein",
        "role": "销售总览数据",
        "prefix": "shein_sales_overview",
        "update_policy": "会更新；新平台导出放入此文件夹后运行索引刷新",
    },
    {
        "path": ROOT / "temu数据源表",
        "domain": "temu",
        "role": "销售总览数据",
        "prefix": "temu_sales_overview",
        "update_policy": "会更新；新平台导出放入此文件夹后运行索引刷新",
    },
    {
        "path": ROOT / "erp数据源",
        "domain": "erp",
        "role": "产品基础数据源",
        "prefix": "erp_base",
        "update_policy": "会更新；ERP产品基础表和组合装表放入此文件夹后运行索引刷新",
    },
]


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def safe_name(text, max_len=96):
    value = re.sub(r"[^\w\u4e00-\u9fff]+", "_", text, flags=re.UNICODE).strip("_")
    value = re.sub(r"_+", "_", value)
    if not value:
        value = "unnamed"
    if len(value) > max_len:
        digest = hashlib.md5(value.encode("utf-8")).hexdigest()[:8]
        value = f"{value[: max_len - 9]}_{digest}"
    return value


def unique_headers(headers):
    result = []
    seen = {}
    for i, h in enumerate(headers, 1):
        name = clean_text(h) or f"空字段_{i}"
        name = name.replace('"', '""')
        base = name
        seen[base] = seen.get(base, 0) + 1
        if seen[base] > 1:
            name = f"{base}_{seen[base]}"
        result.append(name)
    return result


def quote(name):
    return '"' + name.replace('"', '""') + '"'


def file_hash(path):
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def table_name(prefix, file_path, sheet_name):
    stem = safe_name(file_path.stem, 64)
    sheet = safe_name(sheet_name, 28)
    return safe_name(f"{prefix}__{stem}__{sheet}", 120)


def ensure_meta_schema(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS source_folder_index (
            folder_key TEXT PRIMARY KEY,
            folder_path TEXT,
            data_domain TEXT,
            business_role TEXT,
            update_policy TEXT,
            active_flag INTEGER,
            updated_at TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS source_file_index (
            file_key TEXT PRIMARY KEY,
            folder_key TEXT,
            file_name TEXT,
            file_path TEXT,
            data_domain TEXT,
            business_role TEXT,
            file_size INTEGER,
            file_mtime TEXT,
            file_hash TEXT,
            active_flag INTEGER,
            indexed_at TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS hot_item_priority_rules (
            rule_order INTEGER PRIMARY KEY,
            rule_name TEXT,
            applies_to TEXT,
            priority_direction TEXT,
            rule_detail TEXT,
            output_label TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS project_table_index (
            table_name TEXT PRIMARY KEY,
            file_key TEXT,
            file_name TEXT,
            sheet_name TEXT,
            data_domain TEXT,
            business_role TEXT,
            row_count INTEGER,
            column_count INTEGER,
            key_fields TEXT,
            refresh_policy TEXT,
            indexed_at TEXT
        )
        """
    )
    for col, ddl in [
        ("data_domain", "ALTER TABLE data_sources ADD COLUMN data_domain TEXT"),
        ("business_role", "ALTER TABLE data_sources ADD COLUMN business_role TEXT"),
        ("source_folder", "ALTER TABLE data_sources ADD COLUMN source_folder TEXT"),
        ("file_size", "ALTER TABLE data_sources ADD COLUMN file_size INTEGER"),
        ("file_mtime", "ALTER TABLE data_sources ADD COLUMN file_mtime TEXT"),
        ("file_hash", "ALTER TABLE data_sources ADD COLUMN file_hash TEXT"),
        ("active_flag", "ALTER TABLE data_sources ADD COLUMN active_flag INTEGER DEFAULT 1"),
        ("note", "ALTER TABLE data_sources ADD COLUMN note TEXT"),
    ]:
        cols = [r[1] for r in cur.execute("PRAGMA table_info(data_sources)").fetchall()]
        if col not in cols:
            cur.execute(ddl)


def write_rules(cur):
    rules = [
        (1, "同款识别", "爆旺款重复铺货/互相冲突", "先分组", "同款按货品编码一致识别；在SKU层面可由商家编码去掉尺码后缀得到货品编码。涉及爆旺款时必须先按同款分组，再比较优先级。", "同款组"),
        (2, "老品主指标", "上架天数>=30天的爆旺款", "数值越高越优先", "优先保护30天销量更高、30天日均更高的SKC；这是稳定贡献指标。", "高优先级"),
        (3, "新品主指标", "上架天数<30天的爆旺款", "数值越高越优先", "新品优先看7天日均；7天日均更高说明正在爆发，应优先保护。", "高优先级"),
        (4, "新老同款冲突", "新品爆旺款 vs 老品爆旺款", "比较爆发与稳定贡献", "若新品7天日均明显高于老品30天日均，优先保护新品；否则优先保护30天销量更高的老品。", "按指标胜出"),
        (5, "销量接近时", "销量差距不明显的同款爆旺款", "价格越高越优先", "销量接近时，优先保护申报价/供货价更高且未明显影响销量的SKC，避免低价款抢量。", "中高优先级"),
        (6, "库存约束", "重复款仍有平台备货", "库存越高越谨慎", "重复款平台仓库存高时优先调价/控价/自然消化；库存低时可优先下架或停止补货。库存不改变保护对象，只影响处理动作。", "处理方式"),
        (7, "同店冲突", "相同店铺内同款重复", "优先处理", "同店铺内爆旺互冲优先级高于跨店铺冲突，因为更容易互相抢量，也更容易被平台识别。", "高处理优先级"),
        (8, "最终口径", "所有爆旺款核查", "必须输出", "每次涉及爆旺款，都要输出爆旺优先级：高=主指标胜出；中=销量接近但价格/库存更优；低=销量较低、低价或适合作为重复款处理。", "爆旺优先级"),
    ]
    cur.execute("DELETE FROM hot_item_priority_rules")
    cur.executemany(
        """
        INSERT INTO hot_item_priority_rules
        (rule_order, rule_name, applies_to, priority_direction, rule_detail, output_label)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        rules,
    )


def import_sheet(cur, file_conf, path, file_key, file_hash_value):
    wb = load_workbook(path, read_only=False, data_only=True)
    imported = []
    for ws in wb.worksheets:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row < 1 or max_col < 1:
            continue
        headers = unique_headers([ws.cell(1, c).value for c in range(1, max_col + 1)])
        tname = table_name(file_conf["prefix"], path, ws.title)
        cur.execute(f"DROP TABLE IF EXISTS {quote(tname)}")
        col_defs = ", ".join([f"{quote(h)} TEXT" for h in headers])
        cur.execute(f"CREATE TABLE {quote(tname)} (_source_row INTEGER, {col_defs})")
        placeholders = ", ".join(["?"] * (max_col + 1))
        insert_sql = f"INSERT INTO {quote(tname)} VALUES ({placeholders})"
        batch = []
        for r in range(2, max_row + 1):
            row = [r]
            empty = True
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if v is not None and str(v).strip() != "":
                    empty = False
                row.append("" if v is None else str(v))
            if not empty:
                batch.append(row)
            if len(batch) >= 1000:
                cur.executemany(insert_sql, batch)
                batch.clear()
        if batch:
            cur.executemany(insert_sql, batch)
        row_count = cur.execute(f"SELECT COUNT(*) FROM {quote(tname)}").fetchone()[0]
        source_path = str(path)
        indexed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cur.execute("DELETE FROM data_sources WHERE table_name = ?", (tname,))
        cur.execute(
            """
            INSERT INTO data_sources
            (file_name, sheet_name, table_name, row_count, column_count, imported_at, source_path,
             data_domain, business_role, source_folder, file_size, file_mtime, file_hash, active_flag, note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
            """,
            (
                path.name,
                ws.title,
                tname,
                row_count,
                max_col,
                indexed_at,
                source_path,
                file_conf["domain"],
                file_conf["role"],
                str(file_conf["path"]),
                path.stat().st_size,
                datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                file_hash_value,
                file_conf["update_policy"],
            ),
        )
        source_id = cur.lastrowid
        cur.execute("DELETE FROM column_catalog WHERE source_id = ?", (source_id,))
        cur.executemany(
            "INSERT INTO column_catalog (source_id, column_order, column_name) VALUES (?, ?, ?)",
            [(source_id, i + 1, h) for i, h in enumerate(headers)],
        )
        cur.execute("DELETE FROM project_table_index WHERE table_name = ?", (tname,))
        key_fields = infer_key_fields(file_conf["domain"], headers)
        cur.execute(
            """
            INSERT INTO project_table_index
            (table_name, file_key, file_name, sheet_name, data_domain, business_role, row_count,
             column_count, key_fields, refresh_policy, indexed_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                tname,
                file_key,
                path.name,
                ws.title,
                file_conf["domain"],
                file_conf["role"],
                row_count,
                max_col,
                key_fields,
                file_conf["update_policy"],
                indexed_at,
            ),
        )
        imported.append((tname, row_count, max_col))
    wb.close()
    return imported


def infer_key_fields(domain, headers):
    wanted = {
        "shein": ["商家SKU", "SKC", "SPU", "近7天销量", "近30天销量", "SHEIN仓库存", "供货价", "采购价", "上架天数"],
        "temu": ["商家SKU", "SKU", "SKC", "SPU", "近7天销量", "近30天销量", "仓库", "库存", "申报价", "供货价", "采购价"],
        "erp": ["商家编码（新）", "商家编码", "货品编码", "货品名称", "规格名称", "成本价", "批发报价", "可用库存"],
    }
    found = [h for h in wanted.get(domain, []) if h in headers]
    return "、".join(found)


def build_search_index(cur):
    con = cur.connection
    read_cur = con.cursor()
    write_cur = con.cursor()
    cur.execute("DROP TABLE IF EXISTS search_index")
    cur.execute(
        """
        CREATE VIRTUAL TABLE search_index USING fts5(
            source_id UNINDEXED,
            file_name,
            sheet_name,
            table_name,
            source_row UNINDEXED,
            content
        )
        """
    )
    sources = read_cur.execute(
        """
        SELECT id, file_name, sheet_name, table_name
        FROM data_sources
        WHERE active_flag = 1
        """
    ).fetchall()
    for source_id, file_name, sheet_name, tname in sources:
        if not read_cur.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (tname,)).fetchone():
            continue
        cols = [r[1] for r in read_cur.execute(f"PRAGMA table_info({quote(tname)})").fetchall() if r[1] != "_source_row"]
        if not cols:
            continue
        col_expr = " || ' ' || ".join([f"ifnull({quote(c)}, '')" for c in cols[:60]])
        sql = f"SELECT _source_row, {col_expr} AS content FROM {quote(tname)}"
        batch = []
        for source_row, content in read_cur.execute(sql):
            if content and str(content).strip():
                batch.append((source_id, file_name, sheet_name, tname, source_row, content))
            if len(batch) >= 2000:
                write_cur.executemany("INSERT INTO search_index VALUES (?, ?, ?, ?, ?, ?)", batch)
                batch.clear()
        if batch:
            write_cur.executemany("INSERT INTO search_index VALUES (?, ?, ?, ?, ?, ?)", batch)


def update_views(cur):
    cur.execute("DROP VIEW IF EXISTS v_data_sources")
    cur.execute(
        """
        CREATE VIEW v_data_sources AS
        SELECT id, data_domain, business_role, file_name, sheet_name, table_name, row_count,
               column_count, imported_at, source_path, source_folder, active_flag, note
        FROM data_sources
        ORDER BY data_domain, business_role, file_name, sheet_name
        """
    )
    cur.execute("DROP VIEW IF EXISTS v_column_catalog")
    cur.execute(
        """
        CREATE VIEW v_column_catalog AS
        SELECT ds.data_domain, ds.business_role, ds.file_name, ds.sheet_name, ds.table_name,
               cc.column_order, cc.column_name
        FROM column_catalog cc
        JOIN data_sources ds ON ds.id = cc.source_id
        ORDER BY ds.data_domain, ds.file_name, ds.sheet_name, cc.column_order
        """
    )


def write_readme(imported_total):
    text = f"""本项目基础数据库

位置：{DB_PATH}

优先查找规则：
以后本项目需要 ERP 产品基础信息、组合装、Temu/SHEIN 平台销售总览、预警表、店铺表、爆旺款等基础数据时，优先查这里。

三个基础数据源文件夹：
1. shein数据源表：SHEIN 店铺销售总览数据，会更新。
2. temu数据源表：Temu 销售/库存/爆旺相关基础数据，会更新。
3. erp数据源：ERP 产品基础信息和组合装基础信息，会更新。

主要文件：
1. project_base_data.sqlite：本地数据库，已导入并索引三个基础数据源文件夹内的工作表。
2. 数据目录.xlsx：可直接打开查看数据源文件、工作表、数据库表名、行数、字段、更新规则。

常用查询入口：
- v_data_sources：查看有哪些数据表、来源文件、业务分类和行数。
- v_column_catalog：查看每张表的字段。
- project_table_index：规范表索引，包含关键字段和刷新规则。
- source_file_index：源文件索引，包含文件路径、大小、更新时间和哈希。
- hot_item_priority_rules：爆旺款优先级规则。每次涉及爆旺款，都要按这里分优先级。
- search_index：全文搜索入口，保留来源文件、工作表、数据库表名和原始行号。

爆旺款优先级最基础规则：
同款按货品编码一致识别。若同款在不同或相同店铺均为爆旺款，先比较老品30天销量/30天日均，新品比较7天日均；销量接近时优先保护价格更高且销量未受影响的款；库存高的重复款优先调价或控价，库存低的重复款优先下架或停止补货；同店铺冲突优先处理。

生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
本次规范索引导入表数：{imported_total}
"""
    README.write_text(text, encoding="utf-8")


def build_catalog_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheets = [
        ("总览", "SELECT data_domain, business_role, COUNT(*), SUM(row_count) FROM project_table_index GROUP BY data_domain, business_role ORDER BY data_domain, business_role", ["数据域", "业务角色", "表数量", "数据行数"]),
        ("文件索引", "SELECT data_domain, business_role, file_name, file_path, file_size, file_mtime, active_flag FROM source_file_index ORDER BY data_domain, file_name", ["数据域", "业务角色", "文件名", "路径", "大小", "更新时间", "启用"]),
        ("表索引", "SELECT data_domain, business_role, file_name, sheet_name, table_name, row_count, column_count, key_fields, refresh_policy FROM project_table_index ORDER BY data_domain, file_name, sheet_name", ["数据域", "业务角色", "文件名", "工作表", "数据库表名", "行数", "列数", "关键字段", "刷新规则"]),
        ("字段索引", "SELECT data_domain, business_role, file_name, sheet_name, table_name, column_order, column_name FROM v_column_catalog ORDER BY data_domain, file_name, sheet_name, column_order", ["数据域", "业务角色", "文件名", "工作表", "数据库表名", "字段序号", "字段名"]),
        ("爆旺优先级规则", "SELECT rule_order, rule_name, applies_to, priority_direction, rule_detail, output_label FROM hot_item_priority_rules ORDER BY rule_order", ["顺序", "规则", "适用场景", "优先方向", "规则明细", "输出标签"]),
    ]
    for title, sql, headers in sheets:
        ws = wb.create_sheet(title)
        ws.append(headers)
        for row in cur.execute(sql):
            ws.append(list(row))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            values = [str(ws.cell(r, col).value or "") for r in range(1, min(ws.max_row, 200) + 1)]
            width = min(max(max(len(v) for v in values) + 2, 10), 48)
            ws.column_dimensions[get_column_letter(col)].width = width
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    con.close()
    wb.save(INDEX_XLSX)


def main():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    ensure_meta_schema(cur)
    write_rules(cur)
    imported_total = 0
    indexed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for conf in FOLDERS:
        folder_key = safe_name(conf["path"].name)
        cur.execute(
            """
            INSERT OR REPLACE INTO source_folder_index
            (folder_key, folder_path, data_domain, business_role, update_policy, active_flag, updated_at)
            VALUES (?, ?, ?, ?, ?, 1, ?)
            """,
            (folder_key, str(conf["path"]), conf["domain"], conf["role"], conf["update_policy"], indexed_at),
        )
        for path in sorted(conf["path"].glob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            h = file_hash(path)
            file_key = hashlib.md5(str(path).encode("utf-8")).hexdigest()
            cur.execute(
                """
                INSERT OR REPLACE INTO source_file_index
                (file_key, folder_key, file_name, file_path, data_domain, business_role, file_size,
                 file_mtime, file_hash, active_flag, indexed_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
                """,
                (
                    file_key,
                    folder_key,
                    path.name,
                    str(path),
                    conf["domain"],
                    conf["role"],
                    path.stat().st_size,
                    datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                    h,
                    indexed_at,
                ),
            )
            imported = import_sheet(cur, conf, path, file_key, h)
            imported_total += len(imported)
            con.commit()

    update_views(cur)
    build_search_index(cur)
    con.commit()
    con.close()
    build_catalog_xlsx()
    write_readme(imported_total)
    print(f"updated tables={imported_total}")


if __name__ == "__main__":
    main()
