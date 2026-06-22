import copy
import os
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import update_shein_summary_30d_skc as xlsx


ROOT = Path(os.environ.get("DAILYWORK_ROOT", Path(__file__).resolve().parent))
SHEIN_DIR = ROOT / "shein数据源表"
ERP_DIR = ROOT / "erp数据源"
OUTPUT = ROOT / "SHEIN仓备库存异常清单_20260603.xlsx"
ERP_FILES = None
SHEIN_FILES = None

EXCLUDED_PURCHASE_STORES = {"加加", "宝宝"}
STORE_ORDER = ["琪琪", "童话", "牛牛"]
ALL_STORE_NAMES = STORE_ORDER + sorted(EXCLUDED_PURCHASE_STORES)
OWNERS = {
    "琪琪": "胡娟",
    "童话": "胡娟",
    "牛牛": "胡娟",
    "加加": "洁琳",
    "宝宝": "洁琳",
}
SIZE_ORDER = {"XXS": 0, "XS": 1, "S": 2, "M": 3, "MD": 4, "L": 5, "XL": 6, "XXL": 7}


def norm(value):
    return xlsx.norm_text(value)


def num(value):
    return xlsx.to_number(value)


def sku_key(value):
    text = norm(value).upper()
    if "@" in text:
        text = text.split("@", 1)[0]
    return text.strip()


def size_rank(code):
    text = norm(code).upper()
    parts = re.split(r"[-_\s]+", text)
    for part in reversed(parts):
        if part in SIZE_ORDER:
            return SIZE_ORDER[part]
    for size, rank in SIZE_ORDER.items():
        if text.endswith("-" + size) or text.endswith("_" + size):
            return rank
    return 99


def store_from_file(path):
    for store in ALL_STORE_NAMES:
        if store in path.stem:
            return store
    return ""


def latest_store_files():
    if SHEIN_FILES:
        return [Path(path) for path in SHEIN_FILES if store_from_file(Path(path)) not in EXCLUDED_PURCHASE_STORES]
    files = []
    for store in STORE_ORDER:
        candidates = sorted(
            [path for path in SHEIN_DIR.glob("*.xlsx") if store in path.stem],
            key=lambda path: (path.stat().st_mtime, path.name),
        )
        if candidates:
            files.append(candidates[-1])
    return files


def header_map(row):
    return {norm(v): i for i, v in enumerate(row) if norm(v)}


def is_active_listing(row, headers):
    listing = xlsx.norm_text(cell(row, headers, "上架状态"))
    return listing == "已上架"


def cell(row, headers, *names):
    for name in names:
        if name in headers and headers[name] < len(row):
            return row[headers[name]]
    return None


def summarize_source_files(files):
    summary = {
        store: {"active_skc": set(), "active_skc_count": 0}
        for store in STORE_ORDER
    }
    for path in files:
        path = Path(path)
        store = store_from_file(path)
        if store not in summary:
            continue
        rows = xlsx.read_xlsx_rows(path)
        if not rows:
            continue
        hm = header_map(rows[0])
        skc_i = hm.get("SKC")
        if skc_i is None:
            continue
        for row in rows[1:]:
            skc = norm(row[skc_i] if skc_i < len(row) else "")
            if skc and is_active_listing(row, hm):
                summary[store]["active_skc"].add(skc)
    for item in summary.values():
        item["active_skc_count"] = len(item["active_skc"])
    return summary


def load_erp_base():
    records = {}
    source_rows = 0
    for path in (ERP_FILES or sorted(ERP_DIR.glob("erp产品基础信息表*.xlsx"))):
        rows = xlsx.read_xlsx_rows(path)
        if not rows:
            continue
        hm = header_map(rows[0])
        required = ["商家编码（新）", "货品名称", "规格名称"]
        if not all(k in hm for k in required):
            continue
        for row in rows[1:]:
            source_rows += 1
            code = norm(row[hm["商家编码（新）"]] if hm["商家编码（新）"] < len(row) else "")
            key = sku_key(code)
            if not key or key in records:
                continue
            records[key] = {
                "商家编码": code,
                "货品名称": norm(row[hm["货品名称"]] if hm["货品名称"] < len(row) else ""),
                "货品规格": norm(row[hm["规格名称"]] if hm["规格名称"] < len(row) else ""),
                "来源": path.name,
            }
    return records, source_rows


def load_combo_base(records):
    path = ERP_DIR / "erp产品组合装基础信息表.xlsx"
    combo_rows = 0
    if not path.exists():
        return combo_rows
    rows = xlsx.read_xlsx_rows(path)
    if not rows:
        return combo_rows
    hm = header_map(rows[0])
    code_i = hm.get("商家编码（新）")
    name_i = hm.get("组合装名称")
    spec_i = hm.get("组合装简称")
    if code_i is None:
        return combo_rows
    for row in rows[1:]:
        combo_rows += 1
        code = norm(row[code_i] if code_i < len(row) else "")
        key = sku_key(code)
        if not key or key in records:
            continue
        records[key] = {
            "商家编码": code,
            "货品名称": norm(row[name_i] if name_i is not None and name_i < len(row) else ""),
            "货品规格": norm(row[spec_i] if spec_i is not None and spec_i < len(row) else ""),
            "来源": path.name,
        }
    return combo_rows


def read_shein(records):
    summary = {
        store: {
            "skc": set(),
            "hot_skc": set(),
            "gt_2x_skc": set(),
            "gt_1x_skc": set(),
            "rows": 0,
        }
        for store in STORE_ORDER
    }
    gt_2x = []
    gt_1x = []
    skipped_unmatched = 0
    source_rows = 0

    for path in latest_store_files():
        store = store_from_file(path)
        if store not in summary:
            continue
        rows = xlsx.read_xlsx_rows(path)
        if not rows:
            continue
        hm = header_map(rows[0])
        skc_i = hm.get("SKC")
        sku_i = hm.get("商家SKU")
        stock_i = hm.get("SHEIN仓库存")
        sales30_i = hm.get("近30天销量")
        sales7_i = hm.get("近7天销量")
        tag_i = hm.get("商品标签")
        if None in [skc_i, sku_i, stock_i, sales30_i, sales7_i]:
            raise RuntimeError(f"{path.name} 缺少 SKC/商家SKU/SHEIN仓库存/近30天销量/近7天销量 字段")

        for ridx, row in enumerate(rows[1:], start=2):
            source_rows += 1
            summary[store]["rows"] += 1
            skc = norm(row[skc_i] if skc_i < len(row) else "")
            sku = norm(row[sku_i] if sku_i < len(row) else "")
            stock = num(row[stock_i] if stock_i < len(row) else "")
            sales30 = num(row[sales30_i] if sales30_i < len(row) else "")
            sales7 = num(row[sales7_i] if sales7_i < len(row) else "")
            tag = norm(row[tag_i] if tag_i is not None and tag_i < len(row) else "")

            if not is_active_listing(row, hm):
                continue
            if skc:
                summary[store]["skc"].add(skc)
            if skc and any(word in tag for word in ["高销款", "爆", "旺"]):
                summary[store]["hot_skc"].add(skc)

            erp = records.get(sku_key(sku))
            if not erp:
                skipped_unmatched += 1
                continue

            out = {
                "店铺": store,
                "SKC": skc,
                "商家编码": erp["商家编码"],
                "货品名称": erp["货品名称"],
                "货品规格": erp["货品规格"],
                "仓备可用": stock,
                "30天销量": sales30,
                "7天销量": sales7,
                "负责人": OWNERS.get(store, ""),
                "源SKU": sku,
                "源文件": path.name,
                "源行": ridx,
            }

            if stock > sales30 * 2:
                gt_2x.append(out)
                if skc:
                    summary[store]["gt_2x_skc"].add(skc)
            if stock > sales30:
                gt_1x.append(out)
                if skc:
                    summary[store]["gt_1x_skc"].add(skc)

    return summary, gt_2x, gt_1x, source_rows, skipped_unmatched


def sort_rows(rows):
    store_rank = {s: i for i, s in enumerate(STORE_ORDER)}
    return sorted(
        rows,
        key=lambda r: (
            store_rank.get(r["店铺"], 99),
            norm(r["SKC"]),
            re.sub(r"[-_\s]+(XXS|XS|S|M|MD|L|XL|XXL)$", "", norm(r["商家编码"]).upper()),
            size_rank(r["商家编码"]),
            norm(r["商家编码"]).upper(),
        ),
    )


def clone_style(src, dst):
    if src.has_style:
        dst._style = copy.copy(src._style)
    dst.font = copy.copy(src.font)
    dst.fill = copy.copy(src.fill)
    dst.border = copy.copy(src.border)
    dst.alignment = copy.copy(src.alignment)
    dst.number_format = src.number_format


def setup_sheet(ws, title, headers, fill_color):
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="44546A")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor=fill_color)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"


def style_table(ws):
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")
    for col in range(1, ws.max_column + 1):
        max_len = 8
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 34)


def write_detail(wb, sheet_name, title, rows, fill_color):
    headers = ["店铺", "SKC", "商家编码", "货品名称", "货品规格", "仓备可用", "30天销量", "7天销量", "负责人", "源SKU", "源文件", "源行"]
    ws = wb.create_sheet(sheet_name)
    setup_sheet(ws, title, headers, fill_color)
    for r in sort_rows(rows):
        ws.append([r[h] for h in headers])
    for row in ws.iter_rows(min_row=3, min_col=6, max_col=8):
        for cell in row:
            cell.number_format = "0"
    style_table(ws)


def build_workbook(summary, gt_2x, gt_1x, source_rows, erp_rows, combo_rows, skipped):
    wb = Workbook()
    ws = wb.active
    ws.title = "总览表"
    headers = ["店铺", "负责人", "在售SKC数量", "爆旺款数量", "仓备大于30天销量2倍SKC数", "仓备大于30天销量SKC数"]
    setup_sheet(ws, "SHEIN仓备库存异常清单", headers, "5B9BD5")
    for store in STORE_ORDER:
        s = summary[store]
        ws.append([
            store,
            OWNERS.get(store, ""),
            len(s["skc"]),
            len(s["hot_skc"]),
            len(s["gt_2x_skc"]),
            len(s["gt_1x_skc"]),
        ])
    style_table(ws)

    write_detail(wb, "大于30天销量2倍", "Shein仓库可用库存 > 30天销量 * 2", gt_2x, "C00000")
    write_detail(wb, "大于30天销量", "Shein仓库可用库存 > 30天销量", gt_1x, "ED7D31")

    check = wb.create_sheet("数据校验")
    check.append(["项目", "数值"])
    check.append(["Shein源明细行数", source_rows])
    check.append(["ERP基础表读取行数", erp_rows])
    check.append(["ERP组合装读取行数", combo_rows])
    check.append(["未匹配ERP/组合装跳过行数", skipped])
    check.append(["大于30天销量2倍明细行数", len(gt_2x)])
    check.append(["大于30天销量明细行数", len(gt_1x)])
    check.append(["剔除店铺", "加加、宝宝"])
    check.append(["库存口径", "SHEIN仓库存"])
    check.append(["爆旺款口径", "商品标签包含高销款/爆/旺的唯一SKC数"])
    check.append(["排序规则", "店铺、SKC、商家编码尺码（XXS-XS-S-M-MD-L-XL-XXL）"])
    style_table(check)

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    wb.save(OUTPUT)


if __name__ == "__main__":
    erp_records, erp_rows = load_erp_base()
    combo_rows = load_combo_base(erp_records)
    summary, gt_2x, gt_1x, source_rows, skipped = read_shein(erp_records)
    build_workbook(summary, gt_2x, gt_1x, source_rows, erp_rows, combo_rows, skipped)
    print(OUTPUT)
    print("gt_2x", len(gt_2x))
    print("gt_1x", len(gt_1x))
    print("skipped", skipped)
    for store in STORE_ORDER:
        print(store, len(summary[store]["skc"]), len(summary[store]["hot_skc"]), len(summary[store]["gt_2x_skc"]), len(summary[store]["gt_1x_skc"]))
