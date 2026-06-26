import hashlib
import json
import os
import threading
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


STATUS_PENDING_PUSH = "待推送"
STATUS_PENDING_OWNER = "待店长处理"
STATUS_OWNER_SUBMITTED = "店长已填写"
STATUS_PENDING_REVIEW = "待管理员审核"
STATUS_APPROVED = "已通过"
STATUS_REJECTED = "已驳回"
STATUS_DONE = "已完成"
OWNER_OVERDUE_DAYS = 3
REVIEW_OVERDUE_DAYS = 1

TASK_COLUMNS = [
    ("id", "任务ID"),
    ("platform", "平台"),
    ("task_type", "任务类型"),
    ("status", "任务状态"),
    ("next_handler", "下一步处理人"),
    ("next_action", "下一步动作"),
    ("priority", "处理优先级"),
    ("priority_reason", "优先级原因"),
    ("store", "店铺"),
    ("owner", "负责人"),
    ("merchant_code", "商家编码"),
    ("skc", "SKC"),
    ("spu", "SPU"),
    ("product_name", "货品名称"),
    ("system_action", "系统建议动作"),
    ("task_detail", "任务详情"),
    ("owner_action", "店长处理动作"),
    ("owner_remark", "店长备注"),
    ("owner_proof", "店长处理凭证"),
    ("owner_submitted_by", "店长提交人"),
    ("owner_submitted_at", "店长提交时间"),
    ("admin_decision", "管理员审核结果"),
    ("admin_remark", "管理员备注"),
    ("rejection_count", "驳回次数"),
    ("last_rejection_reason", "最近驳回原因"),
    ("admin_reviewed_by", "管理员审核人"),
    ("admin_reviewed_at", "管理员审核时间"),
    ("completed_by", "完成确认人"),
    ("completed_at", "完成时间"),
    ("completed_remark", "完成说明"),
    ("source_report", "来源报表"),
    ("source_file", "来源文件"),
    ("source_sheet", "来源页签"),
    ("source_row", "来源行"),
    ("source_batch_id", "任务生成批次"),
    ("is_overdue", "是否超时"),
    ("overdue_days", "超时天数"),
    ("created_at", "创建时间"),
    ("updated_at", "更新时间"),
]

TASK_HISTORY_COLUMNS = [
    ("task_id", "任务ID"),
    ("platform", "平台"),
    ("task_type", "任务类型"),
    ("status", "当前状态"),
    ("next_handler", "下一步处理人"),
    ("next_action", "下一步动作"),
    ("store", "店铺"),
    ("owner", "负责人"),
    ("product_name", "货品名称"),
    ("time", "操作时间"),
    ("event", "事件"),
    ("actor", "操作人"),
    ("action", "动作"),
    ("remark", "备注"),
    ("proof", "处理凭证"),
    ("status_after", "动作后状态"),
    ("next_handler_after", "动作后下一步处理人"),
    ("next_action_after", "动作后下一步动作"),
]

TASK_UPDATE_LABELS = dict(TASK_COLUMNS)
TASK_STORE_LOCKS = {}
TASK_STORE_LOCKS_GUARD = threading.Lock()


def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def norm(value):
    if value is None:
        return ""
    return str(value).strip()


def task_store_lock(path):
    key = str(Path(path).resolve())
    with TASK_STORE_LOCKS_GUARD:
        if key not in TASK_STORE_LOCKS:
            TASK_STORE_LOCKS[key] = threading.RLock()
        return TASK_STORE_LOCKS[key]


def parse_time(value):
    text = norm(value)
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def task_overdue(row, now=None):
    now = now or datetime.now()
    status = norm(row.get("status"))
    if status == STATUS_PENDING_PUSH:
        return False
    if status == STATUS_PENDING_OWNER:
        start = parse_time(row.get("created_at")) or parse_time(row.get("updated_at"))
        return bool(start and (now - start).total_seconds() >= OWNER_OVERDUE_DAYS * 86400)
    if status == STATUS_PENDING_REVIEW:
        start = parse_time(row.get("owner_submitted_at")) or parse_time(row.get("updated_at"))
        return bool(start and (now - start).total_seconds() >= REVIEW_OVERDUE_DAYS * 86400)
    if status == STATUS_REJECTED:
        start = parse_time(row.get("admin_reviewed_at")) or parse_time(row.get("updated_at"))
        return bool(start and (now - start).total_seconds() >= OWNER_OVERDUE_DAYS * 86400)
    return False


def task_age_days(row, now=None):
    now = now or datetime.now()
    status = norm(row.get("status"))
    if status == STATUS_PENDING_REVIEW:
        start = parse_time(row.get("owner_submitted_at")) or parse_time(row.get("updated_at"))
    elif status == STATUS_REJECTED:
        start = parse_time(row.get("admin_reviewed_at")) or parse_time(row.get("updated_at"))
    else:
        start = parse_time(row.get("created_at")) or parse_time(row.get("updated_at"))
    if not start:
        return ""
    return max(0, int((now - start).total_seconds() // 86400))


def task_identity(row):
    task_type = norm(row.get("task_type"))
    if task_type == "价格异常":
        parts = [
            norm(row.get("platform")),
            task_type,
            norm(row.get("store")),
            norm(row.get("skc")) or norm(row.get("product_name")) or norm(row.get("merchant_code")),
            norm(row.get("product_name")),
            price_action_bucket(row.get("system_action") or row.get("source_sheet")),
        ]
    else:
        parts = [
            norm(row.get("platform")),
            task_type,
            norm(row.get("store")),
            norm(row.get("merchant_code")),
            norm(row.get("skc")),
            norm(row.get("spu")),
            norm(row.get("source_report")),
            norm(row.get("source_file")),
            norm(row.get("source_sheet")),
            norm(row.get("source_row")),
        ]
    raw = "|".join(parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def merge_unique_text(*values, separator="、"):
    result = []
    seen = set()
    for value in values:
        for part in str(value or "").replace(",", separator).split(separator):
            item = norm(part)
            if item and item not in seen:
                seen.add(item)
                result.append(item)
    return separator.join(result)


def price_action_bucket(value):
    text = norm(value)
    if "低于成本" in text or "亏损" in text:
        return "低于成本价"
    if "低于批发价80" in text or "批发价80%" in text:
        return "低于批发价80%"
    return text


def task_next_step(row, now=None):
    status = norm(row.get("status"))
    if status == STATUS_PENDING_PUSH:
        if not norm(row.get("owner")):
            return "管理员", "指派负责人后推送"
        return "管理员", "确认推送给店长"
    if status == STATUS_PENDING_OWNER:
        if not norm(row.get("owner")):
            return "管理员", "指派负责人"
        if task_overdue(row, now):
            return "管理员", "跟进超时店长处理"
        return "店长", "填写处理结果"
    if status == STATUS_PENDING_REVIEW:
        if task_overdue(row, now):
            return "管理员", "处理超时审核"
        return "管理员", "审核通过或驳回"
    if status == STATUS_REJECTED:
        if task_overdue(row, now):
            return "管理员", "跟进驳回返工超时"
        return "店长", "按驳回原因重新处理"
    if status == STATUS_APPROVED:
        return "管理员", "标记完成或归档"
    if status == STATUS_DONE:
        return "无需处理", "已完成"
    return "管理员", "确认任务状态"


def task_priority(row, now=None):
    status = norm(row.get("status"))
    if status == STATUS_DONE:
        return "低", "已完成"
    task_type = norm(row.get("task_type"))
    business_text = " ".join(
        norm(row.get(key))
        for key in ["task_type", "system_action", "task_detail", "source_report", "source_sheet"]
    )
    if task_type == "价格异常" and ("低于成本" in business_text or "亏损" in business_text):
        return "高", "低于成本价亏损销售"
    if task_type == "爆旺冲突":
        return "高", "爆旺冲突抢占资源"
    if task_type == "低分预警":
        return "高", "低分产品增多"
    if task_type == "价格异常" and ("低于批发价80" in business_text or "批发价80%" in business_text):
        return "中", "低于80%申报价在售"
    if task_type == "滞销处理":
        return "中", "滞销品催下架"
    return "低", "其他低级处理"


def task_sort_key(row):
    priority_order = {"高": 0, "中": 1, "低": 2}
    updated_at = parse_time(row.get("updated_at")) or parse_time(row.get("created_at"))
    timestamp = updated_at.timestamp() if updated_at else 0
    product_name = norm(row.get("product_name")) or norm(row.get("merchant_code")) or norm(row.get("skc")) or norm(row.get("spu"))
    return (
        priority_order.get(norm(row.get("priority")), 9),
        product_name,
        norm(row.get("platform")),
        norm(row.get("store")),
        norm(row.get("task_type")),
        -timestamp,
        norm(row.get("id")),
    )


def task_package_id(row):
    parts = [
        norm(row.get("owner")) or "未分配",
        norm(row.get("platform")),
        norm(row.get("store")),
        norm(row.get("task_type")),
        norm(row.get("system_action")) or norm(row.get("next_action")),
    ]
    return hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()[:16]


def task_package_sort_key(package):
    priority_order = {"高": 0, "中": 1, "低": 2}
    status_order = {
        STATUS_PENDING_REVIEW: 0,
        STATUS_PENDING_PUSH: 1,
        STATUS_REJECTED: 2,
        STATUS_PENDING_OWNER: 3,
        STATUS_APPROVED: 4,
        STATUS_DONE: 5,
    }
    updated_at = parse_time(package.get("updated_at")) or parse_time(package.get("created_at"))
    timestamp = updated_at.timestamp() if updated_at else 0
    return (
        priority_order.get(norm(package.get("priority")), 9),
        status_order.get(norm(package.get("main_status")), 9),
        -int(package.get("total") or 0),
        -timestamp,
        norm(package.get("owner")),
        norm(package.get("store")),
        norm(package.get("task_type")),
        norm(package.get("system_action")),
    )


def admin_queue_summary(rows, now=None):
    specs = [
        ("assign_owner", "指派负责人", "高", {"unassigned": "1", "open_only": "1"}),
        ("review_overdue", "处理超时审核", "高", {"status": STATUS_PENDING_REVIEW, "overdue": "1", "open_only": "1"}),
        ("owner_overdue", "跟进超时店长处理", "高", {"status": STATUS_PENDING_OWNER, "overdue": "1", "open_only": "1"}),
        ("rejected_overdue", "跟进驳回返工超时", "高", {"status": STATUS_REJECTED, "overdue": "1", "open_only": "1"}),
        ("review_pending", "审核通过或驳回", "中", {"status": STATUS_PENDING_REVIEW, "open_only": "1"}),
        ("push_pending", "确认推送给店长", "中", {"status": STATUS_PENDING_PUSH, "open_only": "1"}),
        ("mark_done", "标记完成或归档", "中", {"status": STATUS_APPROVED, "open_only": "1"}),
    ]
    counts = {key: 0 for key, _action, _priority, _filters in specs}
    for row in rows:
        status = norm(row.get("status"))
        if status == STATUS_DONE:
            continue
        overdue = task_overdue(row, now)
        if status in {STATUS_PENDING_PUSH, STATUS_PENDING_OWNER} and not norm(row.get("owner")):
            counts["assign_owner"] += 1
        elif status == STATUS_PENDING_REVIEW and overdue:
            counts["review_overdue"] += 1
        elif status == STATUS_PENDING_OWNER and overdue:
            counts["owner_overdue"] += 1
        elif status == STATUS_REJECTED and overdue:
            counts["rejected_overdue"] += 1
        elif status == STATUS_PENDING_REVIEW:
            counts["review_pending"] += 1
        elif status == STATUS_PENDING_PUSH:
            counts["push_pending"] += 1
        elif status == STATUS_APPROVED:
            counts["mark_done"] += 1
    return [
        {"key": key, "action": action, "priority": priority, "count": counts[key], "filters": dict(filters)}
        for key, action, priority, filters in specs
        if counts[key] > 0
    ]


def task_filter_text(filters):
    return "; ".join(f"{key}={value}" for key, value in (filters or {}).items() if norm(value))


def task_rejection_info(row):
    rejection_count = 0
    last_reason = ""
    for item in row.get("history") or []:
        if norm(item.get("event")) in {"管理员审核", "管理员批量审核"} and norm(item.get("action")) == "驳回":
            rejection_count += 1
            last_reason = norm(item.get("remark")) or last_reason
    return rejection_count, last_reason


def public_task(row, now=None):
    item = dict(row)
    item.setdefault("history", [])
    next_handler, next_action = task_next_step(item, now=now)
    priority, priority_reason = task_priority(item, now=now)
    rejection_count, last_rejection_reason = task_rejection_info(item)
    item["next_handler"] = next_handler
    item["next_action"] = next_action
    item["priority"] = priority
    item["priority_reason"] = priority_reason
    item["rejection_count"] = rejection_count
    item["last_rejection_reason"] = last_rejection_reason
    return item


def can_update_generated_owner(task):
    if task.get("status") not in {STATUS_PENDING_PUSH, STATUS_PENDING_OWNER}:
        return False
    for item in task.get("history") or []:
        if norm(item.get("event")) in {"任务指派", "自动指派"}:
            return False
    return True


def generated_task_remark(row):
    parts = []
    for key, label in [
        ("source_report", "来源报表"),
        ("source_file", "来源文件"),
        ("source_sheet", "来源页签"),
        ("source_row", "来源行"),
        ("source_batch_id", "任务生成批次"),
    ]:
        value = norm(row.get(key))
        if value:
            parts.append(f"{label}：{value}")
    owner_source = norm(row.get("owner_source"))
    if owner_source:
        parts.append(f"负责人来源：{owner_source}")
    return "；".join(parts)


def task_source_batch_label(row):
    parts = [
        norm(row.get("source_report")),
        norm(row.get("source_file")),
        norm(row.get("source_sheet")),
        norm(row.get("source_batch_id")),
    ]
    parts = [part for part in parts if part]
    return " / ".join(parts) or "未填写"


def history_entry(task, actor, event, action, remark="", time="", proof=""):
    next_handler, next_action = task_next_step(task)
    return {
        "time": time or now_text(),
        "actor": norm(actor),
        "event": norm(event),
        "action": norm(action),
        "remark": norm(remark),
        "proof": norm(proof),
        "status_after": norm(task.get("status")),
        "next_handler_after": next_handler,
        "next_action_after": next_action,
    }


class OperationTaskStore:
    def __init__(self, path):
        self.path = Path(path)
        self._lock = task_store_lock(self.path)

    def load(self):
        with self._lock:
            if not self.path.exists():
                return {"tasks": []}
            try:
                payload = json.loads(self.path.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError):
                return {"tasks": []}
            if not isinstance(payload, dict):
                return {"tasks": []}
            tasks = payload.get("tasks")
            return {"tasks": tasks if isinstance(tasks, list) else []}

    def save(self, payload):
        with self._lock:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            tmp = self.path.with_name(f".{self.path.name}.{threading.get_ident()}.tmp")
            tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            os.replace(tmp, self.path)

    def list_tasks(self, role="admin", user="", status="", task_type="", store="", platform="", overdue="", unassigned="", next_handler="", priority="", reworked="", open_only="", now=None):
        role = norm(role) or "admin"
        user = norm(user)
        rows = [public_task(row, now=now) for row in self.load()["tasks"]]
        if role != "admin":
            if not user:
                return []
            rows = [row for row in rows if norm(row.get("owner")) == user]
            rows = [row for row in rows if norm(row.get("status")) != STATUS_PENDING_PUSH]
        if status:
            rows = [row for row in rows if norm(row.get("status")) == norm(status)]
        if task_type:
            rows = [row for row in rows if norm(row.get("task_type")) == norm(task_type)]
        if store:
            rows = [row for row in rows if norm(row.get("store")) == norm(store)]
        if platform:
            rows = [row for row in rows if norm(row.get("platform")) == norm(platform)]
        if norm(overdue) in {"1", "true", "是", "超时"}:
            rows = [row for row in rows if task_overdue(row, now)]
        if norm(unassigned) in {"1", "true", "是", "未分配"}:
            rows = [row for row in rows if not norm(row.get("owner"))]
        if next_handler:
            rows = [row for row in rows if norm(row.get("next_handler")) == norm(next_handler)]
        if priority:
            rows = [row for row in rows if norm(row.get("priority")) == norm(priority)]
        if norm(reworked) in {"1", "true", "是", "返工"}:
            rows = [row for row in rows if int(row.get("rejection_count") or 0) > 0]
        if norm(open_only) in {"1", "true", "是", "未完成", "待办"}:
            rows = [row for row in rows if norm(row.get("status")) != STATUS_DONE]
            if role != "admin":
                rows = [row for row in rows if norm(row.get("status")) != STATUS_APPROVED]
        return sorted(rows, key=task_sort_key)

    def summary(self, rows=None, now=None):
        rows = list(rows) if rows is not None else self.load()["tasks"]
        by_status = {}
        by_type = {}
        by_owner = {}
        by_next_handler = {}
        by_next_action = {}
        owner_status = {}
        overdue = {"total": 0, "by_status": {}}
        unassigned = 0
        for row in rows:
            status = norm(row.get("status"))
            by_status[status] = by_status.get(status, 0) + 1
            by_type[norm(row.get("task_type"))] = by_type.get(norm(row.get("task_type")), 0) + 1
            next_handler, next_action = task_next_step(row, now=now)
            by_next_handler[next_handler] = by_next_handler.get(next_handler, 0) + 1
            by_next_action[next_action] = by_next_action.get(next_action, 0) + 1
            owner = norm(row.get("owner"))
            if owner:
                by_owner[owner] = by_owner.get(owner, 0) + 1
            else:
                owner = "未分配"
                unassigned += 1
            item = owner_status.setdefault(owner, {"owner": owner, "total": 0, "by_status": {}, "reworked": 0})
            item["total"] += 1
            item["by_status"][status] = item["by_status"].get(status, 0) + 1
            item.setdefault("overdue", 0)
            if task_rejection_info(row)[0] > 0:
                item["reworked"] += 1
            if task_overdue(row, now):
                overdue["total"] += 1
                overdue["by_status"][status] = overdue["by_status"].get(status, 0) + 1
                item["overdue"] += 1
        return {
            "total": len(rows),
            "by_status": by_status,
            "by_type": by_type,
            "by_owner": by_owner,
            "by_next_handler": by_next_handler,
            "by_next_action": by_next_action,
            "owner_status": owner_status,
            "overdue": overdue,
            "unassigned": unassigned,
            "admin_queue": admin_queue_summary(rows, now=now),
        }

    def task_packages(self, rows=None, now=None):
        rows = list(rows) if rows is not None else self.list_tasks(now=now)
        grouped = {}
        for row in rows:
            item = grouped.setdefault(task_package_id(row), {
                "id": task_package_id(row),
                "owner": norm(row.get("owner")) or "未分配",
                "platform": norm(row.get("platform")),
                "store": norm(row.get("store")),
                "task_type": norm(row.get("task_type")),
                "system_action": norm(row.get("system_action")) or norm(row.get("next_action")),
                "total": 0,
                "task_ids": [],
                "pushable_task_ids": [],
                "submittable_task_ids": [],
                "reviewable_task_ids": [],
                "done_task_ids": [],
                "by_status": {},
                "source_reports": set(),
                "source_files": set(),
                "sample_tasks": [],
                "created_at": "",
                "updated_at": "",
                "overdue_count": 0,
                "reworked_count": 0,
                "priority": "低",
                "priority_reason": "",
                "main_status": "",
                "next_handler": "",
                "next_action": "",
            })
            status = norm(row.get("status"))
            item["total"] += 1
            item["task_ids"].append(row.get("id", ""))
            item["by_status"][status] = item["by_status"].get(status, 0) + 1
            if row.get("source_report"):
                item["source_reports"].add(norm(row.get("source_report")))
            if row.get("source_file"):
                item["source_files"].add(norm(row.get("source_file")))
            if len(item["sample_tasks"]) < 8:
                item["sample_tasks"].append(row)
            if status == STATUS_PENDING_PUSH and norm(row.get("owner")):
                item["pushable_task_ids"].append(row.get("id", ""))
            if status in {STATUS_PENDING_OWNER, STATUS_REJECTED} and norm(row.get("owner")):
                item["submittable_task_ids"].append(row.get("id", ""))
            if status == STATUS_PENDING_REVIEW:
                item["reviewable_task_ids"].append(row.get("id", ""))
            if status == STATUS_APPROVED:
                item["done_task_ids"].append(row.get("id", ""))
            if task_overdue(row, now):
                item["overdue_count"] += 1
            if task_rejection_info(row)[0] > 0:
                item["reworked_count"] += 1
            created = norm(row.get("created_at"))
            updated = norm(row.get("updated_at"))
            if created and (not item["created_at"] or created < item["created_at"]):
                item["created_at"] = created
            if updated and updated > item["updated_at"]:
                item["updated_at"] = updated
            priority, priority_reason = task_priority(row, now=now)
            if task_package_sort_key({"priority": priority, "main_status": status, "total": item["total"], "updated_at": updated}) < task_package_sort_key(item):
                item["priority"] = priority
                item["priority_reason"] = priority_reason
            if not item["main_status"] or item["by_status"].get(status, 0) > item["by_status"].get(item["main_status"], 0):
                item["main_status"] = status
            next_handler, next_action = task_next_step(row, now=now)
            if not item["next_handler"] or next_handler == "管理员":
                item["next_handler"] = next_handler
                item["next_action"] = next_action

        result = []
        for item in grouped.values():
            item["pending_push_count"] = item["by_status"].get(STATUS_PENDING_PUSH, 0)
            item["pending_owner_count"] = item["by_status"].get(STATUS_PENDING_OWNER, 0) + item["by_status"].get(STATUS_REJECTED, 0)
            item["pending_review_count"] = item["by_status"].get(STATUS_PENDING_REVIEW, 0)
            item["approved_count"] = item["by_status"].get(STATUS_APPROVED, 0)
            item["done_count"] = item["by_status"].get(STATUS_DONE, 0)
            item["source_reports"] = sorted(item["source_reports"])
            item["source_files"] = sorted(item["source_files"])
            item["task_ids"] = [task_id for task_id in item["task_ids"] if task_id]
            item["pushable_task_ids"] = [task_id for task_id in item.get("pushable_task_ids", []) if task_id]
            item["submittable_task_ids"] = [task_id for task_id in item["submittable_task_ids"] if task_id]
            item["reviewable_task_ids"] = [task_id for task_id in item["reviewable_task_ids"] if task_id]
            item["done_task_ids"] = [task_id for task_id in item["done_task_ids"] if task_id]
            result.append(item)
        return sorted(result, key=task_package_sort_key)

    def owner_directory(self):
        owners = {}
        for row in self.load()["tasks"]:
            owner = norm(row.get("owner"))
            if not owner:
                continue
            item = owners.setdefault(owner, {"owner": owner, "stores": set(), "platforms": set(), "task_count": 0})
            store = norm(row.get("store"))
            platform = norm(row.get("platform"))
            if store:
                item["stores"].add(store)
            if platform:
                item["platforms"].add(platform)
            item["task_count"] += 1
        result = []
        for item in owners.values():
            result.append({
                "owner": item["owner"],
                "stores": sorted(item["stores"]),
                "platforms": sorted(item["platforms"]),
                "task_count": item["task_count"],
            })
        return sorted(result, key=lambda row: (-row["task_count"], row["owner"]))

    def upsert_generated_tasks(self, rows, default_status=STATUS_PENDING_OWNER, replace_source_report=""):
        payload = self.load()
        tasks = payload["tasks"]
        existing = {row.get("id"): row for row in tasks}
        created = 0
        updated = 0
        archived = 0
        timestamp = now_text()
        current_ids = set()
        reset_source_ids = set()
        for source in rows:
            row = {key: norm(value) for key, value in dict(source).items()}
            row_id = task_identity(row)
            current_ids.add(row_id)
            if row_id in existing:
                task = existing[row_id]
                if norm(task.get("status")) == STATUS_DONE:
                    continue
                changed_labels = []
                for key in [
                    "platform",
                    "task_type",
                    "store",
                    "merchant_code",
                    "skc",
                    "spu",
                    "product_name",
                    "system_action",
                    "task_detail",
                    "source_report",
                    "source_file",
                    "source_sheet",
                    "source_row",
                    "source_batch_id",
                ]:
                    next_value = row.get(key, task.get(key, ""))
                    if key in {"merchant_code", "source_file", "source_row", "source_batch_id"}:
                        if replace_source_report and row_id not in reset_source_ids:
                            task[key] = ""
                        next_value = merge_unique_text(task.get(key, ""), next_value)
                    if norm(task.get(key)) != norm(next_value):
                        changed_labels.append(TASK_UPDATE_LABELS.get(key, key))
                    task[key] = next_value
                reset_source_ids.add(row_id)
                if row.get("owner") and can_update_generated_owner(task):
                    if norm(task.get("owner")) != row.get("owner", ""):
                        changed_labels.append(TASK_UPDATE_LABELS.get("owner", "负责人"))
                    task["owner"] = row.get("owner", "")
                if changed_labels:
                    task.setdefault("history", []).append(history_entry(task, "系统", "系统更新", "更新任务明细", "更新字段：" + "、".join(changed_labels), time=timestamp))
                task["updated_at"] = timestamp
                updated += 1
            else:
                task = {
                    "id": row_id,
                    "platform": row.get("platform", ""),
                    "task_type": row.get("task_type", ""),
                    "status": row.get("status") or default_status,
                    "store": row.get("store", ""),
                    "owner": row.get("owner", ""),
                    "merchant_code": row.get("merchant_code", ""),
                    "skc": row.get("skc", ""),
                    "spu": row.get("spu", ""),
                    "product_name": row.get("product_name", ""),
                    "system_action": row.get("system_action", ""),
                    "task_detail": row.get("task_detail", ""),
                    "owner_action": "",
                    "owner_remark": "",
                    "owner_proof": "",
                    "owner_submitted_by": "",
                    "owner_submitted_at": "",
                    "admin_decision": "",
                    "admin_remark": "",
                    "admin_reviewed_by": "",
                    "admin_reviewed_at": "",
                    "source_report": row.get("source_report", ""),
                    "source_file": row.get("source_file", ""),
                    "source_sheet": row.get("source_sheet", ""),
                    "source_row": row.get("source_row", ""),
                    "source_batch_id": row.get("source_batch_id", ""),
                    "history": [],
                    "created_at": timestamp,
                    "updated_at": timestamp,
                }
                task["history"].append(history_entry(task, "系统", "系统生成", "生成待处理任务", generated_task_remark(row), time=timestamp))
                tasks.append(task)
                existing[row_id] = task
                created += 1
        replace_source_report = norm(replace_source_report)
        if replace_source_report:
            for task in tasks:
                if norm(task.get("source_report")) != replace_source_report:
                    continue
                if norm(task.get("id")) in current_ids:
                    continue
                if norm(task.get("status")) == STATUS_DONE:
                    continue
                task["status"] = STATUS_DONE
                task["completed_by"] = "系统"
                task["completed_at"] = timestamp
                task["completed_remark"] = "最新报表重算后未再出现，自动归档"
                task["updated_at"] = timestamp
                task.setdefault("history", []).append(history_entry(
                    task,
                    "系统",
                    "最新报表重算",
                    "自动归档",
                    "最新报表重算后未再出现，退出当前待处理区",
                    time=timestamp,
                ))
                archived += 1
        self.save(payload)
        return {"created": created, "updated": updated, "archived": archived, "total": len(tasks)}

    def push_tasks(self, task_ids, actor="管理员", remark=""):
        ids = []
        seen = set()
        for task_id in task_ids or []:
            task_id = norm(task_id)
            if task_id and task_id not in seen:
                ids.append(task_id)
                seen.add(task_id)
        if not ids:
            raise ValueError("请选择要推送给店长的任务")
        payload = self.load()
        by_id = {row.get("id"): row for row in payload["tasks"]}
        tasks = []
        for task_id in ids:
            task = by_id.get(task_id)
            if not task:
                raise KeyError("任务不存在")
            if norm(task.get("status")) != STATUS_PENDING_PUSH:
                raise ValueError("只有待推送任务可以推送给店长")
            if not norm(task.get("owner")):
                raise ValueError("待推送任务必须先指派负责人")
            tasks.append(task)
        timestamp = now_text()
        for task in tasks:
            task["status"] = STATUS_PENDING_OWNER
            task["next_handler"] = "店长"
            task["next_action"] = "填写处理结果"
            task["updated_at"] = timestamp
            task.setdefault("history", []).append(history_entry(task, actor, "管理员推送", "推送给店长", norm(remark), time=timestamp))
        self.save(payload)
        return {"count": len(tasks), "tasks": [public_task(task) for task in tasks]}

    def require_task(self, task_id):
        payload = self.load()
        for row in payload["tasks"]:
            if row.get("id") == task_id:
                return payload, row
        raise KeyError("任务不存在")

    def assign_task(self, task_id, actor, owner, remark=""):
        payload, task = self.require_task(task_id)
        owner = norm(owner)
        if not owner:
            raise ValueError("负责人不能为空")
        if norm(task.get("status")) == STATUS_DONE:
            raise ValueError("已完成任务不能重新指派")
        timestamp = now_text()
        previous_owner = norm(task.get("owner"))
        task["owner"] = owner
        task["updated_at"] = timestamp
        task.setdefault("history", []).append(history_entry(task, actor, "任务指派", f"指派给 {owner}", norm(remark) or (f"原负责人：{previous_owner}" if previous_owner else ""), time=timestamp))
        self.save(payload)
        return public_task(task)

    def submit_owner_action(self, task_id, actor, action, remark="", proof=""):
        payload, task = self.require_task(task_id)
        action = norm(action)
        if not action:
            raise ValueError("店长处理动作不能为空")
        remark = norm(remark)
        proof = norm(proof)
        if not remark and not proof:
            raise ValueError("店长提交必须填写处理依据：备注或处理凭证至少填一个")
        owner = norm(task.get("owner"))
        actor = norm(actor)
        if not owner:
            raise ValueError("未分配负责人任务不能填写处理结果，请先指派负责人")
        if actor != owner:
            raise ValueError("只能由任务负责人填写处理结果")
        if task.get("status") not in {STATUS_PENDING_OWNER, STATUS_REJECTED}:
            raise ValueError("只有待店长处理或已驳回的任务可以由店长填写")
        timestamp = now_text()
        task["owner_action"] = action
        task["owner_remark"] = remark
        task["owner_proof"] = proof
        task["owner_submitted_by"] = actor
        task["owner_submitted_at"] = timestamp
        task["status"] = STATUS_PENDING_REVIEW
        task["admin_decision"] = ""
        task["admin_remark"] = ""
        task["admin_reviewed_by"] = ""
        task["admin_reviewed_at"] = ""
        task["updated_at"] = timestamp
        task.setdefault("history", []).append(history_entry(task, actor, "店长提交", action, remark, time=timestamp, proof=proof))
        self.save(payload)
        return public_task(task)

    def submit_owner_actions(self, task_ids, actor, action, remark="", proof=""):
        ids = []
        seen = set()
        for task_id in task_ids or []:
            task_id = norm(task_id)
            if task_id and task_id not in seen:
                ids.append(task_id)
                seen.add(task_id)
        if not ids:
            raise ValueError("请选择要批量处理的任务")
        action = norm(action)
        if not action:
            raise ValueError("店长处理动作不能为空")
        remark = norm(remark)
        proof = norm(proof)
        if not remark and not proof:
            raise ValueError("店长提交必须填写处理依据：备注或处理凭证至少填一个")
        actor = norm(actor)
        payload = self.load()
        by_id = {row.get("id"): row for row in payload["tasks"]}
        tasks = []
        for task_id in ids:
            task = by_id.get(task_id)
            if not task:
                raise KeyError("任务不存在")
            owner = norm(task.get("owner"))
            if not owner:
                raise ValueError("未分配负责人任务不能填写处理结果，请先指派负责人")
            if actor != owner:
                raise ValueError("只能由任务负责人填写处理结果")
            if task.get("status") not in {STATUS_PENDING_OWNER, STATUS_REJECTED}:
                raise ValueError("只有待店长处理或已驳回的任务可以由店长填写")
            tasks.append(task)
        timestamp = now_text()
        for task in tasks:
            task["owner_action"] = action
            task["owner_remark"] = remark
            task["owner_proof"] = proof
            task["owner_submitted_by"] = actor
            task["owner_submitted_at"] = timestamp
            task["status"] = STATUS_PENDING_REVIEW
            task["admin_decision"] = ""
            task["admin_remark"] = ""
            task["admin_reviewed_by"] = ""
            task["admin_reviewed_at"] = ""
            task["updated_at"] = timestamp
            task.setdefault("history", []).append(history_entry(task, actor, "店长批量提交", action, remark, time=timestamp, proof=proof))
        self.save(payload)
        return {"count": len(tasks), "tasks": [public_task(task) for task in tasks]}

    def review_task(self, task_id, admin, decision, remark=""):
        payload, task = self.require_task(task_id)
        decision = norm(decision)
        if task.get("status") != STATUS_PENDING_REVIEW:
            raise ValueError("只有待管理员审核的任务可以审核")
        if decision not in {"通过", "驳回"}:
            raise ValueError("管理员审核结果只能是通过或驳回")
        remark = norm(remark)
        if not remark:
            raise ValueError("管理员审核必须填写说明")
        timestamp = now_text()
        task["admin_decision"] = decision
        task["admin_remark"] = remark
        task["admin_reviewed_by"] = norm(admin)
        task["admin_reviewed_at"] = timestamp
        task["status"] = STATUS_APPROVED if decision == "通过" else STATUS_REJECTED
        task["updated_at"] = timestamp
        task.setdefault("history", []).append(history_entry(task, admin, "管理员审核", decision, remark, time=timestamp))
        self.save(payload)
        return public_task(task)

    def review_tasks(self, task_ids, admin, decision, remark=""):
        ids = []
        seen = set()
        for task_id in task_ids or []:
            task_id = norm(task_id)
            if task_id and task_id not in seen:
                ids.append(task_id)
                seen.add(task_id)
        if not ids:
            raise ValueError("请选择要批量审核的任务")
        decision = norm(decision)
        if decision not in {"通过", "驳回"}:
            raise ValueError("管理员审核结果只能是通过或驳回")
        remark = norm(remark)
        if not remark:
            raise ValueError("批量审核必须填写说明")
        payload = self.load()
        by_id = {row.get("id"): row for row in payload["tasks"]}
        tasks = []
        for task_id in ids:
            task = by_id.get(task_id)
            if not task:
                raise KeyError("任务不存在")
            if task.get("status") != STATUS_PENDING_REVIEW:
                raise ValueError("只有待管理员审核的任务可以审核")
            tasks.append(task)
        timestamp = now_text()
        for task in tasks:
            task["admin_decision"] = decision
            task["admin_remark"] = norm(remark)
            task["admin_reviewed_by"] = norm(admin)
            task["admin_reviewed_at"] = timestamp
            task["status"] = STATUS_APPROVED if decision == "通过" else STATUS_REJECTED
            task["updated_at"] = timestamp
            task.setdefault("history", []).append(history_entry(task, admin, "管理员批量审核", decision, remark, time=timestamp))
        self.save(payload)
        return {"count": len(tasks), "tasks": [public_task(task) for task in tasks]}

    def confirm_review_tasks(self, task_ids, admin, remark=""):
        ids = []
        seen = set()
        for task_id in task_ids or []:
            task_id = norm(task_id)
            if task_id and task_id not in seen:
                ids.append(task_id)
                seen.add(task_id)
        if not ids:
            raise ValueError("请选择要确认完成的任务")
        remark = norm(remark) or "管理员确认店长已处理"
        payload = self.load()
        by_id = {row.get("id"): row for row in payload["tasks"]}
        tasks = []
        for task_id in ids:
            task = by_id.get(task_id)
            if not task:
                raise KeyError("任务不存在")
            if task.get("status") != STATUS_PENDING_REVIEW:
                raise ValueError("只有待管理员审核的任务可以确认完成")
            tasks.append(task)
        timestamp = now_text()
        for task in tasks:
            task["admin_decision"] = "确认完成"
            task["admin_remark"] = remark
            task["admin_reviewed_by"] = norm(admin)
            task["admin_reviewed_at"] = timestamp
            task["status"] = STATUS_DONE
            task["completed_by"] = norm(admin)
            task["completed_at"] = timestamp
            task["completed_remark"] = remark
            task["updated_at"] = timestamp
            task.setdefault("history", []).append(history_entry(task, admin, "管理员确认", "确认完成", remark, time=timestamp))
        self.save(payload)
        return {"count": len(tasks), "tasks": [public_task(task) for task in tasks]}

    def mark_done(self, task_id, actor, remark=""):
        payload, task = self.require_task(task_id)
        if task.get("status") != STATUS_APPROVED:
            raise ValueError("只有已通过的任务可以标记完成")
        if not norm(remark):
            raise ValueError("标记完成必须填写确认说明")
        timestamp = now_text()
        task["status"] = STATUS_DONE
        task["completed_by"] = norm(actor)
        task["completed_at"] = timestamp
        task["completed_remark"] = norm(remark)
        task["updated_at"] = timestamp
        task.setdefault("history", []).append(history_entry(task, actor, "标记完成", STATUS_DONE, remark, time=timestamp))
        self.save(payload)
        return public_task(task)

    def mark_done_tasks(self, task_ids, actor, remark=""):
        ids = []
        seen = set()
        for task_id in task_ids or []:
            task_id = norm(task_id)
            if task_id and task_id not in seen:
                ids.append(task_id)
                seen.add(task_id)
        if not ids:
            raise ValueError("请选择要标记完成的任务")
        remark = norm(remark)
        if not remark:
            raise ValueError("标记完成必须填写确认说明")
        payload = self.load()
        by_id = {row.get("id"): row for row in payload["tasks"]}
        rows = []
        for task_id in ids:
            task = by_id.get(task_id)
            if not task:
                raise KeyError("任务不存在")
            if task.get("status") != STATUS_APPROVED:
                raise ValueError("只有已通过的任务可以标记完成")
            rows.append(task)
        timestamp = now_text()
        actor = norm(actor)
        for task in rows:
            task["status"] = STATUS_DONE
            task["completed_by"] = actor
            task["completed_at"] = timestamp
            task["completed_remark"] = remark
            task["updated_at"] = timestamp
            task.setdefault("history", []).append(history_entry(task, actor, "批量标记完成", STATUS_DONE, remark, time=timestamp))
        self.save(payload)
        return {"count": len(rows), "tasks": [public_task(task) for task in rows]}

    def export_tasks(self, output_path, tasks=None, filters=None, now=None):
        rows = list(tasks) if tasks is not None else self.list_tasks()
        filters = dict(filters or {})
        history_rows = sum(len(row.get("history") or []) for row in rows)
        summary = self.summary(rows, now=now)
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        ws = workbook.active
        ws.title = "任务台账"
        ws.append([label for _key, label in TASK_COLUMNS])
        for row in rows:
            export_row = dict(row)
            export_row["is_overdue"] = "是" if task_overdue(row, now) else "否"
            export_row["overdue_days"] = task_age_days(row, now)
            export_row["next_handler"], export_row["next_action"] = task_next_step(row, now=now)
            export_row["priority"], export_row["priority_reason"] = task_priority(row, now=now)
            export_row["rejection_count"], export_row["last_rejection_reason"] = task_rejection_info(row)
            ws.append([export_row.get(key, "") for key, _label in TASK_COLUMNS])
        style_task_sheet(ws)
        log_ws = workbook.create_sheet("操作记录")
        log_ws.append([label for _key, label in TASK_HISTORY_COLUMNS])
        for row in rows:
            next_handler, next_action = task_next_step(row, now=now)
            for item in row.get("history") or []:
                log_ws.append([
                    row.get("id", ""),
                    row.get("platform", ""),
                    row.get("task_type", ""),
                    row.get("status", ""),
                    next_handler,
                    next_action,
                    row.get("store", ""),
                    row.get("owner", ""),
                    row.get("product_name", ""),
                    item.get("time", ""),
                    item.get("event", ""),
                    item.get("actor", ""),
                    item.get("action", ""),
                    item.get("remark", ""),
                    item.get("proof", ""),
                    item.get("status_after", row.get("status", "")),
                    item.get("next_handler_after", next_handler),
                    item.get("next_action_after", next_action),
                ])
        style_task_sheet(log_ws)
        owner_ws = workbook.create_sheet("负责人汇总")
        owner_ws.append(["负责人", "任务总数", STATUS_PENDING_PUSH, STATUS_PENDING_OWNER, STATUS_PENDING_REVIEW, "超时未处理", "返工任务", STATUS_APPROVED, STATUS_REJECTED, STATUS_DONE])
        owner_rows = sorted(summary.get("owner_status", {}).values(), key=lambda item: (-item.get("total", 0), item.get("owner", "")))
        for item in owner_rows:
            status = item.get("by_status", {})
            owner_ws.append([
                item.get("owner", ""),
                item.get("total", 0),
                status.get(STATUS_PENDING_PUSH, 0),
                status.get(STATUS_PENDING_OWNER, 0),
                status.get(STATUS_PENDING_REVIEW, 0),
                item.get("overdue", 0),
                item.get("reworked", 0),
                status.get(STATUS_APPROVED, 0),
                status.get(STATUS_REJECTED, 0),
                status.get(STATUS_DONE, 0),
            ])
        style_task_sheet(owner_ws)
        summary_ws = workbook.create_sheet("状态汇总")
        summary_ws.append(["指标", "数量"])
        summary_ws.append(["任务总数", summary.get("total", 0)])
        summary_ws.append(["超时未处理", summary.get("overdue", {}).get("total", 0)])
        summary_ws.append(["未分配", summary.get("unassigned", 0)])
        for status in [STATUS_PENDING_PUSH, STATUS_PENDING_OWNER, STATUS_PENDING_REVIEW, STATUS_APPROVED, STATUS_REJECTED, STATUS_DONE]:
            summary_ws.append([status, summary.get("by_status", {}).get(status, 0)])
        summary_ws.append(["", ""])
        summary_ws.append(["任务类型", "数量"])
        for task_type, count in sorted(summary.get("by_type", {}).items()):
            summary_ws.append([task_type or "未填写", count])
        summary_ws.append(["", ""])
        summary_ws.append(["下一步处理人", "数量"])
        for handler, count in sorted(summary.get("by_next_handler", {}).items()):
            summary_ws.append([f"下一步处理人：{handler or '未填写'}", count])
        summary_ws.append(["", ""])
        summary_ws.append(["下一步动作", "数量"])
        for action, count in sorted(summary.get("by_next_action", {}).items()):
            summary_ws.append([f"下一步动作：{action or '未填写'}", count])
        source_batches = {}
        for row in rows:
            label = task_source_batch_label(row)
            source_batches[label] = source_batches.get(label, 0) + 1
        summary_ws.append(["", ""])
        summary_ws.append(["来源批次", "数量"])
        for label, count in sorted(source_batches.items()):
            summary_ws.append([f"来源批次：{label}", count])
        style_task_sheet(summary_ws)
        queue_ws = workbook.create_sheet("管理员待办队列")
        queue_ws.append(["处理动作", "优先级", "任务数量", "筛选条件"])
        for item in summary.get("admin_queue", []):
            queue_ws.append([
                item.get("action", ""),
                item.get("priority", ""),
                item.get("count", 0),
                task_filter_text(item.get("filters", {})),
            ])
        style_task_sheet(queue_ws)
        criteria_ws = workbook.create_sheet("导出口径")
        criteria_ws.append(["字段", "值"])
        for key in ["role", "user", "status", "task_type", "store", "platform", "overdue", "unassigned", "next_handler", "priority", "reworked", "open_only"]:
            criteria_ws.append([key, norm(filters.get(key, ""))])
        criteria_ws.append(["rows", len(rows)])
        criteria_ws.append(["history_rows", history_rows])
        criteria_ws.append(["exported_at", now_text()])
        style_task_sheet(criteria_ws)
        workbook.save(output_path)
        return output_path


def locked_task_store_method(method):
    def wrapper(self, *args, **kwargs):
        with self._lock:
            return method(self, *args, **kwargs)
    return wrapper


for _method_name in [
    "upsert_generated_tasks",
    "push_tasks",
    "assign_task",
    "submit_owner_action",
    "review_task",
    "review_tasks",
    "mark_done",
]:
    setattr(OperationTaskStore, _method_name, locked_task_store_method(getattr(OperationTaskStore, _method_name)))


def style_task_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        width = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                width = max(width, len(str(value)) + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(width, 34)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False


def header_map(values):
    return {norm(value): index for index, value in enumerate(values) if norm(value)}


def row_value(row, headers, *names):
    for name in names:
        index = headers.get(name)
        if index is not None and index < len(row):
            return row[index]
    return ""


def report_task_detail(report_id, row, headers):
    parts = []
    seen = set()
    for name in REPORT_DETAIL_FIELDS.get(report_id, []):
        if name in seen:
            continue
        seen.add(name)
        value = norm(row_value(row, headers, name))
        if value:
            parts.append(f"{name}：{value}")
    return "；".join(parts)


REPORT_TASK_TYPE = {
    "temu_price": "价格异常",
    "temu_inventory": "库存异常",
    "temu_hot": "爆旺冲突",
    "shein_hot": "爆旺冲突",
    "shein_price": "价格异常",
    "shein_inventory": "库存异常",
    "low_score_warning": "低分预警",
    "temu_slow": "滞销处理",
    "temu_bargain": "议价审核",
}

REPORT_PLATFORM = {
    "temu_price": "Temu",
    "temu_inventory": "Temu",
    "temu_hot": "Temu",
    "low_score_warning": "Temu",
    "temu_slow": "Temu",
    "temu_bargain": "Temu",
    "shein_price": "Shein",
    "shein_inventory": "Shein",
    "shein_hot": "Shein",
}

REPORT_DETAIL_FIELDS = {
    "temu_price": ["申报价", "成本价", "批发价", "批发价80%", "7天销量", "30天销量"],
    "shein_price": ["申报价", "成本价", "批发价", "批发价80%", "7天销量", "30天销量"],
    "temu_inventory": ["仓备可用", "30天销量", "7天销量", "触发规则"],
    "shein_inventory": ["仓备可用", "30天销量", "7天销量", "触发规则"],
    "temu_hot": ["冲突类型", "爆旺款skc", "爆旺skc", "申报价", "7天销量", "30天销量", "库存", "平台仓备货"],
    "shein_hot": ["冲突类型", "爆旺款skc", "爆旺skc", "申报价", "7天销量", "30天销量", "库存", "平台仓备货"],
    "low_score_warning": ["品质分", "是否已下架", "是否下架", "是否本周新增低分", "低分原因"],
    "temu_slow": ["预警类型", "上架天数", "7天销量", "30天销量", "库存", "建议动作"],
    "temu_bargain": ["建议价格", "是否通过", "申报价", "成本价", "批发价"],
}


def rows_from_report_workbook(report_id, report_name, workbook_path):
    report_id = norm(report_id)
    task_type = REPORT_TASK_TYPE.get(report_id)
    if not task_type:
        return []
    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    tasks = []
    try:
        for sheet in workbook.worksheets:
            values = list(sheet.iter_rows(values_only=True))
            if not values:
                continue
            headers = header_map(values[0])
            if not should_import_sheet(report_id, headers):
                continue
            for row_number, row in enumerate(values[1:], start=2):
                if not any(norm(value) for value in row):
                    continue
                task = map_report_row(report_id, report_name, workbook_path.name, sheet.title, row_number, row, headers)
                if task:
                    tasks.append(task)
    finally:
        workbook.close()
    return tasks


def should_import_sheet(report_id, headers):
    if report_id in {"temu_price", "shein_price"}:
        return any(name in headers for name in ["申报价", "成本价", "批发价", "批发价80%"])
    if report_id in {"temu_inventory", "shein_inventory"}:
        return any(name in headers for name in ["仓备可用", "触发规则"])
    if report_id in {"temu_hot", "shein_hot"}:
        return any(name in headers for name in ["处理意见", "冲突类型"])
    if report_id == "low_score_warning":
        return any(name in headers for name in ["是否下架", "是否已下架", "品质分", "是否本周新增低分"])
    if report_id == "temu_slow":
        return any(name in headers for name in ["建议动作", "操作", "预警类型"])
    if report_id == "temu_bargain":
        return any(name in headers for name in ["是否通过", "建议价格"])
    return False


def map_report_row(report_id, report_name, file_name, sheet_name, row_number, row, headers):
    platform = REPORT_PLATFORM.get(report_id, "")
    task_type = REPORT_TASK_TYPE.get(report_id, "")
    store = row_value(row, headers, "所属店铺", "店铺", "店铺编号")
    owner = row_value(row, headers, "负责人", "产品负责人", "填表人", "业务")
    merchant_code = row_value(row, headers, "商家编码", "SKU货号", "源SKU")
    skc = row_value(row, headers, "SKC", "skc", "爆旺款skc")
    spu = row_value(row, headers, "SPU", "spu", "爆旺skc")
    product_name = row_value(row, headers, "货品名称", "ERP货品名称", "品名", "名称")
    system_action = row_value(row, headers, "处理意见", "建议动作", "操作", "是否通过")
    if not norm(system_action) and report_id in {"temu_price", "shein_price", "temu_inventory", "shein_inventory"}:
        system_action = sheet_name
    task_detail = report_task_detail(report_id, row, headers)
    if not any(norm(value) for value in [store, owner, merchant_code, skc, spu, product_name, system_action]):
        return None
    return {
        "platform": platform,
        "task_type": task_type,
        "store": store,
        "owner": owner,
        "merchant_code": merchant_code,
        "skc": skc,
        "spu": spu,
        "product_name": product_name,
        "system_action": system_action,
        "task_detail": task_detail,
        "source_report": report_name,
        "source_file": file_name,
        "source_sheet": sheet_name,
        "source_row": row_number,
    }
