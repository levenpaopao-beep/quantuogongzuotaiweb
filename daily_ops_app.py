import cgi
import errno
import hashlib
import html
import importlib.util
import json
import mimetypes
import os
import re
import secrets
import shutil
import sqlite3
import sys
import threading
import zipfile
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import daily_ops_tasks
import daily_ops_task_suppression
import daily_ops_master_data
import daily_ops_bargain
from daily_ops_sales_compare import aggregate_source_sales
import update_shein_summary_30d_skc as raw_xlsx


ROOT = Path(__file__).resolve().parent
APP_VERSION = "v2.0"
OUTPUT_DIR = ROOT / "outputs"
TEMU_DIR = ROOT / "temu数据源表"
SHEIN_DIR = ROOT / "shein数据源表"
ERP_DIR = ROOT / "erp数据源"
BARGAIN_DIR = ROOT / "核价输入表"
LOW_SCORE_DIR = ROOT / "低分预警输入表"
DB_PATH = ROOT / "基础数据库" / "project_base_data.sqlite"
DATA_SOURCE_MANIFEST = ROOT / "基础数据库" / "data_source_manifest.json"
RULES_FILE = ROOT / "基础数据库" / "report_rules.json"
ERP_API_LOCAL_FILE = ROOT / "基础数据库" / "erp_api.local.json"
ERP_SYNC_LOCK_FILE = ROOT / "基础数据库" / "erp_sync.lock"
TASK_DB_PATH = ROOT / "基础数据库" / "operation_tasks.json"
TASK_SUPPRESSION_FILE = ROOT / "基础数据库" / "operation_task_suppressions.json"
STORE_OWNER_MAP_FILE = ROOT / "基础数据库" / "store_owner_map.json"
OPERATOR_ACCOUNTS_FILE = ROOT / "基础数据库" / "operator_accounts.json"
DAILY_SALES_FILE = ROOT / "基础数据库" / "daily_sales.json"
BARGAIN_DB_FILE = ROOT / "基础数据库" / "bargain_requests.json"
CLEARANCE_CATALOG_FILE = ROOT / "基础数据库" / "clearance_catalog.json"
MASTER_IMPORT_REVIEW_FILE = Path.home() / "Downloads" / "基础信息导入整理表.xlsx"
OWNER_FILE = ROOT / "店铺负责人对应表.xlsx"

HOST = "127.0.0.1"
PORT = 8876
CLIENT_CLOSE_SHUTDOWN_DELAY_SECONDS = 4
SCHEDULED_SHUTDOWN = None
SCHEDULED_SHUTDOWN_LOCK = threading.Lock()
OPERATOR_SESSIONS = {}
DOWNLOAD_GRANTS = {}
DEFAULT_OPERATOR_SESSION_SECONDS = 12 * 60 * 60

UPLOAD_TARGETS = {
    "temu_platform": ("Temu平台表", TEMU_DIR),
    "temu_hot": ("Temu爆旺款", TEMU_DIR),
    "shein_platform": ("Shein平台表", SHEIN_DIR),
    "erp_base": ("ERP基础信息", ERP_DIR),
    "erp_stock": ("ERP库存数据", ERP_DIR),
    "erp_combo": ("ERP组合装", ERP_DIR),
    "owner": ("店铺负责人表", ROOT),
    "temu_bargain_input": ("Temu核价输入表", BARGAIN_DIR),
    "low_score_input": ("低分预警输入表", LOW_SCORE_DIR),
}

WEEKLY_SOURCE_GROUPS = {
    "erp_base": {
        "name": "ERP 产品数据源",
        "description": "ERP 产品基础信息表，用于成本、批发价、货品名称和规格匹配。",
        "patterns": ["erp产品基础信息表*.xlsx"],
        "folder": ERP_DIR,
    },
    "erp_stock": {
        "name": "ERP 库存数据源",
        "description": "ERP 宠物圈仓库库存表，用于核对商家编码、规格尺码和当前可销库存。",
        "patterns": ["erp库存同步*.xlsx", "erp库存*.xlsx"],
        "folder": ERP_DIR,
    },
    "temu_platform": {
        "name": "Temu 销售表",
        "description": "Temu 仓库销售情况导出表，可多文件组成同一批数据。",
        "patterns": ["*Temu仓库销售情况导出*.xlsx"],
        "folder": TEMU_DIR,
    },
    "shein_platform": {
        "name": "Shein 销售表",
        "description": "Shein 各店铺销售总览表，通常为 琪琪、童话、加加、宝宝、牛牛。",
        "patterns": ["*shein*.xlsx", "*Shein*.xlsx", "*SHEIN*.xlsx"],
        "folder": SHEIN_DIR,
    },
    "temu_hot": {
        "name": "Temu 爆旺款表",
        "description": "Temu 爆旺款基准表，用于重复铺货保护款判断。",
        "patterns": ["*Temu爆旺款*.xlsx"],
        "folder": TEMU_DIR,
    },
    "temu_bargain_input": {
        "name": "Temu 议价申报表",
        "description": "上传需要回复的申报价/议价申报表，生成时按上传表格原始顺序返回回复。",
        "patterns": ["*议价申报*.xlsx", "*申报价*.xlsx", "*新品申报表*.xlsx", "*新品申报价格表*.xlsx"],
        "folder": BARGAIN_DIR,
    },
    "low_score_input": {
        "name": "店铺低分预警表",
        "description": "每周60分以下店铺低分产品清单，支持多文件组成同一批数据。",
        "patterns": ["*低分预警*.xlsx", "*.xlsx"],
        "folder": LOW_SCORE_DIR,
    },
}


def owner_can_upload_category(category):
    return category in WEEKLY_SOURCE_GROUPS


REPORTS = {
    "temu_price": {
        "name": "Temu申报价异常",
        "description": "识别低于ERP成本价、低于批发价80%的Temu商品。",
        "sources": "Temu平台表、ERP基础信息、ERP组合装、店铺负责人表",
    },
    "temu_inventory": {
        "name": "Temu仓备库存异常",
        "description": "识别平台仓备库存超过销售门槛的SKU/SKC。",
        "sources": "Temu平台表、Temu爆旺款、ERP基础信息、ERP组合装、店铺负责人表",
    },
    "temu_hot": {
        "name": "Temu爆旺款重复预警",
        "description": "按爆旺款基准表和货品编码识别重复铺货风险。",
        "sources": "Temu平台表、Temu爆旺款、ERP基础信息、ERP组合装、店铺负责人表",
    },
    "temu_slow": {
        "name": "Temu滞销品每周报表",
        "description": "从Temu店铺预警表提取新品、老款无销量明细。",
        "sources": "Temu店铺预警表_最终版",
    },
    "shein_price": {
        "name": "Shein申报价异常",
        "description": "识别低于ERP成本价、低于批发价80%的Shein商品。",
        "sources": "Shein平台表、ERP基础信息、ERP组合装、店铺负责人表",
    },
    "shein_inventory": {
        "name": "Shein仓备库存异常",
        "description": "识别Shein平台仓库存超过销售门槛的SKU/SKC。",
        "sources": "Shein平台表、ERP基础信息、ERP组合装",
    },
    "shein_hot": {
        "name": "Shein爆旺款重复预警",
        "description": "按Shein爆旺款定义识别重复铺货风险。",
        "sources": "Shein平台表、ERP基础信息、ERP组合装",
    },
    "temu_bargain": {
        "name": "Temu议价申报回复",
        "description": "根据申报价/议价申报表和平台在售情况，按上传表格顺序输出回复。",
        "sources": "Temu议价申报表、Temu平台表、Temu爆旺款、ERP基础信息、ERP组合装",
    },
    "low_score_warning": {
        "name": "店铺低分产品预警",
        "description": "对比本周与上周低分SPU，联查Temu销售、ERP和爆旺款，输出低分预警台账。",
        "sources": "低分预警输入表、Temu平台表、Temu爆旺款、ERP基础信息、店铺负责人表",
    },
}

SOURCE_DEPENDENT_REPORTS = {
    "temu_platform": ["temu_price", "temu_inventory", "temu_hot", "temu_slow", "temu_bargain", "low_score_warning"],
    "temu_hot": ["temu_inventory", "temu_hot", "temu_bargain", "low_score_warning"],
    "temu_bargain_input": ["temu_bargain"],
    "low_score_input": ["low_score_warning"],
    "shein_platform": ["shein_price", "shein_inventory", "shein_hot"],
    "erp_base": ["temu_price", "temu_inventory", "temu_hot", "temu_bargain", "low_score_warning", "shein_price", "shein_inventory", "shein_hot"],
    "erp_stock": ["temu_inventory", "shein_inventory"],
    "erp_combo": ["temu_price", "temu_inventory", "temu_hot", "temu_bargain", "shein_price", "shein_inventory", "shein_hot"],
    "owner": ["temu_price", "temu_inventory", "temu_hot", "low_score_warning", "shein_price", "shein_hot"],
}

DEFAULT_RULES = {
    "hot_item": {
        "temu_basis": "Temu爆旺款以最新上传的Temu爆旺款表为准。",
        "shein_new_days_lt": 30,
        "shein_new_7d_daily_gte": 10,
        "shein_old_days_gte": 30,
        "shein_old_30d_daily_gt": 20,
        "keywords": ["高销款", "爆", "旺"],
    },
    "sort": {
        "group_order": ["SPU", "SKC", "商家编码"],
        "size_order": ["XXS", "XS", "S", "M", "L", "XL", "XXL", "XXXL"],
    },
    "slow_moving": {
        "new_product_days_lt": 28,
        "new_slow_min_days": 30,
        "new_slow_max_days": 60,
        "old_slow_min_days": 180,
        "group_by": "店铺+SPU",
        "sales30_total_equals": 0,
    },
    "sales_thresholds": {
        "daily_diff_units": 10,
        "month_diff_percent": 5,
        "year_diff_percent": 5,
        "short_period_unit_per_day": 10,
        "long_period_diff_percent": 5,
        "completion_yellow_percent": 100,
        "completion_red_percent": 90,
        "erp_yellow_days": 1,
        "erp_red_days": 2,
        "platform_batch_days": ["周一", "周二", "周三", "周四", "周五", "周六"],
        "platform_batch_yellow_time": "12:00",
        "platform_batch_red_time": "18:00",
    },
    "erp_api": {
        "provider": "旺店通",
        "enabled": False,
        "auto_sync": True,
        "manual_sync_first": True,
        "environment": "test",
        "base_url": "https://openapi.ali.huice.cc/openapi",
        "product_endpoint": "goods_query.php",
        "stock_endpoint": "stock_query.php",
        "sync_product_archive": True,
        "sync_stock_snapshot": True,
        "sync_sales_outbound": False,
        "shop_id": "",
        "shop_no": "",
        "warehouse_no": "",
        "warehouse_name": "宠物圈仓库",
        "sync_days": 7,
        "page_size": 500,
        "stock_limit": 10000,
        "max_pages": 1000,
        "request_interval_seconds": 1.1,
        "rate_limit_retry_seconds": 300,
        "rate_limit_retries": 2,
        "latest_file_only": True,
        "schedule_enabled": True,
        "schedule_time": "05:00",
        "timezone": "Asia/Shanghai",
        "app_key": "",
        "app_secret": "",
        "sid": "",
        "token": "",
        "sync_scope": ["商品信息", "库存数据", "商家编码", "尺码规格"],
        "last_tested_at": "",
        "last_manual_sync_at": "",
        "last_manual_sync_status": "",
        "last_manual_sync_message": "",
    },
}

LEGACY_ERP_PRODUCT_ENDPOINT = "vip_api_goods_query.php"
LEGACY_ERP_STOCK_ENDPOINT = ""
ERP_LOCAL_CREDENTIAL_FIELDS = {"app_key", "app_secret", "sid", "token"}

OWNER_STORE_CODE_MAP = {
    "弟弟": "1",
    "一弟": "1",
    "二弟": "2",
    "三弟": "3",
    "四弟": "4",
    "五弟": "5",
    "六弟": "6",
    "七弟": "7",
    "八弟": "8",
    "九弟": "9",
    "十弟": "10",
    "十一": "11",
    "十一弟": "11",
    "十二": "12",
    "十二弟": "12",
    "十三": "13",
    "十三弟": "13",
    "十五": "15",
    "十五弟": "15",
}


def owner_lookup_keys(store):
    text = norm(store)
    if not text:
        return []
    keys = {text}
    if re.fullmatch(r"[一二三四五六七八九十]+", text):
        keys.add(f"{text}弟")
    if text.endswith("弟") and re.fullmatch(r"[一二三四五六七八九十]+", text[:-1]):
        keys.add(text[:-1])
    code = OWNER_STORE_CODE_MAP.get(text)
    if code:
        keys.add(code)
    return sorted(keys, key=lambda item: (item != text, item))


def cancel_scheduled_shutdown():
    global SCHEDULED_SHUTDOWN
    with SCHEDULED_SHUTDOWN_LOCK:
        if SCHEDULED_SHUTDOWN is not None:
            SCHEDULED_SHUTDOWN.cancel()
            SCHEDULED_SHUTDOWN = None


def schedule_shutdown_after_client_close(server):
    global SCHEDULED_SHUTDOWN
    with SCHEDULED_SHUTDOWN_LOCK:
        if SCHEDULED_SHUTDOWN is not None:
            SCHEDULED_SHUTDOWN.cancel()
        timer = threading.Timer(CLIENT_CLOSE_SHUTDOWN_DELAY_SECONDS, server.shutdown)
        timer.daemon = True
        SCHEDULED_SHUTDOWN = timer
        timer.start()


def today_code():
    return datetime.now().strftime("%y%m%d")


def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def configured_host():
    value = os.environ.get("DAILY_OPS_HOST", "").strip()
    return value or HOST


def access_hint(host, port):
    if host in {"0.0.0.0", "::"}:
        return f"局域网模式已启动：本机访问 http://127.0.0.1:{port}，店长用本机局域网IP访问同一端口。"
    return f"本机模式已启动：http://{host}:{port}"


def login_operator(role, user, password=""):
    role = norm(role) or "owner"
    if role not in {"admin", "owner"}:
        raise ValueError("身份类型不正确")
    user = norm(user) or ("管理员" if role == "admin" else "")
    if not user:
        raise ValueError("请填写姓名")
    admin_password = os.environ.get("DAILY_OPS_ADMIN_PASSWORD", "")
    owner_password = os.environ.get("DAILY_OPS_OWNER_PASSWORD", "")
    if role == "admin":
        if configured_host() in {"0.0.0.0", "::"} and not admin_password:
            raise PermissionError("局域网模式必须先设置 DAILY_OPS_ADMIN_PASSWORD")
        if admin_password and password != admin_password:
            raise PermissionError("管理员密码不正确")
    else:
        if owner_password and password != owner_password:
            raise PermissionError("店长访问密码不正确")
        owners = {norm(row.get("owner")) for row in operation_owner_directory() if norm(row.get("owner"))}
        if owners and user not in owners:
            raise PermissionError("店长姓名不在负责人名单中，请联系管理员配置店铺负责人")
    token = secrets.token_urlsafe(24)
    session = {"token": token, "role": role, "user": user, "login_at": now_text()}
    OPERATOR_SESSIONS[token] = session
    return session


def operator_session_seconds():
    try:
        return max(1, int(os.environ.get("DAILY_OPS_SESSION_SECONDS", DEFAULT_OPERATOR_SESSION_SECONDS)))
    except ValueError:
        return DEFAULT_OPERATOR_SESSION_SECONDS


def session_is_expired(session, now=None):
    login_at = norm(session.get("login_at"))
    if not login_at:
        return True
    try:
        logged_at = datetime.strptime(login_at, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return True
    now = now or datetime.now()
    return (now - logged_at).total_seconds() > operator_session_seconds()


def operator_from_token(token):
    token = norm(token)
    if not token or token not in OPERATOR_SESSIONS:
        raise PermissionError("请先登录")
    session = OPERATOR_SESSIONS[token]
    if session_is_expired(session):
        logout_operator(token)
        raise PermissionError("登录已过期，请重新登录")
    return session


def logout_operator(token):
    token = norm(token)
    if token:
        OPERATOR_SESSIONS.pop(token, None)
        DOWNLOAD_GRANTS.pop(token, None)
    return {"ok": True}


def handle_session_logout(headers):
    try:
        token = token_from_headers(headers)
        operator_from_token(token)
        logout_operator(token)
        return json_bytes({"ok": True})
    except PermissionError as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=401)


def grant_download(token, filename):
    token = norm(token)
    filename = Path(norm(filename)).name
    if token and filename:
        DOWNLOAD_GRANTS.setdefault(token, set()).add(filename)


def require_download_permission(token, filename):
    operator = operator_from_token(token)
    if operator.get("role") == "admin":
        return operator
    if Path(norm(filename)).name not in DOWNLOAD_GRANTS.get(norm(token), set()):
        raise PermissionError("没有权限下载这个文件")
    return operator


def scoped_task_filters(operator, filters):
    filters = dict(filters or {})
    if operator.get("role") == "admin":
        return {
            "role": filters.get("role", "admin") or "admin",
            "user": filters.get("user", "") or "",
        }
    return {"role": "owner", "user": operator.get("user", "")}


def can_review_tasks(operator):
    return operator.get("role") == "admin"


def task_query_payload(params):
    return {
        "role": params.get("role", ["admin"])[0],
        "user": params.get("user", [""])[0],
        "status": params.get("status", [""])[0],
        "task_type": params.get("task_type", [""])[0],
        "store": params.get("store", [""])[0],
        "platform": params.get("platform", [""])[0],
        "next_handler": params.get("next_handler", [""])[0],
        "priority": params.get("priority", [""])[0],
        "open_only": params.get("open_only", [""])[0],
        "overdue": params.get("overdue", [""])[0],
        "unassigned": params.get("unassigned", [""])[0],
        "reworked": params.get("reworked", [""])[0],
        "search": params.get("search", [""])[0],
    }


def json_bytes(payload, status=200):
    body = json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")
    return status, "application/json; charset=utf-8", body


def safe_name(text, fallback="file"):
    text = re.sub(r'[\\/:*?"<>|\r\n]+', "_", str(text)).strip(" ._")
    return text or fallback


def version_text(value):
    value = safe_name(value or "V1", "V1").upper()
    return value if value.startswith("V") else f"V{value}"


def output_path(project_name, version, suffix=".xlsx"):
    OUTPUT_DIR.mkdir(exist_ok=True)
    name = f"{today_code()}-{safe_name(project_name)}-{version_text(version)}{suffix}"
    path = OUTPUT_DIR / name
    if not path.exists():
        return path
    stamp = datetime.now().strftime("%H%M%S")
    return OUTPUT_DIR / f"{today_code()}-{safe_name(project_name)}-{version_text(version)}-{stamp}{suffix}"


def backup_output_dir():
    path = ROOT / "运营备份"
    path.mkdir(exist_ok=True)
    return path


def backup_source_paths():
    paths = [
        TASK_DB_PATH,
        TASK_SUPPRESSION_FILE,
        DAILY_SALES_FILE,
        OPERATOR_ACCOUNTS_FILE,
        RULES_FILE,
        DATA_SOURCE_MANIFEST,
        STORE_OWNER_MAP_FILE,
        BARGAIN_DB_FILE,
        CLEARANCE_CATALOG_FILE,
        DB_PATH,
        OWNER_FILE,
        TEMU_DIR,
        SHEIN_DIR,
        ERP_DIR,
        BARGAIN_DIR,
        LOW_SCORE_DIR,
    ]
    return [Path(path) for path in paths if Path(path).exists()]


def backup_relative_path(path):
    path = Path(path).resolve()
    try:
        return path.relative_to(ROOT.resolve()).as_posix()
    except ValueError:
        return path.name


def create_operational_backup():
    target = backup_output_dir() / f"{datetime.now():%Y%m%d-%H%M%S}-系统数据备份.zip"
    files = []
    with zipfile.ZipFile(target, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for source in backup_source_paths():
            if source.is_file():
                zf.write(source, backup_relative_path(source))
                files.append(backup_relative_path(source))
            elif source.is_dir():
                for path in sorted(source.rglob("*")):
                    if path.is_file() and "__pycache__" not in path.parts:
                        zf.write(path, backup_relative_path(path))
                        files.append(backup_relative_path(path))
        zf.writestr("backup_manifest.json", json.dumps({
            "created_at": now_text(),
            "app_version": APP_VERSION,
            "files": files,
            "excluded": ["outputs", "build", "打包产物", "虚拟环境", "图片产物"],
        }, ensure_ascii=False, indent=2))
    return {"file": target.name, "path": str(target), "count": len(files)}


def restore_operational_backup(backup_path):
    backup_path = Path(backup_path)
    if not backup_path.exists() or backup_path.suffix.lower() != ".zip":
        raise FileNotFoundError("备份文件不存在")
    restored = []
    allowed_roots = {
        "基础数据库",
        "temu数据源表",
        "shein数据源表",
        "erp数据源",
        "核价输入表",
        "低分预警输入表",
    }
    allowed_files = {"店铺负责人对应表.xlsx"}
    with zipfile.ZipFile(backup_path) as zf:
        for member in zf.infolist():
            name = member.filename
            if member.is_dir() or name == "backup_manifest.json":
                continue
            parts = Path(name).parts
            if not parts:
                continue
            if parts[0] not in allowed_roots and name not in allowed_files:
                continue
            target = (ROOT / name).resolve()
            if ROOT.resolve() not in target.parents and target != ROOT.resolve():
                continue
            target.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(member) as src, target.open("wb") as dst:
                shutil.copyfileobj(src, dst)
            restored.append(name)
    return {"file": backup_path.name, "restored": restored, "count": len(restored)}


def operator_accounts():
    payload = daily_ops_master_data.load_operator_accounts(OPERATOR_ACCOUNTS_FILE)
    accounts = []
    for row in payload.get("accounts", []):
        item = dict(row)
        item.pop("password_hash", None)
        item.pop("password_salt", None)
        accounts.append(item)
    return {"accounts": accounts}


def create_operator_account(owner, username="", password="", enabled=True):
    owner = daily_ops_master_data.norm(owner)
    username = daily_ops_master_data.norm(username) or owner
    if not owner:
        raise ValueError("请填写店长姓名")
    if not username:
        raise ValueError("请填写登录账号")
    payload = daily_ops_master_data.load_operator_accounts(OPERATOR_ACCOUNTS_FILE)
    existing = payload.get("accounts", [])
    if any(daily_ops_master_data.norm(row.get("username")) == username for row in existing):
        raise ValueError("账号已存在")
    account = {"owner": owner, "username": username, "role": "owner", "enabled": enabled is not False}
    password_factory = (lambda _account: password) if password else None
    result = daily_ops_master_data.save_operator_accounts(OPERATOR_ACCOUNTS_FILE, existing + [account], password_factory)
    return {
        "accounts": operator_accounts()["accounts"],
        "username": username,
        "initial_password": result.get("initial_passwords", {}).get(username, ""),
    }


def reset_operator_account_password(username, password=""):
    result = daily_ops_master_data.reset_operator_password(OPERATOR_ACCOUNTS_FILE, username, password or None)
    return result


def import_owner_master_data(source_path):
    parsed = daily_ops_master_data.parse_owner_workbook(source_path)
    saved_assignments = save_store_owner_assignments(parsed.get("assignments", []))
    account_result = daily_ops_master_data.save_operator_accounts(OPERATOR_ACCOUNTS_FILE, parsed.get("accounts", []))
    review_path = MASTER_IMPORT_REVIEW_FILE
    daily_ops_master_data.export_master_import_review(parsed, [], account_result, review_path)
    return {
        "assignments": saved_assignments,
        "accounts": operator_accounts()["accounts"],
        "initial_passwords": account_result.get("initial_passwords", {}),
        "review_file": str(review_path),
        "assignment_count": len(saved_assignments),
        "account_count": len(account_result.get("accounts", [])),
    }


def import_crossborder_sales(source_path):
    assignments = load_store_owner_assignments()
    rows = daily_ops_master_data.parse_crossborder_sales_workbook(source_path, assignments)
    result = daily_ops_master_data.import_history_sales_records(DAILY_SALES_FILE, rows, actor="管理员")
    parsed_owner = {"assignments": assignments}
    account_payload = daily_ops_master_data.load_operator_accounts(OPERATOR_ACCOUNTS_FILE)
    review_path = MASTER_IMPORT_REVIEW_FILE
    daily_ops_master_data.export_master_import_review(parsed_owner, rows, {"accounts": account_payload.get("accounts", []), "initial_passwords": {}}, review_path)
    return {**result, "review_file": str(review_path)}


def sales_report(platform="", store="", date_from="", date_to="", allowed_pairs=None):
    return daily_ops_master_data.query_sales_report(DAILY_SALES_FILE, platform, store, date_from, date_to, allowed_pairs=allowed_pairs)


def export_sales_report(platform="", store="", date_from="", date_to=""):
    report = sales_report(platform, store, date_from, date_to)
    return daily_ops_master_data.export_sales_report(report, OUTPUT_DIR)


def platform_business_rows(assignments, range_key="30d", anchor_date=""):
    anchor_day = daily_ops_master_data.parse_day(anchor_date) or datetime.now().date()
    start_day, end_day = daily_ops_master_data.complete_period(range_key or "30d", anchor_day)
    days = max(1, (end_day - start_day).days + 1)
    source_sales = aggregate_source_sales({
        "Temu": temu_sales_files(),
        "Shein": shein_platform_files(),
    })
    assignment_index = daily_ops_master_data.build_assignment_index(assignments)
    rows = []
    for platform, stores in source_sales.items():
        for store, daily_avg in stores.items():
            clean_store = daily_ops_master_data.clean_store_name(store)
            assignment = assignment_index.get((platform, clean_store), {})
            rows.append({
                "id": f"platform|{platform}|{clean_store}|{end_day.isoformat()}",
                "date": end_day.isoformat(),
                "platform": platform,
                "store": clean_store,
                "owner": assignment.get("owner", ""),
                "sales": int(round(float(daily_avg or 0) * days)),
                "status": "平台导入折算",
                "source": "平台导入销量",
                "source_note": f"按导入表日均销量折算 {days} 天；经营报表默认仍以店长填报销量为准。",
            })
    return rows


def business_report(payload=None):
    payload = payload or {}
    assignments = load_store_owner_assignments()
    rows_override = None
    if payload.get("source") == "platform":
        rows_override = platform_business_rows(assignments, payload.get("range_key", "30d"), payload.get("anchor_date", ""))
    return daily_ops_master_data.business_report(
        DAILY_SALES_FILE,
        assignments=assignments,
        role=payload.get("role", "admin"),
        user=payload.get("user", ""),
        date_from=payload.get("date_from", ""),
        date_to=payload.get("date_to", ""),
        platform=payload.get("platform", ""),
        store=payload.get("store", ""),
        grain=payload.get("grain", "month"),
        range_key=payload.get("range_key", "30d"),
        source=payload.get("source", "manual"),
        anchor_date=payload.get("anchor_date", ""),
        rows_override=rows_override,
    )


def monthly_backup_reminder():
    return daily_ops_master_data.monthly_backup_status(backup_output_dir())


def report_id_for_output(filename):
    stem = Path(filename).stem
    if "temu议价回复" in stem.lower() or "temu核价回复" in stem.lower():
        return "temu_bargain"
    for report_id, report in REPORTS.items():
        if safe_name(report["name"]) in stem:
            return report_id
    return ""


def recent_outputs(limit=20):
    OUTPUT_DIR.mkdir(exist_ok=True)
    files = sorted(
        [p for p in OUTPUT_DIR.iterdir() if p.is_file()],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return [
        {
            "name": p.name,
            "size": p.stat().st_size,
            "modified": datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            "download": f"/download?path={quote(p.name)}",
            "report": report_id_for_output(p.name),
        }
        for p in files[:limit]
    ]


def parse_timestamp(value):
    text = norm(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def latest_output_times_by_report(limit=300):
    latest = {}
    for item in recent_outputs(limit):
        report_id = item.get("report", "")
        timestamp = parse_timestamp(item.get("modified"))
        if report_id and timestamp and (report_id not in latest or timestamp > latest[report_id]):
            latest[report_id] = timestamp
    return latest


def source_recompute_state(category, uploaded_at):
    reports = SOURCE_DEPENDENT_REPORTS.get(category, [])
    report_names = [REPORTS.get(report_id, {}).get("name", report_id) for report_id in reports]
    if not reports:
        return {"reports": [], "report_names": [], "needed": False, "stale_reports": [], "latest_generated_at": "", "message": "该数据源不需要重算任务。"}
    uploaded_time = parse_timestamp(uploaded_at)
    output_times = latest_output_times_by_report()
    latest_generated = max([output_times[report_id] for report_id in reports if report_id in output_times], default=None)
    stale = [
        report_id for report_id in reports
        if report_id not in output_times or (uploaded_time and output_times[report_id] < uploaded_time)
    ]
    stale_names = [REPORTS.get(report_id, {}).get("name", report_id) for report_id in stale]
    needed = bool(uploaded_time and stale)
    message = "最新数据源已生效，建议重算关联任务。" if needed else "关联任务已按当前数据源生成。"
    if not uploaded_time:
        message = "当前还没有正式提交的数据源批次。"
    return {
        "reports": reports,
        "report_names": report_names,
        "needed": needed,
        "stale_reports": stale,
        "stale_report_names": stale_names,
        "latest_generated_at": latest_generated.strftime("%Y-%m-%d %H:%M:%S") if latest_generated else "",
        "message": message,
    }


def latest_by_date(folder, pattern):
    files = sorted(folder.glob(pattern), key=lambda p: p.name)
    dated = {}
    for path in files:
        match = re.search(r"(20\d{6})", path.name)
        dated.setdefault(match.group(1) if match else "00000000", []).append(path)
    if not dated:
        return []
    return sorted(dated[max(dated)], key=lambda p: p.name)


def manifest_path(category):
    raw = load_source_manifest().get("categories", {}).get(category, {}).get("path", "")
    if not raw:
        return None
    path = Path(raw)
    return path if path.exists() else None


def manifest_paths(category):
    item = load_source_manifest().get("categories", {}).get(category, {})
    raw_paths = item.get("paths")
    if isinstance(raw_paths, list):
        paths = [Path(raw) for raw in raw_paths]
        return [path for path in paths if path.exists()]
    single = manifest_path(category)
    return [single] if single else []


def source_files_for_category(category):
    uploaded = manifest_paths(category)
    if uploaded:
        return uploaded
    config = WEEKLY_SOURCE_GROUPS.get(category)
    if not config:
        return []
    return matching_files(config["folder"], config["patterns"])


def erp_base_files():
    latest_sync = ERP_DIR / "erp产品基础信息表_接口同步_最新.xlsx"
    if latest_sync.exists():
        return [latest_sync]
    interface_files = sorted(
        ERP_DIR.glob("erp产品基础信息表_接口同步_*.xlsx"),
        key=lambda path: path.stat().st_mtime if path.exists() else 0,
        reverse=True,
    )
    if interface_files:
        return [interface_files[0]]
    return []


def erp_cost_files():
    files = []
    for pattern in ["erp库存销量*.xlsx", "宠物圈仓_库存查询*.xlsx", "erp库存同步*.xlsx", "erp库存*.xlsx"]:
        files.extend(ERP_DIR.glob(pattern))
    unique = {}
    for path in files:
        if path.exists():
            unique[str(path.resolve())] = path
    return sorted(unique.values(), key=lambda path: path.stat().st_mtime if path.exists() else 0, reverse=True)


def temu_sales_files():
    uploaded = manifest_paths("temu_platform")
    if uploaded:
        return uploaded
    files = latest_by_date(TEMU_DIR, "*Temu仓库销售情况导出*.xlsx")
    if files:
        return files
    return matching_files(TEMU_DIR, ["*temu_inventory_product*.xlsx", "*Temu*.xlsx"])


def temu_hot_files():
    uploaded = manifest_paths("temu_hot")
    if uploaded:
        return uploaded
    return latest_by_date(TEMU_DIR, "*Temu爆旺款*.xlsx")


def shein_platform_files():
    uploaded = manifest_paths("shein_platform")
    if uploaded:
        return uploaded
    return matching_files(SHEIN_DIR, WEEKLY_SOURCE_GROUPS["shein_platform"]["patterns"])


def platform_source_rows_for_bargain(platform=""):
    wanted = norm(platform)
    source_plan = []
    if not wanted or wanted == "Temu":
        source_plan.append(("Temu", temu_sales_files()))
    if not wanted or wanted == "Shein":
        source_plan.append(("Shein", shein_platform_files()))
    result = []
    for platform_name, files in source_plan:
        for path in files:
            rows = raw_xlsx.read_xlsx_rows(path)
            if not rows:
                continue
            headers = header_map(rows[0])
            for row_number, row in enumerate(rows[1:], start=2):
                merchant_code = norm(cell(row, headers, "商家编码", "SKU货号", "商品SKU", "供应商SKU", "小秘商品SKU"))
                if not merchant_code:
                    continue
                store = norm(cell(row, headers, "店铺", "店铺名称", "店铺名"))
                item = {
                    "平台": platform_name,
                    "店铺": store,
                    "商家编码": merchant_code,
                    "货品名称": norm(cell(row, headers, "货品名称", "商品名称", "产品名称")),
                    "规格名称": norm(cell(row, headers, "规格名称", "货品规格", "SKU属性", "规格")),
                    "申报价": to_number(cell(row, headers, "申报价", "申报价格", "售价", "价格")),
                    "7天销量": to_number(cell(row, headers, "7天销量", "近7天销量")),
                    "30天销量": to_number(cell(row, headers, "30天销量", "近30天销量")),
                    "平台库存": to_number(cell(row, headers, "平台仓库备货库存", "平台库存", "总库存", "可用库存")),
                    "源文件": path.name,
                    "源行": row_number,
                }
                result.append(item)
    return result


def erp_combo_files():
    uploaded = manifest_paths("erp_combo")
    if uploaded:
        return uploaded
    combo = ERP_DIR / "erp产品组合装基础信息表.xlsx"
    return [combo] if combo.exists() else []


def owner_files():
    uploaded = manifest_paths("owner")
    if uploaded:
        return uploaded
    return [OWNER_FILE] if OWNER_FILE.exists() else []


def normalize_store_owner_assignment(row):
    platform = norm(row.get("platform", ""))
    store = norm(row.get("store", ""))
    owner = norm(row.get("owner", ""))
    if not platform or not store or not owner:
        return None
    enabled = row.get("enabled", True)
    daily_required = row.get("daily_required", row.get("requires_daily_sales", True))
    return {
        "platform": platform,
        "store": store,
        "owner": owner,
        "enabled": enabled not in {False, "false", "False", "0", 0, "否", "停用"},
        "daily_required": daily_required not in {False, "false", "False", "0", 0, "否", "不填"},
    }


def load_store_owner_assignments():
    path = Path(STORE_OWNER_MAP_FILE)
    if not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    rows = payload.get("assignments", []) if isinstance(payload, dict) else payload
    if not isinstance(rows, list):
        return []
    assignments = []
    seen = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        item = normalize_store_owner_assignment(row)
        if not item:
            continue
        key = (item["platform"], item["store"])
        if key in seen:
            continue
        seen.add(key)
        assignments.append(item)
    return assignments


def validate_store_owner_assignments(assignments):
    store_platforms = {}
    for item in assignments:
        store = norm(item.get("store", ""))
        platform = norm(item.get("platform", ""))
        if not store:
            continue
        existing = store_platforms.get(store)
        if existing and existing != platform:
            raise ValueError(f"店铺“{store}”已经归属平台“{existing}”，不能同时归属“{platform}”")
        store_platforms[store] = platform


def save_store_owner_assignments(rows):
    assignments = []
    seen = set()
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        item = normalize_store_owner_assignment(row)
        if not item:
            continue
        key = (item["platform"], item["store"])
        if key in seen:
            continue
        seen.add(key)
        assignments.append(item)
    validate_store_owner_assignments(assignments)
    path = Path(STORE_OWNER_MAP_FILE)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps({"assignments": assignments, "updated_at": now_text()}, ensure_ascii=False, indent=2), encoding="utf-8")
    return assignments


def configured_owner_for_store(platform, store):
    platform = norm(platform)
    keys = owner_lookup_keys(store) or ([norm(store)] if norm(store) else [])
    for assignment in load_store_owner_assignments():
        assignment_platform = norm(assignment.get("platform"))
        if assignment_platform and platform and assignment_platform != platform:
            continue
        if assignment_platform and not platform:
            continue
        assignment_keys = set(owner_lookup_keys(assignment.get("store")) or [norm(assignment.get("store"))])
        if any(key in assignment_keys for key in keys):
            return norm(assignment.get("owner"))
    return ""


def owner_from_assignments(assignments, platform, store):
    platform = norm(platform)
    keys = owner_lookup_keys(store) or ([norm(store)] if norm(store) else [])
    for assignment in assignments or []:
        assignment_platform = norm(assignment.get("platform"))
        if assignment_platform and platform and assignment_platform != platform:
            continue
        if assignment_platform and not platform:
            continue
        assignment_keys = set(owner_lookup_keys(assignment.get("store")) or [norm(assignment.get("store"))])
        if any(key in assignment_keys for key in keys):
            return norm(assignment.get("owner"))
    return ""


def bargain_input_files():
    uploaded = manifest_paths("temu_bargain_input")
    if uploaded:
        return uploaded
    return matching_files(BARGAIN_DIR, ["*议价申报*.xlsx", "*申报价*.xlsx", "*新品申报表*.xlsx", "*新品申报价格表*.xlsx", "*.xlsx"])


def low_score_input_files():
    uploaded = manifest_paths("low_score_input")
    if uploaded:
        return uploaded
    return matching_files(LOW_SCORE_DIR, ["*低分预警*.xlsx", "*.xlsx"])


def shein_source_map():
    stores = ["琪琪", "童话", "加加", "宝宝", "牛牛"]
    result = {store: [] for store in stores}
    for path in shein_platform_files():
        for store in stores:
            if store in path.stem:
                result[store].append(path)
                break
    fallback = {}
    for store, paths in result.items():
        if paths:
            fallback[store] = sorted(paths, key=lambda p: (p.stat().st_mtime, p.name))
        else:
            candidates = sorted(
                [path for path in SHEIN_DIR.glob("*.xlsx") if store in path.stem],
                key=lambda path: (path.stat().st_mtime, path.name),
            )
            fallback[store] = [candidates[-1]] if candidates else []
    return fallback


def size_order_map(value):
    if isinstance(value, dict):
        return value
    if isinstance(value, (list, tuple)):
        return {str(size): index for index, size in enumerate(value)}
    return {}


def matching_files(folder, patterns):
    files = []
    for pattern in patterns:
        files.extend(folder.glob(pattern))
    seen = {}
    for path in files:
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xls", ".csv"}:
            seen[path.resolve()] = path
    return sorted(seen.values(), key=lambda p: (p.stat().st_mtime, p.name), reverse=True)


def file_sha256(path):
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_source_manifest():
    if not DATA_SOURCE_MANIFEST.exists():
        return {"categories": {}}
    try:
        return json.loads(DATA_SOURCE_MANIFEST.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {"categories": {}}


def save_source_manifest(manifest):
    DATA_SOURCE_MANIFEST.parent.mkdir(exist_ok=True)
    DATA_SOURCE_MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def merge_dict(default, current):
    merged = {}
    for key, value in default.items():
        if isinstance(value, dict):
            merged[key] = merge_dict(value, current.get(key, {}) if isinstance(current, dict) else {})
        else:
            merged[key] = current.get(key, value) if isinstance(current, dict) else value
    if isinstance(current, dict):
        for key, value in current.items():
            if key not in merged:
                merged[key] = value
    return merged


def migrate_rules(rules):
    erp_api = rules.get("erp_api")
    if isinstance(erp_api, dict):
        if erp_api.get("product_endpoint") == LEGACY_ERP_PRODUCT_ENDPOINT:
            erp_api["product_endpoint"] = DEFAULT_RULES["erp_api"]["product_endpoint"]
        if LEGACY_ERP_STOCK_ENDPOINT and erp_api.get("stock_endpoint") == LEGACY_ERP_STOCK_ENDPOINT:
            erp_api["stock_endpoint"] = DEFAULT_RULES["erp_api"]["stock_endpoint"]
    return rules


def load_erp_local_credentials():
    if os.environ.get("DAILY_OPS_IGNORE_LOCAL_ERP_CREDENTIALS") == "1":
        return {}
    try:
        data = json.loads(ERP_API_LOCAL_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    if not isinstance(data, dict):
        return {}
    return {key: str(value) for key, value in data.items() if key in ERP_LOCAL_CREDENTIAL_FIELDS and value}


def split_erp_local_credentials(rules):
    erp_api = rules.get("erp_api")
    if not isinstance(erp_api, dict):
        return
    local = load_erp_local_credentials()
    changed = False
    for key in ERP_LOCAL_CREDENTIAL_FIELDS:
        value = erp_api.get(key)
        if value:
            local[key] = str(value)
            erp_api[key] = ""
            changed = True
    if changed:
        ERP_API_LOCAL_FILE.parent.mkdir(exist_ok=True)
        ERP_API_LOCAL_FILE.write_text(json.dumps(local, ensure_ascii=False, indent=2), encoding="utf-8")


def merge_erp_local_credentials(rules):
    erp_api = rules.get("erp_api")
    if isinstance(erp_api, dict):
        erp_api.update(load_erp_local_credentials())
    return rules


def load_rules():
    if not RULES_FILE.exists():
        return merge_erp_local_credentials(migrate_rules(merge_dict(DEFAULT_RULES, {})))
    try:
        current = json.loads(RULES_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        current = {}
    return merge_erp_local_credentials(migrate_rules(merge_dict(DEFAULT_RULES, current)))


def save_rules(payload):
    if not isinstance(payload, dict):
        raise ValueError("规则内容格式不正确")
    rules = merge_dict(DEFAULT_RULES, payload)
    split_erp_local_credentials(rules)
    RULES_FILE.parent.mkdir(exist_ok=True)
    RULES_FILE.write_text(json.dumps(rules, ensure_ascii=False, indent=2), encoding="utf-8")
    return load_rules()


def sync_erp_base_data():
    import daily_ops_erp

    rules = load_rules()
    settings = rules.get("erp_api", {})
    ERP_SYNC_LOCK_FILE.parent.mkdir(parents=True, exist_ok=True)
    try:
        lock_fd = os.open(ERP_SYNC_LOCK_FILE, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError:
        return {"status": "blocked", "message": "ERP正在同步中，请稍后再试"}
    try:
        with os.fdopen(lock_fd, "w", encoding="utf-8") as lock:
            lock.write(json.dumps({"pid": os.getpid(), "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}, ensure_ascii=False))
        try:
            result = daily_ops_erp.manual_sync(settings, ERP_DIR)
        except Exception as exc:
            result = {"status": "failed", "message": f"ERP同步失败：{exc}"}
        now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        settings["last_manual_sync_at"] = now_text
        settings["last_manual_sync_status"] = result.get("status", "")
        settings["last_manual_sync_message"] = result.get("message", "")
        if result.get("status") == "synced":
            settings["last_success_sync_at"] = now_text
            settings["last_product_count"] = result.get("product_count", 0)
            settings["last_stock_count"] = result.get("stock_count", 0)
            settings["last_product_pages"] = result.get("product_pages", 0)
            settings["last_stock_pages"] = result.get("stock_pages", 0)
            settings["last_product_file"] = result.get("product_file", "")
            settings["last_stock_file"] = result.get("stock_file", "")
        rules["erp_api"] = settings
        save_rules(rules)
        return result
    finally:
        try:
            ERP_SYNC_LOCK_FILE.unlink()
        except FileNotFoundError:
            pass


def readable_row_count(path):
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return ""
    try:
        rows = raw_xlsx.read_xlsx_rows(path)
        return max(len(rows) - 1, 0) if rows else 0
    except Exception:
        return ""


def source_file_state(path):
    return {
        "path": str(path),
        "file": path.name,
        "size": path.stat().st_size,
        "sha256": file_sha256(path),
        "rows": readable_row_count(path),
        "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def upload_batch_id(category):
    return f"{datetime.now().strftime('%Y%m%d%H%M%S')}-{safe_name(category)}"


def record_uploaded_source(category, path):
    manifest = load_source_manifest()
    categories = manifest.setdefault("categories", {})
    pending = manifest.setdefault("pending_batches", {}).setdefault(category, {
        "batch_id": upload_batch_id(category),
        "files": [],
        "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })
    state = source_file_state(path)
    active_hashes = set(categories.get(category, {}).get("sha256_list", []))
    state["changed"] = state["sha256"] not in active_hashes
    pending_files = [item for item in pending.get("files", []) if item.get("path") != str(path)]
    pending_files.append(state)
    pending["files"] = pending_files
    pending["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    save_source_manifest(manifest)
    return {**state, "pending": True, "pending_count": len(pending_files), "batch_id": pending.get("batch_id", "")}


def pending_batch(category):
    return load_source_manifest().get("pending_batches", {}).get(category, {"files": []})


def finish_upload_batch(category):
    if category not in UPLOAD_TARGETS:
        raise ValueError("未知上传分类")
    manifest = load_source_manifest()
    categories = manifest.setdefault("categories", {})
    pending_batches = manifest.setdefault("pending_batches", {})
    pending = pending_batches.get(category, {})
    files = [item for item in pending.get("files", []) if Path(item.get("path", "")).exists()]
    if not files:
        raise ValueError("本次上传还没有Temu销售表文件")
    previous = categories.get(category, {})
    previous_hashes = set(previous.get("sha256_list") or ([previous.get("sha256")] if previous.get("sha256") else []))
    sha256_list = [item["sha256"] for item in files]
    changed = sha256_list != list(previous.get("sha256_list", [])) or any(value not in previous_hashes for value in sha256_list)
    categories[category] = {
        "batch_id": pending.get("batch_id") or upload_batch_id(category),
        "paths": [item["path"] for item in files],
        "path": files[-1]["path"],
        "files": [item["file"] for item in files],
        "file": files[-1]["file"],
        "count": len(files),
        "size": sum(item.get("size", 0) for item in files),
        "rows": sum(item.get("rows") or 0 for item in files),
        "sha256_list": sha256_list,
        "sha256": hashlib.sha256("".join(sha256_list).encode("utf-8")).hexdigest(),
        "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "changed": changed,
        "previous_file": "、".join(previous.get("files", [previous.get("file", "")])),
    }
    pending_batches.pop(category, None)
    save_source_manifest(manifest)
    return categories[category]


def clear_upload_batch(category):
    if category not in UPLOAD_TARGETS:
        raise ValueError("未知上传分类")
    manifest = load_source_manifest()
    pending_batches = manifest.setdefault("pending_batches", {})
    count = len(pending_batches.get(category, {}).get("files", []))
    pending_batches.pop(category, None)
    save_source_manifest(manifest)
    return {"category": category, "cleared": count}


def source_group_status():
    manifest = load_source_manifest().get("categories", {})
    pending_batches = load_source_manifest().get("pending_batches", {})
    groups = []
    for key, config in WEEKLY_SOURCE_GROUPS.items():
        uploaded_paths = [] if key == "erp_base" else manifest_paths(key)
        files = uploaded_paths if uploaded_paths else matching_files(config["folder"], config["patterns"])
        if key == "erp_base":
            files = erp_base_files()
        latest = files[0] if files else None
        if files:
            latest = max(files, key=lambda p: (p.stat().st_mtime, p.name))
        pending = pending_batches.get(key, {}).get("files", [])
        item = {
            "key": key,
            "name": config["name"],
            "description": config["description"],
            "count": len(files),
            "latest": None,
            "status": "缺少数据",
            "changed": False,
            "upload_target": key,
            "batch_files": [],
            "total_rows": "",
            "pending_count": len(pending),
            "pending_files": [item.get("file", "") for item in pending if item.get("file")],
            "pending_batch_id": pending_batches.get(key, {}).get("batch_id", ""),
            "pending_started_at": pending_batches.get(key, {}).get("started_at", ""),
            "batch_id": "",
            "uploaded_at": "",
        }
        if latest:
            stat = latest.stat()
            uploaded = {} if key == "erp_base" else manifest.get(key, {})
            latest_hash = ""
            if uploaded.get("path") == str(latest) and uploaded.get("sha256"):
                latest_hash = uploaded["sha256"]
            item["latest"] = {
                "name": latest.name,
                "path": str(latest),
                "size": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "rows": readable_row_count(latest),
            }
            item["batch_files"] = uploaded.get("files") or [path.name for path in files]
            item["total_rows"] = uploaded.get("rows", "")
            item["batch_id"] = uploaded.get("batch_id", "")
            item["uploaded_at"] = uploaded.get("uploaded_at", "")
            if key == "erp_base":
                item["total_rows"] = readable_row_count(latest)
                item["batch_id"] = latest.stem
                item["uploaded_at"] = item["latest"]["modified"]
            item["recompute"] = source_recompute_state(key, item["uploaded_at"])
            item["count"] = len(files)
            if uploaded_paths:
                item["changed"] = bool(uploaded.get("changed"))
                if item["recompute"].get("needed"):
                    item["status"] = "需重算任务"
                else:
                    item["status"] = "已更新" if item["changed"] else "未变化"
            elif latest_hash and uploaded.get("sha256") == latest_hash:
                item["status"] = "未变化"
            else:
                item["status"] = "已有数据"
        if item["pending_count"]:
            item["status"] = "待结束上传"
        groups.append(item)
    return groups


def unique_upload_path(folder, filename):
    folder.mkdir(exist_ok=True)
    name = safe_name(Path(filename).name, "upload.xlsx")
    path = folder / name
    if not path.exists():
        return path
    stem = Path(name).stem
    suffix = Path(name).suffix
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return folder / f"{stem}_{stamp}{suffix}"


def xlsx_path_for_legacy_xls(path):
    target = path.with_suffix(".xlsx")
    if not target.exists():
        return target
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return path.with_name(f"{path.stem}_{stamp}.xlsx")


def convert_xls_to_xlsx(source, target):
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("读取旧版 .xls 文件需要安装 xlrd：pip install xlrd") from exc

    book = xlrd.open_workbook(str(source))
    output = Workbook()
    default_sheet = output.active
    for sheet_index in range(book.nsheets):
        sheet = book.sheet_by_index(sheet_index)
        ws = default_sheet if sheet_index == 0 else output.create_sheet()
        ws.title = safe_name(sheet.name, f"Sheet{sheet_index + 1}")[:31]
        for row_index in range(sheet.nrows):
            ws.append([sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)])
    output.save(target)
    return target


def normalize_uploaded_workbook(path):
    if path.suffix.lower() == ".xls":
        return convert_xls_to_xlsx(path, xlsx_path_for_legacy_xls(path))
    return path


def data_status():
    db_ok = DB_PATH.exists()
    db_tables = 0
    db_rows = 0
    if db_ok:
        try:
            with sqlite3.connect(DB_PATH) as con:
                db_tables = con.execute("select count(*) from data_sources").fetchone()[0]
                db_rows = con.execute("select coalesce(sum(row_count), 0) from data_sources").fetchone()[0]
        except sqlite3.Error:
            db_ok = False
    return {
        "version": APP_VERSION,
        "temu_files": len(list(TEMU_DIR.glob("*.xlsx"))) if TEMU_DIR.exists() else 0,
        "shein_files": len(list(SHEIN_DIR.glob("*.xlsx"))) if SHEIN_DIR.exists() else 0,
        "erp_files": len(list(ERP_DIR.glob("*.xlsx"))) if ERP_DIR.exists() else 0,
        "source_groups": source_group_status(),
        "outputs": recent_outputs(80),
        "database": {"exists": db_ok, "tables": db_tables, "rows": db_rows, "path": str(DB_PATH)},
        "tasks": operation_task_summary(),
        "overview": operation_overview(),
        "report_tasks": report_task_summary(),
        "reports": REPORTS,
        "rules": load_rules(),
        "upload_targets": {key: label for key, (label, _folder) in UPLOAD_TARGETS.items()},
        "backup_reminder": monthly_backup_reminder(),
    }


def public_status(payload):
    safe = dict(payload)
    safe["source_groups"] = []
    safe["outputs"] = []
    safe["reports"] = {}
    safe["rules"] = {}
    safe["upload_targets"] = {}
    safe["report_tasks"] = {}
    safe["tasks"] = {}
    safe["overview"] = {}
    database = dict(safe.get("database") or {})
    database["path"] = ""
    safe["database"] = database
    return safe


def desktop_status_for_operator(payload, role="admin"):
    if norm(role) == "admin":
        return payload
    safe = public_status(payload)
    safe["reports"] = payload.get("reports", {})
    safe["upload_targets"] = payload.get("upload_targets", {})
    safe["backup_reminder"] = {"backup_exists": True, "message": ""}
    safe["source_groups"] = []
    for group in payload.get("source_groups", []):
        safe["source_groups"].append({
            "key": group.get("key", ""),
            "name": group.get("name", ""),
            "description": group.get("description", ""),
            "count": 0,
            "latest": None,
            "status": "待上传" if not group.get("pending_count") else "待结束上传",
            "changed": False,
            "upload_target": group.get("upload_target") or group.get("key", ""),
            "batch_files": [],
            "total_rows": "",
            "pending_count": group.get("pending_count", 0),
            "pending_files": [],
            "pending_batch_id": "",
            "pending_started_at": "",
            "batch_id": "",
            "uploaded_at": "",
        })
    return safe


def handle_status_api(headers):
    try:
        payload = data_status()
        token = token_from_headers(headers or {})
        if token:
            operator = operator_from_token(token)
            if can_review_tasks(operator):
                return json_bytes({"ok": True, **payload})
        return json_bytes({"ok": True, **public_status(payload)})
    except PermissionError:
        return json_bytes({"ok": True, **public_status(data_status())})
    except Exception as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=500)


def load_module(path, name):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


def run_temu_inventory(output):
    module = load_module(ROOT / "generate_temu_inventory_abnormal.py", "_daily_temu_inventory")
    module.ROOT = ROOT
    module.TEMU_DIR = TEMU_DIR
    module.ERP_DIR = ERP_DIR
    module.ERP_FILES = erp_base_files()
    module.SIZE_ORDER = size_order_map(load_rules().get("sort", {}).get("size_order", DEFAULT_RULES["sort"]["size_order"]))
    module.COMBO_FILE = erp_combo_files()[0] if erp_combo_files() else ERP_DIR / "erp产品组合装基础信息表.xlsx"
    module.OWNER_FILE = OWNER_FILE
    module.OUTPUT = output
    def latest_files_override(pattern):
        if "Temu爆旺款" in pattern:
            return temu_hot_files()
        if "Temu仓库销售情况导出" in pattern:
            return temu_sales_files()
        return latest_by_date(TEMU_DIR, pattern)
    module.latest_files = latest_files_override
    owners = module.load_owner_map()
    erp, erp_rows, combo_rows = module.load_erp_records()
    hot, hot_files = module.load_hot_skc()
    groups, summary, source_rows, skipped, source_files = module.read_temu_rows(erp, owners)
    rows_2x, skc_2x, groups_2x = module.expand_group_alerts(groups, 2)
    rows_1x, skc_1x, groups_1x = module.expand_group_alerts(groups, 1)
    module.build_workbook(
        summary,
        hot,
        owners,
        rows_2x,
        skc_2x,
        rows_1x,
        skc_1x,
        source_rows,
        erp_rows,
        combo_rows,
        skipped,
        groups_2x,
        groups_1x,
        source_files,
        hot_files,
    )
    return {"rows": len(rows_2x) + len(rows_1x), "source_files": source_files, "hot_files": hot_files}


def run_temu_hot(output):
    module = load_module(ROOT / "temu_hot_warning_v13.py", "_daily_temu_hot")
    module.ROOT = ROOT
    module.TEMU_DIR = TEMU_DIR
    module.ERP_DIR = ERP_DIR
    hot_files = temu_hot_files()
    sales_files = temu_sales_files()
    if not hot_files:
        raise FileNotFoundError("未找到Temu爆旺款数据")
    if not sales_files:
        raise FileNotFoundError("未找到Temu仓库销售情况导出数据")
    module.HOT_FILE = hot_files[-1]
    module.SALES_FILES = sales_files
    module.ERP_FILES = erp_base_files()
    module.SIZE_ORDER = load_rules().get("sort", {}).get("size_order", DEFAULT_RULES["sort"]["size_order"])
    module.HOT_FILES = hot_files
    module.COMBO_FILE = erp_combo_files()[0] if erp_combo_files() else ERP_DIR / "erp产品组合装基础信息表.xlsx"
    module.OWNER_FILE = OWNER_FILE
    module.OUTPUT_FILE = output
    owners = module.load_owners()
    erp = module.load_erp()
    hot_rows = module.read_hot_rows()
    hot_groups = module.aggregate_groups(hot_rows)
    champions = module.choose_champions(hot_groups)
    sales_rows = module.read_sales_rows(set(champions))
    sales_groups = module.aggregate_groups(sales_rows)
    operation_rows = module.build_rows(owners, erp, hot_groups, sales_groups, champions)
    overview_rows = module.build_overview(operation_rows, hot_groups, owners)
    module.write_workbook(overview_rows, operation_rows)
    return {"rows": len(operation_rows), "source_files": [p.name for p in sales_files], "hot_file": module.HOT_FILE.name}


def run_temu_slow(output):
    module = load_module(ROOT / "generate_temu_slow_moving_weekly.py", "_daily_temu_slow")
    module.ROOT = ROOT
    module.TEMU_SOURCE_DIR = TEMU_DIR
    module.ERP_DIR = ERP_DIR
    module.TEMU_SALES_FILES = temu_sales_files()
    module.ERP_FILES = erp_base_files()
    module.RULES = load_rules()
    module.OUTPUT_DIR = OUTPUT_DIR
    module.OUTPUT = output
    module.main()
    return {"source_files": [p.name for p in module.TEMU_SALES_FILES]}


def run_temu_bargain(output):
    module = load_module(ROOT / "generate_temu_bargain_reply.py", "_daily_temu_bargain")
    input_files = bargain_input_files()
    sales_files = temu_sales_files()
    hot_files = temu_hot_files()
    if not input_files:
        raise FileNotFoundError("未找到Temu议价申报表")
    if not sales_files:
        raise FileNotFoundError("未找到Temu仓库销售情况导出数据")
    if not hot_files:
        raise FileNotFoundError("未找到Temu爆旺款数据")
    module.INPUT_FILE = input_files[0]
    module.SALES_FILES = sales_files
    module.HOT_FILES = hot_files
    module.ERP_FILES = erp_base_files()
    module.COMBO_FILES = erp_combo_files()
    input_rows = module.load_input_rows(module.INPUT_FILE)
    erp, erp_by_name = module.load_erp()
    input_rows = module.enrich_input_rows(input_rows, erp, erp_by_name)
    links_by_key, sales_source = module.load_temu_links()
    hot_keys, hot_source = module.load_hot_keys()
    output_rows = module.build_output_rows(input_rows, erp, links_by_key, hot_keys)
    module.write_workbook(output_rows, output)
    return {
        "rows": len(output_rows),
        "input_file": Path(module.INPUT_FILE).name,
        "source_files": sales_source,
        "hot_files": hot_source,
        "auto_matched": sum(1 for row in input_rows if row.get("商家编码")),
    }


def run_low_score_warning(output):
    module = load_module(ROOT / "generate_low_score_warning.py", "_daily_low_score")
    current_files = low_score_input_files()
    sales_files = temu_sales_files()
    hot_files = temu_hot_files()
    erp_files = erp_base_files()
    if not current_files:
        raise FileNotFoundError("未找到低分预警输入表")
    if not sales_files:
        raise FileNotFoundError("未找到Temu仓库销售情况导出数据")
    if not hot_files:
        raise FileNotFoundError("未找到Temu爆旺款数据")
    if not erp_files:
        raise FileNotFoundError("未找到ERP基础信息")
    module.CURRENT_FILES = current_files
    module.HISTORY_FILES = module.latest_distinct_history_files(current_files)
    module.TEMU_SALES_FILES = sales_files
    module.TEMU_HOT_FILES = hot_files
    module.ERP_FILES = erp_files
    module.OUTPUT = output
    result = module.main()
    result["sales_files"] = [path.name for path in sales_files]
    result["hot_files"] = [path.name for path in hot_files]
    return result


def run_shein_price(output):
    module = load_module(ROOT / "generate_shein_price_abnormal.py", "_daily_shein_price")
    module.ROOT = ROOT
    module.SHEIN_DIR = SHEIN_DIR
    module.SHEIN_FILES = shein_platform_files()
    module.OUTPUT = output
    module.main()
    return inspect_workbook_summary(output)


def run_shein_inventory(output):
    module = load_module(ROOT / "generate_shein_inventory_abnormal.py", "_daily_shein_inventory")
    module.ROOT = ROOT
    module.SHEIN_DIR = SHEIN_DIR
    module.ERP_DIR = ERP_DIR
    module.ERP_FILES = erp_base_files()
    module.SHEIN_FILES = shein_platform_files()
    module.OUTPUT = output
    module.OWNERS = {**module.OWNERS, **load_owners()}
    erp_records, erp_rows = module.load_erp_base()
    combo_rows = module.load_combo_base(erp_records)
    summary, gt_2x, gt_1x, source_rows, skipped = module.read_shein(erp_records)
    module.build_workbook(summary, gt_2x, gt_1x, source_rows, erp_rows, combo_rows, skipped)
    return {"rows": len(gt_2x) + len(gt_1x), "source_rows": source_rows, "skipped": skipped}


def write_shein_hot_workbook(payload, output):
    wb = Workbook()
    overview_headers = [
        "店铺编号", "店铺", "负责人", "爆款总数", "重复铺货预计总数", "平销冲突数",
        "爆款互相冲突数", "低于爆款报价数", "不低于爆款报价数", "立即下架数",
        "售完备货库存下架数", "30天内限时下架数",
    ]
    ws = wb.active
    ws.title = "总览"
    ws.append(overview_headers)
    for row in payload["overview"]:
        ws.append([row.get(header, "") for header in overview_headers])

    operation_headers = [
        "商家编码", "货品名称", "skc", "所属店铺", "爆旺款skc", "爆旺款店铺",
        "爆款报价", "重复款申报价", "爆款月销件数", "重复款月销件数",
        "是否低于爆款报价", "爆款库存", "重复款库存", "负责人", "冲突类型", "处理意见",
    ]
    ws_op = wb.create_sheet("具体店铺操作表")
    ws_op.append(operation_headers)
    for row in payload["operations"]:
        ws_op.append([row.get(header, "") for header in operation_headers])

    notes = wb.create_sheet("说明")
    notes.append(["项目", "说明"])
    for key, value in payload["notes"]:
        notes.append([key, value])

    for sheet in wb.worksheets:
        style_basic_sheet(sheet)

    if ws_op.max_row >= 2:
        headers = [cell.value for cell in ws_op[1]]
        low_col = headers.index("是否低于爆款报价") + 1
        advice_col = headers.index("处理意见") + 1
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        red_font = Font(color="C00000", bold=True)
        for row in ws_op.iter_rows(min_row=2, max_row=ws_op.max_row):
            if row[low_col - 1].value == "是":
                row[low_col - 1].fill = red_fill
            if row[advice_col - 1].value == "立即下架！":
                row[advice_col - 1].font = red_font
    wb.save(output)


def shein_hot_rule_note(rules):
    hot = rules.get("hot_item", {}) if isinstance(rules, dict) else {}
    new_days = hot.get("shein_new_days_lt", DEFAULT_RULES["hot_item"]["shein_new_days_lt"])
    new_avg = hot.get("shein_new_7d_daily_gte", DEFAULT_RULES["hot_item"]["shein_new_7d_daily_gte"])
    old_days = hot.get("shein_old_days_gte", DEFAULT_RULES["hot_item"]["shein_old_days_gte"])
    old_avg = hot.get("shein_old_30d_daily_gt", DEFAULT_RULES["hot_item"]["shein_old_30d_daily_gt"])
    return f"新品上架天数小于{new_days}天且7天销量日均>={new_avg}；老品上架天数大于等于{old_days}天且30天销量日均>{old_avg}。"


def run_shein_hot(output):
    module = load_module(ROOT / "shein_hot_warning_v11_analysis.py", "_daily_shein_hot")
    module.ROOT = ROOT
    module.ERP_FILES = erp_base_files()
    module.COMBO_FILES = erp_combo_files()
    module.SOURCE_FILES = shein_source_map()
    module.RULES = load_rules()
    module.STORE_OWNER = {**module.STORE_OWNER, **load_owners()}
    records, checks = module.load_sales_records()
    skc, hot, style_combos, combo_rows = module.aggregate(records)
    operations = module.build_operations(skc, hot, style_combos, combo_rows)
    overview = module.build_overview(hot, operations)
    payload = {
        "overview": overview,
        "operations": operations,
        "notes": [
            ("版本", "V2"),
            ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("表格结构", "总览、具体店铺操作表、说明。"),
            ("数据源", "使用 shein数据源表 内各店铺最新销售表，ERP 匹配使用 erp数据源。"),
            ("SKU口径", "商家SKU 匹配时忽略 @ 及后续字符。"),
            ("爆旺款定义", shein_hot_rule_note(module.RULES)),
            ("爆旺款优先级定义", "同款冲突时，30天销量越高优先级越高；销量相同则申报价越高优先级越高；再相同则平台仓备货越多优先级越高。"),
            ("处理意见规则", "低价平销款立即下架；非低价且有库存售完下架；低优先级爆旺款有库存30天内限时下架。"),
            ("SHEIN店铺字段说明", "所属店铺和爆旺款店铺直接展示 SHEIN 店铺名。"),
            ("具体店铺操作表行数", str(len(operations))),
        ],
    }
    write_shein_hot_workbook(payload, output)
    return {**checks, "rows": len(operations), "hot_skc": len(hot)}


def inspect_workbook_summary(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    result = {
        "sheets": wb.sheetnames,
        "rows": {sheet: max((wb[sheet].max_row or 1) - 1, 0) for sheet in wb.sheetnames},
    }
    wb.close()
    return result


def run_weekly_reports():
    weekly = [
        ("temu_price", "V1"),
        ("temu_inventory", "V1"),
        ("temu_hot", "V1.1"),
        ("temu_slow", "V3"),
        ("temu_bargain", "V1"),
        ("low_score_warning", "V1"),
        ("shein_price", "V1"),
        ("shein_inventory", "V1"),
        ("shein_hot", "V2"),
    ]
    results = []
    task_sync_total = {"created": 0, "updated": 0, "imported_rows": 0}
    for report_id, version in weekly:
        try:
            result = run_report(report_id, version)
            result["status"] = "ok"
            result["report"] = report_id
            task_sync = result.get("task_sync") or {}
            for key in task_sync_total:
                task_sync_total[key] += int(task_sync.get(key) or 0)
        except Exception as exc:
            result = {
                "status": "failed",
                "report": report_id,
                "name": REPORTS.get(report_id, {}).get("name", report_id),
                "error": str(exc),
            }
        results.append(result)
    ok_count = sum(1 for item in results if item.get("status") == "ok")
    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": {"total": len(results), "ok": ok_count, "failed": len(results) - ok_count},
        "task_sync": task_sync_total,
        "results": results,
    }


def recompute_reports_for_source(category):
    reports = SOURCE_DEPENDENT_REPORTS.get(category, [])
    if category not in WEEKLY_SOURCE_GROUPS and category not in UPLOAD_TARGETS:
        raise ValueError("未知数据源")
    if not reports:
        return {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "category": category,
            "summary": {"total": 0, "ok": 0, "failed": 0},
            "task_sync": {"created": 0, "updated": 0, "imported_rows": 0},
            "results": [],
        }
    results = []
    task_sync_total = {"created": 0, "updated": 0, "imported_rows": 0}
    for report_id in reports:
        try:
            result = run_report(report_id, "V1")
            result["status"] = "ok"
            result["report"] = report_id
            task_sync = result.get("task_sync") or {}
            for key in task_sync_total:
                task_sync_total[key] += int(task_sync.get(key) or 0)
        except Exception as exc:
            result = {
                "status": "failed",
                "report": report_id,
                "name": REPORTS.get(report_id, {}).get("name", report_id),
                "error": str(exc),
            }
        results.append(result)
    ok_count = sum(1 for item in results if item.get("status") == "ok")
    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "category": category,
        "summary": {"total": len(results), "ok": ok_count, "failed": len(results) - ok_count},
        "task_sync": task_sync_total,
        "results": results,
    }


def norm(value):
    return html.unescape(str(value)).strip() if value is not None else ""


def to_number(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return 0.0


def sku_key(value):
    text = norm(value).upper()
    if "@" in text:
        text = text.split("@", 1)[0]
    return text.strip()


def header_map(row):
    return {norm(v): i for i, v in enumerate(row) if norm(v)}


def cell(row, headers, *names):
    for name in names:
        if name in headers and headers[name] < len(row):
            return row[headers[name]]
    return None


def is_shein_active_listing(supply_status, listing_status):
    listing = norm(listing_status)
    return listing == "已上架"


def load_owners():
    owners = {}
    for path in owner_files():
        if not path.exists():
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            continue
        headers = header_map(rows[0])
        for row in rows[1:]:
            store = norm(cell(row, headers, "店铺", "店铺名称", "店铺名"))
            owner = norm(cell(row, headers, "业务", "负责人"))
            if store and owner:
                owners[store] = owner
                for key in owner_lookup_keys(store):
                    owners.setdefault(key, owner)
    for assignment in load_store_owner_assignments():
        store = assignment["store"]
        owner = assignment["owner"]
        owners[store] = owner
        for key in owner_lookup_keys(store):
            owners.setdefault(key, owner)
    return owners


def load_erp_price_map():
    records = {}
    for path in erp_base_files():
        rows = raw_xlsx.read_xlsx_rows(path)
        if not rows:
            continue
        headers = header_map(rows[0])
        for row in rows[1:]:
            code = sku_key(cell(row, headers, "商家编码（新）"))
            if not code or code in records:
                continue
            records[code] = {
                "货品名称": norm(cell(row, headers, "货品名称")),
                "规格名称": norm(cell(row, headers, "规格名称")),
                "成本价": to_number(cell(row, headers, "成本价")),
                "批发价": to_number(cell(row, headers, "批发报价", "批发价")),
                "来源": path.name,
            }
    for combo in erp_combo_files():
        rows = raw_xlsx.read_xlsx_rows(combo)
        headers = header_map(rows[0]) if rows else {}
        for row in rows[1:]:
            code = sku_key(cell(row, headers, "商家编码（新）"))
            if not code or code in records:
                continue
            records[code] = {
                "货品名称": norm(cell(row, headers, "组合装名称", "货品名称")),
                "规格名称": norm(cell(row, headers, "组合装简称", "规格名称")),
                "成本价": to_number(cell(row, headers, "成本价")),
                "批发价": to_number(cell(row, headers, "批发价", "批发报价")),
                "来源": combo.name,
            }
    return records


def run_temu_price(output):
    files = temu_sales_files()
    if not files:
        raise FileNotFoundError("未找到Temu仓库销售情况导出数据")
    erp = load_erp_price_map()
    if not erp:
        raise FileNotFoundError("未读取到ERP基础信息")
    owners = load_owners()
    below_cost = []
    below_wholesale = []
    summary = {}
    skipped_zero = 0
    skipped_erp = 0
    source_rows = 0
    for path in files:
        rows = raw_xlsx.read_xlsx_rows(path)
        if not rows:
            continue
        headers = header_map(rows[0])
        for row_number, row in enumerate(rows[1:], start=2):
            source_rows += 1
            sku = sku_key(cell(row, headers, "SKU货号", "商家编码", "商品SKU"))
            if not sku:
                continue
            price = to_number(cell(row, headers, "申报价格", "申报价"))
            if price == 0:
                skipped_zero += 1
                continue
            info = erp.get(sku)
            if not info:
                skipped_erp += 1
                continue
            store = norm(cell(row, headers, "店铺"))
            skc = norm(cell(row, headers, "SKC"))
            sales7 = to_number(cell(row, headers, "7天销量"))
            sales30 = to_number(cell(row, headers, "30天销量"))
            summary.setdefault(store, {"skc": set(), "sales7": 0, "sales30": 0, "cost_skc": set(), "wholesale_skc": set()})
            summary[store]["skc"].add(skc)
            summary[store]["sales7"] += sales7
            summary[store]["sales30"] += sales30
            base = {
                "店铺": store,
                "SKC": skc,
                "商家编码": sku,
                "货品名称": info["货品名称"],
                "货品规格": info["规格名称"],
                "申报价": price,
                "成本价": info["成本价"],
                "批发价": info["批发价"],
                "批发价80%": round(info["批发价"] * 0.8, 2),
                "负责人": owners.get(store, ""),
                "7天销量": sales7,
                "30天销量": sales30,
                "源SKU": norm(cell(row, headers, "SKU货号", "商家编码", "商品SKU")),
                "源文件": path.name,
                "源行": row_number,
            }
            if info["成本价"] and price < info["成本价"]:
                below_cost.append(base)
                summary[store]["cost_skc"].add(skc)
            if info["批发价"] and price < info["批发价"] * 0.8:
                below_wholesale.append(base)
                summary[store]["wholesale_skc"].add(skc)

    wb = Workbook()
    ws = wb.active
    ws.title = "总览表"
    ws.append(["店铺", "负责人", "在售SKC数量", "7天销量", "30天销量", "亏损SKC数量", "破价SKC数量"])
    for store in sorted(summary):
        item = summary[store]
        ws.append([store, owners.get(store, ""), len(item["skc"]), item["sales7"], item["sales30"], len(item["cost_skc"]), len(item["wholesale_skc"])])
    write_rows(wb, "低于成本价", below_cost, ["店铺", "SKC", "商家编码", "货品名称", "货品规格", "申报价", "成本价", "负责人", "7天销量", "30天销量", "源SKU", "源文件", "源行"])
    write_rows(wb, "低于批发价80%", below_wholesale, ["店铺", "SKC", "商家编码", "货品名称", "货品规格", "申报价", "成本价", "批发价", "批发价80%", "负责人", "7天销量", "30天销量", "源SKU", "源文件", "源行"])
    check = wb.create_sheet("数据校验")
    for row in [
        ["检查项", "结果"],
        ["Temu源文件", "；".join(p.name for p in files)],
        ["源明细行数", source_rows],
        ["申报价为0跳过", skipped_zero],
        ["未匹配ERP跳过", skipped_erp],
        ["低于成本价明细", len(below_cost)],
        ["低于批发价80%明细", len(below_wholesale)],
    ]:
        check.append(row)
    for sheet in wb.worksheets:
        style_basic_sheet(sheet)
    wb.save(output)
    return {"rows": len(below_cost) + len(below_wholesale), "source_files": [p.name for p in files]}


def write_rows(wb, sheet_name, rows, columns):
    ws = wb.create_sheet(sheet_name)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])


def style_basic_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        width = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                width = max(width, len(str(value)) + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(width, 42)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False


def run_report(report_id, version):
    if report_id not in REPORTS:
        raise ValueError("未知报表类型")
    report_name = REPORTS[report_id]["name"]
    out = output_path(report_name, version)
    if report_id == "temu_price":
        detail = run_temu_price(out)
    elif report_id == "temu_inventory":
        detail = run_temu_inventory(out)
    elif report_id == "temu_hot":
        detail = run_temu_hot(out)
    elif report_id == "temu_slow":
        detail = run_temu_slow(out)
    elif report_id == "temu_bargain":
        detail = run_temu_bargain(out)
    elif report_id == "low_score_warning":
        detail = run_low_score_warning(out)
    elif report_id == "shein_price":
        detail = run_shein_price(out)
    elif report_id == "shein_inventory":
        detail = run_shein_inventory(out)
    elif report_id == "shein_hot":
        detail = run_shein_hot(out)
    else:
        raise ValueError("暂未接入该报表")
    task_sync = sync_report_tasks(report_id, out)
    return {"file": out.name, "download": f"/download?path={quote(out.name)}", "detail": detail, "task_sync": task_sync}


def search_database(query, limit=100):
    query = (query or "").strip()
    if not query:
        return []
    limit = max(1, min(int(limit or 100), 500))
    if not DB_PATH.exists():
        raise FileNotFoundError("基础数据库不存在")
    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        try:
            rows = con.execute(
                """
                select file_name, sheet_name, table_name, source_row, content
                from search_index
                where search_index match ?
                limit ?
                """,
                (f'"{query.replace(chr(34), chr(34) + chr(34))}"', limit),
            ).fetchall()
        except sqlite3.Error:
            rows = []
        if not rows:
            rows = con.execute(
                """
                select file_name, sheet_name, table_name, source_row, content
                from search_index
                where content like ?
                limit ?
                """,
                (f"%{query}%", limit),
            ).fetchall()
    return [dict(row) for row in rows]


def query_erp_product_info(query, limit=100):
    query = (query or "").strip()
    limit = max(1, min(int(limit or 100), 300))
    terms = [item.strip().lower() for item in re.split(r"\s+", query) if item.strip()]
    files = erp_base_files()
    matches = []
    preferred_headers = [
        "商家编码（新）", "商家编码", "货品编号", "货品名称", "规格名称",
        "成本价", "批发价", "零售价", "分类", "品牌", "供应商",
    ]
    for path in files:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        try:
            for ws in wb.worksheets:
                rows = ws.iter_rows(values_only=True)
                try:
                    headers = [daily_ops_master_data.norm(value) or f"列{idx + 1}" for idx, value in enumerate(next(rows))]
                except StopIteration:
                    continue
                for row_index, row in enumerate(rows, start=2):
                    values = [daily_ops_master_data.norm(value) for value in row]
                    haystack = " ".join(values).lower()
                    if terms and not all(term in haystack for term in terms):
                        continue
                    record = {
                        headers[idx]: values[idx]
                        for idx in range(min(len(headers), len(values)))
                        if values[idx]
                    }
                    summary = {key: record.get(key, "") for key in preferred_headers if record.get(key)}
                    if not summary:
                        summary = dict(list(record.items())[:8])
                    matches.append({
                        "file_name": path.name,
                        "sheet_name": ws.title,
                        "source_row": row_index,
                        "summary": summary,
                        "content": "　".join(f"{key}: {value}" for key, value in list(record.items())[:12]),
                    })
                    if len(matches) >= limit:
                        return {"items": matches, "source_files": [file.name for file in files]}
        finally:
            wb.close()
    return {"items": matches, "source_files": [path.name for path in files]}


def export_search(query, limit=500):
    rows = search_database(query, limit)
    out = output_path(f"基础数据查询-{safe_name(query, '关键词')}", "V1")
    wb = Workbook()
    ws = wb.active
    ws.title = "查询结果"
    headers = ["来源文件", "工作表", "数据表名", "原始行号", "命中内容"]
    ws.append(headers)
    for row in rows:
        ws.append([row["file_name"], row["sheet_name"], row["table_name"], row["source_row"], row["content"]])
    style_basic_sheet(ws)
    wb.save(out)
    return {"file": out.name, "download": f"/download?path={quote(out.name)}", "rows": len(rows)}


def handle_search_api(action, headers, payload):
    try:
        operator = operator_from_token(token_from_headers(headers))
        if not can_review_tasks(operator):
            return json_bytes({"ok": False, "error": f"只有管理员可以执行{action}"}, status=403)
        if action == "GET":
            rows = search_database(payload.get("q", ""), payload.get("limit", "100"))
            return json_bytes({"ok": True, "rows": rows})
        return json_bytes({"ok": False, "error": "基础数据查询接口不存在"}, status=404)
    except PermissionError as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=401)
    except Exception as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=500)


def handle_rules_api(action, headers, payload=None):
    try:
        operator = operator_from_token(token_from_headers(headers))
        if not can_review_tasks(operator):
            return json_bytes({"ok": False, "error": "只有管理员可以维护规则"}, status=403)
        if action == "GET":
            return json_bytes({"ok": True, "rules": load_rules()})
        if action == "POST_SAVE":
            return json_bytes({"ok": True, "rules": save_rules(payload or {})})
        return json_bytes({"ok": False, "error": "规则接口不存在"}, status=404)
    except PermissionError as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=401)
    except Exception as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=500)


def bargain_store():
    return daily_ops_bargain.BargainStore(BARGAIN_DB_FILE)


def save_clearance_catalog(catalog):
    CLEARANCE_CATALOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    CLEARANCE_CATALOG_FILE.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    return catalog


def load_clearance_catalog():
    if CLEARANCE_CATALOG_FILE.exists():
        try:
            return json.loads(CLEARANCE_CATALOG_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass
    catalog = daily_ops_bargain.build_clearance_catalog(erp_base_files())
    return save_clearance_catalog(catalog)


def rebuild_clearance_catalog():
    return save_clearance_catalog(daily_ops_bargain.build_clearance_catalog(erp_base_files()))


def bargain_platform_rows(payload=None):
    rows = []
    for row in (payload or {}).get("platform_rows") or []:
        rows.append(row)
    if not rows:
        rows.extend(platform_source_rows_for_bargain((payload or {}).get("platform", "")))
    return rows


def lookup_bargain_staging(payload):
    rows = bargain_store().lookup_staging_rows(
        merchant_code=payload.get("merchant_code", ""),
        request_store=payload.get("store", ""),
        platform=payload.get("platform", ""),
        owner=payload.get("owner", ""),
        erp_files=erp_base_files(),
        cost_files=erp_cost_files(),
        clearance_catalog=load_clearance_catalog(),
        platform_rows=bargain_platform_rows(payload),
    )
    return {"rows": rows, "clearance": load_clearance_catalog().get("summary", {})}


def submit_bargain_batch(payload):
    return bargain_store().submit_batch(
        payload.get("store", ""),
        payload.get("platform", ""),
        payload.get("owner", ""),
        payload.get("lines", []),
    )


def review_bargain_lines(payload):
    return bargain_store().review_lines(
        payload.get("batch_id", ""),
        payload.get("line_ids", []),
        payload.get("decision", ""),
        payload.get("admin", "管理员"),
        payload.get("remark", ""),
    )


def resubmit_bargain_line(payload):
    return bargain_store().resubmit_line(
        payload.get("line_id", ""),
        payload.get("price", ""),
        payload.get("owner", ""),
    )


def bargain_history(filters=None):
    return {"rows": bargain_store().history(filters or {})}


def low_price_trace(payload=None):
    payload = payload or {}
    return {"rows": bargain_store().low_price_trace(payload.get("platform_rows", []), payload.get("tolerance", 0.05))}


def ignore_low_price(payload):
    return bargain_store().ignore_low_price(payload.get("risk_ids", []), payload.get("actor", "管理员"), payload.get("remark", ""))


def handle_bargain_api(action, headers, payload=None):
    payload = payload or {}
    try:
        operator = operator_from_token(token_from_headers(headers))
        role = operator.get("role") or "admin"
        user = operator.get("user") or "管理员"
        if action == "GET_HISTORY":
            if not can_review_tasks(operator):
                payload["owner"] = user
            return json_bytes({"ok": True, **bargain_history(payload)})
        if action == "GET_CLEARANCE":
            if not can_review_tasks(operator):
                return json_bytes({"ok": False, "error": "只有管理员可以查看清仓款式表"}, status=403)
            return json_bytes({"ok": True, **load_clearance_catalog()})
        if action == "POST_REBUILD_CLEARANCE":
            if not can_review_tasks(operator):
                return json_bytes({"ok": False, "error": "只有管理员可以重建清仓款式表"}, status=403)
            return json_bytes({"ok": True, **rebuild_clearance_catalog()})
        if action == "POST_LOOKUP":
            if role == "owner":
                payload["owner"] = user
            return json_bytes({"ok": True, **lookup_bargain_staging(payload)})
        if action == "POST_SUBMIT":
            if role == "owner":
                payload["owner"] = user
            return json_bytes({"ok": True, "batch": submit_bargain_batch(payload)})
        if action == "POST_REVIEW":
            if not can_review_tasks(operator):
                return json_bytes({"ok": False, "error": "只有管理员可以审批议价"}, status=403)
            payload["admin"] = user
            return json_bytes({"ok": True, **review_bargain_lines(payload)})
        if action == "POST_RESUBMIT":
            if role == "owner":
                payload["owner"] = user
            return json_bytes({"ok": True, "batch": resubmit_bargain_line(payload)})
        if action == "POST_LOW_PRICE_TRACE":
            if not can_review_tasks(operator):
                return json_bytes({"ok": False, "error": "只有管理员可以低价回追"}, status=403)
            return json_bytes({"ok": True, **low_price_trace(payload)})
        if action == "POST_IGNORE_LOW_PRICE":
            if not can_review_tasks(operator):
                return json_bytes({"ok": False, "error": "只有管理员可以忽略低价风险"}, status=403)
            payload["actor"] = user
            return json_bytes({"ok": True, **ignore_low_price(payload)})
        return json_bytes({"ok": False, "error": "议价接口不存在"}, status=404)
    except PermissionError as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=401)
    except Exception as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=500)


def operation_task_store():
    return daily_ops_tasks.OperationTaskStore(TASK_DB_PATH)


def task_suppression_store():
    return daily_ops_task_suppression.TaskSuppressionStore(TASK_SUPPRESSION_FILE)


def assign_existing_unassigned_tasks(assignments, actor="管理员"):
    store = operation_task_store()
    with store._lock:
        payload = store.load()
        timestamp = now_text()
        assigned = 0
        for task in payload.get("tasks", []):
            if norm(task.get("owner")):
                continue
            owner = owner_from_assignments(assignments, task.get("platform", ""), task.get("store", ""))
            if not owner:
                continue
            task["owner"] = owner
            task["updated_at"] = timestamp
            task.setdefault("history", []).append(daily_ops_tasks.history_entry(
                task,
                norm(actor) or "管理员",
                "自动指派",
                f"按店铺负责人配置指派给 {owner}",
                "保存店铺负责人配置时自动补齐",
                time=timestamp,
            ))
            assigned += 1
        if assigned:
            store.save(payload)
        return assigned


def apply_store_owner_mapping(rows):
    mapped = []
    for row in rows:
        item = dict(row)
        if not norm(item.get("owner")):
            owner = configured_owner_for_store(item.get("platform", ""), item.get("store", ""))
            if owner:
                item["owner"] = owner
                item["owner_source"] = "店铺负责人配置"
        mapped.append(item)
    return mapped


def sync_report_tasks(report_id, workbook_path):
    report = REPORTS.get(report_id, {})
    rows = daily_ops_tasks.rows_from_report_workbook(report_id, report.get("name", report_id), workbook_path)
    batch_id = Path(workbook_path).stem
    rows = [{**row, "source_batch_id": batch_id} for row in rows]
    rows = apply_store_owner_mapping(rows)
    rows, suppressed = task_suppression_store().filter_rows(rows)
    result = operation_task_store().upsert_generated_tasks(rows, default_status=daily_ops_tasks.STATUS_PENDING_PUSH)
    result["imported_rows"] = len(rows)
    result["suppressed_rows"] = len(suppressed)
    return result


def operation_task_summary():
    return operation_task_store().summary()


def report_id_for_task(row):
    source_report = norm(row.get("source_report"))
    for report_id, report in REPORTS.items():
        if source_report == report.get("name"):
            return report_id
    return ""


def report_task_summary():
    result = {report_id: {"total": 0, "by_status": {}} for report_id in REPORTS}
    for row in operation_task_store().list_tasks():
        report_id = report_id_for_task(row)
        if not report_id:
            continue
        item = result.setdefault(report_id, {"total": 0, "by_status": {}})
        item["total"] += 1
        status = norm(row.get("status"))
        item["by_status"][status] = item["by_status"].get(status, 0) + 1
    return result


def operation_overview():
    rows = operation_task_store().list_tasks()
    summary = operation_task_store().summary(rows)
    packages = operation_task_store().task_packages(rows)
    stores = {}
    owners = {}
    for row in rows:
        store_name = norm(row.get("store")) or "未填写店铺"
        owner = norm(row.get("owner")) or "未分配"
        task_type = norm(row.get("task_type"))
        status = norm(row.get("status"))
        store = stores.setdefault(store_name, {
            "store": store_name,
            "owner": owner,
            "owners": set(),
            "task_total": 0,
            "open_count": 0,
            "pending_push_count": 0,
            "pending_owner_count": 0,
            "pending_review_count": 0,
            "done_count": 0,
            "low_score_count": 0,
            "bargain_count": 0,
            "slow_count": 0,
            "price_count": 0,
            "overdue_count": 0,
        })
        store["owners"].add(owner)
        store["task_total"] += 1
        if status != daily_ops_tasks.STATUS_DONE:
            store["open_count"] += 1
        if status == daily_ops_tasks.STATUS_PENDING_PUSH:
            store["pending_push_count"] += 1
        if status in {daily_ops_tasks.STATUS_PENDING_OWNER, daily_ops_tasks.STATUS_REJECTED}:
            store["pending_owner_count"] += 1
        if status == daily_ops_tasks.STATUS_PENDING_REVIEW:
            store["pending_review_count"] += 1
        if status == daily_ops_tasks.STATUS_DONE:
            store["done_count"] += 1
        if "低分" in task_type:
            store["low_score_count"] += 1
        if "议价" in task_type or "核价" in task_type:
            store["bargain_count"] += 1
        if "滞销" in task_type:
            store["slow_count"] += 1
        if "价格" in task_type or "申报价" in task_type:
            store["price_count"] += 1
        if daily_ops_tasks.task_overdue(row):
            store["overdue_count"] += 1

        owner_item = owners.setdefault(owner, {
            "owner": owner,
            "stores": set(),
            "task_total": 0,
            "open_count": 0,
            "pending_review_count": 0,
            "overdue_count": 0,
        })
        owner_item["stores"].add(store_name)
        owner_item["task_total"] += 1
        if status != daily_ops_tasks.STATUS_DONE:
            owner_item["open_count"] += 1
        if status == daily_ops_tasks.STATUS_PENDING_REVIEW:
            owner_item["pending_review_count"] += 1
        if daily_ops_tasks.task_overdue(row):
            owner_item["overdue_count"] += 1

    package_by_store = {}
    for package in packages:
        store_name = norm(package.get("store")) or "未填写店铺"
        package_by_store[store_name] = package_by_store.get(store_name, 0) + 1
    store_rows = []
    for store in stores.values():
        owners_list = sorted(store.pop("owners"))
        store["owner"] = "、".join([item for item in owners_list if item != "未分配"]) or "未分配"
        store["package_count"] = package_by_store.get(store["store"], 0)
        store_rows.append(store)
    owner_rows = []
    for item in owners.values():
        item["stores"] = sorted(item["stores"])
        item["store_count"] = len(item["stores"])
        owner_rows.append(item)
    store_rows.sort(key=lambda row: (-row["open_count"], -row["task_total"], row["store"]))
    owner_rows.sort(key=lambda row: (-row["open_count"], row["owner"]))
    return {
        "task_total": summary.get("total", 0),
        "package_total": len(packages),
        "open_total": sum(1 for row in rows if norm(row.get("status")) != daily_ops_tasks.STATUS_DONE),
        "pending_push_total": summary.get("by_status", {}).get(daily_ops_tasks.STATUS_PENDING_PUSH, 0),
        "pending_owner_total": summary.get("by_status", {}).get(daily_ops_tasks.STATUS_PENDING_OWNER, 0) + summary.get("by_status", {}).get(daily_ops_tasks.STATUS_REJECTED, 0),
        "pending_review_total": summary.get("by_status", {}).get(daily_ops_tasks.STATUS_PENDING_REVIEW, 0),
        "unassigned_total": summary.get("unassigned", 0),
        "overdue_total": summary.get("overdue", {}).get("total", 0),
        "stores": store_rows,
        "owners": owner_rows,
    }


def summarize_operation_tasks(rows):
    return operation_task_store().summary(rows)


def package_operation_tasks(rows):
    return operation_task_store().task_packages(rows)


def operation_owner_directory():
    owners = {row["owner"]: row for row in operation_task_store().owner_directory()}
    for assignment in load_store_owner_assignments():
        owner = assignment["owner"]
        item = owners.setdefault(owner, {"owner": owner, "stores": [], "platforms": [], "task_count": 0})
        if assignment["store"] not in item["stores"]:
            item["stores"].append(assignment["store"])
        platform = assignment.get("platform", "")
        if platform and platform not in item["platforms"]:
            item["platforms"].append(platform)
    for item in owners.values():
        item["stores"] = sorted(item["stores"])
        item["platforms"] = sorted(item["platforms"])
    return sorted(owners.values(), key=lambda row: (-row["task_count"], row["owner"]))


def filter_operation_task_search(rows, search=""):
    keyword = norm(search)
    if not keyword:
        return rows
    fields = ["store", "owner", "merchant_code", "skc", "spu", "product_name", "task_type", "system_action", "task_detail"]
    return [
        row for row in rows
        if any(keyword in norm(row.get(field)) for field in fields)
    ]


def list_operation_tasks(role="admin", user="", status="", task_type="", store="", platform="", overdue="", unassigned="", next_handler="", priority="", reworked="", open_only="", search=""):
    rows = operation_task_store().list_tasks(
        role=role,
        user=user if norm(role) != "admin" else "",
        status=status,
        task_type=task_type,
        store=store,
        platform=platform,
        overdue=overdue,
        unassigned=unassigned,
        next_handler=next_handler,
        priority=priority,
        reworked=reworked,
        open_only=open_only,
    )
    if norm(role) == "admin" and norm(user):
        rows = [row for row in rows if norm(row.get("owner")) == norm(user)]
    return filter_operation_task_search(rows, search)


def submit_operation_task(task_id, actor, action, remark="", proof=""):
    return operation_task_store().submit_owner_action(task_id, actor, action, remark, proof)


def submit_operation_tasks(task_ids, actor, action, remark="", proof=""):
    return operation_task_store().submit_owner_actions(task_ids, actor, action, remark, proof)


def assign_operation_task(task_id, actor, owner, remark=""):
    return operation_task_store().assign_task(task_id, actor, owner, remark)


def push_operation_tasks(task_ids, actor, remark=""):
    return operation_task_store().push_tasks(task_ids, actor, remark)


def review_operation_task(task_id, admin, decision, remark=""):
    return operation_task_store().review_task(task_id, admin, decision, remark)


def review_operation_tasks(task_ids, admin, decision, remark=""):
    return operation_task_store().review_tasks(task_ids, admin, decision, remark)


def confirm_operation_tasks(task_ids, admin, remark=""):
    return operation_task_store().confirm_review_tasks(task_ids, admin, remark)


def list_task_suppressions():
    return {"items": task_suppression_store().list_items()}


def suppress_operation_tasks(task_ids, actor="管理员", reason="", duration="永久"):
    ids = [norm(task_id) for task_id in task_ids or [] if norm(task_id)]
    if not ids:
        raise ValueError("请选择要屏蔽的任务")
    store = operation_task_store()
    with store._lock:
        payload = store.load()
        by_id = {row.get("id"): row for row in payload.get("tasks", [])}
        rows = []
        for task_id in ids:
            task = by_id.get(task_id)
            if not task:
                raise KeyError("任务不存在")
            rows.append(task)
        suppression = task_suppression_store().add_from_rows(rows, actor, reason, duration)
        timestamp = now_text()
        for task in rows:
            task["status"] = daily_ops_tasks.STATUS_DONE
            task["completed_by"] = norm(actor)
            task["completed_at"] = timestamp
            task["completed_remark"] = norm(reason) or "管理员屏蔽，不再重复提示"
            task["updated_at"] = timestamp
            task.setdefault("history", []).append(daily_ops_tasks.history_entry(
                task,
                norm(actor) or "管理员",
                "任务屏蔽",
                "加入屏蔽清单",
                norm(reason) or "不再重复提示",
                time=timestamp,
            ))
        store.save(payload)
    return {"count": len(rows), "suppression": suppression}


def mark_operation_task_done(task_id, actor, remark=""):
    return operation_task_store().mark_done(task_id, actor, remark)


def mark_operation_tasks_done(task_ids, actor, remark=""):
    return operation_task_store().mark_done_tasks(task_ids, actor, remark)


def operation_task_export_title(filters):
    parts = ["运营任务台账"]
    for key in ["platform", "user", "status", "task_type", "store", "next_handler", "priority"]:
        value = norm(filters.get(key, ""))
        if value:
            parts.append(value)
    if norm(filters.get("overdue")) in {"1", "true", "是", "超时"}:
        parts.append("超时")
    if norm(filters.get("unassigned")) in {"1", "true", "是", "未分配"}:
        parts.append("未分配")
    if norm(filters.get("reworked")) in {"1", "true", "是", "返工"}:
        parts.append("返工")
    if norm(filters.get("open_only")) in {"1", "true", "是", "未完成", "待办"}:
        parts.append("未完成")
    return "-".join(parts)


def export_operation_tasks(role="admin", user="", status="", task_type="", store="", platform="", overdue="", unassigned="", next_handler="", priority="", reworked="", open_only="", search=""):
    rows = list_operation_tasks(role=role, user=user, status=status, task_type=task_type, store=store, platform=platform, overdue=overdue, unassigned=unassigned, next_handler=next_handler, priority=priority, reworked=reworked, open_only=open_only, search=search)
    history_rows = sum(len(row.get("history") or []) for row in rows)
    filters = {
        "role": role,
        "user": user,
        "status": status,
        "task_type": task_type,
        "store": store,
        "platform": platform,
        "overdue": overdue,
        "unassigned": unassigned,
        "next_handler": next_handler,
        "priority": priority,
        "reworked": reworked,
        "open_only": open_only,
        "search": search,
    }
    out = output_path(operation_task_export_title(filters), "V1")
    operation_task_store().export_tasks(out, rows, filters=filters)
    return {"file": out.name, "download": f"/download?path={quote(out.name)}", "rows": len(rows), "history_rows": history_rows}


def token_from_headers(headers):
    if not headers:
        return ""
    return headers.get("X-Operator-Token", "") or headers.get("x-operator-token", "")


def handle_tasks_api(action, headers, payload):
    try:
        token = token_from_headers(headers)
        operator = operator_from_token(token)
        if action == "GET":
            filters = scoped_task_filters(operator, payload)
            rows = list_operation_tasks(
                role=filters["role"],
                user=filters["user"],
                status=payload.get("status", ""),
                task_type=payload.get("task_type", ""),
                store=payload.get("store", ""),
                platform=payload.get("platform", ""),
                overdue=payload.get("overdue", ""),
                unassigned=payload.get("unassigned", ""),
                next_handler=payload.get("next_handler", ""),
                priority=payload.get("priority", ""),
                reworked=payload.get("reworked", ""),
                open_only=payload.get("open_only", ""),
                search=payload.get("search", ""),
            )
            return json_bytes({"ok": True, "operator": operator, "summary": summarize_operation_tasks(rows), "packages": package_operation_tasks(rows), "tasks": rows})
        if action == "POST_SUBMIT":
            task_id = payload.get("id", "")
            if operator.get("role") != "owner":
                return json_bytes({"ok": False, "error": "只有店长可以填写处理结果"}, status=403)
            task = operation_task_store().require_task(task_id)[1]
            if norm(task.get("owner")) != norm(operator.get("user")):
                return json_bytes({"ok": False, "error": "不能处理其他负责人的任务"}, status=403)
            task = submit_operation_task(task_id, operator.get("user", ""), payload.get("action", ""), payload.get("remark", ""), payload.get("proof", ""))
            return json_bytes({"ok": True, "task": task})
        if action == "POST_BATCH_SUBMIT":
            if operator.get("role") != "owner":
                return json_bytes({"ok": False, "error": "只有店长可以批量填写处理结果"}, status=403)
            result = submit_operation_tasks(payload.get("ids", []), operator.get("user", ""), payload.get("action", ""), payload.get("remark", ""), payload.get("proof", ""))
            return json_bytes({"ok": True, **result})
        if action == "POST_ASSIGN":
            if not can_review_tasks(operator):
                return json_bytes({"ok": False, "error": "只有管理员可以指派任务"}, status=403)
            task = assign_operation_task(payload.get("id", ""), operator.get("user", "管理员"), payload.get("owner", ""), payload.get("remark", ""))
            return json_bytes({"ok": True, "task": task})
        if action == "POST_BATCH_PUSH":
            if not can_review_tasks(operator):
                return json_bytes({"ok": False, "error": "只有管理员可以推送任务"}, status=403)
            result = push_operation_tasks(payload.get("ids", []), operator.get("user", "管理员"), payload.get("remark", ""))
            return json_bytes({"ok": True, **result})
        if action == "POST_REVIEW":
            if not can_review_tasks(operator):
                return json_bytes({"ok": False, "error": "只有管理员可以审核"}, status=403)
            task = review_operation_task(payload.get("id", ""), operator.get("user", "管理员"), payload.get("decision", ""), payload.get("remark", ""))
            return json_bytes({"ok": True, "task": task})
        if action == "POST_BATCH_REVIEW":
            if not can_review_tasks(operator):
                return json_bytes({"ok": False, "error": "只有管理员可以批量审核"}, status=403)
            result = review_operation_tasks(payload.get("ids", []), operator.get("user", "管理员"), payload.get("decision", ""), payload.get("remark", ""))
            return json_bytes({"ok": True, **result})
        if action == "POST_DONE":
            if not can_review_tasks(operator):
                return json_bytes({"ok": False, "error": "只有管理员可以标记完成"}, status=403)
            task = mark_operation_task_done(payload.get("id", ""), operator.get("user", "管理员"), payload.get("remark", ""))
            return json_bytes({"ok": True, "task": task})
        if action == "POST_EXPORT":
            filters = scoped_task_filters(operator, payload)
            result = export_operation_tasks(
                role=filters["role"],
                user=filters["user"],
                status=payload.get("status", ""),
                task_type=payload.get("task_type", ""),
                store=payload.get("store", ""),
                platform=payload.get("platform", ""),
                overdue=payload.get("overdue", ""),
                unassigned=payload.get("unassigned", ""),
                next_handler=payload.get("next_handler", ""),
                priority=payload.get("priority", ""),
                reworked=payload.get("reworked", ""),
                open_only=payload.get("open_only", ""),
                search=payload.get("search", ""),
            )
            grant_download(token, result.get("file", ""))
            return json_bytes({"ok": True, **result})
        return json_bytes({"ok": False, "error": "任务接口不存在"}, status=404)
    except PermissionError as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=401)
    except KeyError as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=404)
    except ValueError as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=400)
    except Exception as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=500)


def public_owner_directory():
    return [{"owner": row.get("owner", "")} for row in operation_owner_directory()]


def handle_owners_api(headers=None):
    try:
        token = token_from_headers(headers or {})
        if token:
            operator = operator_from_token(token)
            if can_review_tasks(operator):
                return json_bytes({"ok": True, "owners": operation_owner_directory()})
        return json_bytes({"ok": True, "owners": public_owner_directory()})
    except PermissionError:
        return json_bytes({"ok": True, "owners": public_owner_directory()})
    except Exception as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=500)


def handle_store_owners_api(action, headers, payload=None):
    try:
        operator = operator_from_token(token_from_headers(headers))
        if not can_review_tasks(operator):
            return json_bytes({"ok": False, "error": "只有管理员可以维护店铺负责人"}, status=403)
        if action == "GET":
            return json_bytes({"ok": True, "assignments": load_store_owner_assignments(), "owners": operation_owner_directory()})
        payload = payload or {}
        if action == "POST_SAVE":
            assignments = save_store_owner_assignments(payload.get("assignments", []))
            assigned_existing = assign_existing_unassigned_tasks(assignments, operator.get("user", "管理员"))
            return json_bytes({"ok": True, "assignments": assignments, "assigned_existing": assigned_existing, "owners": operation_owner_directory()})
        return json_bytes({"ok": False, "error": "店铺负责人接口不存在"}, status=404)
    except PermissionError as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=401)
    except Exception as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=500)


def handle_admin_api(action, headers):
    try:
        operator = operator_from_token(token_from_headers(headers))
        if not can_review_tasks(operator):
            return json_bytes({"ok": False, "error": f"只有管理员可以执行{action}"}, status=403)
        return json_bytes({"ok": True, "operator": operator})
    except PermissionError as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=401)
    except Exception as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=500)


def handle_upload_operator_api(action, headers, category=""):
    try:
        operator = operator_from_token(token_from_headers(headers))
        if operator.get("role") not in {"admin", "owner"}:
            return json_bytes({"ok": False, "error": f"{action}需要管理员或店长"}, status=403)
        if operator.get("role") == "owner" and not norm(operator.get("user")):
            return json_bytes({"ok": False, "error": f"{action}需要先填写当前店长"}, status=403)
        if operator.get("role") == "owner" and category and not owner_can_upload_category(category):
            return json_bytes({"ok": False, "error": "店长不能上传该数据源"}, status=403)
        return json_bytes({"ok": True, "operator": operator})
    except PermissionError as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=401)
    except Exception as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=500)


def handle_backup_api(action, headers, payload=None):
    try:
        operator = operator_from_token(token_from_headers(headers))
        if not can_review_tasks(operator):
            return json_bytes({"ok": False, "error": "只有管理员可以执行备份和恢复"}, status=403)
        payload = payload or {}
        if action == "CREATE":
            return json_bytes({"ok": True, **create_operational_backup()})
        if action == "RESTORE":
            return json_bytes({"ok": True, **restore_operational_backup(payload.get("path", ""))})
        return json_bytes({"ok": False, "error": "备份接口不存在"}, status=404)
    except PermissionError as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=401)
    except Exception as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=500)


def handle_business_report_api(headers, payload=None):
    try:
        operator = operator_from_token(token_from_headers(headers))
        if operator.get("role") not in {"admin", "owner"}:
            return json_bytes({"ok": False, "error": "经营报表需要先登录管理员或店长身份"}, status=403)
        payload = payload or {}
        role = operator.get("role", "admin")
        user = operator.get("user", "")
        if role == "owner":
            platform = payload.get("platform", "")
            store = payload.get("store", "")
            if store:
                assignments = load_store_owner_assignments()
                owned_pairs = {
                    (norm(item.get("platform")), daily_ops_master_data.clean_store_name(item.get("store")))
                    for item in assignments
                    if norm(item.get("owner")) == norm(user)
                }
                if (norm(platform), daily_ops_master_data.clean_store_name(store)) not in owned_pairs:
                    return json_bytes({"ok": False, "error": "店长只能查询自己负责店铺的经营报表"}, status=403)
        data = business_report({
            "role": role,
            "user": "" if role == "admin" else user,
            "date_from": payload.get("date_from", ""),
            "date_to": payload.get("date_to", ""),
            "platform": payload.get("platform", ""),
            "store": payload.get("store", ""),
            "grain": payload.get("grain", "month"),
        })
        return json_bytes({"ok": True, "report": data})
    except PermissionError as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=401)
    except Exception as exc:
        return json_bytes({"ok": False, "error": str(exc)}, status=500)


HTML_PAGE = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>PETCIRCLE跨境工作台 v2.0</title>
  <style>
    :root { --bg:#f5f7fb; --panel:#fff; --ink:#1f2937; --muted:#6b7280; --line:#d8dee9; --blue:#1f4e78; --green:#2f7d5b; --red:#b42318; --soft:#eef3f8; --amber:#8a5a00; }
    * { box-sizing:border-box; }
    body { margin:0; font-family:"Microsoft YaHei", Arial, sans-serif; background:var(--bg); color:var(--ink); }
    .app { display:grid; grid-template-columns:220px 1fr; min-height:100vh; }
    aside { background:#172033; color:#eef3ff; padding:20px 14px; }
    .brand { font-size:20px; font-weight:700; margin:4px 10px 22px; }
    nav button { display:block; width:100%; border:0; background:transparent; color:#cbd5e1; text-align:left; padding:12px 14px; border-radius:6px; cursor:pointer; font-size:15px; }
    nav button.active, nav button:hover { background:#24324d; color:#fff; }
    main { padding:24px; }
    header { display:flex; justify-content:space-between; align-items:center; margin-bottom:18px; }
    h1 { margin:0; font-size:24px; }
    h2 { font-size:18px; margin:0 0 12px; }
    .grid { display:grid; gap:14px; }
    .cards { grid-template-columns:repeat(3,minmax(240px,1fr)); align-items:stretch; }
    .card, .panel { background:var(--panel); border:1px solid var(--line); border-radius:8px; padding:16px; }
    .card h3 { margin:0 0 8px; font-size:17px; }
    .report-card { height:360px; display:flex; flex-direction:column; padding:0; overflow:hidden; box-shadow:0 8px 22px rgba(31,78,120,.07); }
    .report-top { padding:15px 15px 10px; border-bottom:1px solid #edf1f6; }
    .report-title-row { display:flex; justify-content:space-between; align-items:flex-start; gap:10px; }
    .report-card h3 { margin:0; font-size:16px; line-height:1.25; }
    .report-index { flex:0 0 auto; min-width:28px; height:24px; border-radius:6px; background:var(--soft); color:#17324d; display:flex; align-items:center; justify-content:center; font-size:12px; font-weight:700; }
    .report-desc { margin-top:8px; min-height:40px; display:-webkit-box; -webkit-line-clamp:2; -webkit-box-orient:vertical; overflow:hidden; }
    .report-sources { margin-top:7px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
    .report-body { padding:12px 15px 14px; display:flex; flex-direction:column; gap:10px; flex:1; min-height:0; }
    .report-actions { display:flex; gap:8px; align-items:center; }
    .report-actions input { width:76px; flex:0 0 auto; }
    .report-actions button { flex:1; }
    .report-footer { border-top:1px solid #edf1f6; padding-top:9px; display:flex; flex-direction:column; flex:1; min-height:0; }
    .report-footer-head { display:flex; justify-content:space-between; align-items:center; margin-bottom:7px; color:#344054; font-size:13px; font-weight:700; }
    .report-output-list { flex:1; min-height:42px; overflow:hidden; display:grid; gap:6px; align-content:start; }
    .report-output { min-height:36px; border:1px solid #e4e9f1; border-radius:6px; background:#fbfcfe; padding:6px 7px; display:grid; grid-template-columns:minmax(0,1fr) auto; gap:8px; align-items:center; }
    .report-output-name { min-width:0; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; font-size:12px; color:#263445; }
    .report-output-time { font-size:11px; color:var(--muted); margin-top:1px; }
    .download-link { border:1px solid #c9d8e8; border-radius:6px; padding:5px 8px; background:#fff; font-size:12px; white-space:nowrap; }
    .report-empty { border:1px dashed #d5dce8; border-radius:6px; color:var(--muted); font-size:12px; padding:10px; text-align:center; }
    .muted { color:var(--muted); font-size:13px; line-height:1.5; }
    button.primary, button.secondary, button.danger { border:0; border-radius:6px; padding:9px 12px; cursor:pointer; font-weight:600; }
    button.primary { background:var(--blue); color:#fff; }
    button.secondary { background:#e8eef6; color:#17324d; }
    button.danger { background:#fce4e4; color:#9f1d1d; }
    button:disabled { opacity:.55; cursor:not-allowed; }
    input, select, textarea { border:1px solid var(--line); border-radius:6px; padding:0 10px; background:#fff; font-family:inherit; }
    input, select { height:36px; }
    textarea { min-height:72px; padding:8px 10px; resize:vertical; }
    input[type=file] { height:auto; padding:8px; }
    .row { display:flex; gap:10px; align-items:center; flex-wrap:wrap; }
    .hidden { display:none !important; }
    table { width:100%; border-collapse:collapse; background:#fff; }
    th, td { border-bottom:1px solid var(--line); padding:9px 10px; text-align:left; vertical-align:top; font-size:13px; }
    th { background:#eef3f8; font-weight:700; }
    .status { margin-top:10px; font-size:13px; color:var(--muted); min-height:20px; }
    .ok { color:var(--green); }
    .bad { color:var(--red); }
    a { color:#1b5f97; text-decoration:none; font-weight:600; }
    .section { display:none; }
    .section.active { display:block; }
    .toolbar { margin-bottom:14px; }
    .weekly-shell { display:grid; gap:14px; }
    .weekly-hero { background:#fff; border:1px solid var(--line); border-radius:8px; padding:16px; display:flex; justify-content:space-between; gap:16px; align-items:center; }
    .weekly-hero h2 { margin-bottom:6px; }
    .source-grid { display:grid; grid-template-columns:repeat(3,minmax(240px,1fr)); gap:14px; margin-bottom:0; align-items:stretch; }
    .source-card, .weekly-source-card { border:1px solid var(--line); border-radius:8px; padding:0; background:#fff; height:330px; display:flex; flex-direction:column; overflow:hidden; box-shadow:0 8px 22px rgba(31,78,120,.06); }
    .source-card h3, .weekly-source-card h3 { margin:0; font-size:16px; line-height:1.25; }
    .weekly-source-head { padding:14px 14px 10px; border-bottom:1px solid #edf1f6; display:grid; gap:7px; }
    .weekly-source-title { display:flex; justify-content:space-between; align-items:flex-start; gap:8px; }
    .weekly-source-desc { min-height:38px; display:-webkit-box; -webkit-line-clamp:2; -webkit-box-orient:vertical; overflow:hidden; }
    .weekly-source-body { padding:12px 14px; display:grid; gap:8px; flex:1; min-height:0; }
    .source-meta { border:1px solid #e8edf5; border-radius:8px; background:#fbfcfe; padding:9px 10px; min-height:82px; display:grid; gap:4px; align-content:start; }
    .source-meta-line { white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
    .source-upload-row { margin-top:auto; display:grid; grid-template-columns:minmax(0,1fr) auto; gap:8px; align-items:center; }
    .source-upload-row input[type=file] { min-width:0; width:100%; }
    .source-actions { display:grid; grid-template-columns:1fr 1fr; gap:8px; }
    .source-actions button { padding-left:8px; padding-right:8px; }
    .badge { display:inline-block; border-radius:999px; padding:3px 8px; font-size:12px; font-weight:700; background:#eef3f8; color:#17324d; }
    .badge.ok { background:#e2f0d9; color:#2f6b3f; }
    .badge.warn { background:#fff2cc; color:#7a5200; }
    .badge.bad { background:#fce4e4; color:#9f1d1d; }
    .result-list { display:grid; gap:8px; margin-top:12px; }
    .result-item { border:1px solid var(--line); border-radius:6px; padding:10px; background:#fff; }
    .form-grid { display:grid; grid-template-columns:repeat(auto-fit,minmax(240px,1fr)); gap:12px; }
    .field label { display:block; font-size:13px; color:var(--muted); margin-bottom:6px; }
    .field input, .field textarea { width:100%; }
    .owner-map-tools textarea { width:100%; min-height:132px; font-family:Consolas, "Microsoft YaHei", monospace; }
    .overview-grid { display:grid; grid-template-columns:repeat(7,minmax(120px,1fr)); gap:8px; margin-bottom:12px; }
    .overview-card { border:1px solid var(--line); border-radius:8px; background:#fff; padding:12px; }
    .overview-card span { display:block; color:var(--muted); font-size:12px; }
    .overview-card strong { display:block; font-size:24px; margin-top:5px; }
    .upgrade-backdrop { position:fixed; inset:0; z-index:20; display:grid; place-items:center; background:rgba(15,23,42,.38); padding:20px; }
    .upgrade-dialog { width:min(620px,100%); border-radius:12px; background:#fff; border:1px solid var(--line); box-shadow:0 24px 70px rgba(15,23,42,.25); padding:22px; }
    .upgrade-dialog h2 { font-size:22px; margin:0 0 8px; }
    .upgrade-list { display:grid; gap:8px; margin:14px 0; color:#334155; }
    .upgrade-list div { border:1px solid #e5ebf3; border-radius:8px; padding:9px 10px; background:#fbfcfe; }
    .business-report-web { display:grid; gap:12px; }
    .business-filter-web { display:grid; grid-template-columns:repeat(7,minmax(110px,1fr)); gap:8px; align-items:center; }
    .business-filter-web input, .business-filter-web select { width:100%; }
    .business-shortcuts { display:flex; flex-wrap:wrap; gap:8px; margin-bottom:10px; }
    .business-shortcuts button { border:1px solid var(--line); border-radius:6px; background:#fff; color:#17324d; padding:7px 10px; cursor:pointer; font-weight:600; }
    .business-shortcuts button.active { background:#e8eef6; border-color:#b9c8da; }
    .business-alert-web { border:1px solid #ffd8a8; background:#fff8ed; border-radius:8px; padding:10px 12px; display:flex; justify-content:space-between; gap:12px; align-items:center; }
    .business-alert-web.ok { border-color:#c7e1cf; background:#f1faf4; }
    .business-kpis-web { display:grid; grid-template-columns:repeat(4,minmax(150px,1fr)); gap:8px; }
    .business-kpi-web { border:1px solid var(--line); border-radius:8px; background:#fff; padding:12px; }
    .business-kpi-web span { display:block; color:var(--muted); font-size:12px; }
    .business-kpi-web strong { display:block; margin-top:5px; font-size:24px; }
    .business-kpi-web small { display:block; margin-top:5px; color:var(--muted); }
    .business-kpi-web.up strong, .business-kpi-web.up small { color:var(--green); }
    .business-kpi-web.down strong, .business-kpi-web.down small { color:var(--red); }
    .business-tabs-web { display:flex; gap:8px; flex-wrap:wrap; }
    .business-tabs-web button { border:1px solid var(--line); border-radius:6px; background:#fff; padding:8px 12px; cursor:pointer; font-weight:700; }
    .business-tabs-web button.active { background:#24324d; color:#fff; border-color:#24324d; }
    .business-table-grid-web { display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:12px; }
    .business-table-wrap-web { overflow:auto; max-height:520px; border:1px solid var(--line); border-radius:8px; background:#fff; }
    .business-table-web { min-width:760px; }
    .business-table-web th { position:sticky; top:0; z-index:1; }
    .business-table-web td.num, .business-table-web th.num { text-align:right; font-variant-numeric:tabular-nums; }
    .business-table-web .up { color:var(--green); font-weight:700; }
    .business-table-web .down { color:var(--red); font-weight:700; }
    .business-trend-panel-web { display:none; }
    .business-trend-panel-web.active { display:block; }
    .overview-layout { display:grid; grid-template-columns:minmax(0,2fr) minmax(280px,1fr); gap:12px; align-items:start; }
    .overview-table-wrap { overflow:auto; max-height:560px; border:1px solid var(--line); border-radius:8px; background:#fff; }
    .overview-table { min-width:980px; table-layout:fixed; }
    .overview-table th { position:sticky; top:0; z-index:1; }
    .overview-table td { white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
    .overview-side { display:grid; gap:12px; }
    .task-dashboard-actions { display:flex; justify-content:flex-end; margin:-4px 0 6px; }
    .task-dashboard-actions button { height:28px; padding:0 10px; }
    .task-summary { display:grid; grid-template-columns:repeat(9,minmax(90px,1fr)); gap:6px; margin-bottom:8px; }
    .task-kpi { border:1px solid var(--line); border-radius:8px; background:#fff; padding:12px; }
    .task-kpi strong { display:block; font-size:22px; margin-top:4px; }
    body.task-dashboard-collapsed #adminTaskQueue,
    body.task-dashboard-collapsed #ownerTaskSummary { display:none !important; }
    body.task-dashboard-collapsed .task-kpi { padding:6px 10px; min-height:34px; }
    body.task-dashboard-collapsed .task-kpi span { font-size:12px; }
    body.task-dashboard-collapsed .task-kpi strong { display:inline; font-size:16px; margin:0 0 0 6px; }
    body:not(.task-dashboard-collapsed) .task-summary { grid-template-columns:repeat(3,minmax(180px,1fr)); }
    .task-filters { display:grid; grid-template-columns:220px 110px 110px 118px 118px 128px repeat(4,auto) 92px 92px 92px 92px; gap:6px; align-items:center; margin-bottom:10px; }
    .task-filters input, .task-filters select { height:30px; padding:0 8px; font-size:13px; }
    .task-filters label { white-space:nowrap; font-size:13px; color:#334155; display:flex; align-items:center; gap:4px; }
    .task-filters button { height:30px; padding:0 10px; }
    .task-table-wrap { overflow:auto; max-height:680px; border:1px solid var(--line); border-radius:8px; background:#fff; }
    .package-toolbar { position:sticky; top:0; z-index:2; display:flex; align-items:center; gap:10px; padding:8px 10px; border-bottom:1px solid var(--line); background:#f8fbff; font-size:13px; }
    .package-board { display:grid; gap:8px; padding:10px; min-width:980px; }
    .package-card { border:1px solid var(--line); border-radius:8px; background:#fff; padding:10px 12px; display:grid; grid-template-columns:34px minmax(220px,1.3fr) minmax(120px,.7fr) minmax(120px,.7fr) minmax(160px,1fr) minmax(170px,1fr); gap:10px; align-items:center; }
    .package-card:hover { border-color:#b9c8da; box-shadow:0 1px 4px rgba(15,23,42,.05); }
    .package-title { display:flex; align-items:center; gap:6px; flex-wrap:wrap; }
    .package-title strong { font-size:15px; }
    .package-metric { display:grid; gap:3px; min-width:0; }
    .package-metric span { color:var(--muted); font-size:12px; }
    .package-metric strong { font-size:18px; }
    .package-meta { min-width:0; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; color:#334155; font-size:13px; }
    .package-card .task-actions { justify-content:flex-end; }
    .task-ledger { min-width:1320px; border:0; table-layout:fixed; }
    .task-ledger th { position:sticky; top:0; z-index:1; white-space:nowrap; height:34px; padding:6px 8px; background:#edf3f9; }
    .task-ledger th, .task-ledger td { vertical-align:middle; line-height:1.25; }
    .task-ledger td { height:36px; padding:5px 8px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
    .task-ledger td.copy-cell { font-family:Consolas, "Microsoft YaHei", monospace; }
    .task-select { width:34px; text-align:center; }
    .task-actions { display:flex; gap:4px; flex-wrap:nowrap; align-items:center; }
    .task-actions button { height:26px; padding:0 7px; border-radius:5px; white-space:nowrap; font-size:12px; }
    .task-actions .muted { white-space:nowrap; }
    .package-main { font-weight:700; }
    .package-sub { color:var(--muted); font-size:12px; margin-top:3px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
    .package-counts { display:flex; gap:4px; flex-wrap:wrap; }
    .package-preview-row td { background:#fbfcfe; height:auto; white-space:normal; padding:8px 10px; }
    .package-preview-card { grid-column:2 / -1; border-top:1px solid var(--line); padding-top:8px; }
    .package-preview { display:grid; gap:4px; }
    .package-preview-line { display:grid; grid-template-columns:110px 120px 120px 1fr 82px; gap:8px; font-size:12px; color:#334155; align-items:center; }
    .package-preview-head { font-weight:700; color:#1f2937; }
    .package-preview-line button { height:24px; padding:0 7px; font-size:12px; border-radius:5px; }
    .login-bar { background:#fff; border:1px solid var(--line); border-radius:8px; padding:10px; margin-bottom:12px; display:grid; grid-template-columns:120px 160px 160px auto auto 1fr; gap:8px; align-items:center; }
    .login-bar .identity { color:var(--muted); font-size:13px; }
    .owner-entry { background:#fff; border:1px solid var(--line); border-radius:8px; padding:10px; margin-bottom:12px; display:grid; grid-template-columns:auto minmax(260px,1fr) auto auto; gap:8px; align-items:center; }
    .owner-entry input { width:100%; }
    .backup-tools { margin-top:14px; display:grid; gap:10px; }
    .backup-tools .row input { min-width:360px; flex:1; }
    @media (max-width:1180px) { .cards, .source-grid { grid-template-columns:repeat(2,minmax(240px,1fr)); } }
    @media (max-width:1180px) { .overview-layout { grid-template-columns:1fr; } .overview-grid { grid-template-columns:repeat(3,minmax(120px,1fr)); } .business-filter-web, .business-table-grid-web { grid-template-columns:1fr; } .business-kpis-web { grid-template-columns:repeat(2,minmax(150px,1fr)); } }
    @media (max-width:760px) { .app { grid-template-columns:1fr; } aside { position:static; } .cards, .source-grid, .task-summary, .task-filters, .owner-entry, .overview-grid, .business-kpis-web { grid-template-columns:1fr; } main { padding:16px; } header, .weekly-hero, .business-alert-web { align-items:flex-start; flex-direction:column; gap:12px; } .report-card { height:380px; } .weekly-source-card { height:350px; } }
  </style>
</head>
<body class="task-dashboard-collapsed">
<div id="upgradeDialog" class="upgrade-backdrop hidden" role="dialog" aria-modal="true" aria-labelledby="upgradeTitle">
  <div class="upgrade-dialog">
    <h2 id="upgradeTitle">Web 版已升级为多角色经营工作台</h2>
    <div class="muted">旧的个人工作台入口已不再作为主流程使用。现在管理员看全局经营报表，店长只看自己负责店铺。</div>
    <div class="upgrade-list">
      <div><strong>经营报表</strong>：支持今日、本月、本年、当前范围、同比和异常提醒。</div>
      <div><strong>多角色</strong>：管理员看平台、业务员、店铺；店长只看自己店铺。</div>
      <div><strong>数据异常</strong>：超过 3 天未更新会进入“需要处理”。</div>
    </div>
    <div class="row" style="justify-content:flex-end;">
      <button class="secondary" onclick="dismissUpgradeDialog(true)">不再提示</button>
      <button class="primary" onclick="dismissUpgradeDialog(false); switchTab('reports');">进入经营报表</button>
    </div>
  </div>
</div>
<div class="app">
  <aside>
    <div class="brand">PETCIRCLE跨境工作台 v2.0</div>
    <nav>
      <button class="active" data-tab="overview" data-admin-only="1">数据总览</button>
      <button data-tab="reports">经营报表</button>
      <button data-tab="tasks">任务包中心</button>
      <button data-tab="weekly" data-admin-only="1">每周工作流</button>
      <button data-tab="rules" data-admin-only="1">规则设置</button>
      <button data-tab="search" data-admin-only="1">基础数据查询</button>
      <button data-tab="files" data-admin-only="1">输出文件</button>
    </nav>
  </aside>
  <main>
    <header>
      <div>
        <h1 id="pageTitle">数据总览</h1>
        <div class="muted" id="statusLine">正在读取状态...</div>
      </div>
      <div class="row">
        <button class="secondary" onclick="refreshStatus()">刷新</button>
        <button class="danger" onclick="shutdownWorkbench()">关闭工作台</button>
      </div>
    </header>
    <div class="login-bar">
      <select id="loginRole"><option value="admin">管理员</option><option value="owner">店长</option></select>
      <input id="loginUser" placeholder="登录身份" list="ownerOptions">
      <datalist id="ownerOptions"></datalist>
      <input id="loginPassword" placeholder="管理员/店长访问密码，可空" type="password">
      <button class="primary" onclick="loginOperator()">登录身份</button>
      <button class="secondary" onclick="logoutOperator()">退出身份</button>
      <div class="identity" id="operatorIdentity">未登录：任务台账需要先登录</div>
    </div>
    <div class="owner-entry" data-admin-only="1">
      <strong>店长入口</strong>
      <input id="ownerEntryLink" readonly placeholder="选择或填写负责人后生成店长入口链接，格式包含 role=owner&user=负责人">
      <button class="secondary" onclick="updateOwnerEntryLink()">生成入口</button>
      <button class="secondary" onclick="copyOwnerEntryLink()">复制入口</button>
    </div>

    <section id="overview" class="section active" data-admin-only="1">
      <div class="overview-grid" id="overviewCards"></div>
      <div class="overview-layout">
        <div class="panel">
          <h2>店铺维度总览</h2>
          <div class="overview-table-wrap">
            <table class="overview-table">
              <thead><tr><th>店铺</th><th>负责人</th><th>任务包</th><th>待办明细</th><th>待店长</th><th>待审核</th><th>低分</th><th>议价</th><th>滞销</th><th>价格异常</th><th>超时</th></tr></thead>
              <tbody id="overviewStoreRows"></tbody>
            </table>
          </div>
        </div>
        <div class="overview-side">
          <div class="panel">
            <h2>负责人待办</h2>
            <div class="task-dashboard-actions">
              <button class="secondary" onclick="switchTab('tasks')">进入任务包中心</button>
            </div>
            <div class="overview-table-wrap">
              <table>
                <thead><tr><th>负责人</th><th>店铺</th><th>待办</th><th>待审核</th><th>超时</th></tr></thead>
                <tbody id="overviewOwnerRows"></tbody>
              </table>
            </div>
          </div>
          <div class="panel">
            <h2>数据源状态</h2>
            <div id="overviewSourceRows" class="result-list"></div>
          </div>
        </div>
      </div>
    </section>

    <section id="reports" class="section">
      <div class="panel business-report-web" id="businessReportShell">
        <div>
          <h2>经营报表</h2>
          <div class="muted">管理员看全局平台、业务员、店铺；店长只看自己负责店铺。报表数据来自基础销量台账。</div>
        </div>
        <div class="business-shortcuts" id="businessShortcuts">
          <button data-range="7d" onclick="applyBusinessRangeWeb('7d')">最近7天</button>
          <button data-range="30d" onclick="applyBusinessRangeWeb('30d')">最近30天</button>
          <button data-range="90d" onclick="applyBusinessRangeWeb('90d')">最近90天</button>
          <button data-range="half-year" onclick="applyBusinessRangeWeb('half-year')">半年</button>
          <button data-range="1y" onclick="applyBusinessRangeWeb('1y')">一年</button>
          <button data-range="month" onclick="applyBusinessRangeWeb('month')">自然月</button>
          <button data-range="year" onclick="applyBusinessRangeWeb('year')">自然年</button>
        </div>
        <div class="business-filter-web">
          <input id="businessDateFrom" type="date">
          <input id="businessDateTo" type="date">
          <select id="businessGrain"><option value="day">按日</option><option value="month" selected>按月</option><option value="year">按年</option></select>
          <select id="businessPlatform"><option value="">全部平台</option><option>Temu</option><option>Shein</option><option>速卖通</option><option>TK</option><option>Ozon</option></select>
          <input id="businessStore" placeholder="店铺">
          <button class="primary" onclick="loadBusinessReportWeb(true)">查询</button>
          <button class="secondary" onclick="loadBusinessReportWeb(true)">刷新</button>
        </div>
        <div id="businessAlertWeb" class="business-alert-web ok"><strong>需要处理</strong><span>请先查询经营报表。</span></div>
        <div class="business-kpis-web" id="businessKpis"></div>
        <div class="business-tabs-web" id="businessTabs">
          <button class="active" data-business-tab="overview" onclick="setBusinessTabWeb('overview')">经营概览</button>
          <button data-business-tab="platform" data-admin-only="1" onclick="setBusinessTabWeb('platform')">平台分析</button>
          <button data-business-tab="owner" data-admin-only="1" onclick="setBusinessTabWeb('owner')">业务员分析</button>
          <button data-business-tab="store" onclick="setBusinessTabWeb('store')">店铺分析</button>
        </div>
        <div id="businessOverviewTables" class="business-table-grid-web">
          <div>
            <h2 data-admin-only="1">平台排行</h2>
            <div class="business-table-wrap-web" data-admin-only="1"><table class="business-table-web" id="businessPlatformTable"></table></div>
          </div>
          <div>
            <h2 data-admin-only="1">业务员排行</h2>
            <div class="business-table-wrap-web" data-admin-only="1"><table class="business-table-web" id="businessOwnerTable"></table></div>
          </div>
          <div>
            <h2 id="businessStoreTitle">店铺排行</h2>
            <div class="business-table-wrap-web"><table class="business-table-web" id="businessStoreTable"></table></div>
          </div>
        </div>
        <div id="businessTrendPanel" class="business-trend-panel-web">
          <h2 id="businessTrendTitle">趋势明细</h2>
          <div class="business-table-wrap-web"><table class="business-table-web" id="businessTrendTable"></table></div>
        </div>
        <div class="status" id="businessReportStatus"></div>
      </div>
    </section>

    <section id="tasks" class="section">
      <div class="task-summary" id="taskSummary"></div>
      <div class="task-dashboard-actions" data-admin-only="1"><button class="secondary" id="taskDashboardToggle" onclick="toggleTaskDashboard()">展开看板</button></div>
      <div class="task-summary" id="adminTaskQueue" data-admin-only="1"></div>
      <div class="task-summary" id="ownerTaskSummary" data-admin-only="1"></div>
      <div class="panel">
        <h2 id="taskPanelTitle">任务包中心</h2>
        <select id="taskRole" class="hidden"><option value="admin">管理员</option><option value="owner">店长</option></select>
        <input id="taskUser" class="hidden" placeholder="负责人姓名">
        <div class="task-filters" data-admin-only="1">
          <input id="taskSearch" placeholder="店铺 / 商家编码 / SKC / SPU / 商品名">
          <input id="taskAdminOwner" placeholder="负责人">
          <input id="taskStore" placeholder="店铺">
          <select id="taskStatus"><option value="">全部状态</option><option>待推送</option><option>待店长处理</option><option>待管理员审核</option><option>已通过</option><option>已驳回</option><option>已完成</option></select>
          <select id="taskPlatform"><option value="">全部平台</option><option>Temu</option><option>Shein</option></select>
          <select id="taskType"><option value="">全部类型</option><option>价格异常</option><option>库存异常</option><option>爆旺冲突</option><option>低分预警</option><option>滞销处理</option><option>议价审核</option></select>
          <label><input id="taskOpenOnly" type="checkbox"> 只看未完成</label>
          <label><input id="taskOverdue" type="checkbox"> 只看超时</label>
          <label><input id="taskUnassigned" type="checkbox"> 只看未分配</label>
          <label><input id="taskReworked" type="checkbox"> 只看返工</label>
          <button class="primary" onclick="loadTasks()">查询</button>
          <button class="primary" data-admin-only="task-push" onclick="batchPushTasks()">批量推送</button>
          <button class="primary" data-admin-only="task-review" onclick="batchReviewTasks('通过')">批量通过</button>
          <button class="danger" data-admin-only="task-review" onclick="batchReviewTasks('驳回')">批量驳回</button>
          <button class="secondary" onclick="exportTasks()">导出</button>
        </div>
        <div class="task-filters hidden" data-owner-only="1">
          <input id="taskOwnerSearch" placeholder="店铺 / 商家编码 / SKC / SPU / 商品名">
          <input id="taskOwnerStore" placeholder="店铺">
          <select id="taskOwnerType"><option value="">全部类型</option><option>价格异常</option><option>库存异常</option><option>爆旺冲突</option><option>低分预警</option><option>滞销处理</option><option>议价审核</option></select>
          <select id="taskOwnerPlatform"><option value="">全部平台</option><option>Temu</option><option>Shein</option></select>
          <select id="taskOwnerStatus"><option value="">全部状态</option><option>待店长处理</option><option>待管理员审核</option><option>已通过</option><option>已驳回</option><option>已完成</option></select>
          <label><input id="taskOwnerOpenOnly" type="checkbox" checked> 只看未完成</label>
          <button class="primary" onclick="loadTasks()">查询</button>
          <button class="primary" onclick="batchSubmitTasks()">批量标记已处理</button>
          <button class="secondary" onclick="exportTasks()">导出</button>
        </div>
        <div class="status" id="taskStatusLine"></div>
        <div class="task-table-wrap">
          <div class="package-toolbar">
            <label><input type="checkbox" id="taskSelectAll" onclick="toggleAllTaskSelection(this.checked)"> 全选当前可操作包</label>
            <span class="muted">一块就是一个店铺、负责人、任务类型、处理动作的任务包；点“导出”下载完整处理表。</span>
          </div>
          <div id="taskRows" class="package-board"></div>
        </div>
      </div>
    </section>

    <section id="weekly" class="section">
      <div class="weekly-shell">
        <div class="weekly-hero">
          <div>
            <h2>每周数据源</h2>
            <div class="muted">上传、结束本次上传、生成每周报表都在这里完成。每个数据源卡片会显示最新文件、待提交文件和当前状态。</div>
          </div>
          <div class="row">
            <button class="primary" onclick="runWeeklyReports()">生成本周报表</button>
            <button class="secondary" onclick="refreshStatus()">刷新状态</button>
          </div>
        </div>
        <div class="source-grid" id="weeklySources"></div>
        <div class="status" id="weeklyStatus"></div>
        <div class="result-list" id="weeklyResults"></div>
      </div>
    </section>

    <section id="rules" class="section">
      <div class="panel">
        <h2>字段与判断规则</h2>
        <div class="muted" style="margin-bottom:12px;">这里保存后，后续生成报表会按最新规则执行。</div>
        <div class="form-grid">
          <div class="field"><label>Temu爆旺款口径</label><textarea id="rule_temu_basis"></textarea></div>
          <div class="field"><label>爆旺关键词</label><input id="rule_hot_keywords" placeholder="高销款, 爆, 旺"></div>
          <div class="field"><label>表格排序层级</label><input id="rule_group_order" placeholder="SPU, SKC, 商家编码"></div>
          <div class="field"><label>尺码排序</label><input id="rule_size_order" placeholder="XXS, XS, S, M, L, XL, XXL, XXXL"></div>
          <div class="field"><label>新品定义：上架天数小于</label><input id="rule_new_product_days_lt" type="number" min="0"></div>
          <div class="field"><label>新品滞销：上架天数超过</label><input id="rule_new_slow_min_days" type="number" min="0"></div>
          <div class="field"><label>新品滞销：上架天数小于</label><input id="rule_new_slow_max_days" type="number" min="0"></div>
          <div class="field"><label>老品滞销：上架天数超过</label><input id="rule_old_slow_min_days" type="number" min="0"></div>
          <div class="field"><label>滞销判断分组</label><input id="rule_group_by" placeholder="店铺+SPU"></div>
          <div class="field"><label>SHEIN新品爆旺：上架天数小于</label><input id="rule_shein_new_days_lt" type="number" min="0"></div>
          <div class="field"><label>SHEIN新品爆旺：7天日均不低于</label><input id="rule_shein_new_7d_daily_gte" type="number" min="0" step="0.1"></div>
          <div class="field"><label>SHEIN老品爆旺：上架天数不低于</label><input id="rule_shein_old_days_gte" type="number" min="0"></div>
          <div class="field"><label>SHEIN老品爆旺：30天日均大于</label><input id="rule_shein_old_30d_daily_gt" type="number" min="0" step="0.1"></div>
        </div>
        <div class="row" style="margin-top:14px;">
          <button class="primary" onclick="saveRules()">保存规则</button>
          <button class="secondary" onclick="renderRules()">恢复当前已保存内容</button>
        </div>
        <div class="status" id="rulesStatus"></div>
      </div>
      <div class="panel owner-map-tools" style="margin-top:14px;">
        <h2>店铺负责人配置</h2>
        <div class="muted" style="margin-bottom:12px;">用于 Shein 表格没有负责人、或导入表缺负责人时自动分配任务。每行格式：平台，店铺，负责人。</div>
        <textarea id="storeOwnerMapText" placeholder="Temu，7，小琴&#10;Shein，琪琪，洁琳"></textarea>
        <div class="row" style="margin-top:12px;">
          <button class="primary" onclick="saveStoreOwners()">保存负责人配置</button>
          <button class="secondary" onclick="loadStoreOwners()">重新读取</button>
        </div>
        <div class="status" id="storeOwnerStatus"></div>
      </div>
    </section>

    <section id="search" class="section">
      <div class="panel">
        <h2>基础数据库关键词搜索</h2>
        <div class="row toolbar">
          <input id="searchQuery" style="min-width:320px" placeholder="输入 SKU、SKC、商品名、店铺名">
          <select id="searchLimit"><option>100</option><option>200</option><option>500</option></select>
          <button class="primary" onclick="runSearch()">搜索</button>
          <button class="secondary" onclick="exportSearch()">导出Excel</button>
        </div>
        <div class="status" id="searchStatus"></div>
        <div style="overflow:auto; max-height:560px;">
          <table>
            <thead><tr><th>来源文件</th><th>工作表</th><th>原始行</th><th>命中内容</th></tr></thead>
            <tbody id="searchRows"></tbody>
          </table>
        </div>
      </div>
    </section>

    <section id="files" class="section">
      <div class="panel">
        <h2>最近输出</h2>
        <table>
          <thead><tr><th>文件名</th><th>大小</th><th>更新时间</th><th>操作</th></tr></thead>
          <tbody id="outputRows"></tbody>
        </table>
      </div>
      <div class="panel backup-tools">
        <h2>数据备份</h2>
        <div class="muted">备份只包含运营状态、规则、数据源和任务台账，不包含输出表、图片产物、打包产物、虚拟环境产物。</div>
        <div class="row">
          <button class="primary" onclick="createBackup()">生成备份</button>
          <input id="restoreBackupPath" placeholder="粘贴要恢复的 .zip 备份文件完整路径">
          <button class="danger" onclick="restoreBackup()">恢复备份</button>
        </div>
        <div class="status" id="backupStatus"></div>
      </div>
    </section>
  </main>
</div>
<script>
let appStatus = null;
let userRequestedShutdown = false;
const reportRunMessages = {};
let taskState = {summary:{}, packages:[], tasks:[]};
let expandedTaskPackageId = "";
let ownerOptions = [];
let operatorToken = localStorage.getItem('operatorToken') || '';
let operatorSession = JSON.parse(localStorage.getItem('operatorSession') || 'null');
let businessReport = null;
let businessTab = 'overview';
const titles = {overview:'数据总览', reports:'经营报表', tasks:'任务包中心', weekly:'每周工作流', rules:'规则设置', search:'基础数据查询', files:'输出文件'};
document.querySelectorAll('nav button').forEach(btn => {
  btn.addEventListener('click', () => switchTab(btn.dataset.tab));
});
function fmtSize(n){ if(n>1048576) return (n/1048576).toFixed(1)+' MB'; if(n>1024) return (n/1024).toFixed(1)+' KB'; return n+' B'; }
async function api(url, opts={}){
  const next = {...opts, headers:{...(opts.headers || {})}};
  if(operatorToken) next.headers['X-Operator-Token'] = operatorToken;
  const r = await fetch(url, next);
  const j = await r.json();
  if(!r.ok || j.ok===false) {
    const error = new Error(j.error || '请求失败');
    error.status = r.status;
    if(error.status === 401) clearOperatorSession();
    throw error;
  }
  return j;
}
function authDownload(url){
  if(!operatorToken) return url;
  const next = new URL(url, window.location.origin);
  next.searchParams.set('token', operatorToken);
  return next.pathname + next.search;
}
function renderOperator(){
  const el = document.getElementById('operatorIdentity');
  if(!el) return;
  const taskRole = document.getElementById('taskRole');
  if(operatorSession){
    el.textContent = `当前身份：${operatorSession.role === 'admin' ? '管理员' : '店长'} · ${operatorSession.user}`;
    taskRole.value = operatorSession.role;
    taskRole.disabled = operatorSession.role === 'owner';
    document.getElementById('taskUser').value = operatorSession.user;
    const adminOwner = document.getElementById('taskAdminOwner');
    if(adminOwner && operatorSession.role === 'admin' && !adminOwner.value) adminOwner.value = '';
    defaultOpenTasksForOwner();
  } else {
    el.textContent = '未登录：任务台账需要先登录';
    if(taskRole) taskRole.disabled = false;
  }
  applyRoleVisibility();
  updateOwnerEntryLink();
}
function showUpgradeDialog(){
  if(localStorage.getItem('dailyOpsWebUpgradeDismissed') === '1') return;
  const dialog = document.getElementById('upgradeDialog');
  if(dialog) dialog.classList.remove('hidden');
}
function dismissUpgradeDialog(remember){
  if(remember) localStorage.setItem('dailyOpsWebUpgradeDismissed', '1');
  const dialog = document.getElementById('upgradeDialog');
  if(dialog) dialog.classList.add('hidden');
}
function defaultOpenTasksForOwner(){
  const openOnly = document.getElementById('taskOpenOnly');
  const ownerOpenOnly = document.getElementById('taskOwnerOpenOnly');
  if(operatorSession?.role === 'owner'){
    if(openOnly) openOnly.checked = true;
    if(ownerOpenOnly) ownerOpenOnly.checked = true;
  }
}
function clearOperatorSession(){
  localStorage.removeItem('operatorSession');
  localStorage.removeItem('operatorToken');
  operatorSession = null;
  operatorToken = '';
  renderOperator();
}
function switchTab(tab){
  if(!operatorSession && ['overview', 'reports'].includes(tab)) tab = 'tasks';
  if(operatorSession?.role === 'owner' && !['tasks', 'reports'].includes(tab)) tab = 'reports';
  document.querySelectorAll('nav button').forEach(b => b.classList.toggle('active', b.dataset.tab === tab));
  document.querySelectorAll('.section').forEach(s => s.classList.toggle('active', s.id === tab));
  document.getElementById('pageTitle').textContent = titles[tab] || titles.tasks;
  if(tab === 'reports') loadBusinessReportWeb(false);
}
function applyRoleVisibility(){
  const ownerMode = operatorSession?.role === 'owner';
  document.querySelectorAll('[data-admin-only]').forEach(el => { el.classList.toggle('hidden', ownerMode); });
  document.querySelectorAll('[data-owner-only]').forEach(el => { el.classList.toggle('hidden', !ownerMode); });
  const title = document.getElementById('taskPanelTitle');
  if(title) title.textContent = ownerMode ? '我的任务包' : '管理员任务包中心';
  if(ownerMode) document.body.classList.add('task-dashboard-collapsed');
  updateTaskDashboardToggle();
  if(ownerMode) switchTab('reports');
}
function updateTaskDashboardToggle(){
  const btn = document.getElementById('taskDashboardToggle');
  if(btn) btn.textContent = document.body.classList.contains('task-dashboard-collapsed') ? '展开看板' : '收起看板';
}
function toggleTaskDashboard(){
  document.body.classList.toggle('task-dashboard-collapsed');
  updateTaskDashboardToggle();
}
function renderOverview(){
  const overview = appStatus?.overview || {};
  const cards = document.getElementById('overviewCards');
  if(cards){
    const items = [
      ['任务包', overview.package_total || 0],
      ['明细总数', overview.task_total || 0],
      ['未完成', overview.open_total || 0],
      ['待店长', overview.pending_owner_total || 0],
      ['待审核', overview.pending_review_total || 0],
      ['未分配', overview.unassigned_total || 0],
      ['超时', overview.overdue_total || 0],
    ];
    cards.innerHTML = items.map(([label, value]) => `<div class="overview-card"><span>${label}</span><strong>${value}</strong></div>`).join('');
  }
  const storeRows = document.getElementById('overviewStoreRows');
  if(storeRows){
    const rows = overview.stores || [];
    storeRows.innerHTML = rows.length ? rows.map(row => `<tr>
      <td>${esc(row.store || '-')}</td><td>${esc(row.owner || '-')}</td><td>${row.package_count || 0}</td><td>${row.open_count || 0}</td><td>${row.pending_owner_count || 0}</td><td>${row.pending_review_count || 0}</td><td>${row.low_score_count || 0}</td><td>${row.bargain_count || 0}</td><td>${row.slow_count || 0}</td><td>${row.price_count || 0}</td><td>${row.overdue_count || 0}</td>
    </tr>`).join('') : '<tr><td colspan="11" class="muted">暂无店铺任务数据。</td></tr>';
  }
  const ownerRows = document.getElementById('overviewOwnerRows');
  if(ownerRows){
    const rows = overview.owners || [];
    ownerRows.innerHTML = rows.length ? rows.map(row => `<tr><td>${esc(row.owner || '-')}</td><td>${row.store_count || 0}</td><td>${row.open_count || 0}</td><td>${row.pending_review_count || 0}</td><td>${row.overdue_count || 0}</td></tr>`).join('') : '<tr><td colspan="5" class="muted">暂无负责人待办。</td></tr>';
  }
  const sourceRows = document.getElementById('overviewSourceRows');
  if(sourceRows){
    const groups = appStatus?.source_groups || [];
    sourceRows.innerHTML = groups.length ? groups.map(item => `<div class="result-item"><strong>${esc(item.name || '')}</strong><div class="muted">${esc(item.status || '')} · 文件 ${item.count || 0} · 记录 ${item.total_rows || '-'}</div></div>`).join('') : '<div class="muted">暂无数据源状态。</div>';
  }
}
function localDateText(date){
  const y = date.getFullYear();
  const m = String(date.getMonth() + 1).padStart(2, '0');
  const d = String(date.getDate()).padStart(2, '0');
  return `${y}-${m}-${d}`;
}
function initializeBusinessDatesWeb(){
  const endInput = document.getElementById('businessDateTo');
  const startInput = document.getElementById('businessDateFrom');
  if(!endInput || !startInput || startInput.value || endInput.value) return;
  const end = new Date();
  const start = new Date(end);
  start.setDate(end.getDate() - 29);
  startInput.value = localDateText(start);
  endInput.value = localDateText(end);
  setBusinessRangeActiveWeb('30d');
}
function setBusinessRangeActiveWeb(range){
  document.querySelectorAll('#businessShortcuts [data-range]').forEach(btn => {
    btn.classList.toggle('active', btn.dataset.range === range);
  });
}
function applyBusinessRangeWeb(range){
  const end = new Date();
  const start = new Date(end);
  if(range === '7d') start.setDate(end.getDate() - 6);
  if(range === '30d') start.setDate(end.getDate() - 29);
  if(range === '90d') start.setDate(end.getDate() - 89);
  if(range === 'half-year') start.setMonth(end.getMonth() - 6);
  if(range === '1y') start.setFullYear(end.getFullYear() - 1);
  if(range === 'month') start.setDate(1);
  if(range === 'year'){ start.setMonth(0); start.setDate(1); }
  document.getElementById('businessDateFrom').value = localDateText(start);
  document.getElementById('businessDateTo').value = localDateText(end);
  setBusinessRangeActiveWeb(range);
  loadBusinessReportWeb(true);
}
function signedNumber(value){
  const n = Number(value || 0);
  return n > 0 ? `+${n}` : String(n);
}
function signedRate(value){
  if(value === null || value === undefined) return '-';
  const n = Number(value || 0);
  return `${n > 0 ? '+' : ''}${n}%`;
}
function deltaClass(value){
  const n = Number(value || 0);
  if(n > 0) return 'up';
  if(n < 0) return 'down';
  return '';
}
function kpiWeb(label, item, compareLabel){
  const delta = Number(item?.delta || 0);
  const rate = item?.rate;
  return `<div class="business-kpi-web ${deltaClass(delta)}"><span>${esc(label)}</span><strong>${esc(item?.sales || 0)}</strong><small>${esc(compareLabel)} ${esc(signedNumber(delta))}${rate === null || rate === undefined ? '' : ` (${esc(signedRate(rate))})`}</small></div>`;
}
function renderBusinessKpisWeb(report){
  const box = document.getElementById('businessKpis');
  if(!box) return;
  const s = report?.summary || {};
  const rangeLabel = `最近${String(report?.filters?.range_key || '30d').replace('d', '')}日销量`;
  box.innerHTML = [
    kpiWeb(rangeLabel, s.previous_range || s.range || {}, '较上期'),
    kpiWeb('上期对比', s.previous_range || {}, '上期销量'),
    kpiWeb('去年同期', s.range || {}, '去年同期销量'),
    kpiWeb('本月累计', s.month || {}, '较上月同期'),
    kpiWeb('本年累计', s.year || {}, '较去年同期'),
  ].join('');
}
function renderBusinessAlertWeb(report){
  const box = document.getElementById('businessAlertWeb');
  if(!box) return;
  const anomalies = report?.anomalies || [];
  box.classList.toggle('ok', !anomalies.length);
  box.innerHTML = anomalies.length
    ? `<strong>需要处理</strong><span>${anomalies.length} 条异常：${esc(anomalies.slice(0, 3).map(item => item.message).join('；'))}</span>`
    : '<strong>需要处理</strong><span>暂无超过阈值的数据异常。</span>';
}
function businessRows(dimension){
  return businessReport?.dimensions?.[dimension] || [];
}
function renderBusinessTableWeb(id, rows, emptyText){
  const table = document.getElementById(id);
  if(!table) return;
  if(!rows.length){
    table.innerHTML = `<tbody><tr><td>${esc(emptyText)}</td></tr></tbody>`;
    return;
  }
  table.innerHTML = `<thead><tr><th>名称</th><th>平台</th><th>负责人</th><th class="num">当前销量</th><th class="num">去年同期</th><th class="num">同比件数</th><th class="num">同比率</th><th class="num">占比</th><th>状态</th></tr></thead><tbody>${rows.map(row => `<tr>
    <td><strong>${esc(row.name || '-')}</strong></td>
    <td>${esc(row.platform || '-')}</td>
    <td>${esc(row.owner || '-')}</td>
    <td class="num"><strong>${esc(row.sales || 0)}</strong></td>
    <td class="num">${esc(row.compare_sales || 0)}</td>
    <td class="num ${deltaClass(row.yoy_delta)}">${esc(signedNumber(row.yoy_delta || 0))}</td>
    <td class="num ${deltaClass(row.yoy_delta)}">${esc(row.base_too_small ? '基数小' : signedRate(row.yoy_rate))}</td>
    <td class="num">${esc(row.share || 0)}%</td>
    <td>${esc(row.status || '正常')}</td>
  </tr>`).join('')}</tbody>`;
}
function renderBusinessTrendWeb(dimension){
  const table = document.getElementById('businessTrendTable');
  const trend = businessReport?.trends?.[dimension] || {};
  const buckets = trend.buckets || [];
  const rows = trend.rows || [];
  if(!table) return;
  if(!rows.length){
    table.innerHTML = '<tbody><tr><td>暂无趋势数据。</td></tr></tbody>';
    return;
  }
  table.innerHTML = `<thead><tr><th>名称</th><th>平台</th><th>负责人</th><th class="num">合计</th>${buckets.map(bucket => `<th class="num">${esc(bucket)}</th>`).join('')}</tr></thead><tbody>${rows.map(row => `<tr>
    <td><strong>${esc(row.name || '-')}</strong></td>
    <td>${esc(row.platform || '-')}</td>
    <td>${esc(row.owner || '-')}</td>
    <td class="num"><strong>${esc(row.total || 0)}</strong></td>
    ${buckets.map(bucket => `<td class="num">${row.values?.[bucket] ? esc(row.values[bucket]) : ''}</td>`).join('')}
  </tr>`).join('')}</tbody>`;
}
function setBusinessTabWeb(tab){
  businessTab = tab || 'overview';
  if(operatorSession?.role === 'owner' && ['platform', 'owner'].includes(businessTab)) businessTab = 'store';
  document.querySelectorAll('#businessTabs [data-business-tab]').forEach(btn => {
    btn.classList.toggle('active', btn.dataset.businessTab === businessTab);
  });
  const overview = document.getElementById('businessOverviewTables');
  const trend = document.getElementById('businessTrendPanel');
  if(overview) overview.classList.toggle('hidden', businessTab !== 'overview');
  if(trend) trend.classList.toggle('active', businessTab !== 'overview');
  const dimension = businessTab === 'overview' ? 'platform' : businessTab;
  const titleMap = {platform:'平台趋势明细', owner:'业务员趋势明细', store:'店铺趋势明细'};
  const title = document.getElementById('businessTrendTitle');
  if(title) title.textContent = titleMap[dimension] || '趋势明细';
  renderBusinessTrendWeb(dimension);
}
function renderBusinessReportWeb(){
  const ownerMode = operatorSession?.role === 'owner';
  renderBusinessKpisWeb(businessReport);
  renderBusinessAlertWeb(businessReport);
  renderBusinessTableWeb('businessPlatformTable', businessRows('platform'), '暂无平台数据。');
  renderBusinessTableWeb('businessOwnerTable', businessRows('owner'), '暂无业务员数据。');
  renderBusinessTableWeb('businessStoreTable', businessRows('store'), '暂无店铺数据。');
  const storeTitle = document.getElementById('businessStoreTitle');
  if(storeTitle) storeTitle.textContent = ownerMode ? '我的店铺排行' : '店铺排行';
  if(ownerMode && ['platform', 'owner'].includes(businessTab)) businessTab = 'store';
  setBusinessTabWeb(businessTab);
}
async function loadBusinessReportWeb(showMessage=false){
  if(!operatorSession) return;
  initializeBusinessDatesWeb();
  const st = document.getElementById('businessReportStatus');
  const params = new URLSearchParams({
    date_from: document.getElementById('businessDateFrom')?.value || '',
    date_to: document.getElementById('businessDateTo')?.value || '',
    grain: document.getElementById('businessGrain')?.value || 'month',
    platform: document.getElementById('businessPlatform')?.value || '',
    store: document.getElementById('businessStore')?.value || '',
  });
  if(st) st.textContent = '正在读取经营报表...';
  try {
    const res = await api('/api/business-report?' + params.toString());
    businessReport = res.report;
    renderBusinessReportWeb();
    if(st) st.innerHTML = `<span class="ok">经营报表已更新：</span>${esc(businessReport.filters.date_from)} 至 ${esc(businessReport.filters.date_to)}`;
  } catch(e) {
    if(st) st.innerHTML = `<span class="bad">${esc(e.message)}</span>`;
  }
}
function ownerEntryUrl(owner){
  const url = new URL(window.location.href);
  url.search = '';
  url.searchParams.set('role', 'owner');
  url.searchParams.set('user', owner);
  return url.toString();
}
function updateOwnerEntryLink(){
  const input = document.getElementById('ownerEntryLink');
  if(!input) return;
  const owner = document.getElementById('taskUser')?.value.trim() || document.getElementById('loginUser')?.value.trim() || '';
  input.value = owner ? ownerEntryUrl(owner) : '';
}
async function copyOwnerEntryLink(){
  updateOwnerEntryLink();
  const link = document.getElementById('ownerEntryLink')?.value || '';
  const el = document.getElementById('operatorIdentity');
  if(!link){ if(el) el.textContent = '请先填写负责人姓名，再生成店长入口'; return; }
  if(navigator.clipboard) await navigator.clipboard.writeText(link);
  if(el) el.textContent = '店长入口已复制';
}
function applyEntryParams(){
  const params = new URLSearchParams(window.location.search);
  const role = params.get('role');
  const user = params.get('user');
  if(role !== 'owner' || !user) return;
  localStorage.removeItem('operatorSession');
  localStorage.removeItem('operatorToken');
  operatorSession = null;
  operatorToken = '';
  document.getElementById('loginRole').value = 'owner';
  document.getElementById('loginUser').value = user;
  document.getElementById('taskRole').value = 'owner';
  document.getElementById('taskRole').disabled = true;
  document.getElementById('taskUser').value = user;
  document.getElementById('taskOpenOnly').checked = true;
  const el = document.getElementById('operatorIdentity');
  if(el) el.textContent = `请以店长身份登录：${user}`;
  updateOwnerEntryLink();
}
async function loginOperator(){
  const el = document.getElementById('operatorIdentity');
  if(el) el.textContent = '正在登录...';
  try {
    const payload = {
      role: document.getElementById('loginRole').value,
      user: document.getElementById('loginUser').value.trim(),
      password: document.getElementById('loginPassword').value
    };
    const res = await api('/api/session/login', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(payload)});
    operatorSession = res.session;
    operatorToken = res.session.token;
    localStorage.setItem('operatorSession', JSON.stringify(operatorSession));
    localStorage.setItem('operatorToken', operatorToken);
    document.getElementById('loginPassword').value = '';
    renderOperator();
    await refreshStatus();
  } catch(e) {
    document.getElementById('loginPassword').value = '';
    if(el) el.innerHTML = `<span class="bad">登录失败：${esc(e.message)}</span>`;
  }
}
async function logoutOperator(){
  try {
    if(operatorToken){
      await api('/api/session/logout', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({})});
    }
  } catch(e) {
    // 本地仍清掉身份，避免旧 token 继续留在共享浏览器里。
  }
  clearOperatorSession();
  document.getElementById('loginPassword').value = '';
  renderTaskSummary();
  const tbody = document.getElementById('taskRows');
  if(tbody) tbody.innerHTML = '<tr><td colspan="9" class="muted">请先登录身份。</td></tr>';
}
async function loadOwnerOptions(){
  try {
    const res = await api('/api/owners');
    ownerOptions = res.owners || [];
    const list = document.getElementById('ownerOptions');
    if(list){
      list.innerHTML = ownerOptions.map(item => `<option value="${esc(item.owner)}">${esc((item.stores || []).join('、'))}</option>`).join('');
    }
  } catch(e) {
    ownerOptions = [];
  }
}
function renderStoreOwners(assignments){
  const el = document.getElementById('storeOwnerMapText');
  if(!el) return;
  el.value = (assignments || []).map(item => [item.platform || '', item.store || '', item.owner || ''].join('，')).join('\n');
}
function parseStoreOwnerText(){
  const text = document.getElementById('storeOwnerMapText')?.value || '';
  return text.split(/\n+/).map(line => {
    const parts = line.split(/[,，\t]/).map(item => item.trim());
    return {platform: parts[0] || '', store: parts[1] || '', owner: parts[2] || ''};
  }).filter(item => item.store && item.owner);
}
async function loadStoreOwners(){
  const st = document.getElementById('storeOwnerStatus');
  try {
    const res = await api('/api/store-owners');
    renderStoreOwners(res.assignments || []);
    if(st) st.textContent = `已读取 ${res.assignments?.length || 0} 条负责人配置`;
  } catch(e){
    if(st) st.innerHTML = `<span class="bad">${e.message}</span>`;
  }
}
async function saveStoreOwners(){
  const st = document.getElementById('storeOwnerStatus');
  const assignments = parseStoreOwnerText();
  if(st) st.textContent = '正在保存负责人配置...';
  try {
    const res = await api('/api/store-owners', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({assignments})});
    renderStoreOwners(res.assignments || []);
    await loadOwnerOptions();
    updateOwnerEntryLink();
    if(st) st.innerHTML = `<span class="ok">负责人配置已保存：</span>${res.assignments?.length || 0} 条；已补齐 ${res.assigned_existing || 0} 条未分配任务。后续生成任务会自动使用。`;
  } catch(e){
    if(st) st.innerHTML = `<span class="bad">${e.message}</span>`;
  }
}
async function refreshStatus(){
  try {
    appStatus = await api('/api/status');
    document.getElementById('statusLine').textContent = `当前版本 ${appStatus.version || 'v2.0'}，Temu数据源 ${appStatus.temu_files} 个，Shein数据源 ${appStatus.shein_files} 个，ERP数据源 ${appStatus.erp_files} 个，基础库 ${appStatus.database.tables} 表 / ${appStatus.database.rows} 行`;
    renderReports(); renderOutputs(); renderWeeklySources(); renderRules(); renderOverview(); renderOperator(); loadTasks(false);
    loadBusinessReportWeb(false);
    loadOwnerOptions();
    loadStoreOwners();
    updateReportOutputCapacity();
  } catch(e) { document.getElementById('statusLine').textContent = e.message; }
}
function splitList(value){ return String(value || '').split(/[,，\s]+/).map(s => s.trim()).filter(Boolean); }
function joinList(value){ return Array.isArray(value) ? value.join(', ') : String(value || ''); }
function numValue(id, fallback){ const v = document.getElementById(id).value; return v === '' ? fallback : Number(v); }
function renderRules(){
  if(!appStatus || !appStatus.rules) return;
  const r = appStatus.rules;
  document.getElementById('rule_temu_basis').value = r.hot_item?.temu_basis || '';
  document.getElementById('rule_hot_keywords').value = joinList(r.hot_item?.keywords);
  document.getElementById('rule_group_order').value = joinList(r.sort?.group_order);
  document.getElementById('rule_size_order').value = joinList(r.sort?.size_order);
  document.getElementById('rule_new_product_days_lt').value = r.slow_moving?.new_product_days_lt ?? 28;
  document.getElementById('rule_new_slow_min_days').value = r.slow_moving?.new_slow_min_days ?? 30;
  document.getElementById('rule_new_slow_max_days').value = r.slow_moving?.new_slow_max_days ?? 60;
  document.getElementById('rule_old_slow_min_days').value = r.slow_moving?.old_slow_min_days ?? 180;
  document.getElementById('rule_group_by').value = r.slow_moving?.group_by || '店铺+SPU';
  document.getElementById('rule_shein_new_days_lt').value = r.hot_item?.shein_new_days_lt ?? 30;
  document.getElementById('rule_shein_new_7d_daily_gte').value = r.hot_item?.shein_new_7d_daily_gte ?? 10;
  document.getElementById('rule_shein_old_days_gte').value = r.hot_item?.shein_old_days_gte ?? 30;
  document.getElementById('rule_shein_old_30d_daily_gt').value = r.hot_item?.shein_old_30d_daily_gt ?? 20;
}
async function saveRules(){
  const st = document.getElementById('rulesStatus');
  const current = appStatus.rules || {};
  const payload = {
    hot_item: {
      ...(current.hot_item || {}),
      temu_basis: document.getElementById('rule_temu_basis').value.trim(),
      keywords: splitList(document.getElementById('rule_hot_keywords').value),
      shein_new_days_lt: numValue('rule_shein_new_days_lt', 30),
      shein_new_7d_daily_gte: numValue('rule_shein_new_7d_daily_gte', 10),
      shein_old_days_gte: numValue('rule_shein_old_days_gte', 30),
      shein_old_30d_daily_gt: numValue('rule_shein_old_30d_daily_gt', 20)
    },
    sort: {
      group_order: splitList(document.getElementById('rule_group_order').value),
      size_order: splitList(document.getElementById('rule_size_order').value)
    },
    slow_moving: {
      ...(current.slow_moving || {}),
      new_product_days_lt: numValue('rule_new_product_days_lt', 28),
      new_slow_min_days: numValue('rule_new_slow_min_days', 30),
      new_slow_max_days: numValue('rule_new_slow_max_days', 60),
      old_slow_min_days: numValue('rule_old_slow_min_days', 180),
      group_by: document.getElementById('rule_group_by').value.trim() || '店铺+SPU',
      sales30_total_equals: 0
    }
  };
  st.textContent = '正在保存...';
  try {
    const res = await api('/api/rules', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(payload)});
    appStatus.rules = res.rules;
    renderRules();
    st.innerHTML = '<span class="ok">规则已保存，后续生成报表会按新规则执行。</span>';
  } catch(e){ st.innerHTML = `<span class="bad">${e.message}</span>`; }
}
async function shutdownWorkbench(){
  userRequestedShutdown = true;
  document.getElementById('statusLine').textContent = '正在关闭工作台...';
  try {
    await api('/api/shutdown', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({reason:'button'})});
    document.body.innerHTML = '<div style="padding:28px;font-family:Microsoft YaHei,Arial,sans-serif;"><h2>工作台已关闭</h2><p>端口已释放。下次双击启动文件即可重新打开。</p></div>';
  } catch(e) {
    document.getElementById('statusLine').textContent = e.message;
  }
}
function renderReports(){
  const wrap = document.getElementById('reportCards'); if(!wrap) return; wrap.innerHTML = '';
  Object.entries(appStatus.reports).forEach(([id, r], index) => {
    const card = document.createElement('div'); card.className='card report-card';
    card.innerHTML = `<div class="report-top">
        <div class="report-title-row"><h3>${esc(r.name)}</h3><div class="report-index">${index + 1}</div></div>
        <div class="muted report-desc">${esc(r.description)}</div>
        <div class="muted report-sources" title="${esc(r.sources)}">所需数据源：${esc(r.sources)}</div>
        <div class="muted">${esc(reportTaskSummary(id))}</div>
      </div>
      <div class="report-body">
        <div class="report-actions"><input id="ver_${id}" value="V1"><button class="primary" onclick="runReport('${id}')">生成表格</button></div>
        <div class="status" id="st_${id}">${reportRunMessages[id] || ''}</div>
        <div class="report-footer">
          <div class="report-footer-head"><span>最近生成</span><span class="muted">${reportOutputs(id).length} 个</span></div>
          <div class="report-output-list" id="outs_${id}" data-report="${id}"></div>
        </div>
      </div>`;
    wrap.appendChild(card);
    renderReportOutputItems(id);
  });
  updateReportOutputCapacity();
}
function reportOutputs(id){
  return (appStatus.outputs || []).filter(f => f.report === id);
}
function reportTaskSummary(id){
  const item = appStatus?.report_tasks?.[id] || {};
  const status = item.by_status || {};
  const pendingReview = status['待管理员审核'] || 0;
  const unhandled = status['待店长处理'] || 0;
  return `已生成任务 ${item.total || 0} 条，待店长 ${unhandled} 条，待审核 ${pendingReview} 条`;
}
function shortModified(value){
  const text = String(value || '');
  return text.length > 11 ? text.slice(5, 16) : text;
}
function renderReportOutputItems(id){
  const list = document.getElementById('outs_' + id);
  if(!list || !appStatus) return;
  const capacity = Math.max(1, Number(list.dataset.capacity || 3));
  const files = reportOutputs(id).slice(0, capacity);
  if(!files.length){
    list.innerHTML = '<div class="report-empty">暂无已生成表格</div>';
    return;
  }
  list.innerHTML = files.map(f => `<div class="report-output">
    <div>
      <div class="report-output-name" title="${esc(f.name)}">${esc(f.name)}</div>
      <div class="report-output-time">${esc(shortModified(f.modified))} · ${fmtSize(f.size)}</div>
    </div>
    <a class="download-link" href="${authDownload(f.download)}">下载</a>
  </div>`).join('');
}
function updateReportOutputCapacity(){
  document.querySelectorAll('.report-output-list').forEach(list => {
    const rows = Math.max(1, Math.floor(list.clientHeight / 44));
    if(String(rows) !== list.dataset.capacity){
      list.dataset.capacity = String(rows);
      renderReportOutputItems(list.dataset.report);
    }
  });
}
function taskSyncSummary(sync){
  sync = sync || {};
  return `新增任务 ${sync.created || 0} 条，更新任务 ${sync.updated || 0} 条，导入明细 ${sync.imported_rows || 0} 行`;
}
async function runReport(id){
  const st = document.getElementById('st_'+id); st.textContent='正在生成...';
  try {
    const version = document.getElementById('ver_'+id).value || 'V1';
    const res = await api('/api/reports/run', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({report:id, version})});
    reportRunMessages[id] = `<span class="ok">生成完成：</span><a href="${authDownload(res.result.download)}">${esc(res.result.file)}</a><div class="muted">${esc(taskSyncSummary(res.result.task_sync))}</div>`;
    st.innerHTML = reportRunMessages[id];
    await refreshStatus();
  } catch(e) { st.innerHTML = `<span class="bad">${e.message}</span>`; }
}
window.addEventListener('resize', updateReportOutputCapacity);
async function uploadFile(category=null, inputId=null, statusId=null){
  const fileInput = document.getElementById(inputId || 'uploadFile');
  const files = Array.from(fileInput.files || []);
  const st = document.getElementById(statusId || 'uploadStatus');
  if(!files.length){ st.textContent='请选择文件'; return; }
  if(!category){ st.textContent='请选择每周工作流里的数据源模块上传'; return; }
  const targetCategory = category;
  st.textContent=`正在上传 ${files.length} 个文件...`;
  try {
    let last = null;
    for(const file of files){
      const form = new FormData(); form.append('category', targetCategory); form.append('file', file);
      last = await api('/api/upload', {method:'POST', body:form});
    }
    st.innerHTML=`<span class="ok">已加入本次上传：</span>${files.length} 个文件。确认全部子表上传完后，请点“结束本次上传”。`;
    fileInput.value = '';
    await refreshStatus();
  }
  catch(e){ st.innerHTML=`<span class="bad">${e.message}</span>`; }
}
async function finishBatch(category, statusId){
  const st = document.getElementById(statusId);
  st.textContent = '正在结束本次上传...';
  try {
    const res = await api('/api/upload/finish-batch', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({category})});
    st.innerHTML = `<span class="ok">本次上传已结束：</span>${res.source_state.count || 0} 个文件将作为最新分析数据源。`;
    await refreshStatus();
  } catch(e){ st.innerHTML=`<span class="bad">${e.message}</span>`; }
}
async function clearBatch(category, statusId){
  const st = document.getElementById(statusId);
  st.textContent = '正在清空本次上传...';
  try {
    const res = await api('/api/upload/clear-batch', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({category})});
    st.innerHTML = `<span class="ok">已清空：</span>${res.cleared || 0} 个待提交文件。`;
    await refreshStatus();
  } catch(e){ st.innerHTML=`<span class="bad">${e.message}</span>`; }
}
function badgeClass(status){ if(status==='已更新') return 'ok'; if(status==='缺少数据') return 'bad'; if(status==='已有数据' || status==='待结束上传') return 'warn'; return ''; }
function renderWeeklySources(){
  const wrap = document.getElementById('weeklySources'); if(!wrap || !appStatus) return;
  wrap.innerHTML = '';
  (appStatus.source_groups || []).forEach(g => {
    const latest = g.latest;
    const id = 'weekly_file_' + g.key;
    const stid = 'weekly_st_' + g.key;
    const batchNames = (g.batch_files || []).length ? (g.batch_files || []).map(esc).join('、') : (latest ? esc(latest.name) : '暂无');
    const pendingFiles = (g.pending_files || []).length ? `（${(g.pending_files || []).map(esc).join('、')}）` : '';
    const pendingText = g.pending_count ? `待结束上传：${g.pending_count} 个文件${pendingFiles}` : '无待提交文件';
    const rowsText = g.total_rows !== '' ? g.total_rows : (latest && latest.rows !== '' ? latest.rows : '-');
    const batchText = g.batch_id ? `批次：${esc(g.batch_id)}` : '批次：-';
    const uploadText = g.uploaded_at ? `上传时间：${esc(g.uploaded_at)}` : '上传时间：-';
    const card = document.createElement('div'); card.className = 'weekly-source-card';
    card.innerHTML = `<div class="weekly-source-head">
        <div class="weekly-source-title"><h3>${esc(g.name)}</h3><span class="badge ${badgeClass(g.status)}">${esc(g.status)}</span></div>
        <div class="muted weekly-source-desc">${esc(g.description)}</div>
      </div>
      <div class="weekly-source-body">
        <div class="source-meta">
          <div class="muted source-meta-line" title="${batchNames}">最新文件：${batchNames}</div>
          <div class="muted source-meta-line">更新时间：${latest ? esc(latest.modified) : '-'}</div>
          <div class="muted source-meta-line">${batchText}</div>
          <div class="muted source-meta-line">${uploadText}</div>
          <div class="muted source-meta-line">记录数：${rowsText}</div>
          <div class="muted source-meta-line">${pendingText}</div>
        </div>
        <div class="source-upload-row"><input id="${id}" type="file" accept=".xlsx,.xls,.csv" multiple><button class="secondary" onclick="uploadFile('${g.upload_target}','${id}','${stid}')">上传</button></div>
        <div class="source-actions"><button class="primary" onclick="finishBatch('${g.key}','${stid}')">结束上传</button><button class="secondary" onclick="clearBatch('${g.key}','${stid}')">清空待提交</button></div>
        <div class="status" id="${stid}"></div>
      </div>`;
    wrap.appendChild(card);
  });
}
async function runWeeklyReports(){
  const st = document.getElementById('weeklyStatus');
  const list = document.getElementById('weeklyResults');
  st.textContent = '正在生成本周报表...'; list.innerHTML = '';
  try {
    const res = await api('/api/weekly/run', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({})});
    const results = res.result.results || [];
    const okCount = results.filter(r => r.status === 'ok').length;
    st.innerHTML = `<span class="ok">生成完成：</span>${okCount}/${results.length} 份成功；${esc(taskSyncSummary(res.result.task_sync))}`;
    results.forEach(r => {
      const item = document.createElement('div'); item.className = 'result-item';
      if(r.status === 'ok'){
        item.innerHTML = `<span class="ok">${esc(r.file)}</span><div class="muted">${esc(appStatus.reports[r.report]?.name || r.report)}</div><div class="muted">${esc(taskSyncSummary(r.task_sync))}</div><a href="${authDownload(r.download)}">下载</a>`;
      } else {
        item.innerHTML = `<span class="bad">${esc(r.name || r.report)} 生成失败</span><div class="muted">${esc(r.error)}</div>`;
      }
      list.appendChild(item);
    });
    await refreshStatus();
  } catch(e){ st.innerHTML = `<span class="bad">${e.message}</span>`; }
}
function taskQuery(){
  const ownerMode = operatorSession?.role === 'owner';
  const params = new URLSearchParams();
  if(ownerMode){
    params.set('role', 'owner');
    params.set('user', operatorSession?.user || '');
    params.set('status', document.getElementById('taskOwnerStatus')?.value || '');
    params.set('platform', document.getElementById('taskOwnerPlatform')?.value || '');
    params.set('task_type', document.getElementById('taskOwnerType')?.value || '');
    params.set('store', document.getElementById('taskOwnerStore')?.value.trim() || '');
    params.set('open_only', document.getElementById('taskOwnerOpenOnly')?.checked ? '1' : '');
    params.set('search', document.getElementById('taskOwnerSearch')?.value.trim() || '');
  } else {
    params.set('role', 'admin');
    params.set('user', document.getElementById('taskAdminOwner')?.value.trim() || '');
    params.set('status', document.getElementById('taskStatus')?.value || '');
    params.set('platform', document.getElementById('taskPlatform')?.value || '');
    params.set('task_type', document.getElementById('taskType')?.value || '');
    params.set('store', document.getElementById('taskStore')?.value.trim() || '');
    params.set('open_only', document.getElementById('taskOpenOnly')?.checked ? '1' : '');
    params.set('overdue', document.getElementById('taskOverdue')?.checked ? '1' : '');
    params.set('unassigned', document.getElementById('taskUnassigned')?.checked ? '1' : '');
    params.set('reworked', document.getElementById('taskReworked')?.checked ? '1' : '');
    params.set('search', document.getElementById('taskSearch')?.value.trim() || '');
  }
  return params;
}
async function loadTasks(showMessage=true){
  const tbody = document.getElementById('taskRows');
  if(!tbody) return;
  const st = document.getElementById('taskStatusLine');
  if(!operatorToken){ renderTaskSummary(); tbody.innerHTML = '<div class="muted">请先登录身份。</div>'; return; }
  if(showMessage) st.textContent = '正在读取任务...';
  try {
    const res = await api('/api/tasks?' + taskQuery().toString());
    taskState = {summary:res.summary || {}, packages:res.packages || [], tasks:res.tasks || []};
    renderTaskSummary();
    renderTaskRows();
    st.textContent = `当前筛选 ${taskState.packages.length} 个任务包，包含 ${taskState.tasks.length} 条明细`;
  } catch(e){
    st.innerHTML = `<span class="bad">${e.message}</span>`;
  }
}
function renderTaskSummary(){
  const wrap = document.getElementById('taskSummary'); if(!wrap) return;
  const summary = taskState.summary || {};
  const status = summary.by_status || {};
  const overdue = summary.overdue || {};
  const nextHandler = summary.by_next_handler || {};
  const cards = [
    ['全部任务', summary.total || 0],
    ['管理员待办', nextHandler['管理员'] || 0],
    ['待推送', status['待推送'] || 0],
    ['店长待办', nextHandler['店长'] || 0],
    ['待店长处理', status['待店长处理'] || 0],
    ['待管理员审核', status['待管理员审核'] || 0],
    ['超时未处理', overdue.total || 0],
    ['已通过', status['已通过'] || 0],
    ['未分配', summary.unassigned || 0],
  ];
  wrap.innerHTML = cards.map(([label, value]) => `<div class="task-kpi"><span class="muted">${label}</span><strong>${value}</strong></div>`).join('');
  renderAdminTaskQueue();
  renderOwnerTaskSummary();
}
function renderAdminTaskQueue(){
  const wrap = document.getElementById('adminTaskQueue'); if(!wrap) return;
  const rows = taskState.summary?.admin_queue || [];
  if(!rows.length){ wrap.innerHTML = ''; return; }
  wrap.innerHTML = rows.map((item, index) => `<button class="task-kpi" type="button" data-queue-index="${index}" onclick="applyAdminQueueFilter(${index})"><span class="muted">管理员待办队列</span><strong>${item.count || 0}</strong><div>${esc(item.action || '')}</div><div class="muted">优先级：${esc(item.priority || '')}</div></button>`).join('');
}
function setTaskField(id, value){
  const field = document.getElementById(id);
  if(field) field.value = value || '';
}
function setTaskCheck(id, value){
  const field = document.getElementById(id);
  if(field) field.checked = Boolean(value);
}
function applyAdminQueueFilter(index){
  const item = (taskState.summary?.admin_queue || [])[index];
  if(!item) return;
  const filters = item.filters || {};
  setTaskField('taskRole', 'admin');
  setTaskField('taskUser', '');
  setTaskField('taskAdminOwner', '');
  setTaskField('taskSearch', '');
  setTaskField('taskStatus', filters.status || '');
  setTaskField('taskNextHandler', '');
  setTaskField('taskPriority', '');
  setTaskField('taskPlatform', '');
  setTaskField('taskType', '');
  setTaskField('taskStore', '');
  setTaskCheck('taskOpenOnly', filters.open_only === '1');
  setTaskCheck('taskOverdue', filters.overdue === '1');
  setTaskCheck('taskUnassigned', filters.unassigned === '1');
  setTaskCheck('taskReworked', filters.reworked === '1');
  loadTasks();
}
function renderOwnerTaskSummary(){
  const wrap = document.getElementById('ownerTaskSummary'); if(!wrap) return;
  const ownerStatus = taskState.summary?.owner_status || {};
  const rows = Object.values(ownerStatus).sort((a, b) => (b.total || 0) - (a.total || 0));
  if(!rows.length){ wrap.innerHTML = ''; return; }
  wrap.innerHTML = rows.map((item, index) => {
    const status = item.by_status || {};
    return `<button class="task-kpi" type="button" data-owner-index="${index}" onclick="applyOwnerSummaryFilter(${index})"><span class="muted">负责人待办：${esc(item.owner)}</span><strong>${item.total || 0}</strong><div class="muted">待推送 ${status['待推送'] || 0} / 待店长 ${status['待店长处理'] || 0} / 待审核 ${status['待管理员审核'] || 0} / 超时 ${item.overdue || 0} / 返工 ${item.reworked || 0} / 已完成 ${status['已完成'] || 0}</div></button>`;
  }).join('');
}
function applyOwnerSummaryFilter(index){
  const ownerStatus = taskState.summary?.owner_status || {};
  const rows = Object.values(ownerStatus).sort((a, b) => (b.total || 0) - (a.total || 0));
  const item = rows[index];
  if(!item) return;
  setTaskField('taskRole', 'admin');
  setTaskField('taskUser', item.owner === '未分配' ? '' : item.owner || '');
  setTaskField('taskAdminOwner', item.owner === '未分配' ? '' : item.owner || '');
  setTaskField('taskSearch', '');
  setTaskField('taskStatus', '');
  setTaskField('taskNextHandler', '');
  setTaskField('taskPriority', '');
  setTaskField('taskPlatform', '');
  setTaskField('taskType', '');
  setTaskField('taskStore', '');
  setTaskCheck('taskOpenOnly', true);
  setTaskCheck('taskOverdue', false);
  setTaskCheck('taskUnassigned', item.owner === '未分配');
  setTaskCheck('taskReworked', false);
  loadTasks();
}
function taskSourceText(task){
  const source = [task.source_report, task.source_file].filter(Boolean).join(' / ');
  const row = task.source_row ? ` #${task.source_row}` : '';
  return `来源：${source || '-'}${row}`;
}
function canSubmitOwnerTask(task){
  return task.owner && (task.status === '待店长处理' || task.status === '已驳回');
}
function canReviewTask(task){
  return task.status === '待管理员审核';
}
function canMarkDoneTask(task){
  return task.status === '已通过';
}
function canAssignTask(task){
  return task.status !== '已完成';
}
function packageSourceText(pkg){
  const reports = (pkg.source_reports || []).slice(0, 2).join('、');
  const files = (pkg.source_files || []).slice(0, 2).join('、');
  return [reports, files].filter(Boolean).join(' / ') || '-';
}
function packageProgressText(pkg){
  const status = pkg.by_status || {};
  return `待推送 ${pkg.pending_push_count || 0} / 待店长 ${pkg.pending_owner_count || 0} / 待审核 ${pkg.pending_review_count || 0} / 已通过 ${pkg.approved_count || 0} / 已完成 ${pkg.done_count || 0}${pkg.overdue_count ? ' / 超时 ' + pkg.overdue_count : ''}${pkg.reworked_count ? ' / 返工 ' + pkg.reworked_count : ''}`;
}
function packageStatusBadges(pkg){
  const status = pkg.by_status || {};
  const pairs = [['待推送', status['待推送'] || 0], ['待店长处理', status['待店长处理'] || 0], ['待管理员审核', status['待管理员审核'] || 0], ['已驳回', status['已驳回'] || 0], ['已通过', status['已通过'] || 0], ['已完成', status['已完成'] || 0]].filter(([_label, count]) => count);
  return pairs.map(([label, count]) => `<span class="badge ${label === '待推送' || label === '待管理员审核' ? 'warn' : label === '已通过' || label === '已完成' ? 'ok' : label === '已驳回' ? 'bad' : ''}">${esc(label)} ${count}</span>`).join('');
}
function taskPackageById(id){
  return (taskState.packages || []).find(item => item.id === id);
}
function packageActionButtons(pkg){
  const preview = `<button class="secondary" onclick="togglePackagePreview('${pkg.id}')">${expandedTaskPackageId === pkg.id ? '收起' : '预览'}</button>`;
  if(operatorSession && operatorSession.role === 'owner'){
    const submit = (pkg.submittable_task_ids || []).length ? `<button class="primary" onclick="submitTaskPackage('${pkg.id}')">整包已处理</button>` : '';
    return `${preview}${submit}`;
  }
  const push = (pkg.pushable_task_ids || []).length ? `<button class="primary" onclick="pushTaskPackage('${pkg.id}')">推送店长</button>` : '';
  const approve = (pkg.reviewable_task_ids || []).length ? `<button class="primary" onclick="reviewTaskPackage('${pkg.id}','通过')">整包通过</button>` : '';
  const reject = (pkg.reviewable_task_ids || []).length ? `<button class="danger" onclick="reviewTaskPackage('${pkg.id}','驳回')">整包驳回</button>` : '';
  return `${preview}${push}${approve}${reject}`;
}
function taskActionButtons(task){
  const historyButton = `<button class="secondary" onclick="showTaskHistory('${task.id}')" title="查看操作记录">记录</button>`;
  const submitButton = !task.owner ? '<span class="muted">待指派</span>' : canSubmitOwnerTask(task) ? `<button class="secondary" onclick="submitTask('${task.id}')" title="店长填写处理结果">填写</button>` : '<span class="muted">-</span>';
  if(operatorSession && operatorSession.role === 'owner'){
    return `${historyButton}${submitButton}`;
  }
  const reviewButtons = canReviewTask(task) ? `<button class="primary" onclick="reviewTask('${task.id}','通过')" title="审核通过">通过</button><button class="danger" onclick="reviewTask('${task.id}','驳回')" title="审核驳回">驳回</button>` : '';
  const doneButton = canMarkDoneTask(task) ? `<button class="secondary" onclick="doneTask('${task.id}')" title="标记完成">完成</button>` : '';
  const assignButton = canAssignTask(task) ? `<button class="secondary" onclick="assignTask('${task.id}')" title="指派负责人">指派</button>` : '';
  return `${historyButton}${assignButton}${submitButton}${reviewButtons}${doneButton}`;
}
function selectedTaskIds(){
  const ids = [];
  const seen = new Set();
  Array.from(document.querySelectorAll('.task-check:checked')).forEach(input => {
    const values = (input.dataset.ids || input.value || '').split(',').map(item => item.trim()).filter(Boolean);
    values.forEach(value => {
      if(!seen.has(value)){
        seen.add(value);
        ids.push(value);
      }
    });
  });
  return ids;
}
function toggleAllTaskSelection(checked){
  document.querySelectorAll('.task-check').forEach(input => { input.checked = checked; });
}
function renderPackagePreview(pkg){
  const rows = pkg.sample_tasks || [];
  if(!rows.length) return '<div class="muted">暂无可预览明细。</div>';
  return `<div class="package-preview"><div class="package-preview-line package-preview-head"><span>商家编码</span><span>SKC ID</span><span>SPU ID</span><span>商品 / 异常明细</span><span>单条</span></div>${rows.map(task => {
    const pushButton = operatorSession?.role !== 'owner' && task.status === '待推送' && task.owner ? `<button class="secondary" onclick="pushSingleTask('${task.id}')">推送</button>` : '<span class="muted">-</span>';
    return `<div class="package-preview-line"><span>${esc(task.merchant_code || '-')}</span><span>${esc(task.skc || '-')}</span><span>${esc(task.spu || '-')}</span><span>${esc(task.product_name || task.task_detail || '-')}</span><span>${pushButton}</span></div>`;
  }).join('')}${pkg.total > rows.length ? `<div class="muted">还有 ${pkg.total - rows.length} 条明细，可点导出查看完整表格。</div>` : ''}</div>`;
}
function togglePackagePreview(id){
  expandedTaskPackageId = expandedTaskPackageId === id ? '' : id;
  renderTaskRows();
}
function renderTaskRows(){
  const tbody = document.getElementById('taskRows'); if(!tbody) return;
  const selectAll = document.getElementById('taskSelectAll');
  if(selectAll) selectAll.checked = false;
  if(!taskState.packages.length){
    tbody.innerHTML = '<div class="muted">暂无任务包。导入数据或店长填报后，系统会按店铺、负责人、类型生成任务包。</div>';
    return;
  }
  tbody.innerHTML = taskState.packages.map(pkg => {
    const actionable = operatorSession?.role === 'owner'
      ? (pkg.submittable_task_ids || [])
      : ((pkg.pushable_task_ids || []).length ? pkg.pushable_task_ids : (pkg.reviewable_task_ids || []));
    const ids = actionable;
    const priorityClass = pkg.priority === '高' ? 'bad' : pkg.priority === '中' ? 'warn' : pkg.priority === '低' ? 'ok' : '';
    const preview = expandedTaskPackageId === pkg.id ? `<div class="package-preview-card">${renderPackagePreview(pkg)}</div>` : '';
    return `<div class="package-card">
      <div class="task-select"><input class="task-check" type="checkbox" ${ids.length ? '' : 'disabled'} data-ids="${esc((ids || []).join(','))}"></div>
      <div class="package-metric">
        <div class="package-title"><strong>${esc(pkg.store || '-')}</strong><span class="badge ${priorityClass}">${esc(pkg.priority || '普通')}</span>${packageStatusBadges(pkg) || `<span class="badge">${esc(pkg.main_status || '-')}</span>`}</div>
        <div class="package-meta" title="${esc(pkg.system_action || '')}">${esc(pkg.system_action || '-')}</div>
      </div>
      <div class="package-metric"><span>负责人 / 平台</span><strong>${esc(pkg.owner || '-')}</strong><div class="package-sub">${esc(pkg.platform || '-')}</div></div>
      <div class="package-metric"><span>任务类型</span><strong>${esc(pkg.task_type || '-')}</strong><div class="package-sub">${esc(pkg.next_action || '')}</div></div>
      <div class="package-metric"><span>明细数量</span><strong>${pkg.total || 0}</strong><div class="package-sub">${esc(packageProgressText(pkg))}</div></div>
      <div class="package-metric"><span>来源</span><div class="package-meta" title="${esc(packageSourceText(pkg))}">${esc(packageSourceText(pkg))}</div><div class="task-actions">${packageActionButtons(pkg)}</div></div>
      ${preview}
    </div>`;
  }).join('');
}
function showTaskHistory(id){
  const task = (taskState.tasks || []).find(item => item.id === id);
  if(!task) return;
  const history = task.history || [];
  const title = `操作记录：${task.product_name || task.merchant_code || task.skc || task.spu || task.id}`;
  if(!history.length){ alert(`${title}\n暂无操作记录`); return; }
  const lines = history.map(item => {
    const nextAfter = [item.next_handler_after, item.next_action_after].filter(Boolean).join(' / ') || '-';
    return `${item.time || ''} ${item.event || ''}\n操作人：${item.actor || '-'}\n动作：${item.action || '-'}\n备注：${item.remark || '-'}\n处理凭证：${item.proof || '-'}\n动作后状态：${item.status_after || '-'}\n动作后下一步：${nextAfter}`;
  });
  alert(`${title}\n\n${lines.join('\n\n')}`);
}
function showTaskError(e){
  const st = document.getElementById('taskStatusLine');
  if(st) st.innerHTML = `<span class="bad">${esc(e.message || '任务操作失败')}</span>`;
}
async function submitTask(id){
  try {
    const actor = operatorSession?.user || document.getElementById('taskUser').value.trim() || prompt('填写人') || '';
    if(!actor) return;
    const action = prompt('处理动作，例如：已下架、申请退货、继续观察、同意议价');
    if(!action) return;
    const remark = prompt('处理备注，和处理凭证至少填一个') || '';
    const proof = prompt('处理凭证，例如截图链接、后台单号，和备注至少填一个') || '';
    if(!remark.trim() && !proof.trim()){ document.getElementById('taskStatusLine').textContent = '店长提交必须填写处理依据：备注或处理凭证至少填一个'; return; }
    await api('/api/tasks/submit', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({id, actor, action, remark, proof})});
    await loadTasks();
  } catch(e){ showTaskError(e); }
}
async function batchSubmitTasks(){
  try {
    if(operatorSession?.role !== 'owner'){ document.getElementById('taskStatusLine').textContent = '只有店长可以批量填写处理结果'; return; }
    const ids = selectedTaskIds();
    if(!ids.length){ document.getElementById('taskStatusLine').textContent = '请先勾选要批量处理的任务'; return; }
    const action = prompt('批量处理动作，例如：已下架、申请退货、继续观察、同意议价');
    if(!action) return;
    const remark = prompt('批量处理备注，和处理凭证至少填一个') || '';
    const proof = prompt('批量处理凭证，例如截图链接、后台单号，和备注至少填一个') || '';
    if(!remark.trim() && !proof.trim()){ document.getElementById('taskStatusLine').textContent = '店长提交必须填写处理依据：备注或处理凭证至少填一个'; return; }
    const res = await api('/api/tasks/batch-submit', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({ids, action, remark, proof})});
    document.getElementById('taskStatusLine').textContent = `已批量提交 ${res.count || 0} 条任务，等待管理员审核`;
    await loadTasks(false);
  } catch(e){ showTaskError(e); }
}
async function submitTaskPackage(packageId){
  const pkg = taskPackageById(packageId);
  if(!pkg) return;
  const ids = pkg.submittable_task_ids || [];
  if(!ids.length){ document.getElementById('taskStatusLine').textContent = '这个任务包当前没有可提交的明细'; return; }
  try {
    if(operatorSession?.role !== 'owner'){ document.getElementById('taskStatusLine').textContent = '只有店长可以提交任务包'; return; }
    const action = prompt(`整包处理动作：${pkg.store || ''} / ${pkg.task_type || ''} / ${pkg.system_action || ''}`, '已处理');
    if(!action) return;
    const remark = prompt('整包处理备注，和处理凭证至少填一个') || '';
    const proof = prompt('整包处理凭证，例如截图链接、后台单号，和备注至少填一个') || '';
    if(!remark.trim() && !proof.trim()){ document.getElementById('taskStatusLine').textContent = '店长提交必须填写处理依据：备注或处理凭证至少填一个'; return; }
    const res = await api('/api/tasks/batch-submit', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({ids, action, remark, proof})});
    document.getElementById('taskStatusLine').textContent = `任务包已提交 ${res.count || 0} 条，等待管理员审核`;
    await loadTasks(false);
  } catch(e){ showTaskError(e); }
}
async function batchPushTasks(){
  try {
    if(operatorSession?.role === 'owner'){ document.getElementById('taskStatusLine').textContent = '店长不能推送任务包'; return; }
    const ids = selectedTaskIds();
    if(!ids.length){ document.getElementById('taskStatusLine').textContent = '请先勾选要推送的待推送任务包'; return; }
    const remark = prompt('推送说明，可空', '管理员确认推送') || '';
    const res = await api('/api/tasks/batch-push', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({ids, remark})});
    document.getElementById('taskStatusLine').textContent = `已推送 ${res.count || 0} 条任务给店长`;
    await loadTasks(false);
  } catch(e){ showTaskError(e); }
}
async function pushTaskPackage(packageId){
  const pkg = taskPackageById(packageId);
  if(!pkg) return;
  const ids = pkg.pushable_task_ids || [];
  if(!ids.length){ document.getElementById('taskStatusLine').textContent = '这个任务包当前没有待推送明细'; return; }
  try {
    const remark = prompt(`推送给 ${pkg.owner || '负责人'}：${pkg.store || ''} / ${pkg.task_type || ''}`, '管理员确认推送') || '';
    const res = await api('/api/tasks/batch-push', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({ids, remark})});
    document.getElementById('taskStatusLine').textContent = `任务包已推送 ${res.count || 0} 条给店长`;
    await loadTasks(false);
  } catch(e){ showTaskError(e); }
}
async function pushSingleTask(id){
  try {
    const remark = prompt('单条推送说明，可空', '管理员确认推送') || '';
    const res = await api('/api/tasks/batch-push', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({ids:[id], remark})});
    document.getElementById('taskStatusLine').textContent = `已单条推送 ${res.count || 0} 条任务`;
    await loadTasks(false);
  } catch(e){ showTaskError(e); }
}
async function assignTask(id){
  try {
    const owner = prompt('指派给负责人');
    if(!owner) return;
    const remark = prompt('指派备注') || '';
    await api('/api/tasks/assign', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({id, owner, remark})});
    await loadTasks();
  } catch(e){ showTaskError(e); }
}
async function reviewTask(id, decision){
  try {
    const admin = document.getElementById('taskUser').value.trim() || prompt('管理员') || '管理员';
    const remark = prompt(decision === '驳回' ? '管理员审核：驳回原因（必填）' : `管理员审核：${decision}说明（必填）`) || '';
    if(!remark.trim()){ document.getElementById('taskStatusLine').textContent = '管理员审核必须填写说明'; return; }
    await api('/api/tasks/review', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({id, admin, decision, remark})});
    await loadTasks();
  } catch(e){ showTaskError(e); }
}
async function batchReviewTasks(decision){
  try {
    const ids = selectedTaskIds();
    if(!ids.length){ document.getElementById('taskStatusLine').textContent = '请先勾选要批量审核的任务'; return; }
    const remark = prompt(decision === '驳回' ? '批量驳回原因（必填）' : `批量${decision}说明（必填）`) || '';
    if(!remark.trim()){ document.getElementById('taskStatusLine').textContent = '批量审核必须填写说明'; return; }
    const res = await api('/api/tasks/batch-review', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({ids, decision, remark})});
    document.getElementById('taskStatusLine').textContent = `已批量${decision} ${res.count || 0} 条任务`;
    await loadTasks(false);
  } catch(e){ showTaskError(e); }
}
async function reviewTaskPackage(packageId, decision){
  const pkg = taskPackageById(packageId);
  if(!pkg) return;
  const ids = pkg.reviewable_task_ids || [];
  if(!ids.length){ document.getElementById('taskStatusLine').textContent = '这个任务包当前没有待审核明细'; return; }
  try {
    const remark = prompt(decision === '驳回' ? '整包驳回原因（必填）' : `整包${decision}说明（必填）`) || '';
    if(!remark.trim()){ document.getElementById('taskStatusLine').textContent = '整包审核必须填写说明'; return; }
    const res = await api('/api/tasks/batch-review', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({ids, decision, remark})});
    document.getElementById('taskStatusLine').textContent = `任务包已${decision} ${res.count || 0} 条`;
    await loadTasks(false);
  } catch(e){ showTaskError(e); }
}
async function doneTask(id){
  try {
    const remark = prompt('完成确认说明（必填）') || '';
    if(!remark.trim()){ document.getElementById('taskStatusLine').textContent = '标记完成必须填写确认说明'; return; }
    await api('/api/tasks/done', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({id, remark})});
    await loadTasks();
  } catch(e){ showTaskError(e); }
}
async function exportTasks(){
  const st = document.getElementById('taskStatusLine');
  try {
    if(st) st.textContent = '正在导出任务表，请稍等...';
    const payload = Object.fromEntries(taskQuery().entries());
    const res = await api('/api/tasks/export', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(payload)});
    const downloadUrl = authDownload(res.download);
    if(st) st.innerHTML = `<span class="ok">已导出 ${res.rows} 条，正在下载：</span><a href="${downloadUrl}">${esc(res.file)}</a>`;
    const link = document.createElement('a');
    link.href = downloadUrl;
    link.download = res.file || '';
    link.style.display = 'none';
    document.body.appendChild(link);
    link.click();
    link.remove();
    await refreshStatus();
  } catch(e){ showTaskError(e); }
}
async function createBackup(){
  const st = document.getElementById('backupStatus');
  st.textContent = '正在生成备份...';
  try {
    const res = await api('/api/backup/create', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({})});
    st.innerHTML = `<span class="ok">备份已生成：</span>${esc(res.path)}，共 ${res.count} 个文件。`;
  } catch(e){ st.innerHTML = `<span class="bad">${e.message}</span>`; }
}
async function restoreBackup(){
  const st = document.getElementById('backupStatus');
  const path = document.getElementById('restoreBackupPath').value.trim();
  if(!path){ st.textContent = '请先填写备份文件路径'; return; }
  if(!confirm('恢复备份会覆盖当前运营状态和数据源，确认继续？')) return;
  st.textContent = '正在恢复备份...';
  try {
    const res = await api('/api/backup/restore', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({path})});
    st.innerHTML = `<span class="ok">恢复完成：</span>${res.count} 个文件。`;
    await refreshStatus();
  } catch(e){ st.innerHTML = `<span class="bad">${e.message}</span>`; }
}
async function runSearch(){
  const q = document.getElementById('searchQuery').value.trim();
  const limit = document.getElementById('searchLimit').value;
  const st = document.getElementById('searchStatus'); const body = document.getElementById('searchRows');
  if(!q){ st.textContent='请输入关键词'; return; }
  st.textContent='正在搜索...'; body.innerHTML='';
  try {
    const res = await api(`/api/search?q=${encodeURIComponent(q)}&limit=${limit}`);
    st.textContent = `找到 ${res.rows.length} 条`;
    res.rows.forEach(row => {
      const tr=document.createElement('tr');
      tr.innerHTML=`<td>${esc(row.file_name)}</td><td>${esc(row.sheet_name)}</td><td>${row.source_row}</td><td>${esc(row.content)}</td>`;
      body.appendChild(tr);
    });
  } catch(e){ st.innerHTML=`<span class="bad">${e.message}</span>`; }
}
async function exportSearch(){
  const q = document.getElementById('searchQuery').value.trim();
  const st = document.getElementById('searchStatus');
  if(!q){ st.textContent='请输入关键词'; return; }
  try {
    const res = await api('/api/search/export', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({q, limit:500})});
    st.innerHTML = `<span class="ok">已导出 ${res.rows} 条：</span><a href="${authDownload(res.download)}">${res.file}</a>`;
    await refreshStatus();
  } catch(e){ st.innerHTML=`<span class="bad">${e.message}</span>`; }
}
function renderOutputs(){
  const body = document.getElementById('outputRows'); body.innerHTML='';
  appStatus.outputs.forEach(f => {
    const tr=document.createElement('tr');
    tr.innerHTML=`<td>${esc(f.name)}</td><td>${fmtSize(f.size)}</td><td>${f.modified}</td><td><a href="${authDownload(f.download)}">下载</a></td>`;
    body.appendChild(tr);
  });
}
function esc(s){ return String(s ?? '').replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c])); }
applyEntryParams();
initializeBusinessDatesWeb();
showUpgradeDialog();
refreshStatus();
</script>
</body>
</html>"""


class DailyOpsHandler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        print("[%s] %s" % (datetime.now().strftime("%H:%M:%S"), fmt % args))

    def client_disconnected(self, exc):
        if isinstance(exc, (BrokenPipeError, ConnectionResetError)):
            return True
        return isinstance(exc, OSError) and exc.errno in {errno.EPIPE, errno.ECONNRESET}

    def send_payload(self, status, content_type, body):
        try:
            self.send_response(status)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except OSError as exc:
            if self.client_disconnected(exc):
                return
            raise

    def send_json(self, payload, status=200):
        self.send_payload(*json_bytes(payload, status=status))

    def do_GET(self):
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/health":
                self.send_json({"ok": True, "status": "running"})
                return
            cancel_scheduled_shutdown()
            if parsed.path == "/":
                self.send_payload(200, "text/html; charset=utf-8", HTML_PAGE.encode("utf-8"))
            elif parsed.path == "/api/status":
                self.send_payload(*handle_status_api(self.headers))
            elif parsed.path == "/api/rules":
                self.send_payload(*handle_rules_api("GET", self.headers))
            elif parsed.path == "/api/search":
                params = parse_qs(parsed.query)
                self.send_payload(*handle_search_api("GET", self.headers, {
                    "q": params.get("q", [""])[0],
                    "limit": params.get("limit", ["100"])[0],
                }))
            elif parsed.path == "/api/tasks":
                params = parse_qs(parsed.query)
                status, content_type, body = handle_tasks_api("GET", self.headers, task_query_payload(params))
                self.send_payload(status, content_type, body)
            elif parsed.path == "/api/bargain/history":
                params = parse_qs(parsed.query)
                self.send_payload(*handle_bargain_api("GET_HISTORY", self.headers, {
                    "merchant_code": params.get("merchant_code", [""])[0],
                    "goods_code": params.get("goods_code", [""])[0],
                    "store": params.get("store", [""])[0],
                    "platform": params.get("platform", [""])[0],
                    "owner": params.get("owner", [""])[0],
                    "status": params.get("status", [""])[0],
                }))
            elif parsed.path == "/api/bargain/clearance":
                self.send_payload(*handle_bargain_api("GET_CLEARANCE", self.headers))
            elif parsed.path == "/api/owners":
                self.send_payload(*handle_owners_api(self.headers))
            elif parsed.path == "/api/store-owners":
                self.send_payload(*handle_store_owners_api("GET", self.headers))
            elif parsed.path == "/api/business-report":
                params = parse_qs(parsed.query)
                self.send_payload(*handle_business_report_api(self.headers, {
                    "date_from": params.get("date_from", [""])[0],
                    "date_to": params.get("date_to", [""])[0],
                    "platform": params.get("platform", [""])[0],
                    "store": params.get("store", [""])[0],
                    "grain": params.get("grain", ["month"])[0],
                }))
            elif parsed.path == "/download":
                self.handle_download(parsed)
            else:
                self.send_json({"ok": False, "error": "接口不存在"}, 404)
        except PermissionError as exc:
            self.send_json({"ok": False, "error": str(exc)}, 401)
        except OSError as exc:
            if self.client_disconnected(exc):
                return
            self.send_json({"ok": False, "error": str(exc)}, 500)
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, 500)

    def do_POST(self):
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/upload":
                cancel_scheduled_shutdown()
                if not self.require_upload_operator_request("上传数据源"):
                    return
                self.handle_upload()
            elif parsed.path == "/api/session/login":
                cancel_scheduled_shutdown()
                payload = self.read_json()
                session = login_operator(payload.get("role", ""), payload.get("user", ""), payload.get("password", ""))
                self.send_json({"ok": True, "session": session})
            elif parsed.path == "/api/session/logout":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_session_logout(self.headers))
            elif parsed.path == "/api/reports/run":
                cancel_scheduled_shutdown()
                if not self.require_admin_request("生成报表"):
                    return
                payload = self.read_json()
                result = run_report(payload.get("report"), payload.get("version", "V1"))
                self.send_json({"ok": True, "result": result})
            elif parsed.path == "/api/weekly/run":
                cancel_scheduled_shutdown()
                if not self.require_admin_request("生成本周报表"):
                    return
                self.send_json({"ok": True, "result": run_weekly_reports()})
            elif parsed.path == "/api/rules":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_rules_api("POST_SAVE", self.headers, self.read_json()))
            elif parsed.path == "/api/store-owners":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_store_owners_api("POST_SAVE", self.headers, self.read_json()))
            elif parsed.path == "/api/upload/finish-batch":
                cancel_scheduled_shutdown()
                payload = self.read_json()
                if not self.require_upload_operator_request("结束上传", payload.get("category", "")):
                    return
                self.send_json({"ok": True, "source_state": finish_upload_batch(payload.get("category", ""))})
            elif parsed.path == "/api/upload/clear-batch":
                cancel_scheduled_shutdown()
                payload = self.read_json()
                if not self.require_upload_operator_request("清空待提交文件", payload.get("category", "")):
                    return
                self.send_json({"ok": True, **clear_upload_batch(payload.get("category", ""))})
            elif parsed.path == "/api/search/export":
                cancel_scheduled_shutdown()
                if not self.require_admin_request("导出基础数据查询"):
                    return
                payload = self.read_json()
                result = export_search(payload.get("q", ""), payload.get("limit", 500))
                self.send_json({"ok": True, **result})
            elif parsed.path == "/api/tasks/submit":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_tasks_api("POST_SUBMIT", self.headers, self.read_json()))
            elif parsed.path == "/api/tasks/batch-submit":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_tasks_api("POST_BATCH_SUBMIT", self.headers, self.read_json()))
            elif parsed.path == "/api/tasks/assign":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_tasks_api("POST_ASSIGN", self.headers, self.read_json()))
            elif parsed.path == "/api/tasks/batch-push":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_tasks_api("POST_BATCH_PUSH", self.headers, self.read_json()))
            elif parsed.path == "/api/tasks/review":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_tasks_api("POST_REVIEW", self.headers, self.read_json()))
            elif parsed.path == "/api/tasks/batch-review":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_tasks_api("POST_BATCH_REVIEW", self.headers, self.read_json()))
            elif parsed.path == "/api/tasks/done":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_tasks_api("POST_DONE", self.headers, self.read_json()))
            elif parsed.path == "/api/tasks/export":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_tasks_api("POST_EXPORT", self.headers, self.read_json()))
            elif parsed.path == "/api/bargain/clearance/rebuild":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_bargain_api("POST_REBUILD_CLEARANCE", self.headers, self.read_json()))
            elif parsed.path == "/api/bargain/lookup":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_bargain_api("POST_LOOKUP", self.headers, self.read_json()))
            elif parsed.path == "/api/bargain/submit":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_bargain_api("POST_SUBMIT", self.headers, self.read_json()))
            elif parsed.path == "/api/bargain/review":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_bargain_api("POST_REVIEW", self.headers, self.read_json()))
            elif parsed.path == "/api/bargain/resubmit":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_bargain_api("POST_RESUBMIT", self.headers, self.read_json()))
            elif parsed.path == "/api/bargain/low-price-trace":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_bargain_api("POST_LOW_PRICE_TRACE", self.headers, self.read_json()))
            elif parsed.path == "/api/bargain/low-price-ignore":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_bargain_api("POST_IGNORE_LOW_PRICE", self.headers, self.read_json()))
            elif parsed.path == "/api/backup/create":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_backup_api("CREATE", self.headers, self.read_json()))
            elif parsed.path == "/api/backup/restore":
                cancel_scheduled_shutdown()
                self.send_payload(*handle_backup_api("RESTORE", self.headers, self.read_json()))
            elif parsed.path == "/api/client-close":
                self.send_json({"ok": True, "message": "页面关闭通知已忽略，工作台保持运行"})
            elif parsed.path == "/api/shutdown":
                cancel_scheduled_shutdown()
                if not self.require_admin_request("关闭工作台"):
                    return
                self.send_json({"ok": True, "message": "工作台正在关闭"})
                threading.Thread(target=self.server.shutdown, daemon=True).start()
            else:
                self.send_json({"ok": False, "error": "接口不存在"}, 404)
        except OSError as exc:
            if self.client_disconnected(exc):
                return
            self.send_json({"ok": False, "error": str(exc)}, 500)
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc)}, 500)

    def read_json(self):
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length) if length else b"{}"
        return json.loads(raw.decode("utf-8") or "{}")

    def require_admin_request(self, action):
        status, content_type, body = handle_admin_api(action, self.headers)
        if status == 200:
            return True
        self.send_payload(status, content_type, body)
        return False

    def require_upload_operator_request(self, action, category=""):
        status, content_type, body = handle_upload_operator_api(action, self.headers, category)
        if status == 200:
            return True
        self.send_payload(status, content_type, body)
        return False

    def handle_upload(self):
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type")})
        category = form.getfirst("category", "")
        if category not in UPLOAD_TARGETS:
            raise ValueError("未知上传分类")
        if not self.require_upload_operator_request("上传数据源", category):
            return
        file_item = form["file"] if "file" in form else None
        if file_item is None or not file_item.filename:
            raise ValueError("未收到上传文件")
        label, folder = UPLOAD_TARGETS[category]
        target = unique_upload_path(folder, file_item.filename)
        with target.open("wb") as fh:
            shutil.copyfileobj(file_item.file, fh)
        target = normalize_uploaded_workbook(target)
        source_state = record_uploaded_source(category, target)
        self.send_json({"ok": True, "category": label, "saved": str(target), "source_state": source_state})

    def handle_download(self, parsed):
        params = parse_qs(parsed.query)
        token = token_from_headers(getattr(self, "headers", {})) or params.get("token", [""])[0]
        name = unquote(params.get("path", [""])[0])
        require_download_permission(token, name)
        path = (OUTPUT_DIR / Path(name).name).resolve()
        if OUTPUT_DIR.resolve() not in path.parents and path != OUTPUT_DIR.resolve():
            raise ValueError("下载路径不允许")
        if not path.exists() or not path.is_file():
            raise FileNotFoundError("文件不存在")
        data = path.read_bytes()
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        encoded = quote(path.name)
        self.send_header("Content-Disposition", f'attachment; filename="download{path.suffix}"; filename*=UTF-8\'\'{encoded}')
        self.end_headers()
        try:
            self.wfile.write(data)
        except OSError as exc:
            if self.client_disconnected(exc):
                return
            raise


class ReusableThreadingHTTPServer(ThreadingHTTPServer):
    allow_reuse_address = True


def main():
    server = None
    try:
        OUTPUT_DIR.mkdir(exist_ok=True)
        host = configured_host()
        server = ReusableThreadingHTTPServer((host, PORT), DailyOpsHandler)
        print(f"PETCIRCLE跨境工作台已启动：http://{host}:{PORT}")
        print(access_hint(host, PORT))
        print("按 Ctrl+C 停止服务")
        server.serve_forever()
    except Exception as exc:
        OUTPUT_DIR.mkdir(exist_ok=True)
        log = OUTPUT_DIR / "daily_ops_app_startup_error.log"
        log.write_text(f"{datetime.now():%Y-%m-%d %H:%M:%S} {type(exc).__name__}: {exc}\n", encoding="utf-8")
        raise
    finally:
        if server is not None:
            server.server_close()


if __name__ == "__main__":
    main()
