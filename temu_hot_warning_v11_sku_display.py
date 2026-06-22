from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import temu_hot_warning_v13 as base


VERSION = "爆旺款重复预警V1.1正式版（SPU定义爆款，SKU维度展示）"
OUTPUT_FILE = base.ROOT / "Temu爆旺款重复预警表_V1.1正式版.xlsx"


def action_for_sku(row, champion, is_bomb_conflict):
    if is_bomb_conflict:
        return base.ACT_BOMB_STOCK if row["stock"] > 0 else base.ACT_NOW
    if base.is_low(row["price"], champion["price"]) == base.YES:
        return base.ACT_NOW
    if row["stock"] > 0:
        return base.ACT_SELL_OUT
    return base.ACT_NOW


def build_sku_rows(owners, erp, hot_groups, sales_rows, champions):
    rows = []
    seen = set()
    hot_group_keys = set(hot_groups)

    for row in sales_rows:
        product = row["product_code"]
        champion = champions.get(product)
        if not champion:
            continue

        row_group_key = (product, row["shop"], row["skc"])
        champion_key = (product, champion["shop"], champion["skc"])
        if row_group_key == champion_key:
            continue

        owner = owners.get(row["shop"], "")
        if owner not in base.OWNER_KEEP:
            continue

        is_bomb_conflict = row_group_key in hot_group_keys
        conflict_type = base.TYPE_BOMB if is_bomb_conflict else base.TYPE_FLAT
        action = action_for_sku(row, champion, is_bomb_conflict)

        erp_info = erp.get(row["sku_code"]) or erp.get(product) or {}
        dedupe = (
            row["sku_code"],
            row["shop"],
            row["skc"],
            champion["shop"],
            champion["skc"],
            action,
        )
        if dedupe in seen:
            continue
        seen.add(dedupe)

        rows.append(
            {
                "商家编码": row["sku_code"],
                "货品名称": erp_info.get("goods", ""),
                "skc": row["skc"],
                "所属店铺": base.shop_no(row["shop"]),
                "爆旺款skc": champion["skc"],
                "爆旺款店铺": champion["shop_raw"],
                "爆旺款申报价": champion["price"],
                "重复款申报价": row["price"],
                "爆旺款月销": champion["sales30"],
                "重复款月销": row["sales30"],
                "是否低于爆旺款申报价": base.is_low(row["price"], champion["price"]),
                "爆旺款平台仓备货库存": champion["stock"],
                "重复款备货库存": row["stock"],
                "负责人": owner,
                "冲突类型": conflict_type,
                "处理意见": action,
            }
        )

    rows.sort(
        key=lambda r: (
            base.sort_shop_no(r["所属店铺"]),
            r["skc"],
            r["商家编码"],
            r["爆旺款skc"],
        )
    )
    return rows


def build_overview(operation_rows, hot_groups, owners):
    store_names = {}
    for group in hot_groups.values():
        store_names.setdefault(base.shop_no(group["shop"]), group["shop_raw"])

    overview = defaultdict(
        lambda: {
            "bomb": set(),
            "repeat": 0,
            "flat": 0,
            "bomb_conflict": 0,
            "low": 0,
            "not_low": 0,
            "now": 0,
            "sell_out": 0,
            "clear_30": 0,
        }
    )

    for group in hot_groups.values():
        owner = owners.get(group["shop"], "")
        if owner in base.OWNER_KEEP:
            no = base.shop_no(group["shop"])
            overview[(no, group["shop_raw"], owner)]["bomb"].add(
                (group["product_code"], group["skc"])
            )

    for row in operation_rows:
        no = str(row["所属店铺"])
        key = (no, store_names.get(no, no), row["负责人"])
        item = overview[key]
        item["repeat"] += 1
        if row["冲突类型"] == base.TYPE_FLAT:
            item["flat"] += 1
        else:
            item["bomb_conflict"] += 1
        if row["是否低于爆旺款申报价"] == base.YES:
            item["low"] += 1
        elif row["是否低于爆旺款申报价"] == base.NO:
            item["not_low"] += 1
        if row["处理意见"] == base.ACT_NOW:
            item["now"] += 1
        elif row["处理意见"] == base.ACT_SELL_OUT:
            item["sell_out"] += 1
        elif row["处理意见"] == base.ACT_BOMB_STOCK:
            item["clear_30"] += 1

    rows = []
    for (no, shop, owner), item in sorted(
        overview.items(), key=lambda x: (base.sort_shop_no(x[0][0]), x[0][1])
    ):
        rows.append(
            {
                "店铺编号": no,
                "店铺": shop,
                "负责人": owner,
                "爆款总数": len(item["bomb"]),
                "重复铺货预计总数": item["repeat"],
                "平销冲突数": item["flat"],
                "爆款互相冲突数": item["bomb_conflict"],
                "低于爆旺款申报价数": item["low"],
                "不低于爆旺款申报价数": item["not_low"],
                "立即下架数": item["now"],
                "售完备货库存下架数": item["sell_out"],
                "30天内限时下架数": item["clear_30"],
            }
        )
    return rows


def style_sheet(ws, table_name, low_header=None, action_header=None):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    line = Side(style="hair", color="D9E2F3")
    alert_font = Font(color="C00000", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color="D9E2F3"))

    headers = [cell.value for cell in ws[1]]
    low_col = headers.index(low_header) + 1 if low_header in headers else 0
    action_col = headers.index(action_header) + 1 if action_header in headers else 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = white_fill
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=line)
        if low_col and row[low_col - 1].value == base.YES:
            row[low_col - 1].fill = red_fill
        if action_col and row[action_col - 1].value == base.ACT_NOW:
            row[action_col - 1].font = alert_font

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter][: min(ws.max_row, 250)]:
            max_len = max(max_len, len("" if cell.value is None else str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 46)

    for header in ["爆旺款申报价", "重复款申报价"]:
        if header in headers:
            col = headers.index(header) + 1
            for cells in ws.iter_cols(
                min_col=col, max_col=col, min_row=2, max_row=ws.max_row
            ):
                for cell in cells:
                    cell.number_format = "0.00"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row >= 2:
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=False, showColumnStripes=False
        )
        ws.add_table(table)


def write_workbook(overview_rows, operation_rows):
    wb = Workbook()

    overview_headers = [
        "店铺编号",
        "店铺",
        "负责人",
        "爆款总数",
        "重复铺货预计总数",
        "平销冲突数",
        "爆款互相冲突数",
        "低于爆旺款申报价数",
        "不低于爆旺款申报价数",
        "立即下架数",
        "售完备货库存下架数",
        "30天内限时下架数",
    ]
    ws = wb.active
    ws.title = "总览"
    ws.append(overview_headers)
    for row in overview_rows:
        ws.append([row.get(header) for header in overview_headers])

    operation_headers = [
        "商家编码",
        "货品名称",
        "skc",
        "所属店铺",
        "爆旺款skc",
        "爆旺款店铺",
        "爆旺款申报价",
        "重复款申报价",
        "爆旺款月销",
        "重复款月销",
        "是否低于爆旺款申报价",
        "爆旺款平台仓备货库存",
        "重复款备货库存",
        "负责人",
        "冲突类型",
        "处理意见",
    ]
    ws_op = wb.create_sheet("具体店铺操作表")
    ws_op.append(operation_headers)
    for row in operation_rows:
        ws_op.append([row.get(header) for header in operation_headers])

    ws_note = wb.create_sheet("说明")
    note_rows = [
        ["项目", "说明"],
        ["版本", VERSION],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["表格结构", "1总览；2具体店铺操作表；3说明"],
        [
            "爆旺款定义",
            "爆旺款链接按SPU/货品编码判断：商家编码去掉尺码后得到货品编码，同一SKC内同一货品编码所有尺码销量和库存合并，再比较各SKC链接的汇总销量。",
        ],
        [
            "展示维度",
            "具体店铺操作表按SKU维度逐尺码展示，商家编码保留尺码；爆旺款月销和库存使用SPU/SKC汇总值。",
        ],
        [
            "优先级规则",
            "同一货品编码多链接冲突时，30天销量汇总越高优先级越高；汇总销量相同则申报价越高优先级越高；再相同则平台仓备货库存汇总越多优先级越高。",
        ],
        ["具体店铺操作表行数", len(operation_rows)],
    ]
    for row in note_rows:
        ws_note.append(row)

    style_sheet(ws, "OverviewTable")
    style_sheet(ws_op, "OperationTable", "是否低于爆旺款申报价", "处理意见")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws_note[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws_note.column_dimensions["A"].width = 30
    ws_note.column_dimensions["B"].width = 140
    ws_note.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)


def main():
    owners = base.load_owners()
    erp = base.load_erp()
    hot_rows = base.read_hot_rows()
    hot_groups = base.aggregate_groups(hot_rows)
    champions = base.choose_champions(hot_groups)
    sales_rows = base.read_sales_rows(set(champions))
    operation_rows = build_sku_rows(owners, erp, hot_groups, sales_rows, champions)
    overview_rows = build_overview(operation_rows, hot_groups, owners)
    write_workbook(overview_rows, operation_rows)

    print(f"OUTPUT={OUTPUT_FILE}")
    print(f"HOT_ROWS={len(hot_rows)}")
    print(f"HOT_PRODUCT_CODES={len(champions)}")
    print(f"HOT_GROUPS={len(hot_groups)}")
    print(f"SALES_ROWS={len(sales_rows)}")
    print(f"OPERATION_ROWS={len(operation_rows)}")
    print(
        "LOW_ROWS="
        + str(sum(1 for row in operation_rows if row["是否低于爆旺款申报价"] == base.YES))
    )
    print(
        "BOMB_CONFLICT_ROWS="
        + str(sum(1 for row in operation_rows if row["冲突类型"] == base.TYPE_BOMB))
    )


if __name__ == "__main__":
    main()
