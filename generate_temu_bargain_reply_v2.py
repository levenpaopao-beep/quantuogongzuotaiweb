import html
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(os.environ.get("DAILYWORK_ROOT", Path(__file__).resolve().parent))
SOURCE_DIR = Path(os.environ.get("TEMU_BARGAIN_SOURCE_DIR", ROOT / "核价输入表"))
DOC_FILE = SOURCE_DIR / "260605-temu议价回复V2.md"
APPROVAL_FILE = SOURCE_DIR / "20260605-新品申报表.xlsx"
OUTPUT_FILE = ROOT / "outputs" / "260605-temu议价回复V2.xlsx"
PROJECT_DOC_FILE = ROOT / "260605-temu议价回复V2.md"

HEADERS = [
    "商家编码",
    "货品名称",
    "规格名称",
    "店铺（显示编号）",
    "申报价",
    "建议价格",
    "是否通过",
    "平台在售链接数",
    "平台在售最低申报价",
    "平台在售最高月销量",
    "平台在售上架最长时间",
]


def text(value):
    return "" if value is None else html.unescape(str(value)).strip()


def number(value):
    raw = text(value).replace(",", "")
    if raw == "":
        return ""
    try:
        value = float(raw)
    except ValueError:
        return raw
    return int(value) if value.is_integer() else value


def parse_markdown_rows(path):
    rows = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.startswith("| ") or line.startswith("| ---") or "商家编码" in line:
            continue
        safe = line.replace("\\|", "<<PIPE>>")
        cells = [cell.strip().replace("<<PIPE>>", "|") for cell in safe.strip("|").split("|")]
        if len(cells) != len(HEADERS):
            continue
        rows.append(
            [
                cells[0],
                cells[1],
                cells[2],
                number(cells[3]),
                number(cells[4]),
                number(cells[5]),
                cells[6],
                number(cells[7]),
                number(cells[8]),
                number(cells[9]),
                number(cells[10]),
            ]
        )
    return rows


def approval_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(value is not None for value in row)]
    wb.close()
    return rows


def assert_approval_matches(doc_rows, approval):
    if len(doc_rows) != len(approval):
        raise RuntimeError(f"执行文档明细 {len(doc_rows)} 行，审批表 {len(approval)} 行，行数不一致")
    for idx, (doc, raw) in enumerate(zip(doc_rows, approval), start=2):
        checks = [
            text(doc[1]) == text(raw[1]),
            text(doc[3]) == text(number(raw[3])),
            float(doc[4]) == float(raw[4]),
            float(doc[5]) == float(raw[5]),
        ]
        if not all(checks):
            raise RuntimeError(f"第 {idx} 行与审批表不一致：{doc[:6]} / {raw[:6]}")


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    agree_fill = PatternFill("solid", fgColor="E2F0D9")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
        if len(row) >= 7 and row[6].value == "同意议价":
            row[6].fill = agree_fill

    widths = [20, 30, 20, 14, 12, 12, 18, 16, 20, 20, 22]
    for idx, width in enumerate(widths[: ws.max_column], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for col in [5, 6, 9]:
        if col > ws.max_column:
            continue
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2, max_row=ws.max_row):
            for item in cell:
                item.number_format = "0.00"
    for col in [4, 8, 10, 11]:
        if col > ws.max_column:
            continue
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2, max_row=ws.max_row):
            for item in cell:
                item.number_format = "0"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False


def write_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "议价回复"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    style_sheet(ws)

    check = wb.create_sheet("数据校验")
    check.append(["检查项", "结果"])
    check.append(["执行文档", DOC_FILE.name])
    check.append(["审批表", APPROVAL_FILE.name])
    check.append(["明细行数", len(rows)])
    check.append(["同意议价", sum(1 for row in rows if row[6] == "同意议价")])
    style_sheet(check)

    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    wb.save(OUTPUT_FILE)


def main():
    rows = parse_markdown_rows(DOC_FILE)
    approval = approval_rows(APPROVAL_FILE)
    assert_approval_matches(rows, approval)
    PROJECT_DOC_FILE.write_text(DOC_FILE.read_text(encoding="utf-8"), encoding="utf-8")
    write_workbook(rows)
    print(OUTPUT_FILE)
    print(f"rows={len(rows)}")
    print(f"agree={sum(1 for row in rows if row[6] == '同意议价')}")


if __name__ == "__main__":
    main()
