import hashlib
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


STATUS_PENDING_REVIEW = "待管理员审核"
STATUS_APPROVED = "已通过"
STATUS_REJECTED = "不通过"
STATUS_REWORK = "需调整"


def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def text(value):
    return "" if value is None else str(value).strip()


def number(value, default=0.0):
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return default


def read_workbook_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [text(cell) for cell in next(rows)]
        except StopIteration:
            return []
        result = []
        for row in rows:
            item = {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}
            if any(text(value) for value in item.values()):
                result.append(item)
        return result
    finally:
        wb.close()


def first_value(row, *names):
    for name in names:
        if name in row and text(row.get(name)):
            return row.get(name)
    return ""


def size_from_spec(spec_name, merchant_code=""):
    spec = text(spec_name)
    if "/" in spec:
        return spec.rsplit("/", 1)[-1].strip()
    code = text(merchant_code)
    if "-" in code:
        return code.rsplit("-", 1)[-1].strip()
    return spec or code


def goods_code_from_merchant(merchant_code):
    code = text(merchant_code)
    code = code.split("@", 1)[0]
    if "-" in code:
        return code.rsplit("-", 1)[0]
    return code


def canonical_merchant_code(merchant_code):
    return text(merchant_code).split("@", 1)[0]


def standard_size_merchant_code(merchant_code, goods_code):
    code = canonical_merchant_code(merchant_code)
    prefix = text(goods_code)
    if not code or not prefix or not code.startswith(f"{prefix}-"):
        return False
    suffix = code.rsplit("-", 1)[-1].strip().upper()
    return suffix in {"XXS", "XS", "S", "M", "L", "XL", "XXL", "XXXL"}


def load_erp_items(erp_files):
    items = []
    seen = set()
    for path in erp_files or []:
        path = Path(path)
        if not path.exists():
            continue
        for row in read_workbook_rows(path):
            merchant_code = text(first_value(row, "商家编码", "商家编码（新）", "spec_no"))
            goods_code = text(first_value(row, "货品编码", "goods_no"))
            if not merchant_code and not goods_code:
                continue
            goods_code = goods_code or goods_code_from_merchant(merchant_code)
            if standard_size_merchant_code(merchant_code, goods_code):
                merchant_code = canonical_merchant_code(merchant_code)
            spec_name = text(first_value(row, "规格名称", "spec_name"))
            item = {
                "货品编码": goods_code,
                "货品名称": text(first_value(row, "货品名称", "goods_name")),
                "商家编码": merchant_code,
                "尺码": size_from_spec(spec_name, merchant_code),
                "规格名称": spec_name,
                "货品分类": text(first_value(row, "货品分类", "category_name", "class_name")),
                "成本价": number(first_value(row, "成本价", "ref_cost_price", "sku_default_purchase_price")),
                "批发价": number(first_value(row, "批发价", "批发报价", "wholesale_price")),
            }
            key = (item["货品编码"], item["商家编码"])
            if key in seen:
                continue
            seen.add(key)
            items.append(item)
    return items


def platform_item_from_row(row, fallback_code):
    merchant_code = text(row.get("商家编码") or row.get("merchant_code") or row.get("SKU货号") or row.get("商品SKU") or row.get("供应商SKU"))
    if not merchant_code:
        merchant_code = text(fallback_code)
    goods_code = goods_code_from_merchant(merchant_code)
    if standard_size_merchant_code(merchant_code, goods_code):
        merchant_code = canonical_merchant_code(merchant_code)
    spec_name = text(row.get("规格名称") or row.get("货品规格") or row.get("SKU属性") or row.get("规格"))
    return {
        "货品编码": goods_code,
        "货品名称": text(row.get("货品名称") or row.get("商品名称") or row.get("product_name")),
        "商家编码": merchant_code,
        "尺码": size_from_spec(spec_name, merchant_code),
        "规格名称": spec_name,
        "货品分类": "",
        "成本价": "",
        "批发价": "",
        "missing_erp": True,
    }


def build_clearance_catalog(erp_files):
    rows = []
    goods_codes = set()
    for item in load_erp_items(erp_files):
        category = item["货品分类"]
        if "清仓" not in category:
            continue
        rows.append({
            "货品编码": item["货品编码"],
            "货品名称": item["货品名称"],
            "商家编码": item["商家编码"],
            "尺码": item["尺码"],
            "规格名称": item["规格名称"],
            "清仓分类": category,
        })
        goods_codes.add(item["货品编码"])
    return {
        "rows": rows,
        "goods_codes": sorted(goods_codes),
        "merchant_codes": sorted({row["商家编码"] for row in rows if row["商家编码"]}),
        "summary": {"goods_count": len(goods_codes), "sku_count": len(rows)},
    }


def is_clearance_goods(catalog, goods_code):
    return text(goods_code) in set(catalog.get("goods_codes") or [])


def is_clearance_merchant(catalog, merchant_code):
    return text(merchant_code) in set(catalog.get("merchant_codes") or [])


def row_id(prefix, parts):
    raw = "|".join(text(part) for part in parts)
    return f"{prefix}-{hashlib.sha1(raw.encode('utf-8')).hexdigest()[:16]}"


class BargainStore:
    def __init__(self, path):
        self.path = Path(path)

    def load(self):
        if not self.path.exists():
            return {"batches": [], "ignored_low_prices": []}
        try:
            data = json.loads(self.path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            data = {}
        data.setdefault("batches", [])
        data.setdefault("ignored_low_prices", [])
        return data

    def save(self, data):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    def list_batches(self):
        return self.load()["batches"]

    def lookup_staging_rows(self, merchant_code, request_store, platform, owner, erp_files, clearance_catalog=None, platform_rows=None):
        erp_items = load_erp_items(erp_files)
        by_code = {item["商家编码"]: item for item in erp_items if item["商家编码"]}
        platform_rows = platform_rows or []
        wanted_code = text(merchant_code)
        source = by_code.get(wanted_code)
        missing_erp = False
        if not source:
            prefix = goods_code_from_merchant(wanted_code)
            fallback_rows = [
                row for row in platform_rows
                if standard_size_merchant_code(
                    row.get("商家编码") or row.get("merchant_code") or row.get("SKU货号") or row.get("商品SKU") or row.get("供应商SKU"),
                    prefix,
                )
            ]
            if not fallback_rows:
                raise ValueError("未在 ERP 商品基础信息里找到该商家编码")
            erp_items = [platform_item_from_row(row, wanted_code) for row in fallback_rows]
            by_code = {item["商家编码"]: item for item in erp_items if item["商家编码"]}
            source = by_code.get(wanted_code) or erp_items[0]
            missing_erp = True
        goods_code = source["货品编码"]
        variants = [item for item in erp_items if item["货品编码"] == goods_code]
        standard_variants = [item for item in variants if standard_size_merchant_code(item.get("商家编码"), goods_code)]
        if standard_variants:
            variants = standard_variants
        clearance_catalog = clearance_catalog or build_clearance_catalog(erp_files)
        variant_codes = {item["商家编码"] for item in variants if item.get("商家编码")}
        best_store = best_store_for_goods(goods_code, platform_rows, variant_codes)
        result = []
        for item in variants:
            metrics = platform_metrics_for_merchant(item["商家编码"], platform_rows)
            suggest_price = metrics["最低申报价"]
            wholesale = item["批发价"]
            cost = item["成本价"]
            clearance = is_clearance_goods(clearance_catalog, goods_code)
            risk_tags = []
            if missing_erp or item.get("missing_erp") or (not cost and not wholesale):
                risk_tags.append("ERP成本缺失")
            if missing_erp or item.get("missing_erp") or not wholesale:
                risk_tags.append("ERP批发价缺失")
            risk_level = "review" if risk_tags else price_risk_level(suggest_price, cost, clearance)
            result.append({
                "货品编码": goods_code,
                "货品名称": item["货品名称"],
                "议价申请店铺": text(request_store),
                "平台": text(platform),
                "卖得最好的店铺": best_store,
                "尺码": item["尺码"],
                "商家编码": item["商家编码"],
                "本次议价": suggest_price,
                "成本价": cost,
                "批发价": wholesale,
                "建议申报价/批发价占比": round(suggest_price / wholesale * 100, 2) if suggest_price and wholesale else "",
                "在线销售链接数": metrics["在线链接数"],
                "在售最低申报价": metrics["最低申报价"],
                "Temu 30天最高销量": metrics["Temu 30天最高销量"],
                "Shein 30天最高销量": metrics["Shein 30天最高销量"],
                "平台仓库备货库存": metrics["平台仓库备货库存"],
                "ERP 返回库存": metrics["ERP 返回库存"],
                "清仓款": clearance,
                "风险等级": risk_level,
                "风险标签": "、".join(risk_tags),
            })
        return sorted(result, key=lambda row: size_sort_key(row["尺码"]))

    def submit_batch(self, store, platform, owner, lines):
        data = self.load()
        timestamp = now_text()
        batch_id = row_id("bargain-batch", [store, platform, owner, timestamp, len(data["batches"])])
        prepared = []
        for index, line in enumerate(lines or [], start=1):
            if not text(line.get("本次议价")) or number(line.get("本次议价")) <= 0:
                raise ValueError("每个尺码都必须填写本次议价")
            line_id = row_id("bargain-line", [batch_id, line.get("商家编码"), index])
            item = dict(line)
            item.update({
                "id": line_id,
                "batch_id": batch_id,
                "store": text(store),
                "platform": text(platform or line.get("平台")),
                "owner": text(owner),
                "status": STATUS_PENDING_REVIEW,
                "version": int(line.get("version") or 1),
                "submitted_price": number(line.get("本次议价")),
                "submitted_at": timestamp,
                "reviewed_by": "",
                "reviewed_at": "",
                "review_remark": "",
                "parent_line_id": text(line.get("parent_line_id")),
            })
            prepared.append(item)
        batch = {
            "id": batch_id,
            "store": text(store),
            "platform": text(platform),
            "owner": text(owner),
            "status": STATUS_PENDING_REVIEW,
            "created_at": timestamp,
            "updated_at": timestamp,
            "lines": prepared,
        }
        data["batches"].append(batch)
        self.save(data)
        return batch

    def find_line(self, data, line_id):
        for batch in data["batches"]:
            for line in batch.get("lines", []):
                if line.get("id") == line_id:
                    return batch, line
        raise ValueError("议价记录不存在")

    def review_lines(self, batch_id, line_ids, decision, admin, remark=""):
        if decision != "通过" and not text(remark):
            raise ValueError("拒绝议价必须填写原因")
        data = self.load()
        wanted = set(line_ids or [])
        count = 0
        timestamp = now_text()
        for batch in data["batches"]:
            if batch.get("id") != batch_id:
                continue
            for line in batch.get("lines", []):
                if line.get("id") not in wanted:
                    continue
                line["status"] = STATUS_APPROVED if decision == "通过" else STATUS_REJECTED
                line["reviewed_by"] = text(admin)
                line["reviewed_at"] = timestamp
                line["review_remark"] = text(remark)
                count += 1
            batch["updated_at"] = timestamp
            statuses = {line.get("status") for line in batch.get("lines", [])}
            if statuses and statuses.issubset({STATUS_APPROVED, STATUS_REJECTED}):
                batch["status"] = "已审批"
        self.save(data)
        return {"count": count}

    def rework_lines(self, owner):
        rows = []
        for batch in self.load()["batches"]:
            for line in batch.get("lines", []):
                if line.get("owner") == owner and line.get("status") == STATUS_REJECTED:
                    rows.append(line)
        return rows

    def resubmit_line(self, line_id, new_price, owner):
        data = self.load()
        _batch, line = self.find_line(data, line_id)
        if line.get("status") != STATUS_REJECTED:
            raise ValueError("只有不通过的议价可以重新提交")
        new_line = dict(line)
        new_line["id"] = row_id("bargain-line", [line_id, new_price, now_text()])
        new_line["parent_line_id"] = line_id
        new_line["version"] = int(line.get("version") or 1) + 1
        new_line["本次议价"] = number(new_price)
        new_line["submitted_price"] = number(new_price)
        new_line["status"] = STATUS_PENDING_REVIEW
        new_line["reviewed_by"] = ""
        new_line["reviewed_at"] = ""
        new_line["review_remark"] = ""
        self.save(data)
        return self.submit_batch(line.get("store"), line.get("platform"), owner, [new_line])

    def history(self, filters=None):
        filters = filters or {}
        rows = []
        for batch in self.load()["batches"]:
            for line in batch.get("lines", []):
                if filters.get("merchant_code") and line.get("商家编码") != filters["merchant_code"]:
                    continue
                rows.append(line)
        return sorted(rows, key=lambda row: (row.get("商家编码", ""), int(row.get("version") or 0), row.get("submitted_at", "")))

    def approved_lines(self):
        return [line for line in self.history() if line.get("status") == STATUS_APPROVED]

    def low_price_trace(self, platform_rows, tolerance=0.05):
        data = self.load()
        ignores = data.get("ignored_low_prices", [])
        approved = {}
        for line in self.approved_lines():
            key = (line.get("platform"), line.get("store"), line.get("商家编码"))
            approved[key] = max(number(line.get("submitted_price")), number(approved.get(key, {}).get("submitted_price")))
        risks = []
        for row in platform_rows or []:
            platform = text(row.get("平台") or row.get("platform"))
            store = text(row.get("店铺") or row.get("store"))
            merchant_code = text(row.get("商家编码") or row.get("merchant_code"))
            price = number(row.get("申报价") or row.get("price"))
            if platform not in {"Temu", "Shein"}:
                continue
            if ignored_low_price(ignores, platform, store, merchant_code, price):
                continue
            key = (platform, store, merchant_code)
            approved_price = approved.get(key)
            if approved_price is None or price < number(approved_price) - number(tolerance):
                risks.append({
                    "id": row_id("low-price", [platform, store, merchant_code, price]),
                    "平台": platform,
                    "店铺": store,
                    "商家编码": merchant_code,
                    "当前申报价": price,
                    "历史审批价": approved_price if approved_price is not None else "",
                    "风险原因": "未匹配到已通过议价记录" if approved_price is None else "低于历史审批价",
                })
        data["last_low_price_risks"] = risks
        self.save(data)
        return risks

    def ignore_low_price(self, risk_ids, actor, remark=""):
        data = self.load()
        risks = {risk.get("id"): risk for risk in data.get("last_low_price_risks", [])}
        timestamp = now_text()
        count = 0
        for risk_id in risk_ids or []:
            risk = risks.get(risk_id)
            if not risk:
                continue
            data["ignored_low_prices"].append({
                "id": risk_id,
                "平台": risk["平台"],
                "店铺": risk["店铺"],
                "商家编码": risk["商家编码"],
                "忽略底价": risk["当前申报价"],
                "actor": text(actor),
                "remark": text(remark),
                "created_at": timestamp,
            })
            count += 1
        self.save(data)
        return {"count": count}


def size_sort_key(size):
    order = ["XXS", "XS", "S", "M", "L", "XL", "XXL", "2XL", "XXXL", "3XL"]
    text_size = text(size).upper()
    try:
        return (order.index(text_size), text_size)
    except ValueError:
        return (len(order), text_size)


def best_store_for_goods(goods_code, platform_rows, merchant_codes=None):
    totals = {}
    prefix = text(goods_code)
    allowed_codes = {canonical_merchant_code(code) for code in (merchant_codes or []) if text(code)}
    for row in platform_rows or []:
        merchant_code = text(row.get("商家编码") or row.get("merchant_code"))
        compare_code = canonical_merchant_code(merchant_code)
        if allowed_codes:
            if compare_code not in allowed_codes:
                continue
        elif prefix and not standard_size_merchant_code(merchant_code, prefix):
            continue
        store = text(row.get("店铺") or row.get("store"))
        if not store:
            continue
        totals[store] = totals.get(store, 0) + number(row.get("30天销量") or row.get("sales30"))
    if not totals:
        return ""
    top = max(totals.values())
    stores = sorted(store for store, value in totals.items() if value == top)
    return " / ".join(stores) + (" 并列" if len(stores) > 1 else "")


def platform_metrics_for_merchant(merchant_code, platform_rows):
    wanted_code = canonical_merchant_code(merchant_code)
    rows = [row for row in platform_rows or [] if canonical_merchant_code(row.get("商家编码") or row.get("merchant_code")) == wanted_code]
    prices = [number(row.get("申报价") or row.get("price")) for row in rows if number(row.get("申报价") or row.get("price")) > 0]
    temu_sales = [number(row.get("30天销量") or row.get("sales30")) for row in rows if text(row.get("平台") or row.get("platform")) == "Temu"]
    shein_sales = [number(row.get("30天销量") or row.get("sales30")) for row in rows if text(row.get("平台") or row.get("platform")) == "Shein"]
    return {
        "在线链接数": sum(int(number(row.get("在线链接数"), 1) or 1) for row in rows),
        "最低申报价": min(prices) if prices else 0,
        "Temu 30天最高销量": max(temu_sales) if temu_sales else 0,
        "Shein 30天最高销量": max(shein_sales) if shein_sales else 0,
        "平台仓库备货库存": sum(number(row.get("平台仓库备货库存") or row.get("平台库存")) for row in rows),
        "ERP 返回库存": sum(number(row.get("ERP 返回库存") or row.get("ERP库存")) for row in rows),
    }


def price_risk_level(price, cost, clearance):
    if price and cost and price < cost:
        return "orange" if clearance else "red"
    return "green"


def ignored_low_price(ignores, platform, store, merchant_code, price):
    for item in ignores or []:
        if item.get("平台") == platform and item.get("店铺") == store and item.get("商家编码") == merchant_code and price >= number(item.get("忽略底价")):
            return True
    return False
