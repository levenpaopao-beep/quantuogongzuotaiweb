from pathlib import Path

from openpyxl import load_workbook


def norm(value):
    if value is None:
        return ""
    return str(value).strip()


def number(value):
    text = norm(value).replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def store_from_filename(path):
    name = Path(path).stem
    for store in ["琪琪", "童话", "加加", "宝宝", "牛牛"]:
        if store in name:
            return store
    return ""


def find_header_row(ws, required_any):
    max_row = min(ws.max_row or 0, 40)
    max_col = min(ws.max_column or 0, 120)
    for row_index in range(1, max_row + 1):
        values = [norm(ws.cell(row_index, col).value) for col in range(1, max_col + 1)]
        if any(field in values for field in required_any):
            return row_index, values
    return None, []


def read_source_daily_average(path, platform):
    path = Path(path)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    result = {}
    for ws in wb.worksheets:
        header_row, headers = find_header_row(ws, ["7天销量", "近7天销量", "30天销量", "近30天销量"])
        if not header_row:
            continue
        header_map = {name: index + 1 for index, name in enumerate(headers) if name}
        store_col = header_map.get("店铺") or header_map.get("店铺名称") or header_map.get("店铺名")
        sales7_col = header_map.get("7天销量") or header_map.get("近7天销量")
        sales30_col = header_map.get("30天销量") or header_map.get("近30天销量")
        if not sales7_col and not sales30_col:
            continue
        fallback_store = store_from_filename(path) if platform == "Shein" else ""
        for row_index in range(header_row + 1, (ws.max_row or header_row) + 1):
            store = norm(ws.cell(row_index, store_col).value) if store_col else fallback_store
            if not store:
                continue
            if sales7_col:
                daily = number(ws.cell(row_index, sales7_col).value) / 7
            else:
                daily = number(ws.cell(row_index, sales30_col).value) / 30
            if daily:
                result[store] = result.get(store, 0.0) + daily
    return result


def aggregate_source_sales(files_by_platform):
    result = {}
    for platform, files in (files_by_platform or {}).items():
        platform_bucket = result.setdefault(platform, {})
        for path in files or []:
            for store, value in read_source_daily_average(path, platform).items():
                platform_bucket[store] = platform_bucket.get(store, 0.0) + value
    return result


def compare_sales(manual_entries, source_sales, percent_threshold=0.05, unit_threshold=20):
    rows = []
    for entry in manual_entries or []:
        if not entry.get("submitted"):
            continue
        platform = norm(entry.get("platform"))
        store = norm(entry.get("store"))
        manual = number(entry.get("sales"))
        imported = float((source_sales.get(platform) or {}).get(store) or 0)
        if not imported:
            continue
        diff = manual - imported
        diff_abs = abs(diff)
        percent = diff_abs / max(imported, 1)
        if diff_abs >= unit_threshold and percent >= percent_threshold:
            rows.append({
                "platform": platform,
                "store": store,
                "owner": norm(entry.get("owner")),
                "manual_sales": round(manual, 2),
                "imported_daily_avg": round(imported, 2),
                "diff": round(diff, 2),
                "diff_percent": round(percent * 100, 2),
                "level": "提醒",
            })
    rows.sort(key=lambda row: (-abs(row["diff"]), row["platform"], row["store"]))
    return rows
