import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook

from daily_ops_sales_compare import aggregate_source_sales, compare_sales, read_source_daily_average


class SalesCompareTest(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.path = Path(self.tmpdir.name) / "temu_sales.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["店铺", "SKC", "7天销量"])
        ws.append(["七弟", "A001", 490])
        ws.append(["七弟", "A002", 210])
        ws.append(["童话", "B001", 70])
        wb.save(self.path)

    def tearDown(self):
        self.tmpdir.cleanup()

    def test_read_source_daily_average_from_7_day_sales(self):
        result = read_source_daily_average(self.path, "Temu")

        self.assertEqual(result["七弟"], 100)
        self.assertEqual(result["童话"], 10)

    def test_compare_sales_returns_soft_alerts_above_threshold(self):
        source_sales = aggregate_source_sales({"Temu": [self.path]})
        rows = compare_sales([
            {"platform": "Temu", "store": "七弟", "owner": "小琴", "sales": 150, "submitted": True},
            {"platform": "Temu", "store": "童话", "owner": "胡娟", "sales": 12, "submitted": True},
        ], source_sales)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["store"], "七弟")
        self.assertEqual(rows[0]["manual_sales"], 150)
        self.assertEqual(rows[0]["imported_daily_avg"], 100)
        self.assertEqual(rows[0]["level"], "提醒")

    def test_unrecognized_workbook_is_ignored(self):
        blank = Path(self.tmpdir.name) / "blank.xlsx"
        wb = Workbook()
        wb.active.append(["只有一个字段"])
        wb.save(blank)

        self.assertEqual(read_source_daily_average(blank, "Temu"), {})

    def test_read_source_daily_average_from_dianxiaomi_dimension_a1_export(self):
        path = Path(self.tmpdir.name) / "shein_dimension_a1.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["店铺", "小秘商品SKU", "7天销量", "30天销量"])
        ws.append(["琪琪", "S001", 70, 300])
        ws.append(["琪琪", "S002", 140, 600])
        wb.save(path)

        with zipfile.ZipFile(path, "r") as src:
            parts = {name: src.read(name) for name in src.namelist()}
        sheet_name = "xl/worksheets/sheet1.xml"
        parts[sheet_name] = parts[sheet_name].replace(b'<dimension ref="A1:D3"/>', b'<dimension ref="A1"/>')
        with zipfile.ZipFile(path, "w") as dst:
            for name, payload in parts.items():
                dst.writestr(name, payload)

        result = read_source_daily_average(path, "Shein")

        self.assertEqual(result["琪琪"], 30)


if __name__ == "__main__":
    unittest.main()
