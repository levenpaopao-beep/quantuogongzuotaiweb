import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

import daily_ops_app
import daily_ops_desktop_adapter
import daily_ops_master_data as master_data
from daily_ops_sales import DailySalesStore


class MasterDataImportTest(unittest.TestCase):
    def build_owner_workbook(self, path):
        wb = Workbook()
        ws = wb.active
        ws.title = "temu"
        ws.append(["店铺", "主账号", "店名", "业务", "仓库", "仓库密码", "上架", "上架账号", "上架密码", "备注", "平台"])
        ws.append(["一弟", "15805182150", "Apetcircle", "小琴", "", "", "小琴", "", "", "", "temu"])
        ws.append(["二弟", "13003246055", "Fur Fit", "洁琳", "", "", "陆宝宝", "", "", "", "temu"])
        ws.append(["三弟", "13585986099", "Beautiful", "小琴", "", "", "小琴", "", "", "", "temu"])
        ws.append(["大额", "", "Ozon Goose", "小琴", "", "", "", "", "", "", "ozon"])
        wb.save(path)

    def build_sales_workbook(self, path):
        wb = Workbook()
        ws = wb.active
        ws.title = "总览"
        ws.append(["跳过"])
        month = wb.create_sheet("2606")
        month.append(["2026年", "汇总", "一弟", "二弟", "琪琪", "美美"])
        month.append([46174, 300, 100, 120, 80, 9])
        month.append([46175, 330, 110, 130, 90, 8])
        wb.save(path)

    def test_parse_owner_workbook_extracts_assignments_and_accounts(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "店铺负责人对应表.xlsx"
            self.build_owner_workbook(path)

            parsed = master_data.parse_owner_workbook(path)

        self.assertEqual(len(parsed["assignments"]), 4)
        self.assertEqual(parsed["assignments"][0]["platform"], "Temu")
        self.assertEqual(parsed["assignments"][0]["store"], "一弟")
        self.assertEqual(parsed["assignments"][0]["store_name"], "Apetcircle")
        self.assertEqual(parsed["assignments"][0]["owner"], "小琴")
        self.assertEqual(parsed["assignments"][3]["platform"], "Ozon")
        self.assertEqual(parsed["assignments"][3]["store"], "大鹅")
        self.assertEqual([account["owner"] for account in parsed["accounts"]], ["小琴", "洁琳"])

    def test_save_operator_accounts_hashes_password_and_returns_initial_password_once(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "operator_accounts.json"

            result = master_data.save_operator_accounts(
                path,
                [{"owner": "小琴", "username": "小琴"}, {"owner": "洁琳", "username": "洁琳"}],
                password_factory=lambda account: f"{account['username']}123456",
            )

            payload = json.loads(path.read_text(encoding="utf-8"))

        self.assertEqual(len(result["accounts"]), 2)
        self.assertEqual(result["initial_passwords"]["小琴"], "小琴123456")
        self.assertNotIn("小琴123456", json.dumps(payload, ensure_ascii=False))
        self.assertEqual(payload["accounts"][0]["username"], "小琴")
        self.assertTrue(payload["accounts"][0]["password_hash"])

    def test_create_operator_account_adds_manual_store_owner_login(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "operator_accounts.json"
            with patch.object(daily_ops_app, "OPERATOR_ACCOUNTS_FILE", path):
                result = daily_ops_app.create_operator_account("胡娟", "", "Hu123456")
                accounts = daily_ops_app.operator_accounts()["accounts"]

        self.assertEqual(result["username"], "胡娟")
        self.assertEqual(result["initial_password"], "Hu123456")
        self.assertEqual(accounts[0]["owner"], "胡娟")
        self.assertNotIn("password_hash", accounts[0])

    def test_query_erp_product_info_reads_latest_erp_product_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            erp_dir = Path(tmp)
            path = erp_dir / "erp产品基础信息表_接口同步_最新.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["商家编码（新）", "货品名称", "规格名称", "成本价", "批发价"])
            ws.append(["SKU-001", "宠物雨衣", "蓝色L", 10, 18])
            wb.save(path)

            with patch.object(daily_ops_app, "ERP_DIR", erp_dir):
                result = daily_ops_app.query_erp_product_info("SKU-001", 20)

        self.assertEqual(result["source_files"], [path.name])
        self.assertEqual(result["items"][0]["summary"]["商家编码（新）"], "SKU-001")
        self.assertEqual(result["items"][0]["summary"]["货品名称"], "宠物雨衣")

    def test_query_erp_product_info_defaults_to_latest_interface_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            erp_dir = Path(tmp)
            legacy = erp_dir / "20260606-全部数据-无图版.xlsx"
            latest = erp_dir / "erp产品基础信息表_接口同步_20260624_161745.xlsx"
            for path, code, name in [(legacy, "OLD-001", "旧表商品"), (latest, "NEW-001", "接口商品")]:
                wb = Workbook()
                ws = wb.active
                ws.append(["商家编码（新）", "货品名称", "规格名称", "成本价", "批发价"])
                ws.append([code, name, "蓝色L", 10, 18])
                wb.save(path)

            with patch.object(daily_ops_app, "ERP_DIR", erp_dir), \
                 patch.object(daily_ops_app, "manifest_paths", return_value=[legacy]):
                result = daily_ops_app.query_erp_product_info("", 20)

        self.assertEqual(result["source_files"], [latest.name])
        self.assertEqual(len(result["items"]), 1)
        self.assertEqual(result["items"][0]["summary"]["商家编码（新）"], "NEW-001")

    def test_parse_crossborder_sales_workbook_normalizes_month_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "跨境运营总表new.xlsx"
            self.build_sales_workbook(path)
            assignments = [
                {"platform": "Temu", "store": "一弟", "owner": "小琴"},
                {"platform": "Temu", "store": "二弟", "owner": "洁琳"},
                {"platform": "Shein", "store": "琪琪", "owner": "胡娟"},
                {"platform": "速卖通", "store": "大美", "owner": "胡娟"},
            ]

            rows = master_data.parse_crossborder_sales_workbook(path, assignments)

        self.assertEqual(len(rows), 8)
        self.assertEqual(rows[0]["date"], "2026-06-01")
        self.assertEqual(rows[0]["platform"], "Temu")
        self.assertEqual(rows[0]["store"], "一弟")
        self.assertEqual(rows[0]["sales"], 100)
        self.assertEqual(rows[2]["platform"], "Shein")
        self.assertEqual(rows[2]["store"], "琪琪")
        self.assertEqual(rows[3]["platform"], "速卖通")
        self.assertEqual(rows[3]["store"], "大美")
        self.assertEqual(rows[3]["owner"], "胡娟")

    def test_import_history_sales_does_not_overwrite_manual_records(self):
        with tempfile.TemporaryDirectory() as tmp:
            sales_path = Path(tmp) / "daily_sales.json"
            store = DailySalesStore(sales_path)
            assignments = [{"platform": "Temu", "store": "一弟", "owner": "小琴"}]
            store.submit(assignments, role="admin", user="管理员", day="2026-06-01", platform="Temu", store="一弟", sales="999")

            result = master_data.import_history_sales_records(
                sales_path,
                [{"date": "2026-06-01", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 100, "source_file": "源.xlsx", "source_sheet": "2606"}],
                actor="管理员",
            )
            payload = DailySalesStore(sales_path).load()

        self.assertEqual(result["created"], 0)
        self.assertEqual(result["skipped_existing"], 1)
        self.assertEqual(payload["records"][0]["sales"], 999)

    def test_query_sales_report_filters_and_summarizes(self):
        with tempfile.TemporaryDirectory() as tmp:
            sales_path = Path(tmp) / "daily_sales.json"
            master_data.import_history_sales_records(
                sales_path,
                [
                    {"date": "2026-06-01", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 100},
                    {"date": "2026-06-02", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 120},
                    {"date": "2026-06-01", "platform": "Shein", "store": "琪琪", "owner": "胡娟", "sales": 80},
                ],
                actor="管理员",
            )

            report = master_data.query_sales_report(sales_path, platform="Temu", store="一弟", date_from="2026-06-01", date_to="2026-06-30")

        self.assertEqual(report["summary"]["total_sales"], 220)
        self.assertEqual(report["summary"]["record_count"], 2)
        self.assertEqual(report["summary"]["daily_average"], 110)
        self.assertEqual(report["by_platform"][0]["platform"], "Temu")
        self.assertEqual(report["by_store"][0]["store"], "一弟")

    def test_query_sales_report_allowed_pairs_limits_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            sales_path = Path(tmp) / "daily_sales.json"
            master_data.import_history_sales_records(
                sales_path,
                [
                    {"date": "2026-06-01", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 100},
                    {"date": "2026-06-01", "platform": "Temu", "store": "二弟", "owner": "洁琳", "sales": 120},
                    {"date": "2026-06-01", "platform": "Shein", "store": "琪琪", "owner": "胡娟", "sales": 80},
                ],
                actor="管理员",
            )

            report = master_data.query_sales_report(
                sales_path,
                date_from="2026-06-01",
                date_to="2026-06-30",
                allowed_pairs={("Temu", "一弟"), ("Shein", "琪琪")},
            )

        self.assertEqual(report["summary"]["total_sales"], 180)
        self.assertEqual({row["store"] for row in report["rows"]}, {"一弟", "琪琪"})
        self.assertNotIn("二弟", {row["store"] for row in report["rows"]})

    def test_owner_sales_report_payload_without_store_is_scoped_to_owned_stores(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            sales_path = root / "daily_sales.json"
            owner_map = root / "store_owner_map.json"
            owner_map.write_text(json.dumps({
                "assignments": [
                    {"platform": "Temu", "store": "一弟", "owner": "小琴", "enabled": True, "daily_required": True},
                    {"platform": "Temu", "store": "二弟", "owner": "洁琳", "enabled": True, "daily_required": True},
                    {"platform": "Shein", "store": "琪琪", "owner": "小琴", "enabled": True, "daily_required": True},
                ],
            }, ensure_ascii=False), encoding="utf-8")
            master_data.import_history_sales_records(
                sales_path,
                [
                    {"date": "2026-06-01", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 100},
                    {"date": "2026-06-01", "platform": "Temu", "store": "二弟", "owner": "洁琳", "sales": 120},
                    {"date": "2026-06-01", "platform": "Shein", "store": "琪琪", "owner": "小琴", "sales": 80},
                ],
                actor="管理员",
            )

            with patch.object(daily_ops_app, "DAILY_SALES_FILE", sales_path), \
                 patch.object(daily_ops_app, "STORE_OWNER_MAP_FILE", owner_map):
                report = daily_ops_desktop_adapter.sales_report_payload({
                    "role": "owner",
                    "user": "小琴",
                    "date_from": "2026-06-01",
                    "date_to": "2026-06-30",
                })

        self.assertEqual(report["summary"]["total_sales"], 180)
        self.assertEqual({row["store"] for row in report["rows"]}, {"一弟", "琪琪"})
        self.assertNotIn("二弟", {row["store"] for row in report["rows"]})

    def test_owner_sales_report_payload_rejects_other_owner_store(self):
        with tempfile.TemporaryDirectory() as tmp:
            owner_map = Path(tmp) / "store_owner_map.json"
            owner_map.write_text(json.dumps({
                "assignments": [
                    {"platform": "Temu", "store": "一弟", "owner": "小琴", "enabled": True, "daily_required": True},
                    {"platform": "Temu", "store": "二弟", "owner": "洁琳", "enabled": True, "daily_required": True},
                ],
            }, ensure_ascii=False), encoding="utf-8")

            with patch.object(daily_ops_app, "STORE_OWNER_MAP_FILE", owner_map):
                with self.assertRaisesRegex(PermissionError, "店长只能查询自己负责店铺"):
                    daily_ops_desktop_adapter.sales_report_payload({
                        "role": "owner",
                        "user": "小琴",
                        "platform": "Temu",
                        "store": "二弟",
                    })

    def test_desktop_business_report_payload_forwards_source_and_range(self):
        with patch.object(daily_ops_desktop_adapter.app, "business_report", return_value={"summary": {}}) as mocked:
            report = daily_ops_desktop_adapter.business_report_payload({
                "role": "admin",
                "user": "管理员",
                "platform": "Temu",
                "store": "一弟",
                "grain": "month",
                "range_key": "14d",
                "source": "platform",
            })

        self.assertEqual(report, {"summary": {}})
        payload = mocked.call_args.args[0]
        self.assertEqual(payload["range_key"], "14d")
        self.assertEqual(payload["source"], "platform")
        self.assertEqual(payload["platform"], "Temu")
        self.assertEqual(payload["store"], "一弟")

    def test_business_report_uses_current_owner_assignment_and_comparisons(self):
        with tempfile.TemporaryDirectory() as tmp:
            sales_path = Path(tmp) / "daily_sales.json"
            assignments = [
                {"platform": "Ozon", "store": "大鹅", "owner": "小琴", "enabled": True, "daily_required": True},
                {"platform": "速卖通", "store": "大美", "owner": "胡娟", "enabled": True, "daily_required": True},
            ]
            master_data.import_history_sales_records(
                sales_path,
                [
                    {"date": "2026-06-01", "platform": "Ozon", "store": "大鹅", "owner": "旧负责人", "sales": 100},
                    {"date": "2025-06-01", "platform": "Ozon", "store": "大鹅", "owner": "旧负责人", "sales": 60},
                    {"date": "2026-06-01", "platform": "速卖通", "store": "大美", "owner": "胡娟", "sales": 50},
                ],
                actor="管理员",
            )

            report = master_data.business_report(
                sales_path,
                assignments=assignments,
                date_from="2026-06-01",
                date_to="2026-06-01",
                grain="day",
            )

        owner_rows = {row["name"]: row for row in report["dimensions"]["owner"]}
        platform_rows = {row["name"]: row for row in report["dimensions"]["platform"]}
        self.assertEqual(owner_rows["小琴"]["sales"], 100)
        self.assertNotIn("旧负责人", owner_rows)
        self.assertEqual(platform_rows["Ozon"]["compare_sales"], 60)
        self.assertEqual(platform_rows["Ozon"]["yoy_delta"], 40)
        self.assertEqual(report["summary"]["range"]["sales"], 150)

    def test_business_report_defaults_to_last_30_complete_days_and_reports_completeness(self):
        with tempfile.TemporaryDirectory() as tmp:
            sales_path = Path(tmp) / "daily_sales.json"
            assignments = [
                {"platform": "Temu", "store": "一弟", "owner": "小琴", "enabled": True, "daily_required": True},
                {"platform": "Shein", "store": "琪琪", "owner": "胡娟", "enabled": True, "daily_required": True},
            ]
            master_data.import_history_sales_records(
                sales_path,
                [
                    {"date": "2026-05-25", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 999},
                    {"date": "2026-05-26", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 10},
                    {"date": "2026-06-24", "platform": "Shein", "store": "琪琪", "owner": "胡娟", "sales": 20},
                    {"date": "2026-06-25", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 777},
                    {"date": "2025-05-26", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 5},
                    {"date": "2025-06-24", "platform": "Shein", "store": "琪琪", "owner": "胡娟", "sales": 6},
                ],
                actor="管理员",
            )

            report = master_data.business_report(sales_path, assignments=assignments, range_key="30d", anchor_date="2026-06-25")

        self.assertEqual(report["filters"]["date_from"], "2026-05-26")
        self.assertEqual(report["filters"]["date_to"], "2026-06-24")
        self.assertEqual(report["summary"]["range"]["sales"], 30)
        self.assertEqual(report["summary"]["range"]["compare_sales"], 11)
        self.assertEqual(report["summary"]["previous_range"]["compare_period"], {"date_from": "2026-04-26", "date_to": "2026-05-25"})
        self.assertEqual(report["summary"]["range"]["compare_period"], {"date_from": "2025-05-26", "date_to": "2025-06-24"})
        self.assertEqual(report["summary"]["completion"]["required"], 60)
        self.assertEqual(report["summary"]["completion"]["submitted"], 2)
        self.assertLess(report["summary"]["completion"]["rate"], 90)
        self.assertEqual(report["summary"]["completion"]["level"], "red")
        self.assertIn("店长填报销量", report["definitions"]["range"])
        self.assertIn("2026-05-26 至 2026-06-24", report["definitions"]["range"])

    def test_business_report_month_and_year_totals_end_yesterday(self):
        with tempfile.TemporaryDirectory() as tmp:
            sales_path = Path(tmp) / "daily_sales.json"
            master_data.import_history_sales_records(
                sales_path,
                [
                    {"date": "2026-01-01", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 100},
                    {"date": "2026-06-01", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 20},
                    {"date": "2026-06-24", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 30},
                    {"date": "2026-06-25", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 999},
                ],
                actor="管理员",
            )

            report = master_data.business_report(sales_path, range_key="7d", anchor_date="2026-06-25")

        self.assertEqual(report["summary"]["month"]["sales"], 50)
        self.assertEqual(report["summary"]["year"]["sales"], 150)
        self.assertIn("2026-06-01 至 2026-06-24", report["definitions"]["month"])
        self.assertIn("2026-01-01 至 2026-06-24", report["definitions"]["year"])

    def test_business_report_supports_90_day_complete_period(self):
        with tempfile.TemporaryDirectory() as tmp:
            sales_path = Path(tmp) / "daily_sales.json"
            master_data.import_history_sales_records(
                sales_path,
                [{"date": "2026-03-27", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 1}],
                actor="管理员",
            )

            report = master_data.business_report(sales_path, range_key="90d", anchor_date="2026-06-25")

        self.assertEqual(report["filters"]["date_from"], "2026-03-27")
        self.assertEqual(report["filters"]["date_to"], "2026-06-24")
        self.assertIn("最近90日销量", report["definitions"]["range"])

    def test_business_report_owner_scope_only_shows_owned_stores(self):
        with tempfile.TemporaryDirectory() as tmp:
            sales_path = Path(tmp) / "daily_sales.json"
            assignments = [
                {"platform": "Temu", "store": "一弟", "owner": "小琴", "enabled": True, "daily_required": True},
                {"platform": "Temu", "store": "二弟", "owner": "洁琳", "enabled": True, "daily_required": True},
            ]
            master_data.import_history_sales_records(
                sales_path,
                [
                    {"date": "2026-06-01", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 100},
                    {"date": "2026-06-01", "platform": "Temu", "store": "二弟", "owner": "洁琳", "sales": 200},
                ],
                actor="管理员",
            )

            report = master_data.business_report(
                sales_path,
                assignments=assignments,
                role="owner",
                user="小琴",
                date_from="2026-06-01",
                date_to="2026-06-01",
            )

        self.assertEqual(report["summary"]["range"]["sales"], 100)
        self.assertEqual({row["store"] for row in report["dimensions"]["store"]}, {"一弟"})

    def test_business_report_action_items_use_store_level_200_unit_decline_threshold(self):
        with tempfile.TemporaryDirectory() as tmp:
            sales_path = Path(tmp) / "daily_sales.json"
            assignments = [
                {"platform": "Temu", "store": "一弟", "owner": "小琴", "enabled": True, "daily_required": True},
                {"platform": "Temu", "store": "二弟", "owner": "洁琳", "enabled": True, "daily_required": True},
                {"platform": "Temu", "store": "三弟", "owner": "小琴", "enabled": True, "daily_required": True},
                {"platform": "Temu", "store": "四弟", "owner": "胡娟", "enabled": True, "daily_required": True},
            ]
            master_data.import_history_sales_records(
                sales_path,
                [
                    {"date": "2026-06-01", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 700},
                    {"date": "2026-05-31", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 1000},
                    {"date": "2025-06-01", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 1000},
                    {"date": "2026-06-01", "platform": "Temu", "store": "三弟", "owner": "小琴", "sales": 90},
                    {"date": "2026-05-31", "platform": "Temu", "store": "三弟", "owner": "小琴", "sales": 250},
                    {"date": "2026-05-20", "platform": "Temu", "store": "四弟", "owner": "胡娟", "sales": 80},
                ],
                actor="管理员",
            )

            report = master_data.business_report(
                sales_path,
                assignments=assignments,
                date_from="2026-06-01",
                date_to="2026-06-01",
                grain="day",
            )

        items = report["action_items"]
        keys = {(item["type"], item.get("store")) for item in items}
        self.assertIn(("missing_sales", "二弟"), keys)
        self.assertIn(("stale_store", "四弟"), keys)
        self.assertIn(("decline_previous", "一弟"), keys)
        self.assertIn(("decline_yoy", "一弟"), keys)
        self.assertNotIn(("decline_previous", "三弟"), keys)
        decline = next(item for item in items if item["type"] == "decline_previous" and item["store"] == "一弟")
        self.assertEqual(decline["delta"], -300)
        self.assertEqual(decline["action"], "trend")

    def test_platform_reference_report_exposes_source_coverage_and_unassigned_stores(self):
        with tempfile.TemporaryDirectory() as tmp:
            sales_path = Path(tmp) / "daily_sales.json"
            assignments = [
                {"platform": "Temu", "store": "一弟", "owner": "小琴", "enabled": True, "daily_required": True},
            ]

            report = master_data.business_report(
                sales_path,
                assignments=assignments,
                source="platform",
                range_key="30d",
                anchor_date="2026-06-25",
                rows_override=[
                    {"date": "2026-06-24", "platform": "Temu", "store": "一弟", "owner": "小琴", "sales": 300},
                    {"date": "2026-06-24", "platform": "Temu", "store": "二弟", "owner": "", "sales": 200},
                ],
            )

        self.assertEqual(report["source_summary"]["label"], "平台导入参考")
        self.assertEqual(report["source_summary"]["covered_stores"], 2)
        self.assertEqual(report["source_summary"]["unassigned_stores"], 1)
        self.assertIn(("unassigned_owner", "二弟"), {(item["type"], item.get("store")) for item in report["action_items"]})


if __name__ == "__main__":
    unittest.main()
