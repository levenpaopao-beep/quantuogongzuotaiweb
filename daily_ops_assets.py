import calendar
import sqlite3
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


def norm(value):
    return "" if value is None else str(value).strip()


def num(value, default=0):
    if value is None or value == "":
        return default
    try:
        return int(float(str(value).replace(",", "").strip()))
    except ValueError:
        return default


def date_text(value=""):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = norm(value)
    if not text:
        return datetime.now().date().isoformat()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text[:10] if fmt != "%Y%m%d" else text[:8], fmt).date().isoformat()
        except ValueError:
            continue
    return text[:10]


def shift_month(day_text, months):
    current = datetime.strptime(date_text(day_text), "%Y-%m-%d").date()
    month_index = current.month - 1 + months
    year = current.year + month_index // 12
    month = month_index % 12 + 1
    day = min(current.day, calendar.monthrange(year, month)[1])
    return date(year, month, day).isoformat()


def metric_label(metric):
    return {
        "temu_hot_skc": "Temu 爆旺款 SKC",
        "shein_hot_skc": "Shein 爆款 SKC",
    }.get(metric, metric)


class AssetSnapshotStore:
    def __init__(self, db_path):
        self.db_path = Path(db_path)

    def connect(self):
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        con = sqlite3.connect(self.db_path)
        self.ensure_schema(con)
        return con

    def ensure_schema(self, con):
        exists = con.execute(
            "select name from sqlite_master where type='table' and name='asset_metrics'"
        ).fetchone()
        if not exists:
            self.create_table(con)
            return
        columns = [row[1] for row in con.execute("pragma table_info(asset_metrics)").fetchall()]
        if "scope_type" in columns and "scope_key" in columns:
            return
        con.execute(
            """
            create table asset_metrics_v2 (
              metric_date text not null,
              metric_key text not null,
              scope_type text not null default 'all',
              scope_key text not null default '',
              metric_label text not null,
              value integer not null default 0,
              source_file text not null default '',
              updated_at text not null default '',
              primary key(metric_date, metric_key, scope_type, scope_key)
            )
            """
        )
        con.execute(
            """
            insert or replace into asset_metrics_v2(metric_date, metric_key, scope_type, scope_key, metric_label, value, source_file, updated_at)
            select metric_date, metric_key, 'all', '', metric_label, value, source_file, updated_at
            from asset_metrics
            """
        )
        con.execute("drop table asset_metrics")
        con.execute("alter table asset_metrics_v2 rename to asset_metrics")

    def create_table(self, con):
        con.execute(
            """
            create table if not exists asset_metrics (
              metric_date text not null,
              metric_key text not null,
              scope_type text not null default 'all',
              scope_key text not null default '',
              metric_label text not null,
              value integer not null default 0,
              source_file text not null default '',
              updated_at text not null default '',
              primary key(metric_date, metric_key, scope_type, scope_key)
            )
            """
        )

    def scope_values(self, owner=""):
        owner_name = norm(owner)
        return ("owner", owner_name) if owner_name else ("all", "")

    def upsert_metric(self, metric_date, metric_key, value, source_file="", updated_at="", owner=""):
        day = date_text(metric_date)
        key = norm(metric_key)
        if not key:
            raise ValueError("指标不能为空")
        scope_type, scope_key = self.scope_values(owner)
        timestamp = norm(updated_at) or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with self.connect() as con:
            con.execute(
                """
                insert into asset_metrics(metric_date, metric_key, scope_type, scope_key, metric_label, value, source_file, updated_at)
                values(?, ?, ?, ?, ?, ?, ?, ?)
                on conflict(metric_date, metric_key, scope_type, scope_key) do update set
                  metric_label=excluded.metric_label,
                  value=excluded.value,
                  source_file=excluded.source_file,
                  updated_at=excluded.updated_at
                """,
                (day, key, scope_type, scope_key, metric_label(key), num(value), norm(source_file), timestamp),
            )
        return {"date": day, "metric": key, "value": num(value), "scope_type": scope_type, "scope_key": scope_key}

    def value_for(self, metric_key, metric_date, owner=""):
        scope_type, scope_key = self.scope_values(owner)
        with self.connect() as con:
            row = con.execute(
                """
                select metric_date, metric_key, scope_type, scope_key, metric_label, value, source_file, updated_at
                from asset_metrics
                where metric_key=? and metric_date=? and scope_type=? and scope_key=?
                """,
                (metric_key, date_text(metric_date), scope_type, scope_key),
            ).fetchone()
        if not row:
            return None
        return {
            "date": row[0],
            "metric": row[1],
            "scope_type": row[2],
            "scope_key": row[3],
            "label": row[4],
            "value": row[5],
            "source_file": row[6],
            "updated_at": row[7],
        }

    def latest_for(self, metric_key, anchor_date="", owner=""):
        anchor = date_text(anchor_date)
        scope_type, scope_key = self.scope_values(owner)
        with self.connect() as con:
            row = con.execute(
                """
                select metric_date, metric_key, scope_type, scope_key, metric_label, value, source_file, updated_at
                from asset_metrics
                where metric_key=? and metric_date<=? and scope_type=? and scope_key=?
                order by metric_date desc
                limit 1
                """,
                (metric_key, anchor, scope_type, scope_key),
            ).fetchone()
        if not row:
            return None
        return {
            "date": row[0],
            "metric": row[1],
            "scope_type": row[2],
            "scope_key": row[3],
            "label": row[4],
            "value": row[5],
            "source_file": row[6],
            "updated_at": row[7],
        }

    def compare(self, metric_key, base, target_date, owner=""):
        target = self.value_for(metric_key, target_date, owner=owner)
        value = target["value"] if target else None
        delta = None if value is None or base is None else base - value
        return {"date": date_text(target_date), "value": value, "delta": delta}

    def overview(self, metric_keys, anchor_date="", owner=""):
        result = {}
        for key in metric_keys:
            latest = self.latest_for(key, anchor_date, owner=owner)
            latest_date = latest["date"] if latest else date_text(anchor_date)
            current_value = latest["value"] if latest else None
            result[key] = {
                "metric": key,
                "label": metric_label(key),
                "scope_type": latest["scope_type"] if latest else self.scope_values(owner)[0],
                "scope_key": latest["scope_key"] if latest else self.scope_values(owner)[1],
                "date": latest_date,
                "value": current_value,
                "source_file": latest["source_file"] if latest else "",
                "updated_at": latest["updated_at"] if latest else "",
                "previous_month": self.compare(key, current_value, shift_month(latest_date, -1), owner=owner),
                "previous_year": self.compare(key, current_value, shift_month(latest_date, -12), owner=owner),
            }
        return result

    def rows(self):
        with self.connect() as con:
            rows = con.execute(
                """
                select metric_date, metric_key, scope_type, scope_key, metric_label, value, source_file, updated_at
                from asset_metrics
                order by metric_date, metric_key, scope_type, scope_key
                """
            ).fetchall()
        return [
            {
                "date": row[0],
                "metric": row[1],
                "scope_type": row[2],
                "scope_key": row[3],
                "label": row[4],
                "value": row[5],
                "source_file": row[6],
                "updated_at": row[7],
            }
            for row in rows
        ]

    def export_archive(self, output_path):
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "重要资产快照"
        ws.append(["日期", "指标", "数量", "来源文件", "更新时间", "指标名称", "范围类型", "范围值"])
        rows = self.rows()
        for row in rows:
            ws.append([row["date"], row["metric"], row["value"], row["source_file"], row["updated_at"], row["label"], row["scope_type"], row["scope_key"]])
        wb.save(output)
        return {"path": str(output), "file": output.name, "rows": len(rows)}

    def import_archive(self, source_path):
        source = Path(source_path)
        if not source.exists():
            raise FileNotFoundError("重要资产存档不存在")
        wb = load_workbook(source, read_only=True, data_only=True)
        count = 0
        try:
            ws = wb.active
            headers = [norm(cell.value) for cell in next(ws.iter_rows(max_row=1))]
            index = {header: idx for idx, header in enumerate(headers)}
            for row in ws.iter_rows(min_row=2, values_only=True):
                day = row[index.get("日期", -1)] if "日期" in index else ""
                metric = row[index.get("指标", -1)] if "指标" in index else ""
                value = row[index.get("数量", -1)] if "数量" in index else ""
                source_file = row[index.get("来源文件", -1)] if "来源文件" in index else source.name
                updated_at = row[index.get("更新时间", -1)] if "更新时间" in index else ""
                scope_type = norm(row[index.get("范围类型", -1)] if "范围类型" in index else "")
                scope_key = norm(row[index.get("范围值", -1)] if "范围值" in index else "")
                if not norm(day) or not norm(metric):
                    continue
                owner = scope_key if scope_type == "owner" else ""
                self.upsert_metric(day, metric, value, source_file, updated_at, owner=owner)
                count += 1
        finally:
            wb.close()
        return {"path": str(source), "file": source.name, "rows": count}
