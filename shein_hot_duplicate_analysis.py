import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path.cwd()
SOURCE_FILES = {
    "牛牛": ROOT / "shein数据源表" / "20260601-shein牛牛.xlsx",
    "宝宝": ROOT / "shein数据源表" / "20260601-shein宝宝.xlsx",
    "加加": ROOT / "shein数据源表" / "20260601-shein加加.xlsx",
    "琪琪": ROOT / "shein数据源表" / "20260601-shein琪琪.xlsx",
    "童话": ROOT / "shein数据源表" / "20260601-shein童话.xlsx",
}
ERP_FILES = sorted((ROOT / "erp数据源").glob("erp产品基础信息表*.xlsx"))
OUT_DIR = ROOT / "outputs" / "shein_hot_audit_20260602"
OUT_JSON = OUT_DIR / "shein_hot_duplicate_data.json"

SIZE_RE = re.compile(r"^(33\d+)-(XS|S|M|L|XL)$", re.I)
STORE_OWNER = {"琪琪": "胡娟", "童话": "胡娟", "加加": "洁琳", "宝宝": "洁琳", "牛牛": ""}


def num(value):
    if value is None or value == "":
        return 0.0
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def text(value):
    return "" if value is None else str(value).strip()


def norm_sku(value):
    raw = text(value)
    base = raw.split("@", 1)[0].strip().upper()
    m = SIZE_RE.match(base)
    if not m:
        return None, None, None
    return base, m.group(1), m.group(2).upper()


def load_erp():
    erp = {}
    for path in ERP_FILES:
        wb = load_workbook(path, read_only=False, data_only=True)
        ws = wb.worksheets[0]
        headers = [text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        idx = {h: i + 1 for i, h in enumerate(headers)}
        code_col = idx.get("商家编码（新）") or idx.get("商家编码")
        if not code_col:
            wb.close()
            continue
        for r in range(2, ws.max_row + 1):
            code = text(ws.cell(r, code_col).value).upper()
            if not code or code in erp:
                continue
            erp[code] = {
                "匹配到编码": code,
                "ERP货品名称": text(ws.cell(r, idx.get("货品名称", 0)).value) if idx.get("货品名称") else "",
                "规格名称": text(ws.cell(r, idx.get("规格名称", 0)).value) if idx.get("规格名称") else "",
                "成本价": num(ws.cell(r, idx.get("成本价", 0)).value) if idx.get("成本价") else 0,
                "批发报价": num(ws.cell(r, idx.get("批发报价", 0)).value) if idx.get("批发报价") else 0,
                "可用库存": num(ws.cell(r, idx.get("可用库存", 0)).value) if idx.get("可用库存") else 0,
            }
        wb.close()
    return erp


def load_records():
    erp = load_erp()
    rows = []
    for store, path in SOURCE_FILES.items():
        if not path.exists():
            continue
        wb = load_workbook(path, read_only=False, data_only=True)
        ws = wb.worksheets[0]
        headers = [text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        idx = {h: i + 1 for i, h in enumerate(headers)}
        price_col = idx.get("供货价") or idx.get("采购价")
        price_field = "供货价" if idx.get("供货价") else "采购价"
        for r in range(2, ws.max_row + 1):
            sku, style, size = norm_sku(ws.cell(r, idx.get("商家SKU", 0)).value if idx.get("商家SKU") else None)
            if not sku:
                continue
            erp_hit = erp.get(sku)
            if not erp_hit:
                continue
            rows.append(
                {
                    "店铺": store,
                    "负责人": STORE_OWNER.get(store, ""),
                    "商品名称": text(ws.cell(r, idx.get("供货方号", 0)).value) if idx.get("供货方号") else "",
                    "SPU": text(ws.cell(r, idx.get("SPU", 0)).value) if idx.get("SPU") else "",
                    "SKC": text(ws.cell(r, idx.get("SKC", 0)).value) if idx.get("SKC") else "",
                    "SKU货号": text(ws.cell(r, idx.get("商家SKU", 0)).value) if idx.get("商家SKU") else "",
                    "商家编码": sku,
                    "款号": style,
                    "尺码": size,
                    "申报价": num(ws.cell(r, price_col).value) if price_col else 0,
                    "申报价字段": price_field,
                    "7天销量": num(ws.cell(r, idx.get("近7天销量", 0)).value) if idx.get("近7天销量") else 0,
                    "30天销量": num(ws.cell(r, idx.get("近30天销量", 0)).value) if idx.get("近30天销量") else 0,
                    "平台仓库可用": num(ws.cell(r, idx.get("SHEIN仓库存", 0)).value) if idx.get("SHEIN仓库存") else 0,
                    "上架天数": num(ws.cell(r, idx.get("上架天数", 0)).value) if idx.get("上架天数") else 0,
                    "ERP货品名称": erp_hit["ERP货品名称"],
                    "规格名称": erp_hit["规格名称"],
                    "匹配方式": "ERP基础表",
                    "成本价": erp_hit["成本价"],
                    "批发报价": erp_hit["批发报价"],
                }
            )
        wb.close()
    return rows


def aggregate(records):
    skc = {}
    skc_styles = defaultdict(set)
    combo_rows = defaultdict(list)
    style_combos = defaultdict(set)

    for r in records:
        skc_key = (r["店铺"], r["SKC"])
        if skc_key not in skc:
            skc[skc_key] = {
                "店铺": r["店铺"],
                "负责人": r["负责人"],
                "SKC": r["SKC"],
                "商品名称": r["商品名称"],
                "7天销量": 0.0,
                "30天销量": 0.0,
                "平台仓库可用": 0.0,
                "上架天数": r["上架天数"],
                "SKU数": 0,
            }
        item = skc[skc_key]
        item["7天销量"] += r["7天销量"]
        item["30天销量"] += r["30天销量"]
        item["平台仓库可用"] += r["平台仓库可用"]
        item["上架天数"] = min(item["上架天数"], r["上架天数"]) if item["上架天数"] else r["上架天数"]
        item["SKU数"] += 1
        skc_styles[skc_key].add(r["款号"])

        combo_key = (r["店铺"], r["SKC"], r["款号"])
        combo_rows[combo_key].append(r)
        style_combos[r["款号"]].add(combo_key)

    hot_skc = {}
    for key, item in skc.items():
        is_new = item["上架天数"] < 30
        avg7 = item["7天销量"] / 7 if is_new else 0
        avg30 = item["30天销量"] / 30 if not is_new else 0
        hot = (is_new and avg7 >= 10) or ((not is_new) and avg30 > 20)
        if hot:
            hot_skc[key] = {**item, "新品/老品": "新品" if is_new else "老品", "7天日均": avg7, "30天日均": avg30}

    combo_summary = {}
    for key, rows in combo_rows.items():
        prices = [r["申报价"] for r in rows if r["申报价"]]
        combo_summary[key] = {
            "店铺": key[0],
            "SKC": key[1],
            "款号": key[2],
            "负责人": rows[0]["负责人"],
            "商品名称": rows[0]["商品名称"],
            "ERP货品名称": rows[0]["ERP货品名称"],
            "规格名称": " / ".join(sorted({r["规格名称"] for r in rows if r["规格名称"]})[:5]),
            "7天销量": sum(r["7天销量"] for r in rows),
            "30天销量": sum(r["30天销量"] for r in rows),
            "平台仓库可用": sum(r["平台仓库可用"] for r in rows),
            "上架天数": min((r["上架天数"] for r in rows if r["上架天数"]), default=0),
            "最低申报价": min(prices) if prices else 0,
            "最高申报价": max(prices) if prices else 0,
            "平均申报价": sum(prices) / len(prices) if prices else 0,
            "SKU数": len(rows),
        }
    return skc, hot_skc, combo_rows, combo_summary, style_combos, skc_styles


def build_conflicts(skc, hot_skc, combo_rows, combo_summary, style_combos, skc_styles):
    conflicts = []
    seen = set()
    for hot_key, hot in hot_skc.items():
        for style in sorted(skc_styles[hot_key]):
            hot_combo = (hot_key[0], hot_key[1], style)
            if len(style_combos[style]) < 2:
                continue
            hot_rows_by_sku = {r["商家编码"]: r for r in combo_rows[hot_combo]}
            for dup_combo in sorted(style_combos[style]):
                if dup_combo == hot_combo:
                    continue
                dup_skc_key = (dup_combo[0], dup_combo[1])
                dup = combo_summary[dup_combo]
                dup_rows_by_sku = {r["商家编码"]: r for r in combo_rows[dup_combo]}
                shared_skus = sorted(set(hot_rows_by_sku) & set(dup_rows_by_sku))
                if not shared_skus:
                    shared_skus = [sorted(hot_rows_by_sku)[0]]
                dup_is_hot = dup_skc_key in hot_skc
                for sku in shared_skus:
                    hot_row = hot_rows_by_sku.get(sku) or next(iter(hot_rows_by_sku.values()))
                    dup_row = dup_rows_by_sku.get(sku) or min(combo_rows[dup_combo], key=lambda r: r["申报价"] or 999999)
                    hot_price = hot_row["申报价"] or combo_summary[hot_combo]["最低申报价"]
                    dup_price = dup_row["申报价"] or dup["最低申报价"]
                    lower = dup_price < hot_price if hot_price and dup_price else False
                    conflict_type = "爆旺款互相冲突" if dup_is_hot else "平销冲突"
                    stock = skc.get(dup_skc_key, {}).get("平台仓库可用", dup["平台仓库可用"])
                    if lower:
                        advice = "!! 立刻下架"
                    elif 30 <= stock <= 100:
                        advice = "备货30-100件，先消化库存，后续降至30件以下要求下架"
                    elif dup_is_hot:
                        advice = "多店铺均为爆旺款，按爆旺优先级保留高优先级款，另一款卖完下架"
                    else:
                        advice = "可以卖完下架"
                    row_key = (hot_key, dup_combo, sku)
                    if row_key in seen:
                        continue
                    seen.add(row_key)
                    conflicts.append(
                        {
                            "爆款店铺": hot["店铺"],
                            "爆旺款SKC": hot["SKC"],
                            "商家编码": sku,
                            "货品名称（ERP货品名称）": hot_row["ERP货品名称"],
                            "规格名称（ERP规格名称）": hot_row["规格名称"],
                            "重复店铺": dup["店铺"],
                            "重复款SKC": dup["SKC"],
                            "重复铺货供货价or采购价是否低于爆旺款": "是" if lower else "否",
                            "爆旺款申报价": hot_price,
                            "重复款申报价（或者供货价）": dup_price,
                            "爆旺款30天销量": hot["30天销量"],
                            "重复款30天销量": skc.get(dup_skc_key, {}).get("30天销量", dup["30天销量"]),
                            "价格差（重复-爆旺）": dup_price - hot_price if hot_price or dup_price else 0,
                            "重复款负责人": dup["负责人"],
                            "重复款平台备货仓库件数": stock,
                            "冲突类型": conflict_type,
                            "处理意见": advice,
                            "爆旺款7天销量": hot["7天销量"],
                            "爆旺款上架天数": hot["上架天数"],
                            "重复款上架天数": skc.get(dup_skc_key, {}).get("上架天数", dup["上架天数"]),
                        }
                    )
    conflicts.sort(key=lambda x: (-x["爆旺款30天销量"], x["爆款店铺"], x["爆旺款SKC"], x["重复店铺"], x["商家编码"]))
    return conflicts


def make_sheets(records, skc, hot_skc, conflicts):
    stores = sorted({r["店铺"] for r in records})
    summary = []
    for store in stores:
        as_hot = [c for c in conflicts if c["爆款店铺"] == store]
        as_dup = [c for c in conflicts if c["重复店铺"] == store]
        store_hot = [v for k, v in hot_skc.items() if k[0] == store]
        summary.append(
            [
                store,
                len(store_hot),
                len({c["爆旺款SKC"] for c in as_hot}),
                len(as_hot),
                len({c["重复款SKC"] for c in as_dup}),
                len(as_dup),
                sum(1 for c in as_dup if c["重复铺货供货价or采购价是否低于爆旺款"] == "是"),
                sum(1 for c in as_dup if c["冲突类型"] == "平销冲突"),
                sum(1 for c in as_dup if c["冲突类型"] == "爆旺款互相冲突"),
                sum(c["重复款平台备货仓库件数"] for c in as_dup),
            ]
        )

    operation_map = {}
    for c in conflicts:
        reason = "冲击爆款" if c["冲突类型"] == "平销冲突" else "爆旺款互相冲突"
        key = (c["重复店铺"], c["重复款SKC"], c["商家编码"], reason)
        if key not in operation_map or c["处理意见"].startswith("!!"):
            operation_map[key] = [
                c["商家编码"],
                c["货品名称（ERP货品名称）"],
                c["重复款SKC"],
                c["重复店铺"],
                reason,
                c["处理意见"],
            ]
    operation = sorted(operation_map.values(), key=lambda r: (r[3], r[2], r[0]))

    detail_headers = [
        "爆款店铺",
        "爆旺款SKC",
        "商家编码",
        "货品名称（ERP货品名称）",
        "规格名称（ERP规格名称）",
        "重复店铺",
        "重复款SKC",
        "重复铺货供货价or采购价是否低于爆旺款",
        "爆旺款申报价",
        "重复款申报价（或者供货价）",
        "爆旺款30天销量",
        "重复款30天销量",
        "价格差（重复-爆旺）",
        "重复款负责人",
        "重复款平台备货仓库件数",
        "冲突类型",
        "处理意见",
    ]
    detail_rows = [[c[h] for h in detail_headers] for c in conflicts]
    flat_rows = [[c[h] for h in detail_headers] for c in conflicts if c["冲突类型"] == "平销冲突"]
    hot_rows = [[c[h] for h in detail_headers] for c in conflicts if c["冲突类型"] == "爆旺款互相冲突"]

    hot_list = []
    for hot in sorted(hot_skc.values(), key=lambda x: -x["30天销量"]):
        hot_list.append(
            [
                hot["店铺"],
                hot["负责人"],
                hot["SKC"],
                hot["商品名称"],
                hot["新品/老品"],
                hot["上架天数"],
                hot["7天销量"],
                hot["7天日均"],
                hot["30天销量"],
                hot["30天日均"],
                hot["平台仓库可用"],
                hot["SKU数"],
            ]
        )

    sheets = [
        {
            "name": "总览表",
            "title": "SHEIN爆旺款重复铺货核查总览",
            "headers": ["店铺", "爆旺SKC数", "作为爆款涉及SKC数", "作为爆款冲突行数", "作为重复款涉及SKC数", "作为重复款冲突行数", "重复款低价行数", "平销冲突行数", "爆旺互冲行数", "重复款平台备货合计"],
            "rows": summary,
        },
        {
            "name": "具体店铺操作表",
            "title": "按重复店铺归集的操作清单",
            "headers": ["商家编码", "货品名称", "skc", "所属店铺", "下架原因", "处理意见"],
            "rows": operation,
        },
        {"name": "平销冲突明细表", "title": "爆旺款 vs 平销重复铺货明细", "headers": detail_headers, "rows": flat_rows},
        {"name": "爆旺款互相冲突明细表", "title": "爆旺款之间重复铺货明细", "headers": detail_headers, "rows": hot_rows},
        {
            "name": "爆旺款清单",
            "title": "按规则识别出的爆旺SKC",
            "headers": ["店铺", "负责人", "SKC", "商品名称", "新品/老品", "上架天数", "7天销量", "7天日均", "30天销量", "30天日均", "平台仓库可用", "SKU数"],
            "rows": hot_list,
        },
        {
            "name": "说明",
            "title": "口径说明",
            "headers": ["项目", "说明"],
            "rows": [
                ["数据源", "使用 20260601-shein牛牛.xlsx、20260601-shein宝宝.xlsx、20260601-shein加加.xlsx、20260601-shein琪琪.xlsx、20260601-shein童话.xlsx 五个店铺销售总览表直接取数。"],
                ["SKU口径", "先取 @ 前编码，只保留 33 开头且以 -XS/-S/-M/-L/-XL 结尾的商家 SKU；未匹配 ERP 信息的 SKU 不计入销量。"],
                ["爆旺款定义", "以店铺+SKC 为维度：新品上架天数小于30天且7天销量日均>=10；老品上架天数大于等于30天且30天销量日均>20。"],
                ["重复铺货口径", "同一商家款号（商家编码去掉尺码后缀）出现在多个店铺/SKC 时，视为重复铺货候选。"],
                ["低价标红", "重复款申报价/供货价低于爆旺款对应 SKU 申报价时标记为“是”，并在明细表中红底提示。"],
                ["价格差", "价格差=重复款申报价（或者供货价）-爆旺款申报价，负数代表重复款更低。"],
            ],
        },
    ]
    return sheets


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    records = load_records()
    skc, hot_skc, combo_rows, combo_summary, style_combos, skc_styles = aggregate(records)
    conflicts = build_conflicts(skc, hot_skc, combo_rows, combo_summary, style_combos, skc_styles)
    payload = {
        "source": "、".join(path.name for path in SOURCE_FILES.values()),
        "record_count": len(records),
        "hot_skc_count": len(hot_skc),
        "conflict_count": len(conflicts),
        "sheets": make_sheets(records, skc, hot_skc, conflicts),
    }
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: payload[k] for k in ["record_count", "hot_skc_count", "conflict_count"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
