import tempfile
import json
import os
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

import daily_ops_erp
import daily_ops_app


class ErpSyncTest(unittest.TestCase):
    def test_wangdian_sign_matches_official_example(self):
        params = {
            "appkey": "test2-xx",
            "page_no": "0",
            "end_time": "2016-08-01 13:00:00",
            "start_time": "2016-08-01 12:00:00",
            "page_size": "40",
            "sid": "test2",
            "timestamp": "1470042310",
        }
        self.assertEqual(daily_ops_erp.wangdian_sign(params, "12345"), "ad4e6fe037ea6e3ba4768317be9d1309")

    def test_missing_credentials_blocks_sync_without_creating_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            result = daily_ops_erp.manual_sync({"base_url": "https://api.wangdian.cn/openapi2"}, Path(tmp))
            self.assertEqual(result["status"], "blocked")
            self.assertEqual(list(Path(tmp).glob("*.xlsx")), [])

    def test_api_base_url_uses_web_open_platform_environment_addresses(self):
        self.assertEqual(
            daily_ops_erp.api_base_url({"environment": "test", "base_url": "https://api.wangdian.cn/openapi2"}),
            "https://openapi.ali.huice.cc/openapi",
        )
        self.assertEqual(
            daily_ops_erp.api_base_url({"environment": "prod", "base_url": "https://sandbox.wangdian.cn/openapi2"}),
            "https://openapi.huice.com/openapi",
        )

    def test_load_rules_migrates_platform_goods_endpoint_to_goods_archive_endpoint(self):
        with tempfile.TemporaryDirectory() as tmp:
            rules_file = Path(tmp) / "report_rules.json"
            rules_file.write_text("""{
              "erp_api": {
                "product_endpoint": "vip_api_goods_query.php",
                "stock_endpoint": "stock_query.php"
              }
            }""", encoding="utf-8")
            with patch.object(daily_ops_app, "RULES_FILE", rules_file):
                rules = daily_ops_app.load_rules()

        self.assertEqual(rules["erp_api"]["product_endpoint"], "goods_query.php")
        self.assertEqual(rules["erp_api"]["stock_endpoint"], daily_ops_erp.STOCK_ENDPOINT)
        self.assertEqual(rules["erp_api"]["warehouse_no"], "3")
        self.assertEqual(rules["erp_api"]["warehouse_name"], "宠物圈仓")
        self.assertEqual(rules["erp_api"]["sync_days"], 30)
        self.assertEqual(rules["erp_api"]["page_size"], 500)
        self.assertEqual(rules["erp_api"]["stock_limit"], 10000)

    def test_load_rules_fills_blank_erp_business_defaults_without_touching_credentials(self):
        with tempfile.TemporaryDirectory() as tmp:
            rules_file = Path(tmp) / "report_rules.json"
            local_file = Path(tmp) / "erp_api.local.json"
            local_file.write_text(json.dumps({
                "app_key": "local-app",
                "app_secret": "local-secret",
                "sid": "local-sid",
            }), encoding="utf-8")
            rules_file.write_text(json.dumps({
                "erp_api": {
                    "sync_days": "",
                    "page_size": "",
                    "stock_limit": "",
                    "warehouse_no": "",
                    "warehouse_name": "",
                    "app_key": "",
                    "app_secret": "",
                    "sid": "",
                }
            }, ensure_ascii=False), encoding="utf-8")
            with patch.object(daily_ops_app, "RULES_FILE", rules_file), patch.object(daily_ops_app, "ERP_API_LOCAL_FILE", local_file):
                rules = daily_ops_app.load_rules()

        erp = rules["erp_api"]
        self.assertEqual(erp["sync_days"], 30)
        self.assertEqual(erp["page_size"], 500)
        self.assertEqual(erp["stock_limit"], 10000)
        self.assertEqual(erp["warehouse_no"], "3")
        self.assertEqual(erp["warehouse_name"], "宠物圈仓")
        self.assertEqual(erp["app_key"], "local-app")
        self.assertEqual(erp["app_secret"], "local-secret")
        self.assertEqual(erp["sid"], "local-sid")

    def test_load_rules_keeps_custom_erp_endpoints(self):
        with tempfile.TemporaryDirectory() as tmp:
            rules_file = Path(tmp) / "report_rules.json"
            rules_file.write_text("""{
              "erp_api": {
                "product_endpoint": "custom_goods.php",
                "stock_endpoint": "custom_stock.php"
              }
            }""", encoding="utf-8")
            with patch.object(daily_ops_app, "RULES_FILE", rules_file):
                rules = daily_ops_app.load_rules()

        self.assertEqual(rules["erp_api"]["product_endpoint"], "custom_goods.php")
        self.assertEqual(rules["erp_api"]["stock_endpoint"], "custom_stock.php")

    def test_save_rules_keeps_erp_credentials_in_local_ignored_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            rules_file = Path(tmp) / "report_rules.json"
            local_file = Path(tmp) / "erp_api.local.json"
            with patch.object(daily_ops_app, "RULES_FILE", rules_file), patch.object(daily_ops_app, "ERP_API_LOCAL_FILE", local_file):
                saved = daily_ops_app.save_rules({
                    "erp_api": {
                        "environment": "test",
                        "base_url": "https://openapi.ali.huice.cc/openapi",
                        "app_key": "local-test-app",
                        "app_secret": "secret-value",
                        "sid": "local-test-sid",
                        "token": "token-value",
                        "shop_id": "1314",
                    }
                })
                stored_rules = json.loads(rules_file.read_text(encoding="utf-8"))
                local_rules = json.loads(local_file.read_text(encoding="utf-8"))
                loaded = daily_ops_app.load_rules()

        self.assertEqual(stored_rules["erp_api"]["app_key"], "")
        self.assertEqual(stored_rules["erp_api"]["app_secret"], "")
        self.assertEqual(stored_rules["erp_api"]["sid"], "")
        self.assertEqual(stored_rules["erp_api"]["token"], "")
        self.assertEqual(local_rules["app_key"], "local-test-app")
        self.assertEqual(local_rules["app_secret"], "secret-value")
        self.assertEqual(local_rules["sid"], "local-test-sid")
        self.assertEqual(saved["erp_api"]["app_secret"], "secret-value")
        self.assertEqual(loaded["erp_api"]["sid"], "local-test-sid")

    def test_load_rules_can_ignore_local_erp_credentials_for_smoke_checks(self):
        with tempfile.TemporaryDirectory() as tmp:
            rules_file = Path(tmp) / "report_rules.json"
            local_file = Path(tmp) / "erp_api.local.json"
            rules_file.write_text("""{
              "erp_api": {
                "base_url": "https://openapi.ali.huice.cc/openapi",
                "shop_id": "1314"
              }
            }""", encoding="utf-8")
            local_file.write_text(json.dumps({
                "app_key": "local-test-app",
                "app_secret": "secret-value",
                "sid": "local-test-sid"
            }), encoding="utf-8")
            with patch.object(daily_ops_app, "RULES_FILE", rules_file), \
                    patch.object(daily_ops_app, "ERP_API_LOCAL_FILE", local_file), \
                    patch.dict(os.environ, {"DAILY_OPS_IGNORE_LOCAL_ERP_CREDENTIALS": "1"}):
                rules = daily_ops_app.load_rules()

        self.assertEqual(rules["erp_api"]["app_key"], "")
        self.assertEqual(rules["erp_api"]["app_secret"], "")
        self.assertEqual(rules["erp_api"]["sid"], "")

    def test_product_rows_map_to_existing_erp_headers(self):
        rows = daily_ops_erp.normalize_product_rows([{
            "shop_no": "PETCIRCLE",
            "shop_name": "宠物圈仓库",
            "goods_no": "SKU",
            "goods_name": "猫抓板",
            "category_name": "宠物玩具",
            "cost_price": "8.5",
            "wholesale_price": "12.8",
            "spec_list": [{
                "spec_no": "SKU-XL",
                "spec_name": "XL",
                "stock_num": 12,
            }],
        }])
        self.assertEqual(rows[0]["商家编码（新）"], "SKU-XL")
        self.assertEqual(rows[0]["货品名称"], "猫抓板")
        self.assertEqual(rows[0]["规格名称"], "XL")
        self.assertEqual(rows[0]["货品分类名称"], "宠物玩具")
        self.assertEqual(rows[0]["成本价"], "8.5")
        self.assertEqual(rows[0]["批发价"], "12.8")

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "erp产品基础信息表_接口同步_test.xlsx"
            daily_ops_erp._write_rows(path, ["商家编码（新）", "货品名称", "规格名称", "货品分类名称", "成本价", "批发价"], rows)
            wb = load_workbook(path)
            ws = wb.active
            self.assertEqual(ws.cell(2, 1).value, "SKU-XL")
            self.assertEqual(ws.cell(2, 4).value, "宠物玩具")
            self.assertEqual(ws.cell(2, 5).value, "8.5")
            wb.close()

    def test_stock_rows_keep_warehouse_and_unified_merchant_code(self):
        rows = daily_ops_erp.normalize_stock_rows([{
            "shop_no": "PETCIRCLE",
            "shop_name": "宠物圈",
            "warehouse_no": "CW001",
            "warehouse_name": "宠物圈仓库",
            "spec_no": "SKU-XL",
            "goods_name": "猫抓板",
            "spec_name": "XL",
            "stock_num": 12,
        }])

        self.assertEqual(rows[0]["仓库编号"], "CW001")
        self.assertEqual(rows[0]["仓库"], "宠物圈仓库")
        self.assertEqual(rows[0]["商家编码（新）"], "SKU-XL")
        self.assertEqual(rows[0]["商家编码"], "SKU-XL")
        self.assertEqual(rows[0]["规格名称"], "XL")
        self.assertEqual(rows[0]["可销库存"], "12")

    def test_stock_rows_filter_to_petcircle_warehouse(self):
        rows = daily_ops_erp.normalize_stock_rows([
            {"warehouse_no": "CW001", "warehouse_name": "宠物圈仓库", "spec_no": "SKU-1", "stock_num": 5},
            {"warehouse_no": "CW002", "warehouse_name": "其他仓库", "spec_no": "SKU-2", "stock_num": 9},
        ], warehouse_name="宠物圈仓库")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["商家编码"], "SKU-1")
        self.assertEqual(rows[0]["仓库"], "宠物圈仓库")

    def test_query_erp_product_info_merges_stock_snapshot_for_dense_table(self):
        with tempfile.TemporaryDirectory() as tmp:
            erp_dir = Path(tmp)
            daily_ops_erp._write_rows(
                erp_dir / "erp产品基础信息表_接口同步_最新.xlsx",
                ["货品编码", "货品名称", "规格名称", "货品分类名称", "商家编码（新）", "批发价", "成本价", "零售价", "修改时间", "来源接口"],
                [{
                    "货品编码": "3303177721",
                    "货品名称": "元气四脚冲锋衣-黄紫",
                    "规格名称": "黄紫/XS",
                    "货品分类名称": "宠物服饰",
                    "商家编码（新）": "3303177721-XS",
                    "批发价": "18.8",
                    "成本价": "12.3",
                    "零售价": "29.9",
                    "修改时间": "2026-05-29 09:55:08",
                    "来源接口": "goods_query.php",
                }],
            )
            daily_ops_erp._write_rows(
                erp_dir / "erp库存同步_最新.xlsx",
                ["平台规格编码", "商家编码（新）", "商家编码", "可销库存", "修改时间", "来源接口"],
                [{
                    "平台规格编码": "3303177721-XS",
                    "商家编码（新）": "3303177721-XS",
                    "商家编码": "3303177721-XS",
                    "可销库存": "211.0000",
                    "修改时间": "2026-06-25 15:56:10",
                    "来源接口": "stock_query.php",
                }],
            )

            with patch.object(daily_ops_app, "ERP_DIR", erp_dir):
                result = daily_ops_app.query_erp_product_info("3303177721-XS", 20)

        self.assertEqual(result["columns"], [
            "货品编码", "货品名称", "规格名称", "货品分类名称", "商家编码（新）", "可销库存",
            "批发价", "成本价", "零售价", "商品资料修改时间", "库存修改时间", "来源接口",
        ])
        self.assertEqual(result["items"][0]["record"]["货品分类名称"], "宠物服饰")
        self.assertEqual(result["items"][0]["record"]["可销库存"], "211.0000")
        self.assertEqual(result["items"][0]["record"]["库存修改时间"], "2026-06-25 15:56:10")
        self.assertEqual(result["items"][0]["record"]["来源接口"], "goods_query.php / stock_query.php")

    def test_query_erp_product_info_hides_products_without_stock_snapshot(self):
        with tempfile.TemporaryDirectory() as tmp:
            erp_dir = Path(tmp)
            daily_ops_erp._write_rows(
                erp_dir / "erp产品基础信息表_接口同步_最新.xlsx",
                ["货品编码", "货品名称", "规格名称", "商家编码（新）", "批发价", "成本价", "零售价", "修改时间", "来源接口"],
                [
                    {
                        "货品编码": "PET-001",
                        "货品名称": "宠物圈仓款",
                        "规格名称": "红色/S",
                        "商家编码（新）": "PET-001-S",
                        "来源接口": "goods_query.php",
                    },
                    {
                        "货品编码": "OTHER-001",
                        "货品名称": "甜心仓库款",
                        "规格名称": "粉色/S",
                        "商家编码（新）": "OTHER-001-S",
                        "来源接口": "goods_query.php",
                    },
                ],
            )
            daily_ops_erp._write_rows(
                erp_dir / "erp库存同步_最新.xlsx",
                ["商家编码（新）", "商家编码", "可销库存", "修改时间", "来源接口"],
                [{
                    "商家编码（新）": "PET-001-S",
                    "商家编码": "PET-001-S",
                    "可销库存": "8",
                    "修改时间": "2026-06-26 10:00:00",
                    "来源接口": "stock_query.php",
                }],
            )

            with patch.object(daily_ops_app, "ERP_DIR", erp_dir):
                result = daily_ops_app.query_erp_product_info("", 20)
                hidden = daily_ops_app.query_erp_product_info(product_name="甜心仓库款", limit=20)

        self.assertEqual([item["record"]["商家编码（新）"] for item in result["items"]], ["PET-001-S"])
        self.assertEqual(hidden["items"], [])

    def test_manual_sync_reads_multiple_product_pages_and_one_stock_snapshot(self):
        settings = {
            "base_url": "https://api.wangdian.cn/openapi2",
            "app_key": "app",
            "app_secret": "secret",
            "sid": "sid",
            "shop_id": "1314",
            "warehouse_name": "",
            "page_size": 2,
            "stock_limit": 3,
            "stock_endpoint": daily_ops_erp.STOCK_CHANGE_ENDPOINT,
        }

        def fake_post(_settings, endpoint, params):
            if endpoint == daily_ops_erp.PRODUCT_ENDPOINT:
                page_no = params["page_no"]
                pages = [
                    {"goods_list": [
                        {"spec_no": "P-1", "goods_name": "猫粮", "spec_name": "1kg"},
                        {"spec_no": "P-2", "goods_name": "猫粮", "spec_name": "2kg"},
                    ], "total_count": 3},
                    {"goods_list": [
                        {"spec_no": "P-3", "goods_name": "猫粮", "spec_name": "3kg"},
                    ], "total_count": 3},
                ]
                return pages[page_no]
            return {"stock_change_list": [
                {"spec_no": "S-1", "stock_num": 11},
                {"spec_no": "S-2", "stock_num": 12},
            ], "current_count": 2}

        with tempfile.TemporaryDirectory() as tmp, patch.object(daily_ops_erp, "post_api", side_effect=fake_post):
            result = daily_ops_erp.manual_sync(settings, Path(tmp))

            self.assertEqual(result["status"], "synced")
            self.assertEqual(result["product_count"], 3)
            self.assertEqual(result["stock_count"], 2)
            self.assertEqual(result["product_pages"], 2)
            self.assertEqual(result["stock_pages"], 1)
            stock_path = Path(result["stock_file"])
            wb = load_workbook(stock_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            self.assertIn("仓库", headers)
            self.assertIn("商家编码（新）", headers)
            self.assertIn("规格名称", headers)
            self.assertEqual(ws.cell(2, headers.index("商家编码（新）") + 1).value, "S-1")
            wb.close()

    def test_manual_sync_respects_selected_erp_pull_content(self):
        settings = {
            "base_url": "https://api.wangdian.cn/openapi2",
            "app_key": "app",
            "app_secret": "secret",
            "sid": "sid",
            "shop_id": "1314",
            "warehouse_name": "",
            "sync_product_archive": False,
            "sync_stock_snapshot": True,
        }
        calls = []

        def fake_post(_settings, endpoint, params):
            calls.append(endpoint)
            return {"stock_list": [{"spec_no": "S-1", "stock_num": 11}], "total_count": 1}

        with tempfile.TemporaryDirectory() as tmp, patch.object(daily_ops_erp, "post_api", side_effect=fake_post):
            result = daily_ops_erp.manual_sync(settings, Path(tmp))

            self.assertEqual(result["status"], "synced")
            self.assertEqual(result["product_count"], 0)
            self.assertEqual(result["stock_count"], 1)
            self.assertEqual(result["product_file"], "")
            self.assertEqual(calls, [daily_ops_erp.STOCK_ENDPOINT])

    def test_manual_sync_can_pull_optional_authorized_interfaces_separately(self):
        settings = {
            "base_url": "https://api.wangdian.cn/openapi2",
            "app_key": "app",
            "app_secret": "secret",
            "sid": "sid",
            "shop_id": "1314",
            "warehouse_name": "",
            "sync_product_archive": False,
            "sync_stock_snapshot": False,
            "sync_available_stock": True,
            "sync_shop_query": True,
            "sync_platform_goods": True,
            "sync_sales_outbound": True,
            "latest_file_only": True,
        }
        calls = []

        def fake_post(_settings, endpoint, params):
            calls.append(endpoint)
            if endpoint == daily_ops_erp.SHOP_ENDPOINT:
                return {"shop_list": [{"shop_no": "1314", "shop_name": "测试店铺"}], "total_count": 1}
            if endpoint == daily_ops_erp.PLATFORM_GOODS_ENDPOINT:
                return {"goods_list": [{"api_goods_no": "P-1", "spec_no": "SKU-1"}], "total_count": 1}
            if endpoint == daily_ops_erp.SALES_OUTBOUND_ENDPOINT:
                return {"trade_list": [{"trade_no": "T-1"}], "total_count": 1}
            return {"stock_change_list": [{"spec_no": "S-1", "stock_num": 11}], "current_count": 1}

        with tempfile.TemporaryDirectory() as tmp, patch.object(daily_ops_erp, "post_api", side_effect=fake_post):
            result = daily_ops_erp.manual_sync(settings, Path(tmp))

        self.assertEqual(result["status"], "synced")
        self.assertEqual(calls, [
            daily_ops_erp.STOCK_CHANGE_ENDPOINT,
            daily_ops_erp.SHOP_ENDPOINT,
            daily_ops_erp.PLATFORM_GOODS_ENDPOINT,
            daily_ops_erp.SALES_OUTBOUND_ENDPOINT,
        ])
        self.assertEqual(Path(result["extra_files"]["platform_goods"]).name, "erp平台货品查询_接口同步_最新.xlsx")
        self.assertFalse(result["product_file"])

    def test_stock_change_sync_uses_documented_limit_without_pagination_or_time_window(self):
        settings = {
            "base_url": "https://api.wangdian.cn/openapi2",
            "app_key": "app",
            "app_secret": "secret",
            "sid": "sid",
            "shop_id": "1314",
            "page_size": 500,
            "stock_limit": 50,
            "stock_endpoint": daily_ops_erp.STOCK_CHANGE_ENDPOINT,
        }
        calls = []

        def fake_post(_settings, endpoint, params):
            calls.append((endpoint, dict(params)))
            if endpoint == daily_ops_erp.PRODUCT_ENDPOINT:
                return {"goods_list": [], "total_count": 0}
            return {"stock_change_list": [{"spec_no": "S-1", "sync_stock": 11}], "current_count": 1}

        with tempfile.TemporaryDirectory() as tmp, patch.object(daily_ops_erp, "post_api", side_effect=fake_post):
            result = daily_ops_erp.manual_sync(settings, Path(tmp))

        self.assertEqual(result["status"], "synced")
        product_params = [params for endpoint, params in calls if endpoint == daily_ops_erp.PRODUCT_ENDPOINT][0]
        stock_params = [params for endpoint, params in calls if endpoint == daily_ops_erp.STOCK_CHANGE_ENDPOINT][0]
        self.assertEqual(product_params["page_size"], 500)
        self.assertEqual(product_params["shop_id"], "1314")
        self.assertEqual(stock_params, {"limit": 100, "shop_id": "1314"})
        self.assertEqual(result["stock_count"], 1)
        self.assertEqual(result["stock_pages"], 1)

    def test_manual_sync_reads_available_stock_with_time_window_and_pagination(self):
        settings = {
            "base_url": "https://api.wangdian.cn/openapi2",
            "app_key": "app",
            "app_secret": "secret",
            "sid": "sid",
            "shop_id": "",
            "warehouse_name": "",
            "page_size": 1,
            "stock_endpoint": "stock_query.php",
            "sync_days": 30,
        }
        calls = []

        def fake_post(_settings, endpoint, params):
            calls.append((endpoint, dict(params)))
            if endpoint == daily_ops_erp.PRODUCT_ENDPOINT:
                return {"goods_list": [], "total_count": 0}
            page_no = params["page_no"]
            pages = [
                {"stock_list": [{"spec_no": "S-1", "stock_num": 11}], "total_count": 2},
                {"stock_list": [{"spec_no": "S-2", "stock_num": 12}], "total_count": 2},
            ]
            return pages[page_no]

        with tempfile.TemporaryDirectory() as tmp, patch.object(daily_ops_erp, "post_api", side_effect=fake_post):
            result = daily_ops_erp.manual_sync(settings, Path(tmp))

        stock_calls = [params for endpoint, params in calls if endpoint == "stock_query.php"]
        self.assertEqual(result["stock_count"], 2)
        self.assertEqual(result["stock_pages"], 2)
        self.assertEqual(stock_calls[0]["page_size"], 1)
        self.assertEqual(stock_calls[0]["page_no"], 0)
        self.assertIn("start_time", stock_calls[0])
        self.assertIn("end_time", stock_calls[0])
        self.assertNotIn("limit", stock_calls[0])
        self.assertEqual(stock_calls[1]["page_no"], 1)

    def test_manual_sync_filters_stock_by_warehouse_no_without_default_name(self):
        settings = {
            "base_url": "https://api.wangdian.cn/openapi2",
            "app_key": "app",
            "app_secret": "secret",
            "sid": "sid",
            "shop_id": "",
            "warehouse_no": "3",
            "warehouse_name": "",
            "sync_product_archive": False,
            "sync_stock_snapshot": True,
            "stock_endpoint": "stock_query.php",
        }

        def fake_post(_settings, _endpoint, _params):
            return {"stocks": [
                {"warehouse_no": "3", "warehouse_name": "宠物圈仓", "spec_no": "S-1", "stock_num": 11},
                {"warehouse_no": "9", "warehouse_name": "其他仓", "spec_no": "S-2", "stock_num": 12},
            ], "total_count": 2}

        with tempfile.TemporaryDirectory() as tmp, patch.object(daily_ops_erp, "post_api", side_effect=fake_post):
            result = daily_ops_erp.manual_sync(settings, Path(tmp))

        self.assertEqual(result["status"], "synced")
        self.assertEqual(result["stock_count"], 1)

    def test_goods_archive_sync_does_not_limit_product_rows_by_time_window(self):
        settings = {
            "base_url": "https://api.wangdian.cn/openapi2",
            "app_key": "app",
            "app_secret": "secret",
            "sid": "sid",
            "shop_id": "",
            "warehouse_name": "",
            "page_size": 1,
            "stock_endpoint": "stock_query.php",
            "sync_days": 30,
        }
        calls = []

        def fake_post(_settings, endpoint, params):
            calls.append((endpoint, dict(params)))
            if endpoint == daily_ops_erp.PRODUCT_ENDPOINT:
                return {"goods_list": [], "total_count": 0}
            return {"stock_list": [], "total_count": 0}

        with tempfile.TemporaryDirectory() as tmp, patch.object(daily_ops_erp, "post_api", side_effect=fake_post):
            daily_ops_erp.manual_sync(settings, Path(tmp))

        product_params = [params for endpoint, params in calls if endpoint == daily_ops_erp.PRODUCT_ENDPOINT][0]
        stock_params = [params for endpoint, params in calls if endpoint == "stock_query.php"][0]
        self.assertNotIn("start_time", product_params)
        self.assertNotIn("end_time", product_params)
        self.assertIn("start_time", stock_params)
        self.assertIn("end_time", stock_params)

    def test_manual_sync_can_read_more_than_default_50_product_pages(self):
        settings = {
            "base_url": "https://api.wangdian.cn/openapi2",
            "app_key": "app",
            "app_secret": "secret",
            "sid": "sid",
            "warehouse_name": "",
            "page_size": 1,
            "max_pages": 52,
            "stock_limit": 100,
        }

        def fake_post(_settings, endpoint, params):
            if endpoint == daily_ops_erp.PRODUCT_ENDPOINT:
                return {"goods_list": [{"spec_no": f"P-{params['page_no']}"}], "total_count": 52}
            return {"stock_list": [], "total_count": 0}

        with tempfile.TemporaryDirectory() as tmp, patch.object(daily_ops_erp, "post_api", side_effect=fake_post):
            result = daily_ops_erp.manual_sync(settings, Path(tmp))

        self.assertEqual(result["product_count"], 52)
        self.assertEqual(result["product_pages"], 52)

    def test_manual_sync_allows_large_available_stock_limit(self):
        settings = {
            "base_url": "https://api.wangdian.cn/openapi2",
            "app_key": "app",
            "app_secret": "secret",
            "sid": "sid",
            "warehouse_name": "",
            "page_size": 1000,
            "stock_limit": 2500,
        }
        seen_stock_params = []

        def fake_post(_settings, endpoint, params):
            if endpoint == daily_ops_erp.PRODUCT_ENDPOINT:
                return {"goods_list": [], "total_count": 0}
            seen_stock_params.append(dict(params))
            start = params["page_no"] * params["page_size"]
            count = max(0, min(params["page_size"], 2500 - start))
            return {"stock_list": [{"spec_no": f"S-{i}"} for i in range(start, start + count)], "total_count": 2500}

        with tempfile.TemporaryDirectory() as tmp, patch.object(daily_ops_erp, "post_api", side_effect=fake_post):
            result = daily_ops_erp.manual_sync(settings, Path(tmp))

        self.assertEqual(result["stock_count"], 2500)
        self.assertEqual(result["stock_pages"], 3)
        self.assertEqual(seen_stock_params[-1]["page_no"], 2)

    def test_fetch_paged_rows_paces_requests_between_pages(self):
        settings = {"request_interval_seconds": 1.2}
        calls = []

        def fake_post(_settings, _endpoint, params):
            calls.append(dict(params))
            return {"goods_list": [{"spec_no": f"P-{params['page_no']}"}], "total_count": 2}

        with patch.object(daily_ops_erp, "post_api", side_effect=fake_post), patch.object(daily_ops_erp.time, "sleep") as sleep:
            result = daily_ops_erp.fetch_paged_rows(settings, daily_ops_erp.PRODUCT_ENDPOINT, {}, ["goods_list"], page_size=1, max_pages=5)

        self.assertEqual(len(result["rows"]), 2)
        self.assertEqual([call["page_no"] for call in calls], [0, 1])
        sleep.assert_called_once_with(1.2)

    def test_fetch_paged_rows_retries_after_rate_limit(self):
        settings = {"rate_limit_retry_seconds": 0}
        responses = [
            {"code": 1, "message": "频率超限，请保持1分钟内60次的频率5分钟后重新调用"},
            {"goods_list": [{"spec_no": "P-1"}], "total_count": 1},
        ]

        def fake_post(_settings, _endpoint, _params):
            return responses.pop(0)

        with patch.object(daily_ops_erp, "post_api", side_effect=fake_post), patch.object(daily_ops_erp.time, "sleep") as sleep:
            result = daily_ops_erp.fetch_paged_rows(settings, daily_ops_erp.PRODUCT_ENDPOINT, {}, ["goods_list"], page_size=1)

        self.assertEqual(len(result["rows"]), 1)
        self.assertEqual(result["messages"], [])
        sleep.assert_called_once_with(0)

    def test_manual_sync_can_overwrite_latest_erp_files(self):
        settings = {
            "base_url": "https://api.wangdian.cn/openapi2",
            "app_key": "app",
            "app_secret": "secret",
            "sid": "sid",
            "warehouse_name": "",
            "page_size": 1,
            "latest_file_only": True,
        }

        def fake_post(_settings, endpoint, _params):
            if endpoint == daily_ops_erp.PRODUCT_ENDPOINT:
                return {"goods_list": [{"spec_no": "P-1"}], "total_count": 1}
            return {"stock_list": [{"spec_no": "S-1", "stock_num": 8}], "total_count": 1}

        with tempfile.TemporaryDirectory() as tmp, patch.object(daily_ops_erp, "post_api", side_effect=fake_post):
            result = daily_ops_erp.manual_sync(settings, Path(tmp))

        self.assertEqual(Path(result["product_file"]).name, "erp产品基础信息表_接口同步_最新.xlsx")
        self.assertEqual(Path(result["stock_file"]).name, "erp库存同步_最新.xlsx")

    def test_sync_erp_base_data_blocks_when_lock_exists(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            rules_file = root / "report_rules.json"
            local_file = root / "erp_api.local.json"
            lock_file = root / "erp_sync.lock"
            rules_file.write_text(json.dumps({"erp_api": {"enabled": True}}, ensure_ascii=False), encoding="utf-8")
            lock_file.write_text("running", encoding="utf-8")
            with patch.object(daily_ops_app, "RULES_FILE", rules_file), \
                patch.object(daily_ops_app, "ERP_API_LOCAL_FILE", local_file), \
                patch.object(daily_ops_app, "ERP_SYNC_LOCK_FILE", lock_file):
                result = daily_ops_app.sync_erp_base_data()

        self.assertEqual(result["status"], "blocked")
        self.assertIn("正在同步", result["message"])

    def test_sync_erp_base_data_records_failure_without_clearing_last_success(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            rules_file = root / "report_rules.json"
            local_file = root / "erp_api.local.json"
            lock_file = root / "erp_sync.lock"
            erp_dir = root / "erp数据源"
            rules_file.write_text(json.dumps({
                "erp_api": {
                    "enabled": True,
                    "last_manual_sync_at": "2026-06-23 05:00:00",
                    "last_manual_sync_status": "synced",
                    "last_manual_sync_message": "已同步商品 10 条、库存 8 条",
                    "last_product_count": 10,
                    "last_stock_count": 8,
                }
            }, ensure_ascii=False), encoding="utf-8")

            def fail_sync(_settings, _erp_dir):
                raise RuntimeError("接口超时")

            with patch.object(daily_ops_app, "RULES_FILE", rules_file), \
                patch.object(daily_ops_app, "ERP_API_LOCAL_FILE", local_file), \
                patch.object(daily_ops_app, "ERP_SYNC_LOCK_FILE", lock_file), \
                patch.object(daily_ops_app, "ERP_DIR", erp_dir), \
                patch("daily_ops_erp.manual_sync", side_effect=fail_sync):
                result = daily_ops_app.sync_erp_base_data()
                saved = json.loads(rules_file.read_text(encoding="utf-8"))["erp_api"]

        self.assertEqual(result["status"], "failed")
        self.assertEqual(saved["last_manual_sync_status"], "failed")
        self.assertEqual(saved["last_product_count"], 10)
        self.assertEqual(saved["last_stock_count"], 8)
        self.assertIn("接口超时", saved["last_manual_sync_message"])

    def test_sync_erp_base_data_records_available_stock_summary(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            rules_file = root / "report_rules.json"
            local_file = root / "erp_api.local.json"
            lock_file = root / "erp_sync.lock"
            erp_dir = root / "erp数据源"
            available_file = erp_dir / "erp可用库存_接口同步_最新.xlsx"
            rules_file.write_text(json.dumps({"erp_api": {"enabled": True}}, ensure_ascii=False), encoding="utf-8")

            def synced(_settings, _erp_dir):
                return {
                    "status": "synced",
                    "message": "已同步货品档案 2 条、库存快照 0 条、可用库存 1 条",
                    "product_count": 2,
                    "stock_count": 0,
                    "available_stock_count": 1,
                    "product_pages": 1,
                    "stock_pages": 1,
                    "product_file": str(erp_dir / "erp产品基础信息表_接口同步_最新.xlsx"),
                    "stock_file": str(erp_dir / "erp库存同步_最新.xlsx"),
                    "extra_files": {"available_stock": str(available_file)},
                }

            with patch.object(daily_ops_app, "RULES_FILE", rules_file), \
                patch.object(daily_ops_app, "ERP_API_LOCAL_FILE", local_file), \
                patch.object(daily_ops_app, "ERP_SYNC_LOCK_FILE", lock_file), \
                patch.object(daily_ops_app, "ERP_DIR", erp_dir), \
                patch("daily_ops_erp.manual_sync", side_effect=synced):
                result = daily_ops_app.sync_erp_base_data()
                saved = json.loads(rules_file.read_text(encoding="utf-8"))["erp_api"]

        self.assertEqual(result["status"], "synced")
        self.assertEqual(saved["last_available_stock_count"], 1)
        self.assertEqual(saved["last_available_stock_file"], str(available_file))

    def test_erp_base_files_prefers_latest_sync_file_without_deleting_history_or_uploaded_manifest(self):
        with tempfile.TemporaryDirectory() as tmp:
            erp_dir = Path(tmp)
            latest = erp_dir / "erp产品基础信息表_接口同步_最新.xlsx"
            old = erp_dir / "erp产品基础信息表_接口同步_20260624_161745.xlsx"
            manual = erp_dir / "erp产品基础信息表_人工.xlsx"
            uploaded = erp_dir / "erp产品基础信息表_上传.xlsx"
            latest.write_text("latest", encoding="utf-8")
            old.write_text("old", encoding="utf-8")
            manual.write_text("manual", encoding="utf-8")
            uploaded.write_text("uploaded", encoding="utf-8")

            with patch.object(daily_ops_app, "ERP_DIR", erp_dir), patch.object(daily_ops_app, "manifest_paths", return_value=[uploaded]):
                files = daily_ops_app.erp_base_files()

            self.assertEqual(files, [latest])
            self.assertTrue(old.exists())
            self.assertTrue(manual.exists())
            self.assertTrue(uploaded.exists())

    def test_erp_base_files_uses_latest_interface_sync_instead_of_uploaded_legacy_source(self):
        with tempfile.TemporaryDirectory() as tmp:
            erp_dir = Path(tmp)
            older_sync = erp_dir / "erp产品基础信息表_接口同步_20260624_160552.xlsx"
            newest_sync = erp_dir / "erp产品基础信息表_接口同步_20260624_161745.xlsx"
            uploaded_legacy = erp_dir / "20260606-全部数据-无图版.xlsx"
            older_sync.write_text("old sync", encoding="utf-8")
            newest_sync.write_text("new sync", encoding="utf-8")
            uploaded_legacy.write_text("legacy", encoding="utf-8")

            with patch.object(daily_ops_app, "ERP_DIR", erp_dir), \
                 patch.object(daily_ops_app, "manifest_paths", return_value=[uploaded_legacy]):
                files = daily_ops_app.erp_base_files()

            self.assertEqual(files, [newest_sync])

if __name__ == "__main__":
    unittest.main()
