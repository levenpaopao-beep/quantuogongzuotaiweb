import re
import html
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import update_shein_summary_30d_skc as raw_xlsx


ROOT = Path(__file__).resolve().parent
TEMU_SOURCE_DIR = ROOT / "temu数据源表"
ERP_DIR = ROOT / "erp数据源"
OWNER_FILE = ROOT / "店铺负责人对应表.xlsx"
OUTPUT_DIR = ROOT / "outputs"
OUTPUT = OUTPUT_DIR / "260603-Temu滞销品每周报表-V3.xlsx"
TEMU_SALES_FILES = None
ERP_FILES = None
RULES = None

DEFAULT_RULES = {
    "slow_moving": {
        "new_product_days_lt": 28,
        "new_slow_min_days": 30,
        "new_slow_max_days": 60,
        "old_slow_min_days": 180,
        "group_by": "店铺+SPU",
        "sales30_total_equals": 0,
    }
}

STORE_ORDER = [
    "弟弟",
    "二弟",
    "三弟",
    "四弟",
    "五弟",
    "六弟",
    "七弟",
    "八弟",
    "九弟（喵喵）",
    "十弟",
    "十一弟",
    "十二弟",
    "十三（节日）",
    "十五弟（毛衣）",
]
STORE_RANK = {name: i for i, name in enumerate(STORE_ORDER)}
STORE_ALIASES = {
    "一弟": "弟弟",
    "九弟": "九弟（喵喵）",
    "十一": "十一弟",
    "十二": "十二弟",
    "十三": "十三（节日）",
    "十五": "十五弟（毛衣）",
}

DETAIL_COLUMNS = [
    "店铺",
    "负责人",
    "ERP货品名称",
    "规格名称",
    "SKC",
    "SKU ID",
    "SKU货号",
    "申报价",
    "7天销量",
    "30天销量",
    "平台仓库可用",
    "上架天数",
    "日均销量",
    "成本价",
    "批发报价",
    "ERP可用库存(仅展示)",
    "预警类型",
    "建议动作",
    "源文件",
    "源行",
]


def to_number(value):
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return 0


def norm(value):
    return html.unescape(str(value or "")).strip()


def canonical_store(value):
    text = norm(value)
    return STORE_ALIASES.get(text, text)


def active_slow_rules():
    source = RULES if isinstance(RULES, dict) else {}
    current = source.get("slow_moving", source) if isinstance(source, dict) else {}
    merged = DEFAULT_RULES["slow_moving"].copy()
    if isinstance(current, dict):
        merged.update(current)
    return merged


SIZE_SUFFIXES = {
    "XXXS", "XXS", "XS", "S", "M", "MD", "L", "XL", "XXL", "XXXL",
    "2XL", "3XL", "4XL", "5XL", "6XL", "7XL",
}


def spu_key(value):
    text = str(value or "").strip().upper()
    if "@" in text:
        text = text.split("@", 1)[0]
    text = text.strip()
    if not text:
        return ""

    if text.startswith("33"):
        match = re.match(r"^(33\d+)", text)
        if match:
            return match.group(1)

    parts = re.split(r"[-_\s]+", text)
    if len(parts) > 1 and parts[-1] in SIZE_SUFFIXES:
        return "-".join(parts[:-1])
    return text


def sku_key(value):
    text = str(value or "").strip().upper()
    if "@" in text:
        text = text.split("@", 1)[0]
    return text.strip()


def header_map(headers):
    return {norm(value): idx for idx, value in enumerate(headers) if norm(value)}


def latest_temu_sales_files():
    if TEMU_SALES_FILES:
        files = [Path(path) for path in TEMU_SALES_FILES]
        return "uploaded", files
    date_re = re.compile(r"(20\d{6})")
    files = sorted(TEMU_SOURCE_DIR.glob("*Temu仓库销售情况导出*.xlsx"))
    by_date = defaultdict(list)
    for path in files:
        match = date_re.search(path.name)
        if match:
            by_date[match.group(1)].append(path)
    if not by_date:
        raise FileNotFoundError(f"{TEMU_SOURCE_DIR} 中未找到 Temu仓库销售情况导出 文件")
    latest_date = max(by_date)
    return latest_date, sorted(by_date[latest_date], key=lambda p: p.name)


def iter_xlsx_rows(path):
    yield from raw_xlsx.read_xlsx_rows(path)


def load_owner_map():
    owners = {}
    if not OWNER_FILE.exists():
        return owners
    wb = load_workbook(OWNER_FILE, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    hm = header_map(headers)
    store_i = hm.get("店铺")
    owner_i = hm.get("业务")
    if store_i is None or owner_i is None:
        return owners
    for row in ws.iter_rows(min_row=2, values_only=True):
        store = canonical_store(row[store_i] if store_i < len(row) else "")
        owner = norm(row[owner_i] if owner_i < len(row) else "")
        if store and owner:
            owners[store] = owner
    return owners


def load_erp_records():
    records = {}
    erp_rows = 0
    for path in (ERP_FILES or sorted(ERP_DIR.glob("erp产品基础信息表*.xlsx"))):
        rows = iter_xlsx_rows(path)
        try:
            headers = next(rows)
        except StopIteration:
            continue
        hm = header_map(headers)
        code_i = hm.get("商家编码（新）")
        if code_i is None:
            continue
        for row in rows:
            erp_rows += 1
            code = norm(row[code_i] if code_i < len(row) else "")
            key = sku_key(code)
            if key and key not in records:
                records[key] = {
                    "ERP货品名称": norm(row[hm["货品名称"]] if hm.get("货品名称") is not None and hm["货品名称"] < len(row) else ""),
                    "规格名称": norm(row[hm["规格名称"]] if hm.get("规格名称") is not None and hm["规格名称"] < len(row) else ""),
                    "成本价": row[hm["成本价"]] if hm.get("成本价") is not None and hm["成本价"] < len(row) else "",
                    "批发报价": row[hm["批发报价"]] if hm.get("批发报价") is not None and hm["批发报价"] < len(row) else "",
                    "ERP可用库存(仅展示)": row[hm["可用库存"]] if hm.get("可用库存") is not None and hm["可用库存"] < len(row) else "",
                }

    combo = ERP_DIR / "erp产品组合装基础信息表.xlsx"
    combo_rows = 0
    if combo.exists():
        rows = iter_xlsx_rows(combo)
        try:
            headers = next(rows)
        except StopIteration:
            headers = []
        hm = header_map(headers)
        code_i = hm.get("商家编码（新）")
        if code_i is not None:
            for row in rows:
                combo_rows += 1
                code = norm(row[code_i] if code_i < len(row) else "")
                key = sku_key(code)
                if key and key not in records:
                    records[key] = {
                        "ERP货品名称": norm(row[hm["组合装名称"]] if hm.get("组合装名称") is not None and hm["组合装名称"] < len(row) else ""),
                        "规格名称": norm(row[hm["组合装简称"]] if hm.get("组合装简称") is not None and hm["组合装简称"] < len(row) else ""),
                        "成本价": row[hm["成本价"]] if hm.get("成本价") is not None and hm["成本价"] < len(row) else "",
                        "批发报价": row[hm["批发价"]] if hm.get("批发价") is not None and hm["批发价"] < len(row) else "",
                        "ERP可用库存(仅展示)": row[hm["可用库存"]] if hm.get("可用库存") is not None and hm["可用库存"] < len(row) else "",
                    }
    return records, erp_rows, combo_rows


def store_sort_key(row):
    return (
        STORE_RANK.get(row.get("店铺", ""), 999),
        str(row.get("负责人", "")),
        str(row.get("_SPU", "")),
        str(row.get("SKC", "")),
        str(row.get("SKU货号", "")),
    )


def read_source_rows():
    rules = active_slow_rules()
    new_slow_min_days = to_number(rules.get("new_slow_min_days", 30))
    new_slow_max_days = to_number(rules.get("new_slow_max_days", 60))
    old_slow_min_days = to_number(rules.get("old_slow_min_days", 180))
    source_date, source_files = latest_temu_sales_files()
    owners = load_owner_map()
    erp_records, erp_rows, combo_rows = load_erp_records()

    candidate_rows = []
    groups = defaultdict(lambda: {"rows": [], "sales30_total": 0, "days": []})
    source_count = 0
    skipped_no_days = 0
    skipped_no_spu = 0
    skipped_missing_required = 0
    erp_matched = 0
    erp_unmatched = 0

    required = {"店铺", "SKC", "SKU ID", "SKU货号", "申报价格", "7天销量", "30天销量", "可用", "加入站点时长"}
    for path in source_files:
        rows = iter_xlsx_rows(path)
        try:
            headers = next(rows)
        except StopIteration:
            continue
        hm = header_map(headers)
        if not required.issubset(hm):
            missing = sorted(required - set(hm))
            raise RuntimeError(f"{path.name} 缺少必要字段：{', '.join(missing)}")
        for source_row_num, values in enumerate(rows, start=2):
            source_count += 1
            store = canonical_store(values[hm["店铺"]] if hm["店铺"] < len(values) else "")
            sku = norm(values[hm["SKU货号"]] if hm["SKU货号"] < len(values) else "")
            days = values[hm["加入站点时长"]] if hm["加入站点时长"] < len(values) else ""
            if not store or not sku:
                skipped_missing_required += 1
                continue
            if days is None or days == "":
                skipped_no_days += 1
                continue
            days_num = to_number(days)
            spu = spu_key(sku)
            if not spu:
                skipped_no_spu += 1
                continue

            sales30 = to_number(values[hm["30天销量"]] if hm["30天销量"] < len(values) else 0)
            sales7 = to_number(values[hm["7天销量"]] if hm["7天销量"] < len(values) else 0)
            stock = to_number(values[hm["可用"]] if hm["可用"] < len(values) else 0)
            erp = erp_records.get(sku_key(sku))
            if erp:
                erp_matched += 1
            else:
                erp_unmatched += 1
                erp = {}

            row = {
                "店铺": store,
                "负责人": owners.get(store, ""),
                "ERP货品名称": erp.get("ERP货品名称") or norm(values[hm.get("产品名称", -1)] if hm.get("产品名称", -1) >= 0 and hm["产品名称"] < len(values) else ""),
                "规格名称": erp.get("规格名称") or norm(values[hm.get("SKU属性", -1)] if hm.get("SKU属性", -1) >= 0 and hm["SKU属性"] < len(values) else ""),
                "SKC": norm(values[hm["SKC"]] if hm["SKC"] < len(values) else ""),
                "SKU ID": norm(values[hm["SKU ID"]] if hm["SKU ID"] < len(values) else ""),
                "SKU货号": sku,
                "申报价": values[hm["申报价格"]] if hm["申报价格"] < len(values) else "",
                "7天销量": sales7,
                "30天销量": sales30,
                "平台仓库可用": stock,
                "上架天数": days_num,
                "日均销量": sales30 / 30 if sales30 else 0,
                "成本价": erp.get("成本价", ""),
                "批发报价": erp.get("批发报价", ""),
                "ERP可用库存(仅展示)": to_number(erp.get("ERP可用库存(仅展示)", 0)),
                "预警类型": "",
                "建议动作": "",
                "源文件": path.name,
                "源行": source_row_num,
                "_SPU": spu,
            }
            candidate_rows.append(row)

            group = groups[(row["店铺"], spu)]
            group["rows"].append(row)
            group["sales30_total"] += sales30
            group["days"].append(days_num)

    new_rows = []
    old_rows = []
    skipped_spu_has_sales = 0
    skipped_spu_mixed_days = 0

    for group in groups.values():
        if group["sales30_total"] != 0:
            skipped_spu_has_sales += len(group["rows"])
            continue
        min_days = min(group["days"]) if group["days"] else 0
        max_days = max(group["days"]) if group["days"] else 0
        if min_days > new_slow_min_days and max_days < new_slow_max_days:
            target = new_rows
            action = f"新品滞销：上架超过{int(new_slow_min_days)}天且小于{int(new_slow_max_days)}天，SPU近30天无销量，检查曝光、价格、标题/主图，必要时优化后继续观察"
        elif min_days > old_slow_min_days:
            target = old_rows
            action = f"老品滞销：上架超过{int(old_slow_min_days)}天且SPU近30天无销量，建议下架、清仓或合并库存处理"
        else:
            skipped_spu_mixed_days += len(group["rows"])
            continue

        for row in group["rows"]:
            row["建议动作"] = action
            target.append(row)

    new_rows.sort(key=store_sort_key)
    old_rows.sort(key=store_sort_key)
    return new_rows, old_rows, {
        "源表记录数": source_count,
        "源数据日期": source_date,
        "源文件": "、".join(path.name for path in source_files),
        "跳过：无上架天数": skipped_no_days,
        "跳过：无SPU前缀": skipped_no_spu,
        "跳过：缺少店铺或SKU": skipped_missing_required,
        "参与SPU判断明细行数": len(candidate_rows),
        "参与SPU判断组数": len(groups),
        "新品滞销上架天数超过": new_slow_min_days,
        "新品滞销上架天数小于": new_slow_max_days,
        "老品滞销上架天数超过": old_slow_min_days,
        "跳过：SPU内30天有销量": skipped_spu_has_sales,
        "跳过：SPU上架天数不在新品/老款范围": skipped_spu_mixed_days,
        "ERP基础表读取行数": erp_rows,
        "ERP组合装读取行数": combo_rows,
        "ERP匹配明细行数": erp_matched,
        "ERP未匹配明细行数": erp_unmatched,
    }


def style_sheet(ws, title_fill="1F4E78"):
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=title_fill)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def set_widths(ws):
    limits = {
        "产品名称": 44,
        "ERP货品名称": 30,
        "规格名称": 22,
        "建议动作": 34,
        "预警类型": 28,
        "源文件": 28,
    }
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        header = column_cells[0].value
        max_len = len(str(header or ""))
        for cell in column_cells[1:120]:
            value = cell.value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        width = min(max(max_len + 2, 10), limits.get(header, 18))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_detail_sheet(wb, title, rows, fill):
    ws = wb.create_sheet(title)
    ws.append(DETAIL_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in DETAIL_COLUMNS])
    style_sheet(ws, fill)
    set_widths(ws)
    return ws


def build_summary_rows(new_rows, old_rows):
    summary = defaultdict(lambda: {
        "负责人": "",
        "新品SKU数": 0,
        "新品SKC": set(),
        "新品平台库存": 0,
        "老款SKU数": 0,
        "老款SKC": set(),
        "老款平台库存": 0,
    })

    for row in new_rows:
        item = summary[row["店铺"]]
        item["负责人"] = item["负责人"] or row.get("负责人", "")
        item["新品SKU数"] += 1
        item["新品SKC"].add(row.get("SKC", ""))
        item["新品平台库存"] += to_number(row.get("平台仓库可用"))

    for row in old_rows:
        item = summary[row["店铺"]]
        item["负责人"] = item["负责人"] or row.get("负责人", "")
        item["老款SKU数"] += 1
        item["老款SKC"].add(row.get("SKC", ""))
        item["老款平台库存"] += to_number(row.get("平台仓库可用"))

    rows = []
    for store, item in summary.items():
        rows.append([
            store,
            item["负责人"],
            item["新品SKU数"],
            len(item["新品SKC"] - {""}),
            item["新品平台库存"],
            item["老款SKU数"],
            len(item["老款SKC"] - {""}),
            item["老款平台库存"],
            item["新品SKU数"] + item["老款SKU数"],
            len((item["新品SKC"] | item["老款SKC"]) - {""}),
            item["新品平台库存"] + item["老款平台库存"],
        ])
    rows.sort(key=lambda row: STORE_RANK.get(row[0], 999))
    return rows


def add_summary_sheet(wb, new_rows, old_rows):
    ws = wb.active
    ws.title = "总览"
    ws.append(["Temu滞销品每周报表", "", "", "", "", "", "", "", "", "", ""])
    ws.merge_cells("A1:K1")
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}", "", "", "", "", "", "", "", "", "", ""])
    ws.merge_cells("A2:K2")
    ws.append([])
    ws.append([
        "店铺",
        "负责人",
        "新品SKU数",
        "新品SKC数",
        "新品平台库存",
        "老款SKU数",
        "老款SKC数",
        "老款平台库存",
        "合计SKU数",
        "合计SKC数",
        "合计平台库存",
    ])
    for row in build_summary_rows(new_rows, old_rows):
        ws.append(row)

    total_row = ws.max_row + 1
    ws.cell(total_row, 1, "合计")
    ws.cell(total_row, 2, "")
    for col in range(3, 12):
        ws.cell(total_row, col, f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{total_row - 1})")

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=4, max_row=total_row, max_col=11):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    for cell in ws[total_row]:
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
        cell.font = Font(bold=True)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:K{total_row}"
    widths = [16, 12, 12, 12, 14, 12, 12, 14, 12, 12, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_validation_sheet(wb, new_rows, old_rows, source_stats):
    ws = wb.create_sheet("数据校验")
    ws.append(["检查项", "结果", "说明"])
    rows = [
        ["数据源日期", source_stats["源数据日期"], "读取 temu数据源表 中最新日期的 Temu仓库销售情况导出"],
        ["数据源文件", source_stats["源文件"], "读取最细 SKU 明细源表"],
        ["新品筛选规则", "同店铺同SPU 30天销量合计=0，且上架天数大于30天、小于60天", "按SKU货号提取SPU前缀后分组判断，明细仍按SKU展开"],
        ["老款筛选规则", "同店铺同SPU 30天销量合计=0，且组内最小上架天数>180", "按SKU货号提取SPU前缀后分组判断，明细仍按SKU展开"],
        ["新品明细行数", len(new_rows), ""],
        ["新品去重SKC数", len({row.get("SKC") for row in new_rows if row.get("SKC")}), ""],
        ["新品去重SPU数", len({row.get("_SPU") for row in new_rows if row.get("_SPU")}), ""],
        ["老款明细行数", len(old_rows), ""],
        ["老款去重SKC数", len({row.get("SKC") for row in old_rows if row.get("SKC")}), ""],
        ["老款去重SPU数", len({row.get("_SPU") for row in old_rows if row.get("_SPU")}), ""],
        ["源表记录数", source_stats["源表记录数"], ""],
        ["跳过：无上架天数", source_stats["跳过：无上架天数"], ""],
        ["跳过：无SPU前缀", source_stats["跳过：无SPU前缀"], ""],
        ["跳过：缺少店铺或SKU", source_stats["跳过：缺少店铺或SKU"], ""],
        ["参与SPU判断明细行数", source_stats["参与SPU判断明细行数"], ""],
        ["参与SPU判断组数", source_stats["参与SPU判断组数"], ""],
        ["新品滞销定义", f"上架天数超过{int(source_stats['新品滞销上架天数超过'])}天且小于{int(source_stats['新品滞销上架天数小于'])}天，且同店铺同SPU近30天总销量为0", ""],
        ["老品滞销定义", f"上架天数超过{int(source_stats['老品滞销上架天数超过'])}天，且同店铺同SPU近30天总销量为0", ""],
        ["跳过：SPU内30天有销量", source_stats["跳过：SPU内30天有销量"], ""],
        ["跳过：SPU上架天数不在新品/老款范围", source_stats["跳过：SPU上架天数不在新品/老款范围"], ""],
        ["ERP基础表读取行数", source_stats["ERP基础表读取行数"], ""],
        ["ERP组合装读取行数", source_stats["ERP组合装读取行数"], ""],
        ["ERP匹配明细行数", source_stats["ERP匹配明细行数"], ""],
        ["ERP未匹配明细行数", source_stats["ERP未匹配明细行数"], "未匹配时用源表产品名称/SKU属性补充展示，成本与ERP库存留空或为0"],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws, "666666")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50


def main():
    new_rows, old_rows, source_stats = read_source_rows()

    wb = Workbook()
    add_summary_sheet(wb, new_rows, old_rows)
    add_detail_sheet(wb, "新品滞销", new_rows, "5B9BD5")
    add_detail_sheet(wb, "老品滞销", old_rows, "A64D79")
    add_validation_sheet(wb, new_rows, old_rows, source_stats)

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb.save(OUTPUT)
    print(OUTPUT)
    print(f"新品行数={len(new_rows)} 老款行数={len(old_rows)}")


if __name__ == "__main__":
    main()
