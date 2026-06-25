import hashlib
import json
import os
import re
import secrets
import calendar
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel

from daily_ops_sales import DailySalesStore, sales_record_id


SHEIN_STORE_NAMES = {"琪琪", "童话", "加加", "宝宝", "牛牛", "二美", "元元"}
CHINESE_NUMERAL_STORES = {"一", "二", "三", "四", "五", "六", "七", "八", "九", "十"}
LONG_NUMERAL_STORES = {"十一", "十二", "十三", "十四", "十五", "十六", "十七", "十八", "十九", "二十"}
STORE_NAME_ALIASES = {"大额": "大鹅", "美美": "大美"}


def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def norm(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def norm_key(value):
    return re.sub(r"\s+", "", norm(value)).lower()


def platform_name(value, fallback="Temu"):
    text = norm(value) or fallback
    lower = text.lower()
    if lower == "temu":
        return "Temu"
    if lower == "shein":
        return "Shein"
    if lower in {"tk", "tiktok"}:
        return "TK"
    if lower == "ozon":
        return "Ozon"
    return text


def clean_store_name(value):
    text = re.sub(r"[（(].*?[）)]", "", norm(value)).strip()
    if not text:
        return ""
    if text in STORE_NAME_ALIASES:
        return STORE_NAME_ALIASES[text]
    if text == "大弟":
        return "一弟"
    if text in CHINESE_NUMERAL_STORES:
        return f"{text}弟"
    if text in LONG_NUMERAL_STORES:
        return text
    return text


def store_aliases(value):
    store = clean_store_name(value)
    aliases = {store} if store else set()
    if store == "一弟":
        aliases.add("大弟")
    if store.endswith("弟"):
        aliases.add(store[:-1])
    if store in LONG_NUMERAL_STORES:
        aliases.add(f"{store}弟")
    return {alias for alias in aliases if alias}


def header_map(row):
    return {norm_key(value): index for index, value in enumerate(row) if norm(value)}


def cell(row, mapping, *names):
    for name in names:
        index = mapping.get(norm_key(name))
        if index is not None and index < len(row):
            return row[index]
    return ""


def parse_owner_workbook(path):
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    assignments = []
    account_names = []
    seen_assignments = set()
    seen_accounts = set()
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration:
                continue
            mapping = header_map(headers)
            if "店铺" not in mapping or "业务" not in mapping:
                continue
            for row in rows:
                store = clean_store_name(cell(row, mapping, "店铺"))
                owner = norm(cell(row, mapping, "业务", "负责人", "店长"))
                if not store or not owner:
                    continue
                platform = platform_name(cell(row, mapping, "平台"), ws.title)
                item = {
                    "platform": platform,
                    "store": store,
                    "store_name": norm(cell(row, mapping, "店名", "店铺名称")),
                    "owner": owner,
                    "enabled": True,
                    "daily_required": True,
                }
                key = (item["platform"], item["store"])
                if key not in seen_assignments:
                    seen_assignments.add(key)
                    assignments.append(item)
                if owner not in seen_accounts:
                    seen_accounts.add(owner)
                    account_names.append({"owner": owner, "username": owner, "role": "owner", "enabled": True})
    finally:
        wb.close()
    return {"assignments": assignments, "accounts": account_names}


def random_password():
    alphabet = "abcdefghijkmnpqrstuvwxyzABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return "".join(secrets.choice(alphabet) for _ in range(10))


def hash_password(password, salt=None):
    salt = salt or secrets.token_hex(16)
    digest = hashlib.sha256(f"{salt}:{password}".encode("utf-8")).hexdigest()
    return salt, digest


def load_operator_accounts(path):
    path = Path(path)
    if not path.exists():
        return {"accounts": []}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"accounts": []}
    accounts = payload.get("accounts") if isinstance(payload, dict) else []
    return {"accounts": accounts if isinstance(accounts, list) else []}


def save_operator_accounts(path, accounts, password_factory=None):
    path = Path(path)
    existing = {norm(row.get("username")): row for row in load_operator_accounts(path).get("accounts", []) if norm(row.get("username"))}
    saved = []
    initial_passwords = {}
    for account in accounts or []:
        username = norm(account.get("username")) or norm(account.get("owner"))
        owner = norm(account.get("owner")) or username
        if not username or username in {row.get("username") for row in saved}:
            continue
        current = dict(existing.get(username) or {})
        if not current.get("password_hash"):
            password = password_factory(account) if password_factory else random_password()
            salt, digest = hash_password(password)
            current["password_salt"] = salt
            current["password_hash"] = digest
            initial_passwords[username] = password
        current.update({
            "role": account.get("role", current.get("role", "owner")),
            "owner": owner,
            "username": username,
            "enabled": account.get("enabled", current.get("enabled", True)) is not False,
            "updated_at": now_text(),
        })
        saved.append(current)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps({"accounts": saved, "updated_at": now_text()}, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"accounts": saved, "initial_passwords": initial_passwords}


def reset_operator_password(path, username, password=None):
    username = norm(username)
    if not username:
        raise ValueError("请填写账号")
    payload = load_operator_accounts(path)
    account = next((row for row in payload["accounts"] if norm(row.get("username")) == username), None)
    if not account:
        raise ValueError("账号不存在")
    password = password or random_password()
    salt, digest = hash_password(password)
    account["password_salt"] = salt
    account["password_hash"] = digest
    account["updated_at"] = now_text()
    Path(path).write_text(json.dumps({"accounts": payload["accounts"], "updated_at": now_text()}, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"username": username, "initial_password": password}


def excel_date(value, sheet_name=""):
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = norm(value)
    if not text:
        return ""
    if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", text):
        return datetime.strptime(text, "%Y-%m-%d").date().isoformat()
    try:
        serial = float(text)
        if serial > 30000:
            return from_excel(serial).date().isoformat()
    except (ValueError, TypeError):
        pass
    match = re.fullmatch(r"(\d{2})(\d{2})", sheet_name)
    if match and text.isdigit():
        return f"20{match.group(1)}-{match.group(2)}-{int(text):02d}"
    return ""


def sales_int(value):
    text = norm(value)
    if not text:
        return 0
    try:
        number = Decimal(text)
    except InvalidOperation:
        return 0
    if number < 0:
        return 0
    return int(number)


def platform_for_store(store, assignments):
    aliases = store_aliases(store)
    for item in assignments or []:
        if aliases & store_aliases(item.get("store")):
            return platform_name(item.get("platform"), "")
    clean = clean_store_name(store)
    if clean in SHEIN_STORE_NAMES:
        return "Shein"
    if clean.endswith("弟") or clean in LONG_NUMERAL_STORES:
        return "Temu"
    return "未设置"


def owner_for_store(store, assignments):
    aliases = store_aliases(store)
    for item in assignments or []:
        if aliases & store_aliases(item.get("store")):
            return norm(item.get("owner"))
    return ""


def parse_crossborder_sales_workbook(path, assignments=None):
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = []
    try:
        for ws in wb.worksheets:
            if not re.fullmatch(r"\d{4}", ws.title):
                continue
            sheet_rows = ws.iter_rows(values_only=True)
            try:
                headers = next(sheet_rows)
            except StopIteration:
                continue
            stores = []
            for index, raw in enumerate(headers):
                store = clean_store_name(raw)
                if index == 0 or not store or store in {"汇总", "时间"}:
                    stores.append("")
                else:
                    stores.append(store)
            for row in sheet_rows:
                day = excel_date(row[0] if row else "", ws.title)
                if not day:
                    continue
                for index, store in enumerate(stores):
                    if not store or index >= len(row):
                        continue
                    sales = sales_int(row[index])
                    rows.append({
                        "date": day,
                        "platform": platform_for_store(store, assignments or []),
                        "store": store,
                        "owner": owner_for_store(store, assignments or []),
                        "sales": sales,
                        "source": "历史导入",
                        "source_file": path.name,
                        "source_sheet": ws.title,
                    })
    finally:
        wb.close()
    return rows


def import_history_sales_records(sales_path, rows, actor="管理员"):
    store = DailySalesStore(sales_path)
    payload = store.load()
    records = payload["records"]
    existing = {row.get("id") for row in records}
    created = 0
    skipped_existing = 0
    imported_at = now_text()
    for row in rows or []:
        record_id = sales_record_id(row.get("date"), row.get("platform"), row.get("store"))
        if record_id in existing:
            skipped_existing += 1
            continue
        item = {
            "id": record_id,
            "date": row.get("date"),
            "platform": row.get("platform"),
            "store": row.get("store"),
            "owner": row.get("owner", ""),
            "sales": int(row.get("sales") or 0),
            "status": "历史导入",
            "abnormal": "",
            "remark": "从跨境运营总表导入",
            "submitted_by": actor,
            "submitted_at": imported_at,
            "updated_at": imported_at,
            "source": row.get("source", "历史导入"),
            "source_file": row.get("source_file", ""),
            "source_sheet": row.get("source_sheet", ""),
            "imported_at": imported_at,
            "history": [{
                "time": imported_at,
                "actor": actor,
                "action": "导入历史销量",
                "old_sales": "",
                "new_sales": int(row.get("sales") or 0),
                "remark": row.get("source_file", ""),
            }],
        }
        records.append(item)
        existing.add(record_id)
        created += 1
    payload["updated_at"] = imported_at
    store.save(payload)
    return {"created": created, "skipped_existing": skipped_existing, "total": len(rows or [])}


def date_in_range(value, date_from="", date_to=""):
    text = norm(value)
    return (not date_from or text >= date_from) and (not date_to or text <= date_to)


def parse_day(value, fallback=None):
    text = norm(value)
    if not text:
        return fallback
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return fallback


def add_months(day, months):
    month_index = day.month - 1 + months
    year = day.year + month_index // 12
    month = month_index % 12 + 1
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, min(day.day, last_day))


def period_tuple(start, end):
    return (start.isoformat(), end.isoformat())


def same_span_previous_year(start, end):
    return period_tuple(add_months(start, -12), add_months(end, -12))


def previous_span(start, end):
    days = (end - start).days + 1
    prev_end = start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=days - 1)
    return period_tuple(prev_start, prev_end)


def same_month_previous_period(end):
    start = end.replace(day=1)
    previous_start = add_months(start, -1)
    previous_end = add_months(end, -1)
    return period_tuple(start, end), period_tuple(previous_start, previous_end)


def year_to_date_period(end):
    start = end.replace(month=1, day=1)
    previous_start = add_months(start, -12)
    previous_end = add_months(end, -12)
    return period_tuple(start, end), period_tuple(previous_start, previous_end)


def build_assignment_index(assignments):
    index = {}
    for item in assignments or []:
        platform = norm(item.get("platform"))
        store = clean_store_name(item.get("store"))
        if not platform or not store:
            continue
        index[(platform, store)] = {
            "platform": platform,
            "store": store,
            "owner": norm(item.get("owner")),
            "enabled": item.get("enabled", True) is not False,
            "daily_required": item.get("daily_required", True) is not False,
        }
    return index


def scoped_sales_rows(sales_path, assignments=None, role="admin", user="", platform="", store="", allowed_pairs=None):
    platform = platform_name(platform, "") if platform else ""
    store = clean_store_name(store)
    user = norm(user)
    assignment_index = build_assignment_index(assignments)
    allowed = None
    if allowed_pairs is not None:
        allowed = {
            (norm(item_platform), clean_store_name(item_store))
            for item_platform, item_store in allowed_pairs
            if norm(item_platform) and clean_store_name(item_store)
        }
    elif norm(role) != "admin":
        allowed = {
            key for key, item in assignment_index.items()
            if item.get("owner") == user and item.get("enabled") and item.get("daily_required")
        }
    rows = []
    for row in DailySalesStore(sales_path).load()["records"]:
        row_platform = platform_name(row.get("platform"), "")
        row_store = clean_store_name(row.get("store"))
        pair = (row_platform, row_store)
        if allowed is not None and pair not in allowed:
            continue
        if platform and row_platform != platform:
            continue
        if store and row_store != store:
            continue
        item = dict(row)
        item["platform"] = row_platform
        item["store"] = row_store
        current_assignment = assignment_index.get(pair)
        if current_assignment and current_assignment.get("owner"):
            item["owner"] = current_assignment["owner"]
        else:
            item["owner"] = norm(item.get("owner")) or "未分配"
        rows.append(item)
    return rows


def sales_between(rows, date_from, date_to):
    return [row for row in rows if date_in_range(row.get("date"), date_from, date_to)]


def sum_sales(rows):
    return sum(int(row.get("sales") or 0) for row in rows)


def pct_change(current, base):
    if not base:
        return None
    return round((current - base) / base * 100, 2)


def comparison_summary(rows, current_period, compare_period):
    current_rows = sales_between(rows, *current_period)
    compare_rows = sales_between(rows, *compare_period)
    current_sales = sum_sales(current_rows)
    compare_sales = sum_sales(compare_rows)
    return {
        "sales": current_sales,
        "compare_sales": compare_sales,
        "delta": current_sales - compare_sales,
        "rate": pct_change(current_sales, compare_sales),
        "record_count": len(current_rows),
        "period": {"date_from": current_period[0], "date_to": current_period[1]},
        "compare_period": {"date_from": compare_period[0], "date_to": compare_period[1]},
    }


def complete_period(range_key, anchor_day):
    end = anchor_day - timedelta(days=1)
    if range_key == "month":
        return end.replace(day=1), end
    if range_key == "year":
        return end.replace(month=1, day=1), end
    if range_key == "7d":
        return end - timedelta(days=6), end
    if range_key == "14d":
        return end - timedelta(days=13), end
    return end - timedelta(days=29), end


def business_definition(label, period, source_label="店长填报销量"):
    return f"{label}：统计 {period[0]} 至 {period[1]} 的{source_label}，不含今日。"


def business_source_label(source):
    return "平台导入销量" if norm(source) == "platform" else "店长填报销量"


def completion_summary(rows, assignments, current_period, role="admin", user=""):
    start = parse_day(current_period[0])
    end = parse_day(current_period[1])
    days = max(0, (end - start).days + 1)
    assignment_index = build_assignment_index(assignments)
    required_pairs = [
        key for key, item in assignment_index.items()
        if item.get("enabled") and item.get("daily_required") and (norm(role) == "admin" or item.get("owner") == norm(user))
    ]
    if not required_pairs:
        required_pairs = sorted({
            (platform_name(row.get("platform"), ""), clean_store_name(row.get("store")))
            for row in rows
            if row.get("platform") and row.get("store")
        })
    submitted = {
        (norm(row.get("date")), platform_name(row.get("platform"), ""), clean_store_name(row.get("store")))
        for row in sales_between(rows, *current_period)
    }
    required = days * len(required_pairs)
    count = sum(1 for day_offset in range(days) for pair in required_pairs if ((start + timedelta(days=day_offset)).isoformat(), pair[0], pair[1]) in submitted)
    rate = round(count / required * 100, 2) if required else 100
    level = "ok" if rate >= 100 else "warn" if rate >= 90 else "red"
    return {"required": required, "submitted": count, "missing": max(0, required - count), "rate": rate, "level": level}


def business_group_key(row, dimension):
    platform = norm(row.get("platform")) or "未设置"
    store = clean_store_name(row.get("store")) or "未设置"
    owner = norm(row.get("owner")) or "未分配"
    if dimension == "platform":
        return platform, {"dimension": "platform", "name": platform, "platform": platform, "owner": "", "store": ""}
    if dimension == "owner":
        return owner, {"dimension": "owner", "name": owner, "platform": "", "owner": owner, "store": ""}
    return f"{platform}::{store}", {"dimension": "store", "name": store, "platform": platform, "owner": owner, "store": store}


def latest_date(rows, date_to=""):
    dates = [
        norm(row.get("date")) for row in rows
        if norm(row.get("date")) and (not date_to or norm(row.get("date")) <= date_to)
    ]
    return max(dates) if dates else ""


def group_business_rows(rows, dimension, current_period, compare_period, previous_period, end_day, stale_days=3, small_base=100):
    groups = {}
    for row in rows:
        key, base = business_group_key(row, dimension)
        groups.setdefault(key, {**base, "rows": []})["rows"].append(row)
    total_sales = sum_sales(sales_between(rows, *current_period))
    result = []
    for item in groups.values():
        group_rows = item.pop("rows")
        current_rows = sales_between(group_rows, *current_period)
        compare_rows = sales_between(group_rows, *compare_period)
        previous_rows = sales_between(group_rows, *previous_period)
        sales = sum_sales(current_rows)
        compare_sales_value = sum_sales(compare_rows)
        previous_sales = sum_sales(previous_rows)
        latest = latest_date(group_rows, end_day.isoformat())
        stale = False
        if latest:
            stale = (end_day - parse_day(latest, end_day)).days > stale_days
        item.update({
            "sales": sales,
            "compare_sales": compare_sales_value,
            "yoy_delta": sales - compare_sales_value,
            "yoy_rate": pct_change(sales, compare_sales_value),
            "previous_sales": previous_sales,
            "mom_delta": sales - previous_sales,
            "mom_rate": pct_change(sales, previous_sales),
            "share": round(sales / total_sales * 100, 2) if total_sales else 0,
            "record_count": len(current_rows),
            "latest_date": latest,
            "status": "超过3天未更新" if stale else "正常",
            "stale": stale,
            "base_too_small": compare_sales_value < small_base,
        })
        result.append(item)
    return sorted(result, key=lambda row: (-row["sales"], row["name"]))


def bucket_key(day_text, grain):
    if grain == "year":
        return day_text[:4]
    if grain == "day":
        return day_text
    return day_text[:7]


def trend_rows(rows, dimension, current_period, grain="month"):
    current_rows = sales_between(rows, *current_period)
    buckets = sorted({bucket_key(norm(row.get("date")), grain) for row in current_rows if norm(row.get("date"))})
    groups = {}
    for row in current_rows:
        key, base = business_group_key(row, dimension)
        groups.setdefault(key, {**base, "total": 0, "values": defaultdict(int)})
        sales = int(row.get("sales") or 0)
        bucket = bucket_key(norm(row.get("date")), grain)
        groups[key]["total"] += sales
        groups[key]["values"][bucket] += sales
    table = []
    for item in groups.values():
        values = {bucket: item["values"].get(bucket, 0) for bucket in buckets}
        item["values"] = values
        table.append(item)
    return {"buckets": buckets, "rows": sorted(table, key=lambda row: (-row["total"], row["name"]))}


def scope_business_rows(records, assignments=None, role="admin", user="", platform="", store=""):
    platform = platform_name(platform, "") if platform else ""
    store = clean_store_name(store)
    user = norm(user)
    assignment_index = build_assignment_index(assignments)
    allowed = None
    if norm(role) != "admin":
        allowed = {
            key for key, item in assignment_index.items()
            if item.get("owner") == user and item.get("enabled") and item.get("daily_required")
        }
    rows = []
    for row in records or []:
        row_platform = platform_name(row.get("platform"), "")
        row_store = clean_store_name(row.get("store"))
        pair = (row_platform, row_store)
        if allowed is not None and pair not in allowed:
            continue
        if platform and row_platform != platform:
            continue
        if store and row_store != store:
            continue
        item = dict(row)
        item["platform"] = row_platform
        item["store"] = row_store
        current_assignment = assignment_index.get(pair)
        if current_assignment and current_assignment.get("owner"):
            item["owner"] = current_assignment["owner"]
        else:
            item["owner"] = norm(item.get("owner")) or "未分配"
        rows.append(item)
    return rows


def business_report(sales_path, assignments=None, role="admin", user="", date_from="", date_to="", platform="", store="", grain="month", stale_days=3, diff_percent=20, diff_units=20, small_base=100, range_key="30d", source="manual", anchor_date="", rows_override=None):
    rows = scope_business_rows(rows_override, assignments, role, user, platform, store) if rows_override is not None else scoped_sales_rows(sales_path, assignments, role, user, platform, store)
    available_days = [parse_day(row.get("date")) for row in rows if parse_day(row.get("date"))]
    anchor_day = parse_day(anchor_date) or date.today()
    if date_from or date_to:
        end_day = parse_day(date_to) or (max(available_days) if available_days else anchor_day - timedelta(days=1))
        default_start = end_day.replace(day=1)
        start_day = parse_day(date_from) or default_start
    else:
        start_day, end_day = complete_period(range_key, anchor_day)
    if start_day > end_day:
        start_day, end_day = end_day, start_day
    current_period = period_tuple(start_day, end_day)
    compare_period = same_span_previous_year(start_day, end_day)
    previous_period = previous_span(start_day, end_day)
    month_period, previous_month_period = same_month_previous_period(end_day)
    year_period, previous_year_period = year_to_date_period(end_day)
    visible_rows = sales_between(rows, *current_period)
    completion = completion_summary(rows, assignments, current_period, role, user)
    source_label = business_source_label(source)
    anomalies = []
    for dimension, label in [("platform", "平台"), ("owner", "负责人"), ("store", "店铺")]:
        for row in group_business_rows(rows, dimension, current_period, compare_period, previous_period, end_day, stale_days, small_base):
            if row.get("stale"):
                anomalies.append({
                    "type": "数据未更新",
                    "dimension": label,
                    "name": row.get("name"),
                    "latest_date": row.get("latest_date"),
                    "message": f"{label}「{row.get('name')}」超过 {stale_days} 天没有销量更新",
                })
    return {
        "filters": {
            "date_from": current_period[0],
            "date_to": current_period[1],
            "platform": platform_name(platform, "") if platform else "",
            "store": clean_store_name(store),
            "grain": grain,
            "role": norm(role) or "admin",
            "user": norm(user),
            "range_key": range_key,
            "source": source,
        },
        "settings": {
            "stale_days": stale_days,
            "diff_percent": diff_percent,
            "diff_units": diff_units,
            "small_base": small_base,
        },
        "summary": {
            "month": comparison_summary(rows, month_period, previous_month_period),
            "year": comparison_summary(rows, year_period, previous_year_period),
            "range": comparison_summary(rows, current_period, compare_period),
            "previous_range": comparison_summary(rows, current_period, previous_period),
            "completion": completion,
            "record_count": len(visible_rows),
            "store_count": len({(row.get("platform"), row.get("store")) for row in visible_rows}),
            "latest_date": latest_date(rows, end_day.isoformat()),
            "anomaly_count": len(anomalies),
        },
        "definitions": {
            "range": business_definition(f"最近{(end_day - start_day).days + 1}日销量", current_period, source_label),
            "previous_range": f"上期对比：当前范围 {current_period[0]} 至 {current_period[1]}，对比上一段同长度完整销售日 {previous_period[0]} 至 {previous_period[1]}。",
            "year_over_year": f"去年同期：当前范围 {current_period[0]} 至 {current_period[1]}，对比去年同日期范围 {compare_period[0]} 至 {compare_period[1]}。",
            "month": business_definition("本月累计", month_period, source_label),
            "year": business_definition("本年累计", year_period, source_label),
            "completion": f"店长填报完整度：当前统计范围内，应填店铺销售日中已填写的比例。当前范围 {current_period[0]} 至 {current_period[1]}。",
        },
        "dimensions": {
            "platform": group_business_rows(rows, "platform", current_period, compare_period, previous_period, end_day, stale_days, small_base),
            "owner": group_business_rows(rows, "owner", current_period, compare_period, previous_period, end_day, stale_days, small_base),
            "store": group_business_rows(rows, "store", current_period, compare_period, previous_period, end_day, stale_days, small_base),
        },
        "trends": {
            "platform": trend_rows(rows, "platform", current_period, grain),
            "owner": trend_rows(rows, "owner", current_period, grain),
            "store": trend_rows(rows, "store", current_period, grain),
        },
        "anomalies": anomalies[:100],
    }


def query_sales_report(sales_path, platform="", store="", date_from="", date_to="", allowed_pairs=None):
    platform = norm(platform)
    store = clean_store_name(store)
    allowed = None
    if allowed_pairs is not None:
        allowed = {
            (norm(item_platform), clean_store_name(item_store))
            for item_platform, item_store in allowed_pairs
            if norm(item_platform) and clean_store_name(item_store)
        }
    rows = []
    for row in DailySalesStore(sales_path).load()["records"]:
        row_pair = (norm(row.get("platform")), clean_store_name(row.get("store")))
        if allowed is not None and row_pair not in allowed:
            continue
        if platform and norm(row.get("platform")) != platform:
            continue
        if store and clean_store_name(row.get("store")) != store:
            continue
        if not date_in_range(row.get("date"), date_from, date_to):
            continue
        rows.append(dict(row))
    rows.sort(key=lambda row: (norm(row.get("date")), norm(row.get("platform")), norm(row.get("store"))))
    total_sales = sum(int(row.get("sales") or 0) for row in rows)
    by_platform = grouped_sales(rows, "platform")
    by_store = grouped_sales(rows, "store")
    days = {row.get("date") for row in rows if row.get("date")}
    return {
        "filters": {"platform": platform, "store": store, "date_from": date_from, "date_to": date_to},
        "summary": {
            "total_sales": total_sales,
            "record_count": len(rows),
            "day_count": len(days),
            "daily_average": round(total_sales / len(days), 2) if days else 0,
        },
        "by_platform": by_platform,
        "by_store": by_store,
        "rows": rows,
    }


def grouped_sales(rows, field):
    groups = defaultdict(lambda: {"sales": 0, "record_count": 0, "days": set()})
    for row in rows:
        key = norm(row.get(field)) or "未设置"
        groups[key]["sales"] += int(row.get("sales") or 0)
        groups[key]["record_count"] += 1
        if row.get("date"):
            groups[key]["days"].add(row.get("date"))
    result = []
    for key, item in groups.items():
        result.append({
            field: key,
            "sales": item["sales"],
            "record_count": item["record_count"],
            "day_count": len(item["days"]),
            "daily_average": round(item["sales"] / len(item["days"]), 2) if item["days"] else 0,
        })
    return sorted(result, key=lambda row: (-row["sales"], row[field]))


def export_sales_report(report, output_dir):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    path = output_dir / f"{stamp}-销量查询报表.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "查询汇总"
    ws.append(["总销量", report["summary"]["total_sales"]])
    ws.append(["记录数", report["summary"]["record_count"]])
    ws.append(["天数", report["summary"]["day_count"]])
    ws.append(["日均销量", report["summary"]["daily_average"]])
    pws = wb.create_sheet("按平台")
    pws.append(["平台", "销量", "记录数", "天数", "日均"])
    for row in report["by_platform"]:
        pws.append([row["platform"], row["sales"], row["record_count"], row["day_count"], row["daily_average"]])
    sws = wb.create_sheet("按店铺")
    sws.append(["店铺", "销量", "记录数", "天数", "日均"])
    for row in report["by_store"]:
        sws.append([row["store"], row["sales"], row["record_count"], row["day_count"], row["daily_average"]])
    dws = wb.create_sheet("每日明细")
    dws.append(["日期", "平台", "店铺", "负责人", "销量", "状态", "来源"])
    for row in report["rows"]:
        dws.append([row.get("date"), row.get("platform"), row.get("store"), row.get("owner"), row.get("sales"), row.get("status"), row.get("source")])
    wb.save(path)
    return {"file": path.name, "path": str(path)}


def monthly_backup_status(backup_dir, now=None):
    now = now or datetime.now()
    backup_dir = Path(backup_dir)
    prefix = now.strftime("%Y%m")
    exists = backup_dir.exists() and any(
        path.is_file() and path.name.startswith(prefix) and ("系统数据备份" in path.name or "运营状态备份" in path.name)
        for path in backup_dir.glob("*.zip")
    )
    return {
        "month": now.strftime("%Y-%m"),
        "backup_exists": bool(exists),
        "message": "" if exists else f"{now.strftime('%Y年%m月')}还没有生成系统数据备份，请管理员先备份当前系统数据。",
    }


def export_master_import_review(owner_data, sales_rows, accounts, output_path):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "店铺负责人"
    ws.append(["平台", "店铺", "店名", "负责人", "启用", "每日填报"])
    for row in owner_data.get("assignments", []):
        ws.append([row.get("platform"), row.get("store"), row.get("store_name"), row.get("owner"), "是", "是"])
    aws = wb.create_sheet("店长账号")
    aws.append(["负责人", "账号", "初始密码", "角色", "状态"])
    initial = accounts.get("initial_passwords", {})
    for row in accounts.get("accounts", []):
        aws.append([row.get("owner"), row.get("username"), initial.get(row.get("username"), "已生成过，请在系统内重置"), row.get("role"), "启用" if row.get("enabled") else "停用"])
    sws = wb.create_sheet("历史销量明细")
    sws.append(["日期", "平台", "店铺", "负责人", "销量", "来源文件", "来源sheet"])
    for row in sales_rows:
        sws.append([row.get("date"), row.get("platform"), row.get("store"), row.get("owner"), row.get("sales"), row.get("source_file"), row.get("source_sheet")])
    sumws = wb.create_sheet("导入统计")
    sumws.append(["项目", "数量"])
    sumws.append(["店铺负责人关系", len(owner_data.get("assignments", []))])
    sumws.append(["店长账号", len(accounts.get("accounts", []))])
    sumws.append(["历史销量明细", len(sales_rows)])
    wb.save(output_path)
    return {"path": str(output_path), "file": output_path.name}
