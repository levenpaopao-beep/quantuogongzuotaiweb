import html
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import update_shein_summary_30d_skc as raw_xlsx


ROOT = Path(__file__).resolve().parent
TEMU_DIR = ROOT / "temu数据源表"
ERP_DIR = ROOT / "erp数据源"
OUTPUT_DIR = ROOT / "outputs"
INPUT_FILE = None
SALES_FILES = None
HOT_FILES = None
ERP_FILES = None
COMBO_FILES = None

OUTPUT_HEADERS = [
    "商家编码",
    "货品名称",
    "规格名称",
    "店铺（显示编号）",
    "申报价",
    "建议价格",
    "是否通过",
    "平台在售链接数",
    "平台在售最低申报价",
    "平台在售最高月销量",
    "平台在售上架最长时间",
]

SIZE_RE = re.compile(r"(?i)([-_\s]*)(XXXXL|XXXL|XXL|XL|XS|2XL|3XL|4XL|5XL|XXS|S|M|L|MD)$")
STORE_NAME_TO_CODE = {
    "弟弟": "1",
    "一弟": "1",
    "二弟": "2",
    "三弟": "3",
    "四弟": "4",
    "五弟": "5",
    "六弟": "6",
    "七弟": "7",
    "八弟": "8",
    "九弟": "9",
    "九弟（喵喵）": "9",
    "十弟": "10",
    "十一弟": "11",
    "十二弟": "12",
    "十三弟": "13",
    "十三（节日）": "13",
    "十五弟": "15",
    "十五弟（毛衣）": "15",
}


def text(value):
    if value is None:
        return ""
    return html.unescape(str(value)).strip()


def number(value):
    raw = text(value).replace(",", "")
    raw = re.sub(r"[^\d.\-]", "", raw)
    if raw in {"", "-", ".", "-."}:
        return 0.0
    try:
        return float(raw)
    except ValueError:
        return 0.0


def money(value):
    value = round(float(value), 2)
    return int(value) if value.is_integer() else value


def money_text(value):
    return f"{float(value):.2f}"


def sku_code(value):
    return text(value).split("@", 1)[0].strip()


def product_code(value):
    code = sku_code(value).upper()
    previous = None
    while previous != code:
        previous = code
        code = SIZE_RE.sub("", code)
    return code.strip("-_ ")


def store_code(value):
    raw = text(value)
    if not raw:
        return ""
    if re.fullmatch(r"\d+(\.0)?", raw):
        return str(int(float(raw)))
    return STORE_NAME_TO_CODE.get(raw, raw)


def header_map(row):
    return {text(value): idx for idx, value in enumerate(row) if text(value)}


def get(row, headers, name):
    idx = headers.get(name)
    if idx is None or idx >= len(row):
        return ""
    return row[idx]


def latest_by_date(pattern, folder):
    files = sorted(folder.glob(pattern), key=lambda path: path.name)
    dated = defaultdict(list)
    for path in files:
        match = re.search(r"(20\d{6})", path.name)
        dated[match.group(1) if match else "00000000"].append(path)
    if not dated:
        return []
    return sorted(dated[max(dated)])


def latest_input_file():
    if INPUT_FILE:
        return Path(INPUT_FILE)
    candidates = sorted(ROOT.glob("*新品申报价格表*.xlsx"), key=lambda path: path.stat().st_mtime)
    if not candidates:
        raise RuntimeError("未找到新品申报价格表")
    return candidates[-1]


def next_output_file():
    today = datetime.now().strftime("%y%m%d")
    OUTPUT_DIR.mkdir(exist_ok=True)
    existing = sorted(OUTPUT_DIR.glob(f"{today}-temu议价回复-V*.xlsx"))
    version = len(existing) + 1
    return OUTPUT_DIR / f"{today}-temu议价回复-V{version}.xlsx"


def size_rank(code):
    raw = sku_code(code).upper()
    size = raw.rsplit("-", 1)[-1] if "-" in raw else ""
    order = {"XXS": 0, "XS": 1, "S": 2, "M": 3, "MD": 4, "L": 5, "XL": 6, "XXL": 7, "XXXL": 8}
    return (order.get(size, 99), raw)


def normalized_name_candidates(name):
    raw = text(name)
    if not raw:
        return []
    variants = [raw]
    replacements = {
        "黄色": "柠檬黄",
        "粉色": "粉",
        "红色": "大红",
        "军绿色": "军绿",
        "黑色": "黑",
        "灰色": "烟灰",
    }
    for source, target in replacements.items():
        if source in raw:
            variants.append(raw.replace(source, target))
    if raw.endswith("蓬蓬"):
        variants.append(raw + "裙")
    if "-" in raw:
        prefix, suffix = raw.split("-", 1)
        variants.append(prefix)
        variants.append(prefix + "-" + suffix.replace("色", ""))
    return list(dict.fromkeys(variants))


def read_raw_rows(path):
    return [[text(cell) for cell in row] for row in raw_xlsx.read_xlsx_rows(path)]


def load_input_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = header_map(rows[0])
    has_code_format = "商家编码" in headers
    required = {"货品名称", "店铺", "申报价"} if not has_code_format else {"商家编码", "货品名称", "店铺", "申报价"}
    missing = sorted(required - set(headers))
    if missing:
        raise RuntimeError(f"{path.name} 缺少必要字段：{', '.join(missing)}")
    result = []
    for source_row, row in enumerate(rows[1:], start=2):
        code = sku_code(get(row, headers, "商家编码")) if has_code_format else ""
        result.append(
            {
                "源行": source_row,
                "商家编码": code,
                "款号": product_code(code) if code else "",
                "货品名称": text(get(row, headers, "货品名称")),
                "规格名称": text(get(row, headers, "规格名称")),
                "店铺": store_code(get(row, headers, "店铺")),
                "原申报价": number(get(row, headers, "申报价")),
                "原建议价格": number(get(row, headers, "建议价格")),
            }
        )
    return result


def load_erp():
    records = {}
    name_index = defaultdict(list)

    def put(code_value, name, spec, cost, wholesale):
        code = sku_code(code_value)
        if not code:
            return
        clean_name = text(name)
        clean_spec = text(spec)
        item = {
            "货品名称": clean_name,
            "规格名称": clean_spec,
            "成本价": number(cost),
            "批发价": number(wholesale),
        }
        for key in {code.upper(), product_code(code)}:
            current = records.setdefault(key, {})
            for field, value in item.items():
                if value not in ("", 0, 0.0) and not current.get(field):
                    current[field] = value
        name_index[clean_name].append(
            {
                "商家编码": code.upper(),
                "货品名称": clean_name,
                "规格名称": clean_spec,
                "成本价": item["成本价"],
                "批发价": item["批发价"],
            }
        )

    for path in (ERP_FILES or sorted(ERP_DIR.glob("erp产品基础信息表*.xlsx"))):
        rows = read_raw_rows(path)
        if not rows:
            continue
        headers = header_map(rows[0])
        code_col = "商家编码（新）"
        wholesale_col = "批发报价" if "批发报价" in headers else "批发价"
        if code_col not in headers:
            continue
        for row in rows[1:]:
            put(
                get(row, headers, code_col),
                get(row, headers, "货品名称"),
                get(row, headers, "规格名称"),
                get(row, headers, "成本价"),
                get(row, headers, wholesale_col),
            )

    for combo in (COMBO_FILES or [ERP_DIR / "erp产品组合装基础信息表.xlsx"]):
        if not combo.exists():
            continue
        rows = read_raw_rows(combo)
        headers = header_map(rows[0]) if rows else {}
        if "商家编码（新）" in headers:
            for row in rows[1:]:
                put(
                    get(row, headers, "商家编码（新）"),
                    get(row, headers, "组合装名称"),
                    get(row, headers, "组合装简称"),
                    get(row, headers, "成本价"),
                    get(row, headers, "批发报价"),
                )
    for name, items in name_index.items():
        uniq = {}
        for item in items:
            uniq[item["商家编码"]] = item
        name_index[name] = sorted(uniq.values(), key=lambda item: size_rank(item["商家编码"]))
    return records, name_index


def aggregate_links(rows):
    grouped = {}
    for row in rows:
        key = (row["店铺"], row["款号"], row["SKC"])
        item = grouped.setdefault(
            key,
            {
                "申报价": None,
                "7天销量": 0.0,
                "30天销量": 0.0,
                "加入站点时长": 0.0,
            },
        )
        price = row["申报价"]
        if price > 0:
            item["申报价"] = price if item["申报价"] is None else min(item["申报价"], price)
        item["7天销量"] += row["7天销量"]
        item["30天销量"] += row["30天销量"]
        item["加入站点时长"] = max(item["加入站点时长"], row["加入站点时长"])

    result = defaultdict(list)
    for (shop, product, _skc), item in grouped.items():
        if item["申报价"] is not None:
            result[(shop, product)].append(item)
    return result


def load_temu_links():
    source_rows = []
    files = [Path(path) for path in SALES_FILES] if SALES_FILES else latest_by_date("*Temu仓库销售情况导出*.xlsx", TEMU_DIR)
    for path in files:
        rows = read_raw_rows(path)
        if not rows:
            continue
        headers = header_map(rows[0])
        required = {"SKC", "SKU货号", "申报价格", "7天销量", "30天销量", "加入站点时长", "店铺"}
        missing = required - set(headers)
        if missing:
            raise RuntimeError(f"{path.name} 缺少必要字段：{', '.join(sorted(missing))}")
        for row in rows[1:]:
            sku = sku_code(get(row, headers, "SKU货号"))
            shop = store_code(get(row, headers, "店铺"))
            skc = text(get(row, headers, "SKC"))
            if not sku or not shop or not skc:
                continue
            price = number(get(row, headers, "申报价格"))
            if price <= 0:
                continue
            source_rows.append(
                {
                    "店铺": shop,
                    "款号": product_code(sku),
                    "SKC": skc,
                    "申报价": price,
                    "7天销量": number(get(row, headers, "7天销量")),
                    "30天销量": number(get(row, headers, "30天销量")),
                    "加入站点时长": number(get(row, headers, "加入站点时长")),
                }
            )
    return aggregate_links(source_rows), [path.name for path in files]


def load_hot_keys():
    hot = set()
    files = [Path(path) for path in HOT_FILES] if HOT_FILES else latest_by_date("*Temu爆旺款*.xlsx", TEMU_DIR)
    if not files:
        files = latest_by_date("*Temu爆旺款*.xlsx", ROOT)
    for path in files:
        rows = read_raw_rows(path)
        if not rows:
            continue
        headers = header_map(rows[0])
        if "SKU货号" not in headers or "店铺" not in headers:
            continue
        for row in rows[1:]:
            sku = sku_code(get(row, headers, "SKU货号"))
            shop = store_code(get(row, headers, "店铺"))
            if sku and shop:
                hot.add((shop, product_code(sku)))
    return hot, [path.name for path in files]


def enrich_input_rows(input_rows, erp_by_code, erp_by_name):
    enriched = []
    for row in input_rows:
        current = dict(row)
        if not current.get("商家编码"):
            candidates = []
            for variant in normalized_name_candidates(current["货品名称"]):
                candidates = erp_by_name.get(variant, [])
                if candidates:
                    break
            if not candidates:
                prefix = current["货品名称"].split("-", 1)[0]
                prefix_hits = []
                for name, items in erp_by_name.items():
                    if name.startswith(prefix):
                        prefix_hits.extend(items)
                uniq = {}
                for item in prefix_hits:
                    uniq[item["商家编码"]] = item
                candidates = sorted(uniq.values(), key=lambda item: size_rank(item["商家编码"]))
            if not candidates and "-" not in current["货品名称"]:
                colors = ["宝蓝", "大红", "粉色", "粉", "军绿色", "军绿", "黑色", "黑", "白色", "白", "麻灰", "烟灰", "豆绿", "枣红", "柠檬黄", "黄色", "雾霾蓝", "粉红", "红粉"]
                for color in colors:
                    if current["货品名称"].endswith(color):
                        prefix = current["货品名称"][: -len(color)]
                        prefix_hits = []
                        for name, items in erp_by_name.items():
                            if name.startswith(prefix):
                                for item in items:
                                    if color.replace("色", "") in item["规格名称"] or color in item["规格名称"]:
                                        prefix_hits.append(item)
                        uniq = {}
                        for item in prefix_hits:
                            uniq[item["商家编码"]] = item
                        candidates = sorted(uniq.values(), key=lambda item: size_rank(item["商家编码"]))
                        if candidates:
                            break
            if candidates:
                chosen = candidates[0]
                current["商家编码"] = chosen["商家编码"]
                current["规格名称"] = current["规格名称"] or chosen["规格名称"]
                current["款号"] = product_code(chosen["商家编码"])
        if current.get("商家编码") and not current.get("款号"):
            current["款号"] = product_code(current["商家编码"])
        proposed_price = number(current.get("原建议价格"))
        if proposed_price <= 0:
            proposed_price = number(current.get("原申报价"))
        current["申报价"] = proposed_price
        enriched.append(current)
    return enriched


def evaluate_offer(offer):
    price = number(offer.get("申报价"))
    cost = number(offer.get("成本价"))
    wholesale = number(offer.get("批发价"))
    links = offer.get("在售链接") or []
    link_count = len(links)
    min_price = min((number(link.get("申报价")) for link in links if number(link.get("申报价")) > 0), default=0)

    if cost > 0 and price < cost:
        return {"是否通过": "拒绝上架-理由 亏损", "建议价格": money(cost)}

    if link_count > 0 and offer.get("有爆旺款"):
        return {"是否通过": "拒绝上架-理由 有爆旺款在售", "建议价格": ""}

    for link in links:
        link_price = number(link.get("申报价"))
        if number(link.get("7天销量")) > 10 and link_price > 0 and price < link_price * 0.95:
            return {"是否通过": "拒绝上架-理由 有预备爆款链接在售", "建议价格": money(link_price * 0.95)}

    if link_count > 7:
        return {"是否通过": "拒绝上架-理由 同时在架产品过多，15天以后再尝试上架", "建议价格": ""}

    if 3 < link_count <= 6 and min_price > 0 and price < min_price * 0.95:
        suggested = min_price * 0.95
        status = f"拒绝上架-理由 报价过低建议价格为{money_text(suggested)}"
        if cost > 0 and suggested < cost:
            status += f"；现在最低申报价是{money_text(min_price)}，接近亏损"
        return {"是否通过": status, "建议价格": money(suggested)}

    if wholesale > 0 and price < wholesale * 0.8:
        return {"是否通过": "拒绝上架-理由 破价", "建议价格": money(wholesale * 0.8)}

    return {"是否通过": "同意议价", "建议价格": ""}


def build_output_rows(input_rows, erp, links_by_key, hot_keys):
    output = []
    for row in input_rows:
        erp_row = erp.get(row["商家编码"].upper()) if row.get("商家编码") else None
        erp_row = erp_row or erp.get(row["款号"]) or {}
        links = links_by_key.get((row["店铺"], row["款号"]), []) if row.get("款号") else []
        min_price = min((number(link.get("申报价")) for link in links), default=0)
        max_month_sales = max((number(link.get("30天销量")) for link in links), default=0)
        max_days = max((number(link.get("加入站点时长")) for link in links), default=0)
        if not row.get("商家编码"):
            output.append(
                [
                    "",
                    row["货品名称"],
                    row["规格名称"],
                    row["店铺"],
                    row["申报价"],
                    "",
                    "需人工核查-未匹配ERP货品",
                    len(links),
                    money(min_price) if min_price else "",
                    money(max_month_sales) if max_month_sales else 0,
                    money(max_days) if max_days else "",
                ]
            )
            continue
        decision = evaluate_offer(
            {
                "申报价": row["申报价"],
                "成本价": erp_row.get("成本价", 0),
                "批发价": erp_row.get("批发价", 0),
                "在售链接": links,
                "有爆旺款": (row["店铺"], row["款号"]) in hot_keys,
            }
        )
        output.append(
            [
                row["商家编码"],
                row["货品名称"] or erp_row.get("货品名称", ""),
                row["规格名称"] or erp_row.get("规格名称", ""),
                row["店铺"],
                row["申报价"],
                decision["建议价格"],
                decision["是否通过"],
                len(links),
                money(min_price) if min_price else "",
                money(max_month_sales) if max_month_sales else 0,
                money(max_days) if max_days else "",
            ]
        )
    return output


def write_workbook(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "议价回复"
    ws.append(OUTPUT_HEADERS)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    widths = [18, 26, 18, 14, 12, 18, 48, 16, 18, 20, 22]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)


def main():
    input_path = latest_input_file()
    input_rows = load_input_rows(input_path)
    erp, erp_by_name = load_erp()
    input_rows = enrich_input_rows(input_rows, erp, erp_by_name)
    links_by_key, sales_files = load_temu_links()
    hot_keys, hot_files = load_hot_keys()
    output_rows = build_output_rows(input_rows, erp, links_by_key, hot_keys)
    output_path = next_output_file()
    write_workbook(output_rows, output_path)
    print(f"input={input_path.name}")
    print(f"rows={len(output_rows)}")
    print(f"sales_files={','.join(sales_files)}")
    print(f"hot_files={','.join(hot_files)}")
    print(f"output={output_path}")


if __name__ == "__main__":
    main()
