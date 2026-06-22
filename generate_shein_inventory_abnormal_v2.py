import os
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import update_shein_summary_30d_skc as xlsx


ROOT = Path(os.environ.get("DAILYWORK_ROOT", Path(__file__).resolve().parent))
SHEIN_DIR = ROOT / "shein数据源表"
ERP_DIR = ROOT / "erp数据源"
OUTPUT = ROOT / "SHEIN仓备库存异常清单_V2_20260603.xlsx"

STORE_ORDER = ["琪琪", "童话", "牛牛"]
OWNERS = {"琪琪": "胡娟", "童话": "胡娟", "牛牛": "胡娟"}
SIZE_ORDER = {"XXS": 0, "XS": 1, "S": 2, "M": 3, "MD": 4, "L": 5, "XL": 6, "XXL": 7}


def norm(value):
    return xlsx.norm_text(value)


def num(value):
    return xlsx.to_number(value)


def sku_key(value):
    text = norm(value).upper()
    if "@" in text:
        text = text.split("@", 1)[0]
    return text.strip()


def size_rank(code):
    text = norm(code).upper()
    parts = re.split(r"[-_\s]+", text)
    for part in reversed(parts):
        if part in SIZE_ORDER:
            return SIZE_ORDER[part]
    return 99


def base_style(ws):
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")
    for col in range(1, ws.max_column + 1):
        max_len = 8
        for row in range(1, min(ws.max_row, 300) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 36)
    ws.sheet_view.showGridLines = False


def setup_sheet(ws, title, headers, color):
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="44546A")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"


def header_map(row):
    return {norm(v): i for i, v in enumerate(row) if norm(v)}


def store_from_file(path):
    for store in STORE_ORDER:
        if store in path.stem:
            return store
    return ""


def load_erp_records():
    records = {}
    erp_rows = 0
    for path in sorted(ERP_DIR.glob("erp产品基础信息表*.xlsx")):
        rows = xlsx.read_xlsx_rows(path)
        if not rows:
            continue
        hm = header_map(rows[0])
        if not {"商家编码（新）", "货品名称", "规格名称"}.issubset(hm):
            continue
        for row in rows[1:]:
            erp_rows += 1
            code = norm(row[hm["商家编码（新）"]] if hm["商家编码（新）"] < len(row) else "")
            key = sku_key(code)
            if key and key not in records:
                records[key] = {
                    "商家编码": code,
                    "货品名称": norm(row[hm["货品名称"]] if hm["货品名称"] < len(row) else ""),
                    "货品规格": norm(row[hm["规格名称"]] if hm["规格名称"] < len(row) else ""),
                }

    combo_rows = 0
    combo = ERP_DIR / "erp产品组合装基础信息表.xlsx"
    if combo.exists():
        rows = xlsx.read_xlsx_rows(combo)
        hm = header_map(rows[0]) if rows else {}
        code_i = hm.get("商家编码（新）")
        name_i = hm.get("组合装名称")
        spec_i = hm.get("组合装简称")
        if code_i is not None:
            for row in rows[1:]:
                combo_rows += 1
                code = norm(row[code_i] if code_i < len(row) else "")
                key = sku_key(code)
                if key and key not in records:
                    records[key] = {
                        "商家编码": code,
                        "货品名称": norm(row[name_i] if name_i is not None and name_i < len(row) else ""),
                        "货品规格": norm(row[spec_i] if spec_i is not None and spec_i < len(row) else ""),
                    }
    return records, erp_rows, combo_rows


def read_shein_rows(erp):
    groups = defaultdict(list)
    summary = {s: {"skc": set(), "hot": set()} for s in STORE_ORDER}
    source_rows = 0
    skipped_unmatched = 0

    for path in sorted(SHEIN_DIR.glob("20260601-shein*.xlsx")):
        store = store_from_file(path)
        if store not in STORE_ORDER:
            continue
        rows = xlsx.read_xlsx_rows(path)
        if not rows:
            continue
        hm = header_map(rows[0])
        required = ["SKC", "SPU", "商家SKU", "SHEIN仓库存", "近30天销量", "近7天销量"]
        if not all(k in hm for k in required):
            raise RuntimeError(f"{path.name} 缺少必要字段")
        tag_i = hm.get("商品标签")
        for ridx, row in enumerate(rows[1:], start=2):
            source_rows += 1
            store = store_from_file(path)
            skc = norm(row[hm["SKC"]] if hm["SKC"] < len(row) else "")
            spu = norm(row[hm["SPU"]] if hm["SPU"] < len(row) else "")
            sku = norm(row[hm["商家SKU"]] if hm["商家SKU"] < len(row) else "")
            stock = num(row[hm["SHEIN仓库存"]] if hm["SHEIN仓库存"] < len(row) else "")
            sales30 = num(row[hm["近30天销量"]] if hm["近30天销量"] < len(row) else "")
            sales7 = num(row[hm["近7天销量"]] if hm["近7天销量"] < len(row) else "")
            tag = norm(row[tag_i] if tag_i is not None and tag_i < len(row) else "")
            if skc:
                summary[store]["skc"].add(skc)
            if skc and any(word in tag for word in ["高销款", "爆", "旺"]):
                summary[store]["hot"].add(skc)
            erp_row = erp.get(sku_key(sku))
            if not erp_row:
                skipped_unmatched += 1
                continue
            group_key = (store, spu or skc)
            groups[group_key].append(
                {
                    "店铺": store,
                    "SPU": spu,
                    "SKC": skc,
                    "商家编码": erp_row["商家编码"],
                    "货品名称": erp_row["货品名称"],
                    "货品规格": erp_row["货品规格"],
                    "仓备可用": stock,
                    "30天销量": sales30,
                    "7天销量": sales7,
                    "负责人": OWNERS.get(store, ""),
                    "源SKU": sku,
                    "源文件": path.name,
                    "源行": ridx,
                }
            )
    return groups, summary, source_rows, skipped_unmatched


def expand_group_alerts(groups, multiplier):
    rows = []
    alert_skc = defaultdict(set)
    alert_groups = set()
    for group_key, items in groups.items():
        link_stock_total = sum(item["仓备可用"] for item in items)
        if link_stock_total <= 30:
            continue
        has_alert = any(item["仓备可用"] > item["30天销量"] * multiplier for item in items)
        if not has_alert:
            continue
        alert_groups.add(group_key)
        store = group_key[0]
        for item in items:
            out = dict(item)
            out["链接备货总数量"] = link_stock_total
            out["触发规则"] = f"组内任一尺码仓备可用 > 30天销量 * {multiplier}，且链接备货总数量 > 30"
            rows.append(out)
            if out["SKC"]:
                alert_skc[store].add(out["SKC"])
    return rows, alert_skc, alert_groups


def sort_rows(rows):
    store_rank = {s: i for i, s in enumerate(STORE_ORDER)}
    return sorted(
        rows,
        key=lambda r: (
            store_rank.get(r["店铺"], 99),
            norm(r["SKC"]),
            re.sub(r"[-_\s]+(XXS|XS|S|M|MD|L|XL|XXL)$", "", norm(r["商家编码"]).upper()),
            size_rank(r["商家编码"]),
            norm(r["商家编码"]).upper(),
        ),
    )


def write_detail(wb, sheet_name, title, rows, color):
    headers = [
        "店铺", "SPU", "SKC", "商家编码", "货品名称", "货品规格", "仓备可用",
        "链接备货总数量", "30天销量", "7天销量", "负责人", "源SKU", "源文件", "源行", "触发规则",
    ]
    ws = wb.create_sheet(sheet_name)
    setup_sheet(ws, title, headers, color)
    for row in sort_rows(rows):
        ws.append([row[h] for h in headers])
    for row in ws.iter_rows(min_row=3, min_col=7, max_col=10):
        for cell in row:
            cell.number_format = "0"
    base_style(ws)


def build_workbook(summary, rows_2x, skc_2x, rows_1x, skc_1x, source_rows, erp_rows, combo_rows, skipped, group_count_2x, group_count_1x):
    wb = Workbook()
    ws = wb.active
    ws.title = "总览表"
    headers = ["店铺", "负责人", "在售SKC数量", "爆旺款数量", "仓备大于30天销量2倍SKC数", "仓备大于30天销量SKC数"]
    setup_sheet(ws, "SHEIN仓备库存异常清单V2", headers, "5B9BD5")
    for store in STORE_ORDER:
        ws.append([
            store,
            OWNERS.get(store, ""),
            len(summary[store]["skc"]),
            len(summary[store]["hot"]),
            len(skc_2x.get(store, set())),
            len(skc_1x.get(store, set())),
        ])
    base_style(ws)

    write_detail(wb, "大于30天销量2倍", "组内任一尺码仓备可用 > 30天销量 * 2，列出整个SPU", rows_2x, "C00000")
    write_detail(wb, "大于30天销量", "组内任一尺码仓备可用 > 30天销量，列出整个SPU", rows_1x, "ED7D31")

    check = wb.create_sheet("数据校验")
    for row in [
        ["项目", "数值"],
        ["Shein源明细行数（仅琪琪/童话/牛牛）", source_rows],
        ["ERP基础表读取行数", erp_rows],
        ["ERP组合装读取行数", combo_rows],
        ["未匹配ERP/组合装跳过行数", skipped],
        ["大于30天销量2倍明细行数（SPU展开后）", len(rows_2x)],
        ["大于30天销量明细行数（SPU展开后）", len(rows_1x)],
        ["大于30天销量2倍触发SPU组数", group_count_2x],
        ["大于30天销量触发SPU组数", group_count_1x],
        ["剔除店铺", "加加、宝宝"],
        ["库存口径", "SHEIN仓库存"],
        ["触发门槛", "同一店铺同一SPU/SKC的链接备货总数量 > 30"],
        ["展开口径", "组内任一尺码命中预警，则列出该SPU内所有可匹配ERP的码数"],
        ["排序规则", "店铺、SKC、商家编码尺码（XXS-XS-S-M-MD-L-XL-XXL）"],
    ]:
        check.append(row)
    base_style(check)

    wb.save(OUTPUT)


if __name__ == "__main__":
    erp, erp_rows, combo_rows = load_erp_records()
    groups, summary, source_rows, skipped = read_shein_rows(erp)
    rows_2x, skc_2x, groups_2x = expand_group_alerts(groups, 2)
    rows_1x, skc_1x, groups_1x = expand_group_alerts(groups, 1)
    build_workbook(
        summary, rows_2x, skc_2x, rows_1x, skc_1x,
        source_rows, erp_rows, combo_rows, skipped, len(groups_2x), len(groups_1x)
    )
    print(OUTPUT)
    print("2x rows", len(rows_2x), "groups", len(groups_2x))
    print("1x rows", len(rows_1x), "groups", len(groups_1x))
    for store in STORE_ORDER:
        print(store, len(summary[store]["skc"]), len(summary[store]["hot"]), len(skc_2x.get(store, set())), len(skc_1x.get(store, set())))
