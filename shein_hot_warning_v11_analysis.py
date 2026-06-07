import json
import os
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path.cwd()
OUT_DIR = ROOT / "outputs" / "shein_hot_warning_v11"
OUT_JSON = OUT_DIR / "shein_hot_warning_v11_data.json"
REPORT_VERSION = os.environ.get("SHEIN_HOT_VERSION", "V1.1正式版")

SOURCE_FILES = {}
STORE_NO = {"琪琪": "琪琪", "童话": "童话", "加加": "加加", "宝宝": "宝宝", "牛牛": "牛牛"}
STORE_OWNER = {"琪琪": "胡娟", "童话": "胡娟", "加加": "洁琳", "宝宝": "洁琳", "牛牛": "胡娟"}
STORE_ORDER = ["琪琪", "童话", "加加", "宝宝", "牛牛"]
ERP_FILES = sorted((ROOT / "erp数据源").glob("erp产品基础信息表*.xlsx"))
COMBO_FILES = sorted((ROOT / "erp数据源").glob("erp产品组合装基础信息表.xlsx"))
RULES = None

SIZE_RE = re.compile(r"^(33\d+)-(XS|S|M|MD|L|XL|XXL)$", re.I)
SIZE_ORDER = {"XS": 1, "S": 2, "M": 3, "MD": 4, "L": 5, "XL": 6, "XXL": 7}

DEFAULT_HOT_RULES = {
    "shein_new_days_lt": 30,
    "shein_new_7d_daily_gte": 10,
    "shein_old_days_gte": 30,
    "shein_old_30d_daily_gt": 20,
}


def txt(value):
    return "" if value is None else str(value).strip()


def num(value):
    if value is None or value == "":
        return 0.0


def active_hot_rules():
    source = RULES if isinstance(RULES, dict) else {}
    current = source.get("hot_item", source) if isinstance(source, dict) else {}
    merged = DEFAULT_HOT_RULES.copy()
    if isinstance(current, dict):
        merged.update(current)
    return merged
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def norm_sku(value):
    raw = txt(value)
    base = raw.split("@", 1)[0].strip().upper()
    m = SIZE_RE.match(base)
    if not m:
        return None, None, None
    return base, m.group(1), m.group(2).upper()


def latest_store_file(store):
    candidates = sorted(
        [path for path in (ROOT / "shein数据源表").glob("*.xlsx") if store in path.stem],
        key=lambda path: (path.stat().st_mtime, path.name),
    )
    return candidates[-1] if candidates else None


def source_files():
    configured = SOURCE_FILES or {store: latest_store_file(store) for store in STORE_ORDER}
    normalized = {}
    for store, value in configured.items():
        values = value if isinstance(value, (list, tuple)) else [value]
        normalized[store] = [Path(path) for path in values if path and Path(path).exists()]
    missing = [store for store, paths in normalized.items() if not paths]
    if missing:
        raise FileNotFoundError("未找到 Shein 店铺数据源：" + "、".join(missing))
    return normalized


def load_erp_file(path):
    data = {}
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb.worksheets[0]
    headers = [txt(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    code_col = idx.get("商家编码（新）") or idx.get("商家编码")
    if not code_col:
        wb.close()
        return data
    for r in range(2, ws.max_row + 1):
        code = txt(ws.cell(r, code_col).value).upper()
        if not code:
            continue
        data[code] = {
            "货品名称": txt(ws.cell(r, idx.get("货品名称", 0)).value) if idx.get("货品名称") else "",
            "规格名称": txt(ws.cell(r, idx.get("规格名称", 0)).value) if idx.get("规格名称") else "",
            "成本价": num(ws.cell(r, idx.get("成本价", 0)).value) if idx.get("成本价") else 0,
            "批发报价": num(ws.cell(r, idx.get("批发报价", 0)).value) if idx.get("批发报价") else 0,
            "来源": path.name,
        }
    wb.close()
    return data


def load_erp():
    erp = {}
    for path in ERP_FILES:
        erp.update({k: v for k, v in load_erp_file(path).items() if k not in erp})
    for path in COMBO_FILES:
        erp.update({k: v for k, v in load_erp_file(path).items() if k not in erp})
    return erp


def load_sales_records():
    erp = load_erp()
    sources = source_files()
    records = []
    unmatched = 0
    invalid_sku = 0
    for store in STORE_ORDER:
        for path in sources[store]:
            wb = load_workbook(path, read_only=False, data_only=True)
            ws = wb.worksheets[0]
            headers = [txt(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
            idx = {h: i + 1 for i, h in enumerate(headers)}
            price_col = idx.get("供货价") or idx.get("采购价")
            price_field = "供货价" if idx.get("供货价") else "采购价"
            for r in range(2, ws.max_row + 1):
                raw_sku = ws.cell(r, idx.get("商家SKU", 0)).value if idx.get("商家SKU") else ""
                sku, style, size = norm_sku(raw_sku)
                if not sku:
                    invalid_sku += 1
                    continue
                hit = erp.get(sku)
                if not hit:
                    unmatched += 1
                    continue
                records.append(
                    {
                        "店铺": store,
                        "店铺编号": STORE_NO[store],
                        "负责人": STORE_OWNER[store],
                        "源文件": path.name,
                        "源行": r,
                        "商品名称": txt(ws.cell(r, idx.get("供货方号", 0)).value) if idx.get("供货方号") else "",
                        "SKC": txt(ws.cell(r, idx.get("SKC", 0)).value) if idx.get("SKC") else "",
                        "SPU": txt(ws.cell(r, idx.get("SPU", 0)).value) if idx.get("SPU") else "",
                        "原始商家SKU": txt(raw_sku),
                        "商家编码": sku,
                        "货品编码": style,
                        "尺码": size,
                        "申报价": num(ws.cell(r, price_col).value) if price_col else 0,
                        "申报价字段": price_field,
                        "7天销量": num(ws.cell(r, idx.get("近7天销量", 0)).value) if idx.get("近7天销量") else 0,
                        "30天销量": num(ws.cell(r, idx.get("近30天销量", 0)).value) if idx.get("近30天销量") else 0,
                        "平台仓备货库存": num(ws.cell(r, idx.get("SHEIN仓库存", 0)).value) if idx.get("SHEIN仓库存") else 0,
                        "上架天数": num(ws.cell(r, idx.get("上架天数", 0)).value) if idx.get("上架天数") else 0,
                        "ERP货品名称": hit["货品名称"],
                        "ERP规格": hit["规格名称"],
                    }
                )
            wb.close()
    return records, {"未匹配ERP": unmatched, "不符合33尺码SKU": invalid_sku, "ERP编码数": len(erp)}


def aggregate(records):
    rules = active_hot_rules()
    new_days_lt = num(rules.get("shein_new_days_lt", 30))
    new_7d_daily_gte = num(rules.get("shein_new_7d_daily_gte", 10))
    old_days_gte = num(rules.get("shein_old_days_gte", 30))
    old_30d_daily_gt = num(rules.get("shein_old_30d_daily_gt", 20))
    skc = {}
    style_combos = defaultdict(set)
    combo_rows = defaultdict(list)
    for r in records:
        skc_key = (r["店铺"], r["SKC"])
        combo_key = (r["店铺"], r["SKC"], r["货品编码"])
        if skc_key not in skc:
            skc[skc_key] = {
                "店铺": r["店铺"],
                "店铺编号": r["店铺编号"],
                "负责人": r["负责人"],
                "SKC": r["SKC"],
                "商品名称": r["商品名称"],
                "7天销量": 0.0,
                "30天销量": 0.0,
                "平台仓备货库存": 0.0,
                "上架天数": r["上架天数"],
                "SKU数": 0,
                "价格列表": [],
            }
        item = skc[skc_key]
        item["7天销量"] += r["7天销量"]
        item["30天销量"] += r["30天销量"]
        item["平台仓备货库存"] += r["平台仓备货库存"]
        item["上架天数"] = min(item["上架天数"], r["上架天数"]) if item["上架天数"] else r["上架天数"]
        item["SKU数"] += 1
        if r["申报价"]:
            item["价格列表"].append(r["申报价"])
        combo_rows[combo_key].append(r)
        style_combos[r["货品编码"]].add(combo_key)
    for item in skc.values():
        item["平均申报价"] = sum(item["价格列表"]) / len(item["价格列表"]) if item["价格列表"] else 0
    hot = {}
    for key, item in skc.items():
        is_new = item["上架天数"] < new_days_lt
        avg7 = item["7天销量"] / 7 if is_new else 0
        avg30 = item["30天销量"] / 30 if not is_new else 0
        if (is_new and avg7 >= new_7d_daily_gte) or ((not is_new) and item["上架天数"] >= old_days_gte and avg30 > old_30d_daily_gt):
            hot[key] = {**item, "新品老品": "新品" if is_new else "老品", "7天日均": avg7, "30天日均": avg30}
    return skc, hot, style_combos, combo_rows


def combo_summary(combo_key, combo_rows, skc):
    rows = combo_rows[combo_key]
    skc_item = skc[(combo_key[0], combo_key[1])]
    prices = [r["申报价"] for r in rows if r["申报价"]]
    return {
        **skc_item,
        "货品编码": combo_key[2],
        "ERP货品名称": rows[0]["ERP货品名称"],
        "平均申报价": sum(prices) / len(prices) if prices else skc_item["平均申报价"],
    }


def hot_priority_key(item):
    return (item["30天销量"], item["平均申报价"], item["平台仓备货库存"])


def sku_sort_key(sku):
    base, style, size = norm_sku(sku)
    return (style or sku, SIZE_ORDER.get(size or "", 99), sku)


def build_operations(skc, hot, style_combos, combo_rows):
    operations = {}
    for style, combos in style_combos.items():
        hot_combos = [c for c in combos if (c[0], c[1]) in hot]
        if not hot_combos or len(combos) < 2:
            continue
        hot_summaries = [(c, combo_summary(c, combo_rows, skc)) for c in hot_combos]
        protected_combo, protected = max(hot_summaries, key=lambda x: hot_priority_key(x[1]))
        protected_rows_by_size = {r["尺码"]: r for r in combo_rows[protected_combo]}
        protected_fallback = min(combo_rows[protected_combo], key=lambda r: r["申报价"] or 999999)
        for dup_combo in sorted(combos, key=lambda x: (STORE_ORDER.index(x[0]) if x[0] in STORE_ORDER else 99, x[1])):
            if dup_combo == protected_combo:
                continue
            dup_skc_key = (dup_combo[0], dup_combo[1])
            dup_is_hot = dup_skc_key in hot
            conflict_type = "多店铺均为爆旺款" if dup_is_hot else "平销款冲突爆款"
            dup_skc = skc[dup_skc_key]
            for dup_row in combo_rows[dup_combo]:
                hot_row = protected_rows_by_size.get(dup_row["尺码"], protected_fallback)
                hot_price = hot_row["申报价"] or protected["平均申报价"]
                dup_price = dup_row["申报价"]
                lower = bool(dup_price and hot_price and dup_price < hot_price)
                dup_stock = dup_row["平台仓备货库存"]
                hot_stock = hot_row["平台仓备货库存"]
                if dup_is_hot:
                    advice = "禁止备货，30天内限时下架，清理库存。" if dup_stock > 0 else "立即下架！"
                elif lower:
                    advice = "立即下架！"
                elif dup_stock > 0:
                    advice = "售完备货库存下架。"
                else:
                    advice = "立即下架！"
                key = (dup_row["商家编码"], dup_combo[0], dup_combo[1], protected_combo[0], protected_combo[1])
                operations[key] = {
                    "商家编码": dup_row["商家编码"],
                    "货品名称": dup_row["ERP货品名称"],
                    "skc": dup_row["SKC"],
                    "所属店铺": dup_row["店铺编号"],
                    "爆旺款skc": protected["SKC"],
                    "爆旺款店铺": protected["店铺编号"],
                    "爆款报价": hot_price,
                    "重复款申报价": dup_price,
                    "爆款月销件数": hot_row["30天销量"],
                    "重复款月销件数": dup_row["30天销量"],
                    "是否低于爆款报价": "是" if lower else "否",
                    "爆款库存": hot_stock,
                    "重复款库存": dup_stock,
                    "负责人": dup_row["负责人"],
                    "冲突类型": conflict_type,
                    "处理意见": advice,
                }
    rows = list(operations.values())
    rows.sort(key=lambda r: (r["skc"], *sku_sort_key(r["商家编码"]), STORE_ORDER.index(r["所属店铺"]) if r["所属店铺"] in STORE_ORDER else 99))
    return rows


def build_overview(hot, operations):
    rows = []
    for store in STORE_ORDER:
        store_no = STORE_NO[store]
        ops = [r for r in operations if r["所属店铺"] == store_no]
        hot_count = sum(1 for (s, _), _item in hot.items() if s == store)
        rows.append(
            {
                "店铺编号": store_no,
                "店铺": store,
                "负责人": STORE_OWNER[store],
                "爆款总数": hot_count,
                "重复铺货预计总数": len(ops),
                "平销冲突数": sum(1 for r in ops if r["冲突类型"] == "平销款冲突爆款"),
                "爆款互相冲突数": sum(1 for r in ops if r["冲突类型"] == "多店铺均为爆旺款"),
                "低于爆款报价数": sum(1 for r in ops if r["是否低于爆款报价"] == "是"),
                "不低于爆款报价数": sum(1 for r in ops if r["是否低于爆款报价"] == "否"),
                "立即下架数": sum(1 for r in ops if r["处理意见"] == "立即下架！"),
                "售完备货库存下架数": sum(1 for r in ops if r["处理意见"] == "售完备货库存下架。"),
                "30天内限时下架数": sum(1 for r in ops if r["处理意见"] == "禁止备货，30天内限时下架，清理库存。"),
            }
        )
    return rows


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    sources = source_files()
    records, checks = load_sales_records()
    skc, hot, style_combos, combo_rows = aggregate(records)
    operations = build_operations(skc, hot, style_combos, combo_rows)
    overview = build_overview(hot, operations)
    payload = {
        "version": REPORT_VERSION,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sources": [p.name for p in sources.values()],
        "checks": {
            **checks,
            "有效匹配SKU": len(records),
            "爆旺SKC": len(hot),
            "具体店铺操作表行数": len(operations),
        },
        "overview": overview,
        "operations": operations,
        "notes": [
            ("版本", REPORT_VERSION),
            ("表格结构", "总览、具体店铺操作表、说明。旧版平销冲突明细表和爆款互相冲突明细表不再单独输出。"),
            ("数据源", "使用 shein数据源表 文件夹内 5 个 SHEIN 店铺销售总览表，ERP 匹配使用 erp数据源 文件夹内 ERP 产品基础信息表和组合装表。"),
            ("SKU口径", "商家SKU 匹配时忽略 @ 及后续字符；本次按 33 开头且以 -XS/-S/-M/-L/-XL 结尾的商家编码进入核查。"),
            ("爆旺款定义", "以店铺+SKC为维度：新品上架天数小于30天且7天销量日均>=10；老品上架天数大于等于30天且30天销量日均>20。"),
            ("爆旺款优先级定义", "同款货品编码冲突时，30天销量越高优先级越高；30天销量相同，申报价越高优先级越高；30天销量和申报价都相同，平台仓备货越多优先级越高。"),
            ("处理意见规则", "平销款低于爆款报价：立即下架！平销款不低于爆款报价且有备货库存：售完备货库存下架。平销款不低于爆款报价且无备货库存：立即下架！同为爆旺款但优先级低且有备货库存：禁止备货，30天内限时下架，清理库存。同为爆旺款但优先级低且无备货库存：立即下架！"),
            ("SHEIN店铺字段说明", "SHEIN 平台不使用数字店铺代码，所属店铺和爆旺款店铺直接展示：琪琪、童话、加加、宝宝、牛牛。"),
            ("具体店铺操作表行数", str(len(operations))),
        ],
    }
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload["checks"], ensure_ascii=False))


if __name__ == "__main__":
    main()
