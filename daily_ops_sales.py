import json
import os
import threading
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SALES_LOCKS = {}
SALES_LOCKS_GUARD = threading.Lock()


def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today_text():
    return date.today().isoformat()


def norm(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_date(value):
    text = norm(value)
    if not text:
        return date.today()
    return datetime.strptime(text, "%Y-%m-%d").date()


def sales_lock(path):
    key = str(Path(path).resolve())
    with SALES_LOCKS_GUARD:
        if key not in SALES_LOCKS:
            SALES_LOCKS[key] = threading.RLock()
        return SALES_LOCKS[key]


def sales_record_id(day, platform, store):
    return "|".join([norm(day), norm(platform), norm(store)])


def sales_number(value):
    text = norm(value)
    if text == "":
        raise ValueError("请填写销售件数")
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError("销售件数必须是数字") from exc
    if number < 0:
        raise ValueError("销售件数不能为负数")
    if number != number.to_integral_value():
        raise ValueError("销售件数必须是整数")
    return int(number)


def owner_visible_assignments(assignments, role="admin", user=""):
    role = norm(role) or "admin"
    user = norm(user)
    rows = []
    seen = set()
    for item in assignments or []:
        platform = norm(item.get("platform"))
        store = norm(item.get("store"))
        owner = norm(item.get("owner"))
        if not platform or not store or not owner:
            continue
        if item.get("enabled") is False or item.get("daily_required") is False:
            continue
        if role != "admin" and owner != user:
            continue
        key = (platform, store)
        if key in seen:
            continue
        seen.add(key)
        rows.append({"platform": platform, "store": store, "owner": owner})
    return rows


def is_manual_sales_record(record):
    if not record:
        return False
    source = norm(record.get("source"))
    status = norm(record.get("status"))
    return source != "历史导入" and status != "历史导入"


def platform_summary(entries):
    summary = {}
    for item in entries:
        platform = norm(item.get("platform")) or "未设置"
        bucket = summary.setdefault(platform, {"platform": platform, "required": 0, "submitted": 0, "missing": 0, "sales": 0, "abnormal": 0})
        bucket["required"] += 1
        if item.get("submitted"):
            bucket["submitted"] += 1
            bucket["sales"] += int(item.get("sales") or 0)
        else:
            bucket["missing"] += 1
        if item.get("abnormal"):
            bucket["abnormal"] += 1
    return sorted(summary.values(), key=lambda row: row["platform"])


SALES_EXPORT_COLUMNS = [
    ("date", "日期"),
    ("platform", "平台"),
    ("store", "店铺"),
    ("owner", "负责人"),
    ("sales", "销售件数"),
    ("status", "状态"),
    ("abnormal", "异常提示"),
    ("remark", "备注"),
    ("updated_at", "更新时间"),
]


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="F4F0FF")
    header_font = Font(bold=True, color="223047")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 3, 42)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


class DailySalesStore:
    def __init__(self, path):
        self.path = Path(path)
        self._lock = sales_lock(self.path)

    def load(self):
        with self._lock:
            if not self.path.exists():
                return {"records": []}
            try:
                payload = json.loads(self.path.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError):
                return {"records": []}
            if not isinstance(payload, dict):
                return {"records": []}
            records = payload.get("records")
            return {"records": records if isinstance(records, list) else []}

    def save(self, payload):
        with self._lock:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            tmp = self.path.with_name(f".{self.path.name}.{threading.get_ident()}.tmp")
            tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            os.replace(tmp, self.path)

    def records_for_store(self, platform, store, before_day=None):
        platform = norm(platform)
        store = norm(store)
        cutoff = parse_date(before_day) if before_day else None
        rows = []
        for row in self.load()["records"]:
            if norm(row.get("platform")) != platform or norm(row.get("store")) != store:
                continue
            row_day = parse_date(row.get("date"))
            if cutoff and row_day >= cutoff:
                continue
            rows.append(row)
        return sorted(rows, key=lambda row: norm(row.get("date")), reverse=True)

    def abnormal_hint(self, platform, store, day, sales):
        history = self.records_for_store(platform, store, before_day=day)
        if not history:
            return ""
        yesterday = parse_date(day) - timedelta(days=1)
        yesterday_row = next((row for row in history if norm(row.get("date")) == yesterday.isoformat()), None)
        recent = history[:7]
        recent_sales = [int(row.get("sales") or 0) for row in recent if norm(row.get("sales")) != ""]
        avg7 = sum(recent_sales) / len(recent_sales) if recent_sales else 0
        hints = []
        if yesterday_row:
            yesterday_sales = int(yesterday_row.get("sales") or 0)
            base = max(yesterday_sales, 1)
            if abs(sales - yesterday_sales) / base > 0.5:
                hints.append(f"较昨日 {yesterday_sales} 件波动超过 50%")
        if avg7 > 0 and abs(sales - avg7) / avg7 > 0.5:
            hints.append(f"较近 7 次均值 {avg7:.1f} 件波动超过 50%")
        if sales == 0 and avg7 > 0:
            hints.append("近 7 次有销量，今日为 0")
        return "；".join(hints)

    def submit(self, assignments, role="admin", user="", day="", platform="", store="", sales="", remark=""):
        day = parse_date(day).isoformat()
        platform = norm(platform)
        store = norm(store)
        actor = norm(user) or "管理员"
        number = sales_number(sales)
        if not platform or not store:
            raise ValueError("请填写平台和店铺")

        assignment = next((item for item in assignments or [] if norm(item.get("platform")) == platform and norm(item.get("store")) == store), None)
        if not assignment:
            raise ValueError("该平台店铺未配置负责人，请先在基础资料维护")
        if assignment.get("enabled") is False or assignment.get("daily_required") is False:
            raise ValueError("该平台店铺已停用或不需要每日销量填报")
        owner = norm(assignment.get("owner"))
        if norm(role) != "admin" and owner != actor:
            raise PermissionError("店长只能填写自己负责店铺的销量")

        payload = self.load()
        records = payload["records"]
        record_id = sales_record_id(day, platform, store)
        existing = next((row for row in records if row.get("id") == record_id), None)
        old_sales = existing.get("sales") if existing else ""
        abnormal = self.abnormal_hint(platform, store, day, number)
        history_item = {
            "time": now_text(),
            "actor": actor,
            "action": "提交销量" if existing is None else "更新销量",
            "old_sales": old_sales,
            "new_sales": number,
            "remark": norm(remark),
        }
        row = {
            "id": record_id,
            "date": day,
            "platform": platform,
            "store": store,
            "owner": owner,
            "sales": number,
            "status": "已填写",
            "abnormal": abnormal,
            "remark": norm(remark),
            "submitted_by": actor,
            "submitted_at": existing.get("submitted_at") if existing else now_text(),
            "updated_at": now_text(),
            "history": (existing.get("history", []) if existing else []) + [history_item],
        }
        if existing:
            existing.update(row)
        else:
            records.append(row)
        payload["updated_at"] = now_text()
        self.save(payload)
        return row

    def daily_payload(self, assignments, role="admin", user="", day=""):
        day = parse_date(day).isoformat()
        visible = owner_visible_assignments(assignments, role, user)
        all_records = self.load()["records"]
        records = {sales_record_id(row.get("date"), row.get("platform"), row.get("store")): row for row in all_records}
        visible_keys = {(item["platform"], item["store"]) for item in visible}
        entries = []
        for item in visible:
            key = sales_record_id(day, item["platform"], item["store"])
            record = records.get(key)
            submitted = is_manual_sales_record(record)
            display_status = "待确认" if record and not submitted else (record.get("status", "未填") if record else "未填")
            entries.append({
                "date": day,
                "platform": item["platform"],
                "store": item["store"],
                "owner": item["owner"],
                "submitted": submitted,
                "sales": record.get("sales", "") if record else "",
                "status": display_status,
                "abnormal": record.get("abnormal", "") if record else "",
                "remark": record.get("remark", "") if record else "",
                "updated_at": record.get("updated_at", "") if record else "",
                "source": record.get("source", "") if record else "",
                "needs_confirmation": bool(record and not submitted),
            })
        required = len(entries)
        submitted = sum(1 for item in entries if item["submitted"])
        abnormal = sum(1 for item in entries if item["abnormal"])
        total_sales = sum(int(item["sales"] or 0) for item in entries if item["submitted"])
        recent_records = [
            row for row in all_records
            if norm(role) == "admin" or (norm(row.get("platform")), norm(row.get("store"))) in visible_keys
        ]
        recent_records = sorted(recent_records, key=lambda row: (norm(row.get("date")), norm(row.get("updated_at"))), reverse=True)[:80]
        return {
            "date": day,
            "summary": {
                "required": required,
                "submitted": submitted,
                "missing": required - submitted,
                "abnormal": abnormal,
                "total_sales": total_sales,
            },
            "entries": entries,
            "platforms": platform_summary(entries),
            "records": recent_records,
        }

    def export_daily_workbook(self, assignments, output_dir, role="admin", user="", day=""):
        payload = self.daily_payload(assignments, role, user, day)
        day = payload["date"]
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{day}-每日销量台账.xlsx"
        if norm(role) != "admin" and norm(user):
            filename = f"{day}-{norm(user)}-每日销量台账.xlsx"
        path = output_dir / filename
        if path.exists():
            stamp = datetime.now().strftime("%H%M%S")
            path = output_dir / f"{path.stem}-{stamp}{path.suffix}"

        wb = Workbook()
        ws = wb.active
        ws.title = "每日销量明细"
        ws.append([label for _key, label in SALES_EXPORT_COLUMNS])
        for row in payload["entries"]:
            ws.append([row.get(key, "") for key, _label in SALES_EXPORT_COLUMNS])
        style_sheet(ws)

        summary = wb.create_sheet("平台汇总")
        summary.append(["平台", "应填店铺", "已填写", "未填写", "总销量", "异常数"])
        for row in payload["platforms"]:
            summary.append([
                row.get("platform", ""),
                row.get("required", 0),
                row.get("submitted", 0),
                row.get("missing", 0),
                row.get("sales", 0),
                row.get("abnormal", 0),
            ])
        style_sheet(summary)

        wb.save(path)
        return {
            "file": path.name,
            "path": str(path),
            "date": day,
            "summary": payload["summary"],
            "platforms": payload["platforms"],
        }
