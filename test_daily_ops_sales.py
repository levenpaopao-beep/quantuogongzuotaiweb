import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from daily_ops_sales import DailySalesStore, sales_number


class DailySalesStoreTest(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.path = Path(self.tmpdir.name) / "daily_sales.json"
        self.store = DailySalesStore(self.path)
        self.assignments = [{"platform": "Temu", "store": "七弟", "owner": "小琴"}]

    def tearDown(self):
        self.tmpdir.cleanup()

    def test_sales_number_rejects_fractional_values(self):
        self.assertEqual(sales_number("12"), 12)
        self.assertEqual(sales_number("12.0"), 12)
        with self.assertRaisesRegex(ValueError, "必须是整数"):
            sales_number("12.8")

    def test_submit_and_update_sales(self):
        first = self.store.submit(self.assignments, role="owner", user="小琴", day="2026-06-23", platform="Temu", store="七弟", sales="12")
        second = self.store.submit(self.assignments, role="owner", user="小琴", day="2026-06-23", platform="Temu", store="七弟", sales="15", remark="后台复核")

        self.assertEqual(first["sales"], 12)
        self.assertEqual(second["sales"], 15)
        self.assertEqual(len(second["history"]), 2)

        payload = self.store.daily_payload(self.assignments, role="owner", user="小琴", day="2026-06-23")
        self.assertEqual(payload["summary"]["required"], 1)
        self.assertEqual(payload["summary"]["submitted"], 1)
        self.assertEqual(payload["summary"]["total_sales"], 15)

    def test_owner_can_only_update_existing_sales_on_submission_day(self):
        with patch("daily_ops_sales.today_text", return_value="2026-06-23"), patch("daily_ops_sales.now_text", return_value="2026-06-23 10:00:00"):
            self.store.submit(self.assignments, role="owner", user="小琴", day="2026-06-18", platform="Temu", store="七弟", sales="12")
            updated = self.store.submit(self.assignments, role="owner", user="小琴", day="2026-06-18", platform="Temu", store="七弟", sales="15")
        self.assertEqual(updated["sales"], 15)

        with patch("daily_ops_sales.today_text", return_value="2026-06-24"), patch("daily_ops_sales.now_text", return_value="2026-06-24 10:00:00"):
            with self.assertRaisesRegex(PermissionError, "店长只能修改当天填写的数据"):
                self.store.submit(self.assignments, role="owner", user="小琴", day="2026-06-18", platform="Temu", store="七弟", sales="18")
            admin_updated = self.store.submit(self.assignments, role="admin", user="管理员", day="2026-06-18", platform="Temu", store="七弟", sales="20")

        self.assertEqual(admin_updated["sales"], 20)

    def test_daily_payload_marks_owner_history_as_locked_after_submission_day(self):
        with patch("daily_ops_sales.today_text", return_value="2026-06-23"), patch("daily_ops_sales.now_text", return_value="2026-06-23 10:00:00"):
            self.store.submit(self.assignments, role="owner", user="小琴", day="2026-06-18", platform="Temu", store="七弟", sales="12")

        with patch("daily_ops_sales.today_text", return_value="2026-06-24"):
            owner_payload = self.store.daily_payload(self.assignments, role="owner", user="小琴", day="2026-06-18")
            admin_payload = self.store.daily_payload(self.assignments, role="admin", user="管理员", day="2026-06-18")

        self.assertFalse(owner_payload["entries"][0]["editable"])
        self.assertIn("已过当天", owner_payload["entries"][0]["locked_reason"])
        self.assertTrue(admin_payload["entries"][0]["editable"])

    def test_history_import_does_not_count_as_owner_daily_submission(self):
        data = self.store.load()
        data["records"].append({
            "id": "2026-06-23|Temu|七弟",
            "date": "2026-06-23",
            "platform": "Temu",
            "store": "七弟",
            "owner": "小琴",
            "sales": 0,
            "status": "历史导入",
            "source": "历史导入",
        })
        self.store.save(data)

        payload = self.store.daily_payload(self.assignments, role="owner", user="小琴", day="2026-06-23")

        self.assertEqual(payload["summary"]["required"], 1)
        self.assertEqual(payload["summary"]["submitted"], 0)
        self.assertEqual(payload["summary"]["missing"], 1)
        self.assertEqual(payload["entries"][0]["status"], "待确认")
        self.assertTrue(payload["entries"][0]["needs_confirmation"])

    def test_owner_can_only_submit_own_store(self):
        with self.assertRaises(PermissionError):
            self.store.submit(self.assignments, role="owner", user="别人", day="2026-06-23", platform="Temu", store="七弟", sales="12")

    def test_owner_recent_records_are_limited_to_own_store(self):
        assignments = [
            {"platform": "Temu", "store": "七弟", "owner": "小琴"},
            {"platform": "Shein", "store": "琪琪", "owner": "胡娟"},
        ]
        self.store.submit(assignments, role="admin", user="管理员", day="2026-06-22", platform="Temu", store="七弟", sales="12")
        self.store.submit(assignments, role="admin", user="管理员", day="2026-06-22", platform="Shein", store="琪琪", sales="88")

        owner_payload = self.store.daily_payload(assignments, role="owner", user="小琴", day="2026-06-22")
        admin_payload = self.store.daily_payload(assignments, role="admin", user="管理员", day="2026-06-22")

        self.assertEqual([row["store"] for row in owner_payload["records"]], ["七弟"])
        self.assertEqual({row["store"] for row in admin_payload["records"]}, {"七弟", "琪琪"})

    def test_disabled_or_not_required_stores_are_not_daily_entries(self):
        assignments = [
            {"platform": "Temu", "store": "七弟", "owner": "小琴", "enabled": True, "daily_required": True},
            {"platform": "Shein", "store": "琪琪", "owner": "小琴", "enabled": False, "daily_required": True},
            {"platform": "Ozon", "store": "店铺 A", "owner": "小琴", "enabled": True, "daily_required": False},
        ]

        payload = self.store.daily_payload(assignments, role="owner", user="小琴", day="2026-06-23")

        self.assertEqual(payload["summary"]["required"], 1)
        self.assertEqual(payload["entries"][0]["store"], "七弟")

    def test_disabled_or_not_required_stores_reject_sales_submit(self):
        assignments = [
            {"platform": "Shein", "store": "琪琪", "owner": "小琴", "enabled": False, "daily_required": True},
            {"platform": "Ozon", "store": "店铺 A", "owner": "小琴", "enabled": True, "daily_required": False},
        ]

        with self.assertRaisesRegex(ValueError, "已停用或不需要每日销量填报"):
            self.store.submit(assignments, role="admin", user="管理员", day="2026-06-23", platform="Shein", store="琪琪", sales="12")
        with self.assertRaisesRegex(ValueError, "已停用或不需要每日销量填报"):
            self.store.submit(assignments, role="owner", user="小琴", day="2026-06-23", platform="Ozon", store="店铺 A", sales="12")
        self.assertEqual(self.store.load()["records"], [])

    def test_abnormal_hint_uses_recent_sales(self):
        self.store.submit(self.assignments, role="admin", user="管理员", day="2026-06-22", platform="Temu", store="七弟", sales="100")
        row = self.store.submit(self.assignments, role="admin", user="管理员", day="2026-06-23", platform="Temu", store="七弟", sales="10")
        self.assertIn("波动超过 50%", row["abnormal"])

    def test_export_daily_workbook(self):
        self.store.submit(self.assignments, role="admin", user="管理员", day="2026-06-23", platform="Temu", store="七弟", sales="12")

        result = self.store.export_daily_workbook(self.assignments, Path(self.tmpdir.name), role="admin", user="管理员", day="2026-06-23")

        workbook = load_workbook(result["path"], data_only=True)
        self.assertIn("每日销量明细", workbook.sheetnames)
        self.assertIn("平台汇总", workbook.sheetnames)
        self.assertEqual(workbook["每日销量明细"]["A2"].value, "2026-06-23")
        self.assertEqual(workbook["每日销量明细"]["E2"].value, 12)

    def test_owner_export_daily_workbook_only_contains_own_store(self):
        assignments = [
            {"platform": "Temu", "store": "七弟", "owner": "小琴"},
            {"platform": "Shein", "store": "琪琪", "owner": "胡娟"},
        ]
        self.store.submit(assignments, role="admin", user="管理员", day="2026-06-23", platform="Temu", store="七弟", sales="12")
        self.store.submit(assignments, role="admin", user="管理员", day="2026-06-23", platform="Shein", store="琪琪", sales="88")

        result = self.store.export_daily_workbook(assignments, Path(self.tmpdir.name), role="owner", user="小琴", day="2026-06-23")

        workbook = load_workbook(result["path"], data_only=True)
        detail = workbook["每日销量明细"]
        summary = workbook["平台汇总"]
        self.assertIn("小琴", result["file"])
        self.assertEqual(detail.max_row, 2)
        self.assertEqual(detail["C2"].value, "七弟")
        self.assertEqual(detail["D2"].value, "小琴")
        self.assertEqual(summary.max_row, 2)
        self.assertEqual(summary["A2"].value, "Temu")


if __name__ == "__main__":
    unittest.main()
