import copy
import html
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(os.environ.get("DAILYWORK_ROOT", Path(__file__).resolve().parent))
SOURCE_DIR = ROOT / "shein数据源表"
INPUT_WB = ROOT / "Shein申报价异常交叉对比表_20260603.xlsx"
OUTPUT_WB = ROOT / "Shein申报价异常交叉对比表_增加30天有销量SKC_20260603.xlsx"
STORE_ORDER = ["琪琪", "童话", "牛牛", "加加", "宝宝"]


def norm_text(value):
    if value is None:
        return ""
    return html.unescape(str(value)).strip()


def to_number(value):
    text = norm_text(value)
    if not text:
        return 0.0
    text = text.replace(",", "").replace("，", "")
    text = re.sub(r"[^\d.\-]", "", text)
    if text in {"", "-", ".", "-."}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def col_to_num(col):
    n = 0
    for ch in col:
        n = n * 26 + ord(ch) - 64
    return n


def read_shared_strings(zf):
    try:
        xml = zf.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    root = ET.fromstring(xml)
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    out = []
    for si in root.findall("a:si", ns):
        parts = [t.text or "" for t in si.findall(".//a:t", ns)]
        out.append("".join(parts))
    return out


def first_sheet_path(zf):
    workbook = ET.fromstring(zf.read("xl/workbook.xml"))
    ns = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    sheet = workbook.find("a:sheets/a:sheet", ns)
    rid = sheet.attrib[f"{{{ns['r']}}}id"]
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
    for rel in rels.findall("r:Relationship", rns):
        if rel.attrib["Id"] == rid:
            target = rel.attrib["Target"].lstrip("/")
            if not target.startswith("xl/"):
                target = "xl/" + target
            return target
    raise RuntimeError("Cannot resolve first worksheet path")


def read_xlsx_rows(path):
    with zipfile.ZipFile(path) as zf:
        shared = read_shared_strings(zf)
        sheet_xml = zf.read(first_sheet_path(zf)).decode("utf-8", errors="ignore")

    rows = {}
    for m in re.finditer(r"<c\b([^>]*)>(.*?)</c>", sheet_xml, re.S):
        attrs, body = m.group(1), m.group(2)
        ref_m = re.search(r'\br="([A-Z]+)(\d+)"', attrs)
        if not ref_m:
            continue
        col, row = ref_m.group(1), int(ref_m.group(2))
        idx = col_to_num(col)
        value = ""
        if re.search(r'\bt="s"', attrs):
            v = re.search(r"<v>(.*?)</v>", body, re.S)
            if v:
                si = int(v.group(1))
                value = shared[si] if 0 <= si < len(shared) else ""
        elif re.search(r'\bt="inlineStr"', attrs):
            parts = re.findall(r"<t[^>]*>(.*?)</t>", body, re.S)
            value = "".join(parts)
        else:
            v = re.search(r"<v>(.*?)</v>", body, re.S)
            value = v.group(1) if v else ""
        rows.setdefault(row, {})[idx] = value

    if not rows:
        return []
    max_row = max(rows)
    max_col = max(max(r.keys()) for r in rows.values())
    return [[rows.get(r, {}).get(c, "") for c in range(1, max_col + 1)] for r in range(1, max_row + 1)]


def store_from_filename(path):
    name = path.stem
    for store in STORE_ORDER:
        if store in name:
            return store
    return name


def compute_counts():
    counts = {store: set() for store in STORE_ORDER}
    for path in SOURCE_DIR.glob("20260601-shein*.xlsx"):
        store = store_from_filename(path)
        if store not in counts:
            continue
        rows = read_xlsx_rows(path)
        if not rows:
            continue
        headers = [norm_text(v) for v in rows[0]]
        header_map = {h: i for i, h in enumerate(headers) if h}
        skc_i = header_map.get("SKC")
        sales_i = header_map.get("近30天销量")
        if skc_i is None or sales_i is None:
            raise RuntimeError(f"{path.name} 缺少 SKC 或 近30天销量 字段")
        for row in rows[1:]:
            skc = norm_text(row[skc_i] if skc_i < len(row) else "")
            sales_30 = to_number(row[sales_i] if sales_i < len(row) else "")
            if skc and sales_30 > 0:
                counts[store].add(skc)
    return {store: len(counts[store]) for store in STORE_ORDER}


def copy_cell_style(src, dst):
    if src.has_style:
        dst._style = copy.copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy.copy(src.alignment)
    if src.font:
        dst.font = copy.copy(src.font)
    if src.fill:
        dst.fill = copy.copy(src.fill)
    if src.border:
        dst.border = copy.copy(src.border)


def update_workbook(counts):
    wb = load_workbook(INPUT_WB)
    ws = wb["总览表"]
    header_row = None
    headers = []
    for row in range(1, min(ws.max_row, 10) + 1):
        row_headers = [norm_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if "店铺" in row_headers:
            header_row = row
            headers = row_headers
            break
    if header_row is None:
        raise RuntimeError("总览表未找到店铺表头行")

    if "30天内有销量的SKC数" in headers:
        target_col = headers.index("30天内有销量的SKC数") + 1
    else:
        try:
            insert_after = headers.index("30天销量") + 1
        except ValueError:
            insert_after = ws.max_column
        target_col = insert_after + 1
        ws.insert_cols(target_col)
        for row in range(1, ws.max_row + 1):
            left = ws.cell(row, target_col - 1)
            cell = ws.cell(row, target_col)
            copy_cell_style(left, cell)
        ws.column_dimensions[get_column_letter(target_col)].width = 18

    ws.cell(header_row, target_col).value = "30天内有销量的SKC数"
    store_col = headers.index("店铺") + 1 if "店铺" in headers else 1
    for row in range(header_row + 1, ws.max_row + 1):
        store = norm_text(ws.cell(row, store_col).value)
        if store in counts:
            ws.cell(row, target_col).value = counts[store]

    if ws.auto_filter and ws.auto_filter.ref:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

    if "数据校验" in wb.sheetnames:
        check = wb["数据校验"]
        next_row = check.max_row + 1
        check.cell(next_row, 1).value = "总览新增字段"
        check.cell(next_row, 2).value = "30天内有销量的SKC数=按店铺统计近30天销量>0的唯一SKC数量"

    wb.save(OUTPUT_WB)


if __name__ == "__main__":
    counts = compute_counts()
    update_workbook(counts)
    print(OUTPUT_WB)
    for store in STORE_ORDER:
        print(f"{store}: {counts[store]}")
