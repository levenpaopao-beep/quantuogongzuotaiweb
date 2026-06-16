import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

import generate_temu_slow_moving_weekly


class OwnerMappingTest(unittest.TestCase):
    def test_temu_slow_owner_map_supports_numbered_store_aliases(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "店铺负责人对应表.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["店铺", "业务"])
            ws.append(["十二", "胡娟"])
            wb.save(path)

            original = generate_temu_slow_moving_weekly.OWNER_FILE
            generate_temu_slow_moving_weekly.OWNER_FILE = path
            try:
                owners = generate_temu_slow_moving_weekly.load_owner_map()
            finally:
                generate_temu_slow_moving_weekly.OWNER_FILE = original

        self.assertEqual(owners["十二"], "胡娟")
        self.assertEqual(owners["十二弟"], "胡娟")
        self.assertEqual(owners["12"], "胡娟")

    def test_temu_slow_owner_map_supports_fifteen_store_aliases(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "店铺负责人对应表.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["店铺", "业务"])
            ws.append(["十五弟", "洁琳"])
            wb.save(path)

            original = generate_temu_slow_moving_weekly.OWNER_FILE
            generate_temu_slow_moving_weekly.OWNER_FILE = path
            try:
                owners = generate_temu_slow_moving_weekly.load_owner_map()
            finally:
                generate_temu_slow_moving_weekly.OWNER_FILE = original

        self.assertEqual(owners["十五弟"], "洁琳")
        self.assertEqual(owners["十五"], "洁琳")
        self.assertEqual(owners["15"], "洁琳")
        self.assertEqual(owners["15di"], "洁琳")

    def test_temu_slow_judges_slow_moving_by_skc_not_sku(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            sales = folder / "temu.xlsx"
            owner = folder / "owner.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["店铺", "SKC", "SKU ID", "SKU货号", "申报价格", "7天销量", "30天销量", "可用", "加入站点时长", "产品名称", "SKU属性"])
            ws.append(["十五", "9795739993", "sku-1", "330000001-S", 10, 0, 1, 5, 45, "有销量SKU", "S"])
            ws.append(["十五", "9795739993", "sku-2", "330000002-M", 10, 0, 0, 5, 45, "无销量SKU", "M"])
            wb.save(sales)

            wb = Workbook()
            ws = wb.active
            ws.append(["店铺", "业务"])
            ws.append(["十五弟", "洁琳"])
            wb.save(owner)

            originals = {
                "TEMU_SALES_FILES": generate_temu_slow_moving_weekly.TEMU_SALES_FILES,
                "ERP_FILES": generate_temu_slow_moving_weekly.ERP_FILES,
                "ERP_DIR": generate_temu_slow_moving_weekly.ERP_DIR,
                "OWNER_FILE": generate_temu_slow_moving_weekly.OWNER_FILE,
                "RULES": generate_temu_slow_moving_weekly.RULES,
            }
            generate_temu_slow_moving_weekly.TEMU_SALES_FILES = [sales]
            generate_temu_slow_moving_weekly.ERP_FILES = []
            generate_temu_slow_moving_weekly.ERP_DIR = folder
            generate_temu_slow_moving_weekly.OWNER_FILE = owner
            generate_temu_slow_moving_weekly.RULES = {
                "slow_moving": {
                    "new_slow_min_days": 30,
                    "new_slow_max_days": 60,
                    "old_slow_min_days": 180,
                }
            }
            try:
                new_rows, old_rows, stats = generate_temu_slow_moving_weekly.read_source_rows()
            finally:
                for key, value in originals.items():
                    setattr(generate_temu_slow_moving_weekly, key, value)

        self.assertEqual(new_rows, [])
        self.assertEqual(old_rows, [])
        self.assertEqual(stats["跳过：SKC内30天有销量"], 2)


if __name__ == "__main__":
    unittest.main()
