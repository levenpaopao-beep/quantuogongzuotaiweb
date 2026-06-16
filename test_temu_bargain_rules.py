import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import load_workbook

import daily_ops_app
import generate_temu_bargain_reply
from generate_temu_bargain_reply import evaluate_offer


def offer(price=20, cost=10, wholesale=20, links=None, has_hot=False):
    return {
        "申报价": price,
        "成本价": cost,
        "批发价": wholesale,
        "在售链接": links or [],
        "有爆旺款": has_hot,
    }


class TemuBargainRuleTest(unittest.TestCase):
    def test_rejects_price_below_cost_first(self):
        result = evaluate_offer(offer(price=9, cost=10, wholesale=8, has_hot=True))
        self.assertEqual(result["是否通过"], "拒绝上架-理由 亏损")

    def test_rejects_when_hot_link_exists(self):
        result = evaluate_offer(offer(has_hot=True, links=[{"申报价": 20}]))
        self.assertEqual(result["是否通过"], "拒绝上架-理由 有爆旺款在售")

    def test_rejects_preview_hot_when_quote_below_95_percent(self):
        result = evaluate_offer(
            offer(price=18, links=[{"申报价": 20, "7天销量": 11, "30天销量": 30}])
        )
        self.assertEqual(result["是否通过"], "拒绝上架-理由 有预备爆款链接在售")

    def test_rejects_when_more_than_seven_links_exist(self):
        result = evaluate_offer(offer(links=[{"申报价": 20}] * 8))
        self.assertEqual(
            result["是否通过"], "拒绝上架-理由 同时在架产品过多，15天以后再尝试上架"
        )

    def test_rejects_four_to_seven_links_when_quote_too_low(self):
        result = evaluate_offer(
            offer(price=18, cost=10, links=[{"申报价": 20}, {"申报价": 22}, {"申报价": 21}, {"申报价": 23}])
        )
        self.assertEqual(result["是否通过"], "拒绝上架-理由 平台建议价低于在售最低价95%")

    def test_rejects_price_below_80_percent_wholesale(self):
        result = evaluate_offer(offer(price=15, cost=10, wholesale=20))
        self.assertEqual(result["是否通过"], "拒绝上架-理由 破价")

    def test_approves_when_no_rule_matches(self):
        result = evaluate_offer(offer(price=20, cost=10, wholesale=20, links=[{"申报价": 20}]))
        self.assertEqual(result["是否通过"], "同意议价")

    def test_weekly_workflow_exposes_bargain_declaration_upload(self):
        group = daily_ops_app.WEEKLY_SOURCE_GROUPS["temu_bargain_input"]
        self.assertIn("议价申报", group["name"])
        self.assertIn("*议价申报*.xlsx", group["patterns"])

    def test_bargain_input_files_match_declaration_name(self):
        with TemporaryDirectory() as tmp:
            folder = Path(tmp)
            target = folder / "20260613-议价申报.xlsx"
            target.write_bytes(b"xlsx")
            with patch.object(daily_ops_app, "BARGAIN_DIR", folder), \
                 patch.object(daily_ops_app, "manifest_paths", return_value=[]):
                files = daily_ops_app.bargain_input_files()
            self.assertEqual(files, [target])

    def test_old_bargain_reply_filename_maps_to_bargain_report_card(self):
        self.assertEqual(daily_ops_app.report_id_for_output("260613-temu议价回复-V1.xlsx"), "temu_bargain")

    def test_bargain_reply_keeps_input_order_and_reference_headers(self):
        rows = [
            {"商家编码": "SKU-2-S", "货品名称": "第二个", "skc": "skc-2", "规格名称": "", "店铺": "2", "申报价": 12, "建议价格": 10, "判断价格": 10, "款号": "SKU-2"},
            {"商家编码": "SKU-1-S", "货品名称": "第一个", "skc": "skc-1", "规格名称": "", "店铺": "1", "申报价": 11, "建议价格": 9, "判断价格": 9, "款号": "SKU-1"},
        ]
        erp = {
            "SKU-2-S": {"成本价": 5, "批发价": 8, "货品名称": "第二个", "规格名称": "S"},
            "SKU-1-S": {"成本价": 5, "批发价": 8, "货品名称": "第一个", "规格名称": "S"},
        }
        output_rows = generate_temu_bargain_reply.build_output_rows(rows, erp, {}, set())
        with TemporaryDirectory() as tmp:
            output = Path(tmp) / "reply.xlsx"
            generate_temu_bargain_reply.write_workbook(output_rows, output)
            wb = load_workbook(output, read_only=True, data_only=True)
            ws = wb.active
            values = list(ws.iter_rows(values_only=True))
            wb.close()

        self.assertEqual(list(values[0]), [
            "商家编码", "货品名称", "skc", "店铺", "申报价", "建议价格", "成本价", "批发价", "是否通过",
            "平台在售链接数", "平台在售最低申报价", "平台在售最高月销量", "平台在售上架最长时间",
        ])
        self.assertEqual(values[1][0], "SKU-2-S")
        self.assertEqual(values[1][2], "skc-2")
        self.assertEqual(values[1][4], 12)
        self.assertEqual(values[1][5], 10)
        self.assertEqual(values[1][6], 5)
        self.assertEqual(values[1][7], 8)
        self.assertEqual(values[2][0], "SKU-1-S")
        self.assertEqual(values[2][2], "skc-1")

    def test_platform_suggested_price_is_used_for_decision_not_overwritten(self):
        input_rows = [{
            "商家编码": "SKU-1-S",
            "货品名称": "第一个",
            "skc": "skc-1",
            "规格名称": "",
            "店铺": "1",
            "原申报价": 20,
            "原建议价格": 9,
            "款号": "SKU-1",
        }]
        enriched = generate_temu_bargain_reply.enrich_input_rows(input_rows, {}, {})
        self.assertEqual(enriched[0]["申报价"], 20)
        self.assertEqual(enriched[0]["建议价格"], 9)
        self.assertEqual(enriched[0]["判断价格"], 9)

        erp = {"SKU-1-S": {"成本价": 10, "批发价": 12, "货品名称": "第一个", "规格名称": "S"}}
        output = generate_temu_bargain_reply.build_output_rows(enriched, erp, {}, set())
        self.assertEqual(output[0][4], 20)
        self.assertEqual(output[0][5], 9)
        self.assertEqual(output[0][8], "拒绝上架-理由 亏损")


if __name__ == "__main__":
    unittest.main()
