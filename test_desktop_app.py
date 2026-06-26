import unittest
import json
import subprocess
import sys
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

import daily_ops_app
from openpyxl import Workbook, load_workbook
from PIL import Image


ROOT = Path(__file__).resolve().parent


class DesktopAppTest(unittest.TestCase):
    def test_electron_desktop_entrypoint_exists_and_does_not_start_web_server(self):
        main_source = (ROOT / "electron" / "main.js").read_text(encoding="utf-8")
        package = json.loads((ROOT / "package.json").read_text(encoding="utf-8"))
        self.assertEqual(package["main"], "electron/main.js")
        self.assertEqual(package["productName"], "PETCIRCLE跨境工作台")
        self.assertIn("BrowserWindow", main_source)
        self.assertIn("daily_ops_cli.py", main_source)
        self.assertNotIn("http://127.0.0.1", main_source)
        self.assertNotIn("listen(", main_source)

    def test_petcircle_branding_and_icons_are_wired_into_desktop_app(self):
        main_source = (ROOT / "electron" / "main.js").read_text(encoding="utf-8")
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        css = (ROOT / "electron" / "renderer.css").read_text(encoding="utf-8")
        self.assertIn("PETCIRCLE跨境工作台", main_source)
        self.assertIn("function appIcon", main_source)
        self.assertIn("petcircle-app-icon.icns", main_source)
        self.assertIn("petcircle-app-icon.png", html)
        self.assertIn("data-fallback-label", html)
        self.assertIn("installImageFallbacks", (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8"))
        self.assertIn("PETCIRCLE跨境工作台", html)
        self.assertIn("全托管日常运营", html)
        self.assertIn("brand-logo", css)

    def test_petcircle_dock_icon_uses_transparent_rounded_corners(self):
        icon_path = ROOT / "electron" / "assets" / "petcircle-app-icon.png"
        if not icon_path.exists():
            self.skipTest("图片产物不要求进入 git；界面已有缺图兜底")
        tracked = subprocess.run(
            ["git", "ls-files", "--error-unmatch", str(icon_path.relative_to(ROOT))],
            cwd=ROOT,
            text=True,
            capture_output=True,
        )
        if tracked.returncode != 0:
            self.skipTest("图片产物不要求进入 git；界面已有缺图兜底")
        with Image.open(icon_path) as image:
            self.assertEqual(image.mode, "RGBA")
            width, height = image.size
            for point in [(0, 0), (width - 1, 0), (0, height - 1), (width - 1, height - 1)]:
                self.assertEqual(image.getpixel(point)[3], 0)

    def test_macos_app_launcher_opens_without_terminal(self):
        app_dir = ROOT / "PETCIRCLE跨境工作台.app"
        plist = app_dir / "Contents" / "Info.plist"
        launcher = app_dir / "Contents" / "MacOS" / "launcher"
        icon = app_dir / "Contents" / "Resources" / "petcircle-app-icon.icns"
        launcher_source = (ROOT / "electron" / "macos_launcher.c").read_text(encoding="utf-8")
        self.assertIn("node_modules/electron", launcher_source)
        self.assertNotIn("Terminal", launcher_source)
        if app_dir.exists():
            self.assertTrue(plist.exists())
            self.assertTrue(launcher.exists())
            self.assertTrue(icon.exists())
            self.assertTrue(launcher.stat().st_mode & 0o111)
            plist_text = plist.read_text(encoding="utf-8")
            self.assertIn("PETCIRCLE跨境工作台", plist_text)
            self.assertIn("petcircle-app-icon", plist_text)

    def test_desktop_adapter_upload_records_pending_batch_without_http(self):
        import daily_ops_desktop_adapter as adapter

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            source.write_bytes(b"xlsx")
            target_dir = root / "uploads"
            target_dir.mkdir()
            manifest = root / "manifest.json"

            with patch.dict(daily_ops_app.UPLOAD_TARGETS, {"demo": ("测试数据源", target_dir)}, clear=True), \
                 patch.object(daily_ops_app, "DATA_SOURCE_MANIFEST", manifest):
                result = adapter.import_source_files("demo", [source])

            self.assertEqual(result["count"], 1)
            self.assertEqual(result["files"][0]["file"], "source.xlsx")
            self.assertTrue((target_dir / "source.xlsx").exists())
            self.assertIn("pending_batches", manifest.read_text(encoding="utf-8"))

    def test_desktop_adapter_owner_export_sales_is_scoped_to_owner(self):
        import daily_ops_desktop_adapter as adapter

        assignments = [
            {"platform": "Temu", "store": "七弟", "owner": "小琴"},
            {"platform": "Shein", "store": "琪琪", "owner": "胡娟"},
        ]
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            sales_db = root / "daily_sales.json"
            output_dir = root / "outputs"
            with patch.object(adapter, "SALES_DB_PATH", sales_db), \
                 patch.object(daily_ops_app, "OUTPUT_DIR", output_dir), \
                 patch.object(daily_ops_app, "load_store_owner_assignments", return_value=assignments):
                adapter.submit_sales_payload({
                    "role": "admin",
                    "user": "管理员",
                    "date": "2026-06-23",
                    "platform": "Temu",
                    "store": "七弟",
                    "sales": "12",
                })
                adapter.submit_sales_payload({
                    "role": "admin",
                    "user": "管理员",
                    "date": "2026-06-23",
                    "platform": "Shein",
                    "store": "琪琪",
                    "sales": "88",
                })
                result = adapter.export_sales_payload({
                    "role": "owner",
                    "user": "小琴",
                    "date": "2026-06-23",
                })

            workbook = load_workbook(result["path"], data_only=True)
            detail = workbook["每日销量明细"]
            summary = workbook["平台汇总"]
            self.assertIn("小琴", result["file"])
            self.assertEqual(detail.max_row, 2)
            self.assertEqual(detail["C2"].value, "七弟")
            self.assertEqual(detail["D2"].value, "小琴")
            self.assertEqual(summary.max_row, 2)
            self.assertEqual(summary["A2"].value, "Temu")

    def test_desktop_adapter_owner_sales_compare_is_scoped_to_owner(self):
        import daily_ops_desktop_adapter as adapter

        assignments = [
            {"platform": "Temu", "store": "七弟", "owner": "小琴"},
            {"platform": "Temu", "store": "童话", "owner": "胡娟"},
        ]
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            sales_db = root / "daily_sales.json"
            source = root / "temu_sales.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["店铺", "SKC", "7天销量"])
            ws.append(["七弟", "A001", 700])
            ws.append(["童话", "B001", 700])
            wb.save(source)

            with patch.object(adapter, "SALES_DB_PATH", sales_db), \
                 patch.object(daily_ops_app, "load_store_owner_assignments", return_value=assignments), \
                 patch.object(daily_ops_app, "temu_sales_files", return_value=[source]), \
                 patch.object(daily_ops_app, "shein_platform_files", return_value=[]):
                adapter.submit_sales_payload({
                    "role": "admin",
                    "user": "管理员",
                    "date": "2026-06-23",
                    "platform": "Temu",
                    "store": "七弟",
                    "sales": "150",
                })
                adapter.submit_sales_payload({
                    "role": "admin",
                    "user": "管理员",
                    "date": "2026-06-23",
                    "platform": "Temu",
                    "store": "童话",
                    "sales": "150",
                })
                result = adapter.sales_compare_payload({
                    "role": "owner",
                    "user": "小琴",
                    "date": "2026-06-23",
                })

            self.assertEqual(result["summary"]["checked"], 1)
            self.assertEqual(result["summary"]["alerts"], 1)
            self.assertEqual([row["store"] for row in result["rows"]], ["七弟"])
            self.assertEqual([row["owner"] for row in result["rows"]], ["小琴"])

    def test_desktop_adapter_imports_master_data_and_sales_history(self):
        import daily_ops_desktop_adapter as adapter

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            owner_file = root / "店铺负责人对应表.xlsx"
            owner_wb = Workbook()
            owner_ws = owner_wb.active
            owner_ws.append(["店铺", "店名", "业务", "平台"])
            owner_ws.append(["一弟", "Apetcircle", "小琴", "temu"])
            owner_wb.save(owner_file)

            sales_file = root / "跨境运营总表new.xlsx"
            sales_wb = Workbook()
            sales_wb.active.title = "总览"
            month = sales_wb.create_sheet("2606")
            month.append(["2026年", "汇总", "大弟"])
            month.append([46174, 100, 100])
            sales_wb.save(sales_file)

            owner_map = root / "store_owner_map.json"
            accounts = root / "operator_accounts.json"
            sales_db = root / "daily_sales.json"
            review_file = root / "基础信息导入整理表.xlsx"

            with patch.object(daily_ops_app, "STORE_OWNER_MAP_FILE", owner_map), \
                 patch.object(daily_ops_app, "OPERATOR_ACCOUNTS_FILE", accounts), \
                 patch.object(daily_ops_app, "DAILY_SALES_FILE", sales_db), \
                 patch.object(daily_ops_app, "MASTER_IMPORT_REVIEW_FILE", review_file), \
                 patch.object(adapter, "SALES_DB_PATH", sales_db):
                owner_result = adapter.import_owner_master_payload({"role": "admin", "user": "管理员", "path": str(owner_file)})
                sales_result = adapter.import_sales_history_payload({"role": "admin", "user": "管理员", "path": str(sales_file)})
                report = adapter.sales_report_payload({"role": "admin", "user": "管理员", "platform": "Temu", "store": "一弟", "date_from": "2026-06-01", "date_to": "2026-06-30"})

            self.assertEqual(owner_result["assignment_count"], 1)
            self.assertEqual(owner_result["account_count"], 1)
            self.assertEqual(sales_result["created"], 1)
            self.assertEqual(report["summary"]["total_sales"], 100)

    def test_mac_launcher_uses_desktop_entrypoint_not_local_url(self):
        text = (ROOT / "启动日常运营工作台.command").read_text(encoding="utf-8")
        self.assertIn("npm start", text)
        self.assertIn("package.json", text)
        self.assertNotIn("http://127.0.0.1", text)
        self.assertNotIn("lsof -ti tcp", text)

    def test_electron_renderer_matches_selected_workbench_reference(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        css = (ROOT / "electron" / "renderer.css").read_text(encoding="utf-8")
        js = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        for text in ["PETCIRCLE跨境工作台", "全托管日常运营", "今日工作台", "销量管理", "商品任务", "数据导入", "经营报表", "输出记录"]:
            self.assertIn(text, html + js)
        self.assertIn("sidebar", html)
        self.assertIn("source-table", html)
        self.assertIn("queue-panel", html)
        self.assertIn("--blue", css)
        self.assertIn("grid-template-columns: 248px minmax(0, 1fr)", css)

    def test_master_data_keeps_builtin_platforms_and_erp_limits_visible(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        js = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        for platform in ["Temu", "Shein", "速卖通", "TK", "Ozon"]:
            self.assertIn(platform, html)
            self.assertIn(platform, js)
        self.assertIn('data-erp-field="page_size" type="number" min="1" max="1000"', html)
        self.assertIn('data-erp-field="stock_limit" type="number" min="100" max="10000"', html)

    def test_master_data_uses_separate_secondary_modules(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        js = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        css = (ROOT / "electron" / "renderer.css").read_text(encoding="utf-8")
        self.assertIn('data-master-module="master-import"', html)
        self.assertIn('data-master-panel="master-import"', html)
        for page in ["operatorAccountsPage", "storeInfoPage", "productInfoPage", "taskSuppressionsPage"]:
            self.assertIn(page, html)
        self.assertIn('data-page="operatorAccounts"', html)
        self.assertIn('data-page="storeInfo"', html)
        self.assertIn('operatorAccounts: "operatorAccountsPage"', js)
        self.assertIn('storeInfo: "storeInfoPage"', js)
        self.assertIn('if (next === "operatorAccounts") loadOperatorAccounts(false);', js)
        self.assertIn('if (next === "storeInfo") loadStoreOwners();', js)
        self.assertIn("masterModuleDialog", html)
        self.assertIn("openMasterModule", js)
        self.assertIn("closeMasterModule", js)
        self.assertIn("master-module-card", css)
        self.assertIn("商品信息查询", html)
        self.assertIn("loadProductInfo", js)
        self.assertNotIn('data-master-panel="product-info"', html)
        self.assertNotIn('data-master-panel="operator-accounts"', html)
        self.assertNotIn('data-master-panel="store-info"', html)
        self.assertNotIn("基础数据查询", html)
        self.assertNotIn("searchBtn", html + js)

    def test_master_data_replaces_large_checks_with_store_toggles_and_manual_accounts(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        js = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        css = (ROOT / "electron" / "renderer.css").read_text(encoding="utf-8")
        self.assertIn("store-toggle", js)
        self.assertIn(".store-toggle input:checked + span", css)
        self.assertIn("newOperatorOwner", html)
        self.assertIn("newOperatorRole", html)
        self.assertIn("newOperatorStoreList", html)
        self.assertIn("data-account-role", js)
        self.assertIn("data-account-store", js)
        self.assertIn("createOperatorAccountBtn", html)
        self.assertIn("createOperatorAccount", js)
        self.assertIn("updateOperatorAccount", js)
        self.assertIn("updateOperatorAccount:", (ROOT / "electron" / "preload.js").read_text(encoding="utf-8"))
        self.assertIn("deleteOperatorAccount", js)
        self.assertIn("deleteOperatorAccount:", (ROOT / "electron" / "preload.js").read_text(encoding="utf-8"))
        self.assertIn("erpProductInfo", (ROOT / "electron" / "preload.js").read_text(encoding="utf-8"))

    def test_system_settings_are_grouped_into_admin_modules_with_thresholds(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        js = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        for module in ["field-rules", "sales-thresholds", "erp-settings", "system-maintenance"]:
            self.assertIn(f'data-settings-module="{module}"', html)
            self.assertIn(f'data-settings-panel="{module}"', html)
        for field in [
            "daily_diff_units",
            "month_diff_percent",
            "year_diff_percent",
            "completion_yellow_percent",
            "completion_red_percent",
            "erp_yellow_days",
            "platform_batch_yellow_time",
            "platform_batch_red_time",
        ]:
            self.assertIn(f'data-rule="{field}"', html + js)
        self.assertIn("备份当前系统数据", html)
        self.assertIn("检查系统是否可正常运行", html)
        self.assertIn("openSettingsModule", js)

    def test_business_report_defaults_to_manual_30_day_period_with_tooltips(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        js = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        self.assertIn('data-business-source="manual"', html)
        self.assertIn('data-business-range="30d"', html)
        self.assertIn("最近30日销量", html + js)
        self.assertIn("不含今日", html + js)
        self.assertIn("businessDefinition", js)
        self.assertIn("店长填报完整度", html + js)
        self.assertNotIn("今日销量", html + js)

    def test_business_report_first_screen_highlights_actions_and_reference_source(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        js = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        for text in [
            "businessActionList",
            "businessMoverGrid",
            "平台导入参考",
            "不作为月结主口径",
            "renderBusinessActionList",
            "renderBusinessMovers",
            "data-business-action",
            "导入覆盖店铺数",
            "未匹配负责人",
        ]:
            self.assertIn(text, html + js)

    def test_today_workbench_puts_daily_flow_first_without_slogan_hero(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        js = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        self.assertIn("homeBusinessOverview", html)
        self.assertIn("todayActionList", html)
        self.assertIn("todayWorkflowSteps", html)
        self.assertIn("今日流程", html + js)
        self.assertLess(html.index("homeBusinessOverview"), html.index("todayActionList"))
        self.assertLess(html.index("todayActionList"), html.index("todayWorkflowSteps"))
        for phrase in ["先填销售日销量", "先确认销量，再处理任务包", "再整包处理任务"]:
            self.assertNotIn(phrase, html + js)

    def test_weekly_workflow_uses_table_workbench_layout(self):
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        self.assertIn("renderSources", source)
        self.assertIn("selectFiles", source)
        self.assertIn("uploadSource", source)
        self.assertIn("finishUpload", source)
        self.assertIn("renderReportQueue", source)

    def test_source_upload_state_is_visible_in_table_rows(self):
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        css = (ROOT / "electron" / "renderer.css").read_text(encoding="utf-8")
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        self.assertIn("sourceProgress", source)
        self.assertIn("renderSourceProgress", source)
        self.assertIn("renderSourceRecompute", source)
        self.assertIn("recomputeSource", source)
        self.assertIn("重算关联报表", source)
        self.assertIn("api.recomputeSource", source)
        self.assertIn("已选择", source)
        self.assertIn("上传成功", source)
        self.assertIn("上传失败", source)
        self.assertIn("待结束上传", source)
        self.assertIn("source-progress", css)
        self.assertIn("source-recompute", css)
        self.assertIn("经营报表仍默认使用店长填报销量", source)
        self.assertIn("summary_only: true", source)
        self.assertNotIn("<h2>平台与店铺</h2>", html)
        self.assertNotIn("<h2>店铺负责人配置</h2>", html)

    def test_report_cards_include_direct_download_actions(self):
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        css = (ROOT / "electron" / "renderer.css").read_text(encoding="utf-8")
        self.assertIn("latestOutputForReport", source)
        self.assertIn("download-report", source)
        self.assertIn("打开表格", source)
        self.assertIn("打开所在文件夹", source)
        self.assertIn("download-actions", css)

    def test_weekly_report_queue_shows_completion_status_compactly(self):
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        css = (ROOT / "electron" / "renderer.css").read_text(encoding="utf-8")
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        self.assertIn("completionSummary", source)
        self.assertIn("queue-item-done", source)
        self.assertIn("queue-item-todo", source)
        self.assertIn("已完成", source)
        self.assertIn("未完成", source)
        self.assertIn("生成就绪报表/任务", html)
        self.assertIn("report-readiness", css)
        self.assertIn("min-height: 62px", css)

    def test_weekly_report_queue_shows_task_badges(self):
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        css = (ROOT / "electron" / "renderer.css").read_text(encoding="utf-8")
        self.assertIn("reportTaskBadges", source)
        self.assertIn("queue-task-badges", source)
        self.assertIn("queue-task-badge", css)
        for text in ["任务", "待店长", "待确认"]:
            self.assertIn(text, source)

    def test_bargain_low_price_trace_tab_has_real_admin_workflow(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        self.assertIn('data-bargain-tab="lowprice"', html)
        self.assertIn("renderBargainLowPriceTrace", source)
        self.assertIn("runBargainLowPriceTrace", source)
        self.assertIn("ignoreBargainLowPrice", source)
        self.assertIn("api.bargainLowPriceTrace", source)
        self.assertIn("api.bargainIgnoreLowPrice", source)
        for text in ["低价回追", "重新检查", "忽略", "历史审批价", "风险原因"]:
            self.assertIn(text, source)

    def test_bargain_form_uses_store_dropdown_price_reuse_and_batch_review(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        self.assertIn('<select id="bargainStore"', html)
        self.assertIn('id="bargainLookupControls"', html)
        self.assertIn("renderBargainStoreOptions", source)
        self.assertIn("bargainLookupControls", source)
        self.assertIn("event.preventDefault()", source)
        self.assertIn("syncBargainPriceToGoods", source)
        self.assertIn("同步到本款全部尺码", source)
        self.assertIn("bargainHistoryDialog", html)
        self.assertIn("subpage-panel", html)
        self.assertIn("third-level-panel", html)
        self.assertIn("openBargainHistoryDialog", source)
        self.assertIn('currentOperator().role !== "owner"', source)
        self.assertIn('openBargainHistoryDialog("pending")', source)
        self.assertIn('data-bargain-tab="pending"', source)
        self.assertIn("openBargainPendingBtn", html)
        self.assertIn("bargainFilterDateFrom", html)
        self.assertIn("bargainFilterDateTo", html)
        self.assertIn("renderBargainPendingTable", source)
        self.assertIn("卖得最好店铺30天销量", source)
        self.assertIn("同款在售链接数", source)
        self.assertIn("bargainSelectAll", html)
        self.assertIn("reviewSelectedBargains", source)
        self.assertIn("可空", source)
        self.assertIn("api.storeOwners(operatorPayload())", source)
        self.assertNotIn("state.storeOwners = [];", source)

    def test_bargain_risk_recomputes_from_typed_price(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        self.assertIn("data-bargain-risk", source)
        self.assertIn("bargainComputedRisk", source)
        self.assertIn("低于批发价80%", source)
        self.assertIn("ERP成本缺失", source)

    def test_owner_sales_entry_defaults_to_yesterday_and_marks_imported_rows_editable(self):
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        css = (ROOT / "electron" / "renderer.css").read_text(encoding="utf-8")
        self.assertIn("function salesDefaultDateText", source)
        self.assertIn("date.setDate(date.getDate() - 1)", source)
        self.assertIn("历史导入待确认", source)
        self.assertIn("sales-day-table", source)
        self.assertIn("salesEditingIndex", source)
        self.assertIn("确认", source)
        self.assertIn("sales-entry-pending", css)

    def test_home_dashboard_shows_manual_7_30_90_day_business_overview(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        self.assertIn("homeBusinessOverview", html)
        self.assertIn("homeBusinessReports", source)
        self.assertIn('"7d", "30d", "90d"', source)
        self.assertIn('source: "manual"', source)
        for text in ["最近7天销量", "最近30天销量", "最近90天销量", "Temu 爆旺款 SKC", "Shein 爆款 SKC"]:
            self.assertIn(text, source + html)

    def test_important_assets_archive_is_available_to_home_and_settings(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        preload = (ROOT / "electron" / "preload.js").read_text(encoding="utf-8")
        main = (ROOT / "electron" / "main.js").read_text(encoding="utf-8")
        cli = (ROOT / "daily_ops_cli.py").read_text(encoding="utf-8")
        adapter = (ROOT / "daily_ops_desktop_adapter.py").read_text(encoding="utf-8")
        for text in ["assetOverview", "exportAssetArchive", "importAssetArchive"]:
            self.assertIn(text, source + preload)
        for channel in ["api:asset-overview", "api:export-asset-archive", "api:import-asset-archive"]:
            self.assertIn(channel, main)
        for command in ["asset-overview", "export-asset-archive", "import-asset-archive"]:
            self.assertIn(command, cli)
        for text in ["important_asset", "重要资产", "导出资产存档", "初始化导入"]:
            self.assertIn(text, html + source + adapter)
        refresh_body = source[source.index("async function refreshAll"):]
        self.assertLess(refresh_body.index("await loadAssetOverview(false)"), refresh_body.index("renderTodayDashboard()"))

    def test_erp_settings_show_manual_pull_choices(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        self.assertIn('data-erp-field="sync_product_archive"', html)
        self.assertIn('data-erp-field="sync_stock_snapshot"', html)
        self.assertIn('data-erp-field="sync_available_stock"', html)
        self.assertIn('data-erp-field="sync_shop_query"', html)
        self.assertIn('data-erp-field="sync_platform_goods"', html)
        self.assertIn('data-erp-field="sync_sales_outbound"', html)
        self.assertIn("销售出库单", html)
        self.assertIn("保存 ERP 设置", html)
        self.assertIn("按选择同步 ERP", html)
        self.assertIn("sync_product_archive", source)

    def test_erp_manual_sync_blocked_status_does_not_show_success_pages(self):
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        start = source.index("async function manualErpSync()")
        end = source.index("\nasync function createBackup()", start)
        body = source[start:end]
        self.assertIn('result.status === "blocked"', body)
        self.assertIn('setErpStatus("running", "ERP 同步未开始"', body)
        self.assertLess(body.index('result.status === "blocked"'), body.index("const pages ="))

    def test_erp_settings_shows_latest_data_overview(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        source = (ROOT / "electron" / "renderer.js").read_text(encoding="utf-8")
        self.assertIn("erpLatestDataOverview", html)
        self.assertIn("最新 ERP 数据", html)
        self.assertIn("renderErpLatestOverview", source)
        for text in ["最新成功", "商品资料", "库存快照", "可用库存", "文件位置"]:
            self.assertIn(text, source + html)

    def test_bargain_draft_table_has_horizontal_scroll_boundary(self):
        html = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        css = (ROOT / "electron" / "renderer.css").read_text(encoding="utf-8")
        self.assertIn("bargain-table-wrap", html)
        self.assertRegex(css, r"\.bargain-table-wrap\s*\{[\s\S]*overflow-x:\s*auto")
        self.assertRegex(css, r"\.bargain-table\s*\{[\s\S]*min-width:\s*1120px")

    def test_desktop_shell_matches_workbench_reference(self):
        source = (ROOT / "electron" / "renderer.html").read_text(encoding="utf-8")
        for page in ["today", "sales", "tasks", "imports", "reports", "masterdata", "rules"]:
            self.assertIn(f'data-page="{page}"', source)

    def test_python_cli_returns_json_for_electron(self):
        result = subprocess.run(
            [sys.executable, str(ROOT / "daily_ops_cli.py"), "reports"],
            cwd=ROOT,
            text=True,
            capture_output=True,
            check=True,
        )
        payload = json.loads(result.stdout)
        self.assertTrue(payload["ok"])
        self.assertIn("data", payload)

    def test_delivery_scripts_use_petcircle_cross_border_name(self):
        import build_windows_install_package_v2 as win_pack

        files = [
            ROOT / "build_portable_package.py",
            ROOT / "build_windows_install_package_v2.py",
            ROOT / "daily_ops_desktop.py",
        ]
        combined = "\n".join(path.read_text(encoding="utf-8") for path in files)
        self.assertIn("PETCIRCLE跨境工作台", combined)
        self.assertIn("PETCIRCLECrossBorderWorkbench", combined)
        for legacy in ["正在安装日常运营工作台", "日常运营工作台 v2.0", "DailyOpsWorkbench_v2.0_Setup"]:
            self.assertNotIn(legacy, combined)

        with TemporaryDirectory() as tmp:
            package_root = Path(tmp)
            win_pack.make_install_scripts(package_root)
            scripts = "\n".join(path.read_text(encoding="utf-8") for path in (package_root / "scripts").glob("*"))
            self.assertIn("PETCIRCLE跨境工作台 v2.0", scripts)
            self.assertIn("PETCIRCLECrossBorderWorkbenchV2", scripts)
            self.assertNotIn("{APP_NAME}", scripts)
            self.assertNotIn("{APP_ID}", scripts)


if __name__ == "__main__":
    unittest.main()
