import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path.cwd()
TEMU_DIR = ROOT / "temu数据源表"
ERP_DIR = ROOT / "erp数据源"

HOT_FILE = TEMU_DIR / "Temu爆旺款.xlsx"
HOT_FILES = None
SALES_FILES = [
    TEMU_DIR / "20260601-Temu仓库销售情况导出 (1).xlsx",
    TEMU_DIR / "20260601-Temu仓库销售情况导出 (2).xlsx",
]
OWNER_FILE = ROOT / "店铺负责人对应表.xlsx"
ERP_FILES = sorted(ERP_DIR.glob("erp产品基础信息表 (*.xlsx)"))
COMBO_FILE = ERP_DIR / "erp产品组合装基础信息表.xlsx"
LEGACY_REPORTS = [
    ROOT / "Temu爆旺款重复预警表_V1正式版.xlsx",
    ROOT / "最新Temu爆旺款重复预警表_20260602.xlsx",
    ROOT / "Temu爆旺款重复铺货表_20260601_5表版.xlsx",
]

OUTPUT_FILE = ROOT / "Temu爆旺款重复预警表_V1.3正式版.xlsx"

VERSION = "爆旺款重复预警V1.3正式版"

SIZE_RE = re.compile(
    r"(?i)([-_\s]*)(XXXXL|XXXL|XXL|XXS|XL|XS|2XL|3XL|4XL|5XL|S|M|L)$"
)
SIZE_ORDER = ["XXS", "XS", "S", "M", "L", "XL", "XXL", "XXXL"]

OWNER_KEEP = {"洁琳", "胡娟", "小琴"}
SHOP_ORDER = {
    "1": 1,
    "2": 2,
    "3": 3,
    "4": 4,
    "5": 5,
    "6": 6,
    "7": 7,
    "8": 8,
    "9": 9,
    "10": 10,
    "11": 11,
    "12": 12,
    "13": 13,
    "15": 15,
}

TYPE_FLAT = "平销款冲突爆款"
TYPE_BOMB = "爆款互相冲突"
YES = "是"
NO = "否"
UNKNOWN = "无法判断"
ACT_NOW = "立即下架！"
ACT_SELL_OUT = "售完备货库存下架。"
ACT_BOMB_STOCK = "禁止备货，30天内限时下架，清理库存。"


def text(value):
    return "" if value is None else str(value).strip()


def number(value):
    if value is None:
        return 0.0
    try:
        raw = str(value).replace(",", "").strip()
        if raw in {"", "--", "-"}:
            return 0.0
        return float(raw)
    except ValueError:
        return 0.0


def sku_code(value):
    return text(value).split("@", 1)[0].strip()


def product_code(value):
    code = sku_code(value)
    previous = None
    while previous != code:
        previous = code
        code = SIZE_RE.sub("", code)
    return code


def size_rank(value):
    sku = sku_code(value)
    match = SIZE_RE.search(sku)
    if not match:
        return (999, "")
    size = match.group(2).upper()
    rank = {name.upper(): idx for idx, name in enumerate(SIZE_ORDER)}
    return (rank.get(size, 999), size)


def normalize_shop(value):
    shop = text(value)
    for suffix in ["（毛衣）", "（喵喵）", "（节日）"]:
        shop = shop.replace(suffix, "")
    return {
        "十一": "十一弟",
        "十二": "十二弟",
        "十三": "十三弟",
        "十五": "十五弟",
    }.get(shop, shop)


def shop_no(value):
    shop = normalize_shop(value)
    mapping = {
        "弟弟": "1",
        "一弟": "1",
        "二弟": "2",
        "三弟": "3",
        "四弟": "4",
        "五弟": "5",
        "六弟": "6",
        "七弟": "7",
        "八弟": "8",
        "九弟": "9",
        "十弟": "10",
        "十一弟": "11",
        "十二弟": "12",
        "十三弟": "13",
        "十五弟": "15",
    }
    return mapping.get(shop, shop)


def sort_shop_no(value):
    return SHOP_ORDER.get(str(value), 999)


def worksheet_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    ws.reset_dimensions()
    rows = ws.iter_rows(values_only=True)
    headers = [text(v) for v in next(rows)]
    idx = {h: i for i, h in enumerate(headers) if h}
    try:
        for row_number, row in enumerate(rows, start=2):
            if not any(row):
                continue
            yield row_number, row, idx
    finally:
        wb.close()


def get(row, idx, header):
    col = idx.get(header)
    if col is None or col >= len(row):
        return None
    return row[col]


def get_by_header_or_pos(row, idx, header, pos_1based):
    value = get(row, idx, header)
    if value not in (None, ""):
        return value
    pos = pos_1based - 1
    if pos < 0 or pos >= len(row):
        return None
    return row[pos]


def load_owners():
    owners = {}
    for _, row, idx in worksheet_rows(OWNER_FILE):
        shop = normalize_shop(get(row, idx, "店铺"))
        owner = text(get(row, idx, "业务"))
        if shop and owner:
            owners[shop] = owner
    owners.setdefault("弟弟", "小琴")
    return owners


def load_erp():
    erp = {}

    def set_erp(code_value, goods, spec):
        raw_code = sku_code(code_value)
        if not raw_code:
            return
        base_code = product_code(raw_code)
        for key in {raw_code, base_code}:
            current = erp.setdefault(key, {"goods": "", "spec": ""})
            if not current["goods"] and text(goods):
                current["goods"] = text(goods)
            if not current["spec"] and text(spec):
                current["spec"] = text(spec)

    for path in ERP_FILES:
        for _, row, idx in worksheet_rows(path):
            set_erp(
                get_by_header_or_pos(row, idx, "商家编码（新）", 8),
                get_by_header_or_pos(row, idx, "货品名称", 6),
                get_by_header_or_pos(row, idx, "规格名称", 7),
            )

    if COMBO_FILE.exists():
        for _, row, idx in worksheet_rows(COMBO_FILE):
            set_erp(
                get_by_header_or_pos(row, idx, "商家编码（新）", 2),
                get_by_header_or_pos(row, idx, "组合装名称", 3),
                "",
            )

    for path in LEGACY_REPORTS:
        if not path.exists():
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                ws.reset_dimensions()
                rows = ws.iter_rows(values_only=True)
                try:
                    headers = [text(v) for v in next(rows)]
                except StopIteration:
                    continue
                idx = {h: i for i, h in enumerate(headers) if h}
                code_col = idx.get("商家编码")
                goods_col = idx.get("货品名称") or idx.get("erp货品名称") or idx.get("ERP货品名称")
                spec_col = idx.get("erp规格") or idx.get("ERP规格名称") or idx.get("规格名称")
                if code_col is None or goods_col is None:
                    continue
                for row in rows:
                    if code_col >= len(row):
                        continue
                    raw_code = text(row[code_col])
                    base_code = product_code(raw_code)
                    if not base_code:
                        continue
                    goods = text(row[goods_col]) if goods_col < len(row) else ""
                    spec = text(row[spec_col]) if spec_col is not None and spec_col < len(row) else ""
                    if not goods:
                        continue
                    current = erp.setdefault(base_code, {"goods": "", "spec": ""})
                    if not current["goods"]:
                        current["goods"] = goods
                    if not current["spec"] and spec:
                        current["spec"] = spec
        finally:
            wb.close()

    return erp


def read_hot_rows():
    rows = []
    for path in [Path(path) for path in (HOT_FILES or [HOT_FILE])]:
        for row_number, row, idx in worksheet_rows(path):
            sku = sku_code(get(row, idx, "SKU货号"))
            base = product_code(sku)
            if not sku or not base:
                continue
            rows.append(
                {
                    "source_row": row_number,
                    "sku_code": sku,
                    "product_code": base,
                    "shop_raw": text(get(row, idx, "店铺")),
                    "shop": normalize_shop(get(row, idx, "店铺")),
                    "skc": text(get(row, idx, "SKC")),
                    "price": number(get(row, idx, "申报价格")),
                    "sales30": number(get(row, idx, "30天销量")),
                    "stock": number(get(row, idx, "可用")),
                }
            )
    return rows


def read_sales_rows(hot_product_codes):
    rows = []
    for path in SALES_FILES:
        for row_number, row, idx in worksheet_rows(path):
            sku = sku_code(get(row, idx, "SKU货号"))
            base = product_code(sku)
            if not sku or base not in hot_product_codes:
                continue
            rows.append(
                {
                    "source_file": path.name,
                    "source_row": row_number,
                    "sku_code": sku,
                    "product_code": base,
                    "shop_raw": text(get(row, idx, "店铺")),
                    "shop": normalize_shop(get(row, idx, "店铺")),
                    "skc": text(get(row, idx, "SKC")),
                    "price": number(get(row, idx, "申报价格")),
                    "sales30": number(get(row, idx, "30天销量")),
                    "stock": number(get(row, idx, "可用")),
                }
            )
    return rows


def aggregate_groups(rows):
    groups = {}
    members = defaultdict(list)
    for row in rows:
        key = (row["product_code"], row["shop"], row["skc"])
        members[key].append(row)

    for key, items in members.items():
        product, shop, skc = key
        prices = [r["price"] for r in items if r["price"] > 0]
        groups[key] = {
            "product_code": product,
            "shop": shop,
            "shop_raw": items[0]["shop_raw"],
            "skc": skc,
            "price": max(prices) if prices else 0.0,
            "sales30": sum(r["sales30"] for r in items),
            "stock": sum(r["stock"] for r in items),
            "rows": sorted(items, key=lambda r: r["sku_code"]),
            "sku_codes": sorted({r["sku_code"] for r in items}),
        }
    return groups


def choose_champions(hot_groups):
    by_product = defaultdict(list)
    for group in hot_groups.values():
        by_product[group["product_code"]].append(group)

    champions = {}
    for product, groups in by_product.items():
        champions[product] = sorted(
            groups,
            key=lambda g: (g["sales30"], g["price"], g["stock"]),
            reverse=True,
        )[0]
    return champions


def is_low(price, reference):
    if price <= 0 or reference <= 0:
        return UNKNOWN
    return YES if price < reference else NO


def action_for(row, champion, is_bomb_conflict):
    if is_bomb_conflict:
        return ACT_BOMB_STOCK if row["stock"] > 0 else ACT_NOW
    if is_low(row["price"], champion["price"]) == YES:
        return ACT_NOW
    if row["stock"] > 0:
        return ACT_SELL_OUT
    return ACT_NOW


def build_rows(owners, erp, hot_groups, sales_groups, champions):
    operation_rows = []
    seen = set()

    hot_group_keys = set(hot_groups)
    for key, row in sales_groups.items():
        product = row["product_code"]
        champion = champions.get(product)
        if not champion:
            continue
        champion_key = (product, champion["shop"], champion["skc"])
        if key == champion_key:
            continue

        owner = owners.get(row["shop"], "")
        if owner not in OWNER_KEEP:
            continue

        is_bomb_conflict = key in hot_group_keys
        conflict_type = TYPE_BOMB if is_bomb_conflict else TYPE_FLAT
        erp_info = erp.get(product) or erp.get(row["sku_codes"][0], {}) or {}
        champion_rows_by_size = {}
        for champion_row in champion.get("rows", []):
            size = re.sub(r"^.*[-_\s]", "", champion_row["sku_code"].upper())
            champion_rows_by_size.setdefault(size, champion_row)
        champion_fallback = (champion.get("rows") or [champion])[0]

        for dup_row in row.get("rows", []):
            dup_sku = dup_row["sku_code"]
            dup_size = re.sub(r"^.*[-_\s]", "", dup_sku.upper())
            champion_row = champion_rows_by_size.get(dup_size, champion_fallback)
            hot_price = number(champion_row.get("price")) or champion["price"]
            hot_sales = number(champion_row.get("sales30")) if "sales30" in champion_row else champion["sales30"]
            hot_stock = number(champion_row.get("stock")) if "stock" in champion_row else champion["stock"]
            dup_price = number(dup_row.get("price"))
            dup_sales = number(dup_row.get("sales30"))
            dup_stock = number(dup_row.get("stock"))
            sku_row = {
                **row,
                "price": dup_price,
                "sales30": dup_sales,
                "stock": dup_stock,
            }
            action = action_for(sku_row, {**champion, "price": hot_price}, is_bomb_conflict)
            sku_erp = erp.get(dup_sku) or erp_info

            dedupe = (dup_sku, row["shop"], row["skc"], champion["shop"], champion["skc"], action)
            if dedupe in seen:
                continue
            seen.add(dedupe)

            operation_rows.append(
                {
                    "商家编码": dup_sku,
                    "货品名称": sku_erp.get("goods", ""),
                    "skc": row["skc"],
                    "所属店铺": shop_no(row["shop"]),
                    "爆旺款skc": champion["skc"],
                    "爆旺款店铺": champion["shop_raw"],
                    "爆旺款申报价": hot_price,
                    "重复款申报价": dup_price,
                    "爆旺款月销": hot_sales,
                    "重复款月销": dup_sales,
                    "是否低于爆旺款申报价": is_low(dup_price, hot_price),
                    "爆旺款平台仓备货库存": hot_stock,
                    "重复款备货库存": dup_stock,
                    "负责人": owner,
                    "冲突类型": conflict_type,
                    "处理意见": action,
                }
            )

    operation_rows.sort(
        key=lambda r: (
            product_code(r["商家编码"]),
            r["skc"],
            size_rank(r["商家编码"]),
            r["商家编码"],
            sort_shop_no(r["所属店铺"]),
            r["爆旺款skc"],
        )
    )
    return operation_rows


def build_overview(operation_rows, hot_groups, owners):
    store_names = {}
    for group in hot_groups.values():
        store_names.setdefault(shop_no(group["shop"]), group["shop_raw"])

    overview = defaultdict(
        lambda: {
            "bomb": set(),
            "repeat": 0,
            "flat": 0,
            "bomb_conflict": 0,
            "low": 0,
            "not_low": 0,
            "now": 0,
            "sell_out": 0,
            "clear_30": 0,
        }
    )

    for group in hot_groups.values():
        owner = owners.get(group["shop"], "")
        if owner in OWNER_KEEP:
            no = shop_no(group["shop"])
            overview[(no, group["shop_raw"], owner)]["bomb"].add(
                (group["product_code"], group["skc"])
            )

    for row in operation_rows:
        no = str(row["所属店铺"])
        shop_name = store_names.get(no, no)
        key = (no, shop_name, row["负责人"])
        item = overview[key]
        item["repeat"] += 1
        if row["冲突类型"] == TYPE_FLAT:
            item["flat"] += 1
        else:
            item["bomb_conflict"] += 1
        if row["是否低于爆旺款申报价"] == YES:
            item["low"] += 1
        elif row["是否低于爆旺款申报价"] == NO:
            item["not_low"] += 1
        if row["处理意见"] == ACT_NOW:
            item["now"] += 1
        elif row["处理意见"] == ACT_SELL_OUT:
            item["sell_out"] += 1
        elif row["处理意见"] == ACT_BOMB_STOCK:
            item["clear_30"] += 1

    rows = []
    for (no, shop, owner), item in sorted(
        overview.items(), key=lambda x: (sort_shop_no(x[0][0]), x[0][1])
    ):
        rows.append(
            {
                "店铺编号": no,
                "店铺": shop,
                "负责人": owner,
                "爆款总数": len(item["bomb"]),
                "重复铺货预计总数": item["repeat"],
                "平销冲突数": item["flat"],
                "爆款互相冲突数": item["bomb_conflict"],
                "低于爆旺款申报价数": item["low"],
                "不低于爆旺款申报价数": item["not_low"],
                "立即下架数": item["now"],
                "售完备货库存下架数": item["sell_out"],
                "30天内限时下架数": item["clear_30"],
            }
        )
    return rows


def style_sheet(ws, table_name, low_header=None, action_header=None):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    line = Side(style="hair", color="D9E2F3")
    alert_font = Font(color="C00000", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color="D9E2F3"))

    headers = [cell.value for cell in ws[1]]
    low_col = headers.index(low_header) + 1 if low_header in headers else 0
    action_col = headers.index(action_header) + 1 if action_header in headers else 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = white_fill
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=line)
        if low_col and row[low_col - 1].value == YES:
            row[low_col - 1].fill = red_fill
        if action_col and row[action_col - 1].value == ACT_NOW:
            row[action_col - 1].font = alert_font

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter][: min(ws.max_row, 250)]:
            max_len = max(max_len, len("" if cell.value is None else str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 46)

    for header in ["爆旺款申报价", "重复款申报价"]:
        if header in headers:
            col = headers.index(header) + 1
            for cells in ws.iter_cols(
                min_col=col, max_col=col, min_row=2, max_row=ws.max_row
            ):
                for cell in cells:
                    cell.number_format = "0.00"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row >= 2:
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=False, showColumnStripes=False
        )
        ws.add_table(table)


def write_workbook(overview_rows, operation_rows):
    wb = Workbook()

    overview_headers = [
        "店铺编号",
        "店铺",
        "负责人",
        "爆款总数",
        "重复铺货预计总数",
        "平销冲突数",
        "爆款互相冲突数",
        "低于爆旺款申报价数",
        "不低于爆旺款申报价数",
        "立即下架数",
        "售完备货库存下架数",
        "30天内限时下架数",
    ]
    ws = wb.active
    ws.title = "总览"
    ws.append(overview_headers)
    for row in overview_rows:
        ws.append([row.get(header) for header in overview_headers])

    operation_headers = [
        "商家编码",
        "货品名称",
        "skc",
        "所属店铺",
        "爆旺款skc",
        "爆旺款店铺",
        "爆旺款申报价",
        "重复款申报价",
        "爆旺款月销",
        "重复款月销",
        "是否低于爆旺款申报价",
        "爆旺款平台仓备货库存",
        "重复款备货库存",
        "负责人",
        "冲突类型",
        "处理意见",
    ]
    ws_op = wb.create_sheet("具体店铺操作表")
    ws_op.append(operation_headers)
    for row in operation_rows:
        ws_op.append([row.get(header) for header in operation_headers])

    ws_note = wb.create_sheet("说明")
    note_rows = [
        ["项目", "说明"],
        ["版本", VERSION],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["表格结构", "1总览；2具体店铺操作表；3说明"],
        [
            "V1.3修正",
            "爆旺款链接按货品编码判断：商家编码去掉尺码后得到货品编码，同一SKC内同一货品编码的所有尺码销量和库存合并，再比较各SKC链接的汇总销量。",
        ],
        [
            "爆旺款优先级定义",
            "同一货品编码在多个SKC/店铺出现时，30天销量汇总越高优先级越高；汇总销量相同则申报价越高优先级越高；再相同则平台仓备货库存汇总越多优先级越高。",
        ],
        [
            "处理意见规则1",
            "申报价低于爆旺款的平销款，不管有没有备货库存，返回“立即下架！”。",
        ],
        [
            "处理意见规则2",
            "申报价高于或等于爆旺款的平销款，有备货库存返回“售完备货库存下架。”；没有备货库存返回“立即下架！”。",
        ],
        [
            "处理意见规则3",
            "同为爆旺款但优先级低的，有备货库存返回“禁止备货，30天内限时下架，清理库存。”；没有备货库存返回“立即下架！”。",
        ],
        [
            "店铺编号",
            "1=弟弟/一弟，2=二弟，3=三弟，4=四弟，5=五弟，6=六弟，7=七弟，8=八弟，9=九弟，10=十弟，11=十一弟，12=十二弟，13=十三弟，15=十五弟。",
        ],
        ["具体店铺操作表行数", len(operation_rows)],
    ]
    for row in note_rows:
        ws_note.append(row)

    style_sheet(ws, "OverviewTable")
    style_sheet(ws_op, "OperationTable", "是否低于爆旺款申报价", "处理意见")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws_note[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws_note.column_dimensions["A"].width = 30
    ws_note.column_dimensions["B"].width = 140
    ws_note.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)


def main():
    owners = load_owners()
    erp = load_erp()
    hot_rows = read_hot_rows()
    hot_groups = aggregate_groups(hot_rows)
    champions = choose_champions(hot_groups)
    sales_rows = read_sales_rows(set(champions))
    sales_groups = aggregate_groups(sales_rows)
    operation_rows = build_rows(owners, erp, hot_groups, sales_groups, champions)
    overview_rows = build_overview(operation_rows, hot_groups, owners)
    write_workbook(overview_rows, operation_rows)

    print(f"OUTPUT={OUTPUT_FILE}")
    print(f"HOT_ROWS={len(hot_rows)}")
    print(f"HOT_PRODUCT_CODES={len(champions)}")
    print(f"HOT_GROUPS={len(hot_groups)}")
    print(f"SALES_ROWS={len(sales_rows)}")
    print(f"SALES_GROUPS={len(sales_groups)}")
    print(f"OPERATION_ROWS={len(operation_rows)}")
    print(
        "LOW_ROWS="
        + str(sum(1 for row in operation_rows if row["是否低于爆旺款申报价"] == YES))
    )
    print(
        "BOMB_CONFLICT_ROWS="
        + str(sum(1 for row in operation_rows if row["冲突类型"] == TYPE_BOMB))
    )


if __name__ == "__main__":
    main()
